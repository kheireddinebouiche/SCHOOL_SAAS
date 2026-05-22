import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import math
from decimal import Decimal
import psutil
import platform
import time
import glob
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.db import connection
from django.db.models.deletion import ProtectedError
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.core.mail import send_mail, EmailMessage
from django.views.decorators.http import require_POST
from app.models import Institut
from django_tenants.utils import tenant_context, schema_context
from t_formations.models import Formation, Specialites, DossierInscription
from t_formations.sync_utils import sync_formation_to_tenant
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth import get_user_model
from .models import DatabaseBackup, SaaSEmailConfiguration, SaaSMaintenanceConfiguration, SaaSGlobalConfiguration
from .utils_backup import perform_backup, perform_restore
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages



def superadmin_only(user):
    return user.is_authenticated and user.is_superuser

def get_schema_size_bytes(schema_name):
    """Calcule la taille du schéma en bytes (bases de données PostgreSQL)."""
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT SUM(pg_total_relation_size(quote_ident(schemaname) || '.' || quote_ident(tablename)))
            FROM pg_tables
            WHERE schemaname = %s;
        """, [schema_name])
        row = cursor.fetchone()
        return row[0] if (row and row[0]) else 0

def format_size(size_bytes):
    """Formate une taille de fichier en format lisible par l'homme."""
    if not size_bytes or size_bytes == 0:
        return "0 MB"
    
    # Convertir en float pour éviter les erreurs avec les types Decimal
    size_bytes = float(size_bytes)
    
    size_name = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    s = round(size_bytes / p, 2)
    return f"{s} {size_name[i]}"

def get_directory_size(directory):
    """Calcule la taille totale d'un dossier en bytes."""
    total_size = 0
    if not os.path.exists(directory):
        return total_size
    for dirpath, _, filenames in os.walk(directory):
        for f in filenames:
            fp = os.path.join(dirpath, f)
            if not os.path.islink(fp):
                try:
                    total_size += os.path.getsize(fp)
                except Exception:
                    pass
    return total_size

def get_unknown_files_count(institut):
    """Calcule le nombre de fichiers média non reliés à un prospect ou document CRM."""
    schema_name = institut.schema_name
    tenant_media_dir = os.path.join(settings.MEDIA_ROOT, schema_name)
    
    if not os.path.exists(tenant_media_dir):
        return 0
        
    # 1. Lister tous les fichiers physiques
    physical_files = set()
    for root, _, files in os.walk(tenant_media_dir):
        for f in files:
            abs_f = os.path.join(root, f)
            rel_f = os.path.relpath(abs_f, tenant_media_dir).replace('\\', '/')
            # Format attendu en base : schema_name/relative/path
            physical_files.add(f"{schema_name}/{rel_f}")
            
    # 2. Récupérer tous les fichiers référencés en base
    documented_files = set()
    try:
        with tenant_context(institut):
            from t_crm.models import DocumentsDemandeInscription, Prospets
            
            # Documents d'inscription
            docs_qs = DocumentsDemandeInscription.objects.filter(file__isnull=False).values_list('file', flat=True)
            for f in docs_qs:
                if f: documented_files.add(str(f))
                
            # Photos et logos des prospects
            prospects_qs = Prospets.objects.all().only('photo', 'logo_entreprise')
            for p in prospects_qs:
                if p.photo: documented_files.add(str(p.photo))
                if p.logo_entreprise: documented_files.add(str(p.logo_entreprise))
    except Exception:
        pass
        
    # 3. La différence donne les fichiers orphelins (Inconnue)
    unknown_files = physical_files - documented_files
    return len(unknown_files)

from django.contrib.auth import authenticate, login, logout
from django.shortcuts import redirect
from django.contrib import messages


def saas_login_view(request):
    """Vue dédiée à la connexion au SaaS Admin Dashboard."""
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect('saas_admin_app:saas_dashboard')
        
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        user = authenticate(request, username=email, password=password)
        
        if user is not None:
            if user.is_superuser:
                login(request, user)
                return redirect('saas_admin_app:saas_dashboard')
            else:
                messages.error(request, 'Accès refusé. Vous devez être un Super Administrateur.')
        else:
            messages.error(request, 'Identifiants invalides.')
            
    return render(request, 'saas_admin_app/login.html')

def saas_logout_view(request):
    """Déconnexion de l'utilisateur superadmin."""
    logout(request)
    messages.success(request, 'Vous avez été déconnecté avec succès.')
    return redirect('saas_admin_app:saas_login')

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_dashboard_view(request):
    """Affiche le tableau de bord avec les statistiques de chaque schéma/tenant."""
    instituts = Institut.objects.all().order_by('nom')
    metrics_list = []
    
    total_db_bytes = 0
    total_media_bytes = 0
    total_users = 0
    total_docs = 0

    User = get_user_model()
    
    for institut in instituts:
        schema_name = institut.schema_name
        
        # 1. Taille de la Base de Données
        db_bytes = get_schema_size_bytes(schema_name)
        total_db_bytes += db_bytes
        db_size_formatted = format_size(db_bytes)
        
        # 2. Statistiques Tenant (Users, Documents)
        try:
            with tenant_context(institut):
                user_count = User.objects.count()
                total_users += user_count
                
                # Nombre de Documents d'Inscription CRM
                try:
                    from t_crm.models import DocumentsDemandeInscription
                    doc_count = DocumentsDemandeInscription.objects.count()
                except Exception:
                    doc_count = 0
                total_docs += doc_count
        except Exception:
            user_count = 0
            doc_count = 0
            
        # 3. Fichiers Orphelins (Inconnue)
        unknown_count = get_unknown_files_count(institut)
        
        # 4. Taille des Fichiers Média (fichiers statiques uploader)
        # Assuming the standard django-tenants structure: MEDIA_ROOT / tenant_schema directories
        tenant_media_dir = os.path.join(settings.MEDIA_ROOT, schema_name)
        media_bytes = get_directory_size(tenant_media_dir)
        total_media_bytes += media_bytes
        media_size_formatted = format_size(media_bytes)
        
        # Extraire d'autres détails utiles
        date_creation = institut.date_creation if hasattr(institut, 'date_creation') else None
        
        metrics_list.append({
            'institut': institut,
            'db_size': db_size_formatted,
            'db_bytes': db_bytes,  # Garder la valeur brute pour tri/jauges
            'user_count': user_count,
            'doc_count': doc_count,
            'unknown_count': unknown_count,
            'media_size': media_size_formatted,
            'media_bytes': media_bytes,
            'date_creation': date_creation,
        })
        
    context = {
        'metrics_list': metrics_list,
        'total_db_size': format_size(total_db_bytes),
        'total_media_size': format_size(total_media_bytes),
        'total_users': total_users,
        'total_docs': total_docs,
        'nombre_instances': len(instituts),
    }
    
    from .models import SaaSMaintenanceConfiguration
    context['maintenance_config'] = SaaSMaintenanceConfiguration.get_solo()

    return render(request, 'saas_admin_app/saas_dashboard.html', context)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_tenant_data_summary_view(request):
    """Affiche une synthèse des données (Prospects, Vœux, Formations, Spécialités) pour chaque tenant."""
    instituts = Institut.objects.all().order_by('nom')
    data_summary = []
    
    total_prospects = 0
    total_voeux = 0
    total_voeux_double = 0
    total_formations = 0
    total_specialites = 0

    for inst in instituts:
        if inst.schema_name == 'public':
            continue
            
        metrics = {
            'institut': inst,
            'prospect_count': 0,
            'voeux_count': 0,
            'voeux_double_count': 0,
            'formation_count': 0,
            'specialite_count': 0,
        }
        
        try:
            with tenant_context(inst):
                # Importation dynamique pour éviter les erreurs si l'app n'est pas installée
                try:
                    from t_crm.models import Prospets, FicheDeVoeux, FicheVoeuxDouble
                    metrics['prospect_count'] = Prospets.objects.count()
                    metrics['voeux_count'] = FicheDeVoeux.objects.count()
                    metrics['voeux_double_count'] = FicheVoeuxDouble.objects.count()
                except ImportError:
                    pass
                
                try:
                    from t_formations.models import Formation, Specialites
                    metrics['formation_count'] = Formation.objects.count()
                    metrics['specialite_count'] = Specialites.objects.count()
                except ImportError:
                    pass
                    
                # Totaux globaux
                total_prospects += metrics['prospect_count']
                total_voeux += metrics['voeux_count']
                total_voeux_double += metrics['voeux_double_count']
                total_formations += metrics['formation_count']
                total_specialites += metrics['specialite_count']
                
        except Exception as e:
            print(f"Erreur lors de la récupération des données pour {inst.schema_name}: {e}")
            
        data_summary.append(metrics)
        
    context = {
        'data_summary': data_summary,
        'totals': {
            'prospects': total_prospects,
            'voeux': total_voeux,
            'voeux_double': total_voeux_double,
            'formations': total_formations,
            'specialites': total_specialites,
        }
    }
    
    return render(request, 'saas_admin_app/saas_tenant_data_summary.html', context)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_tenant_data_explorer_view(request, tenant_id):
    """Affiche les listes détaillées (Prospects, Vœux, Formations, Spécialités) pour un tenant spécifique."""
    try:
        institut = Institut.objects.get(id=tenant_id)
    except Institut.DoesNotExist:
        from django.http import Http404
        raise Http404("Tenant non trouvé")
        
    context = {
        'institut': institut,
        'prospects': [],
        'voeux': [],
        'voeux_double': [],
        'formations': [],
        'specialites': [],
        'promos': [],
        'due_paiements': [],
        'paiements': [],
        'operations_bancaire': [],
        'demande_paiements': [],
        'remboursements': [],
        'autres_produits': [],
        'stats': {},
    }
    
    from django.core.paginator import Paginator
    
    # Page numbers and active tab from GET params
    p_page_num = request.GET.get('p_page', 1)
    v_page_num = request.GET.get('v_page', 1)
    vd_page_num = request.GET.get('vd_page', 1)
    f_page_num = request.GET.get('f_page', 1)
    s_page_num = request.GET.get('s_page', 1)
    dp_page_num = request.GET.get('dp_page', 1)
    pa_page_num = request.GET.get('pa_page', 1)
    ob_page_num = request.GET.get('ob_page', 1)
    dmp_page_num = request.GET.get('dmp_page', 1)
    rem_page_num = request.GET.get('rem_page', 1)
    ap_page_num = request.GET.get('ap_page', 1)
    
    context['active_tab'] = request.GET.get('tab', 'prospects')
    
    items_per_page = 50

    try:
        with tenant_context(institut):
            # Prospects
            try:
                from t_crm.models import Prospets, DocumentsDemandeInscription
                from django.db.models import Count
                prospects_qs = Prospets.objects.all().annotate(
                    doc_count=Count('documentsdemandeinscription')
                ).order_by('-created_at')
                context['stats']['prospects'] = prospects_qs.count()
                p_paginator = Paginator(prospects_qs, items_per_page)
                context['prospects'] = p_paginator.get_page(p_page_num)
            except Exception as e:
                print(f"Error in prospects: {e}")
                context['prospects'] = []
                context['stats']['prospects'] = 0
            
            # Vœux
            try:
                from t_crm.models import FicheDeVoeux
                voeux_qs = FicheDeVoeux.objects.all().select_related('prospect', 'specialite__formation', 'promo').order_by('-created_at')
                
                # We need all voeux to detect duplicates across pages, OR we just detect on current page?
                # Actually, detection should be on the full queryset for accuracy.
                # But that might be slow if there are 10,000 voeux.
                # For now, let's detect on a larger subset or just on the current page.
                # Better: detection on full queryset but limited to recent ones.
                
                # For duplicate detection, we still need a way to know if a prospect has multiple.
                # Let's get the IDs of prospects with > 1 wish in the whole DB.
                from django.db.models import Count
                duplicate_ids_qs = FicheDeVoeux.objects.values('prospect_id').annotate(count=Count('id')).filter(count__gt=1)
                context['duplicate_voeux_prospect_ids'] = [item['prospect_id'] for item in duplicate_ids_qs if item['prospect_id']]
                
                v_paginator = Paginator(voeux_qs, items_per_page)
                context['stats']['voeux'] = voeux_qs.count()
                context['voeux'] = v_paginator.get_page(v_page_num)
            except:
                context['voeux'] = []
                context['stats']['voeux'] = 0
            
            # Vœux Doubles
            try:
                from t_crm.models import FicheVoeuxDouble
                vd_qs = FicheVoeuxDouble.objects.all().select_related('prospect', 'specialite__specialite1__formation', 'specialite__specialite2__formation', 'promo').order_by('-created_at')
                
                # Duplicate detection for doubles
                duplicate_double_ids_qs = FicheVoeuxDouble.objects.values('prospect_id').annotate(count=Count('id')).filter(count__gt=1)
                context['duplicate_double_prospect_ids'] = [item['prospect_id'] for item in duplicate_double_ids_qs if item['prospect_id']]
                
                vd_paginator = Paginator(vd_qs, items_per_page)
                context['stats']['voeux_double'] = vd_qs.count()
                context['voeux_double'] = vd_paginator.get_page(vd_page_num)
            except:
                context['voeux_double'] = []
                context['stats']['voeux_double'] = 0
            
            # Formations
            try:
                from t_formations.models import Formation
                f_qs = Formation.objects.all().order_by('nom')
                context['stats']['formations'] = f_qs.count()
                f_paginator = Paginator(f_qs, items_per_page)
                context['formations'] = f_paginator.get_page(f_page_num)
            except:
                context['formations'] = []
                context['stats']['formations'] = 0
            
            # Spécialités
            try:
                from t_formations.models import Specialites
                s_qs = Specialites.objects.all().select_related('formation').order_by('label')
                context['stats']['specialites'] = s_qs.count()
                s_paginator = Paginator(s_qs, items_per_page)
                context['specialites'] = s_paginator.get_page(s_page_num)
            except:
                context['specialites'] = []
                context['stats']['specialites'] = 0
                
            # Promos (not paginated as they are usually few)
            try:
                from t_formations.models import Promos
                context['promos'] = list(Promos.objects.all().order_by('-created_at'))
            except:
                context['promos'] = []

            # Due Paiements
            try:
                from t_tresorerie.models import DuePaiements
                dp_qs = DuePaiements.objects.all().select_related('client', 'ref_echeancier', 'promo').order_by('-date_echeance')
                context['stats']['due_paiements'] = dp_qs.count()
                dp_paginator = Paginator(dp_qs, items_per_page)
                context['due_paiements'] = dp_paginator.get_page(dp_page_num)
            except:
                context['due_paiements'] = []
                context['stats']['due_paiements'] = 0

            # Paiements
            try:
                from t_tresorerie.models import Paiements
                pa_qs = Paiements.objects.all().select_related('due_paiements', 'prospect', 'promo').order_by('-date_paiement')
                context['stats']['paiements'] = pa_qs.count()
                pa_paginator = Paginator(pa_qs, items_per_page)
                context['paiements'] = pa_paginator.get_page(pa_page_num)
            except:
                context['paiements'] = []
                context['stats']['paiements'] = 0

            # Operations Bancaire
            try:
                from t_tresorerie.models import OperationsBancaire
                ob_qs = OperationsBancaire.objects.all().order_by('-id')
                context['stats']['operations_bancaire'] = ob_qs.count()
                ob_paginator = Paginator(ob_qs, items_per_page)
                context['operations_bancaire'] = ob_paginator.get_page(ob_page_num)
            except:
                context['operations_bancaire'] = []
                context['stats']['operations_bancaire'] = 0

            # Demande de Remboursement
            try:
                from t_tresorerie.models import ClientPaiementsRequest
                dmp_qs = ClientPaiementsRequest.objects.all().order_by('-id')
                context['stats']['demande_paiements'] = dmp_qs.count()
                dmp_paginator = Paginator(dmp_qs, items_per_page)
                context['demande_paiements'] = dmp_paginator.get_page(dmp_page_num)
            except:
                context['demande_paiements'] = []
                context['stats']['demande_paiements'] = 0

            # Remboursements
            try:
                from t_tresorerie.models import Rembourssements
                rem_qs = Rembourssements.objects.all().order_by('-id')
                context['stats']['remboursements'] = rem_qs.count()
                rem_paginator = Paginator(rem_qs, items_per_page)
                context['remboursements'] = rem_paginator.get_page(rem_page_num)
            except:
                context['remboursements'] = []
                context['stats']['remboursements'] = 0

            # Autres Produits
            try:
                from t_tresorerie.models import AutreProduit
                ap_qs = AutreProduit.objects.all().order_by('-id')
                context['stats']['autres_produits'] = ap_qs.count()
                ap_paginator = Paginator(ap_qs, items_per_page)
                context['autres_produits'] = ap_paginator.get_page(ap_page_num)
            except:
                context['autres_produits'] = []
                context['stats']['autres_produits'] = 0
            
    except Exception as e:
        context['error'] = str(e)
        
    with tenant_context(institut):
        return render(request, 'saas_admin_app/saas_tenant_data_explorer.html', context)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_export_prospects_view(request, tenant_id):
    """Génère un fichier Excel contenant la liste des prospects pour un tenant spécifique."""
    institut = get_object_or_404(Institut, id=tenant_id)
    
    with tenant_context(institut):
        from t_crm.models import Prospets, DocumentsDemandeInscription
        from django.db.models import Count
        
        # Récupération de tous les prospects avec le compte de documents
        prospects = Prospets.objects.all().annotate(
            doc_count=Count('documentsdemandeinscription')
        ).order_by('-created_at')
        
        # Création du classeur Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prospects"
        
        # En-têtes
        headers = ["ID", "Nom", "Prénom", "Email", "Téléphone", "État", "Statut", "Documents", "Date de Création"]
        ws.append(headers)
        
        # Style pour les en-têtes
        header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
        # Remplissage des données
        for p in prospects:
            # Nettoyage de la date pour Excel (naïve)
            created_at_naive = p.created_at.replace(tzinfo=None) if p.created_at else ""
            
            ws.append([
                p.id,
                p.nom,
                p.prenom,
                p.email,
                p.telephone,
                p.get_etat_display() if hasattr(p, 'get_etat_display') else p.etat,
                p.get_statut_display() if hasattr(p, 'get_statut_display') else p.statut,
                p.doc_count,
                created_at_naive
            ])
            
        # Ajustement automatique de la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
        # Préparation de la réponse HTTP
        filename = f"prospects_{institut.schema_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        return response

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_update_voeu_action_view(request, tenant_id, voeu_id):
    """Permet de modifier la spécialité ou supprimer une fiche de voeux (standard ou double)."""
    from django.http import JsonResponse
    from app.models import Institut
    
    institut = get_object_or_404(Institut, id=tenant_id)
    action = request.POST.get('action') # 'delete' ou 'update_specialite'
    voeu_type = request.POST.get('voeu_type') # 'standard' ou 'double'
    
    try:
        with tenant_context(institut):
            from t_crm.models import FicheDeVoeux, FicheVoeuxDouble
            from t_formations.models import Specialites
            
            model = FicheDeVoeux if voeu_type == 'standard' else FicheVoeuxDouble
            voeu = get_object_or_404(model, id=voeu_id)
            
            if action == 'delete':
                voeu.delete()
                return JsonResponse({'status': 'success', 'message': 'Fiche de vœux supprimée avec succès.'})
            
            elif action == 'update_specialite':
                new_specialite_id = request.POST.get('specialite_id')
                new_promo_id = request.POST.get('promo_id')
                is_confirmed = request.POST.get('is_confirmed') == 'true'
                
                new_specialite = get_object_or_404(Specialites, id=new_specialite_id)
                voeu.specialite = new_specialite
                voeu.is_confirmed = is_confirmed
                
                if new_promo_id:
                    from t_formations.models import Promos
                    new_promo = get_object_or_404(Promos, id=new_promo_id)
                    voeu.promo = new_promo
                else:
                    voeu.promo = None
                
                voeu.save()
                return JsonResponse({
                    'status': 'success', 
                    'message': 'Fiche de vœux mise à jour avec succès.',
                    'new_label': new_specialite.label,
                    'new_formation': new_specialite.formation.nom,
                    'new_promo': voeu.promo.label if voeu.promo else 'Standard',
                    'new_is_confirmed': voeu.is_confirmed
                })
                
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Action non reconnue.'}, status=400)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_update_due_paiement_action_view(request, tenant_id, dp_id):
    """Permet de modifier ou supprimer un montant dû."""
    from django.http import JsonResponse
    from app.models import Institut
    
    institut = get_object_or_404(Institut, id=tenant_id)
    action = request.POST.get('action') # 'delete' ou 'update'
    
    try:
        with tenant_context(institut):
            from t_tresorerie.models import DuePaiements
            dp = get_object_or_404(DuePaiements, id=dp_id)
            
            if action == 'delete':
                client = dp.client
                dp.delete()
                
                # Check if any other due payments exist for this client
                if not DuePaiements.objects.filter(client=client).exists():
                    from t_tresorerie.models import EcheancierSpecial, ClientPaiementsRequest
                    EcheancierSpecial.objects.filter(prospect=client).update(is_validate=False, is_approuved=False)
                    ClientPaiementsRequest.objects.filter(client=client).delete()
                    
                    # Reset special echeancier flag on prospect
                    if client:
                        client.has_special_echeancier = False
                        client.save()
                    
                return JsonResponse({'status': 'success', 'message': 'Montant dû supprimé avec succès. L\'échéancier est marqué comme non confirmé si c\'était le dernier.'})
            
            elif action == 'update':
                dp.montant_due = request.POST.get('montant_due')
                dp.montant_restant = request.POST.get('montant_restant')
                dp.date_echeance = request.POST.get('date_echeance')
                dp.label = request.POST.get('label')
                dp.save()
                return JsonResponse({'status': 'success', 'message': 'Montant dû mis à jour.'})
                
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Action non reconnue.'}, status=400)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_update_paiement_action_view(request, tenant_id, p_id):
    """Permet de modifier ou supprimer un paiement."""
    from django.http import JsonResponse
    from app.models import Institut
    
    institut = get_object_or_404(Institut, id=tenant_id)
    action = request.POST.get('action') # 'delete' ou 'update'
    
    try:
        with tenant_context(institut):
            from t_tresorerie.models import Paiements
            p = get_object_or_404(Paiements, id=p_id)
            
            if action == 'delete':
                # Restaurer le montant restant dans le DuePaiements associé
                if p.due_paiements:
                    dp = p.due_paiements
                    if dp.montant_restant is not None:
                        dp.montant_restant += Decimal(str(p.montant_paye or 0))
                    else:
                        dp.montant_restant = Decimal(str(p.montant_paye or 0))
                    
                    # Si le montant restant est > 0, l'échéance n'est plus payée
                    if dp.montant_restant > 0:
                        dp.is_done = False
                    
                    dp.save()

                p.delete()
                return JsonResponse({'status': 'success', 'message': 'Paiement supprimé et montant dû rétabli.'})
            
            elif action == 'update':
                p.montant_paye = request.POST.get('montant_paye')
                p.date_paiement = request.POST.get('date_paiement')
                p.observation = request.POST.get('observation')
                p.save()
                return JsonResponse({'status': 'success', 'message': 'Paiement mis à jour.'})
                
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Action non reconnue.'}, status=400)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_bulk_delete_action_view(request, tenant_id):
    """Permet de supprimer plusieurs enregistrements (DuePaiements ou Paiements) en une seule fois."""
    from django.http import JsonResponse
    from app.models import Institut
    from decimal import Decimal
    import json
    
    institut = get_object_or_404(Institut, id=tenant_id)
    model_name = request.POST.get('model') # 'due_paiement' ou 'paiement'
    ids_json = request.POST.get('ids', '[]')
    
    try:
        ids = json.loads(ids_json)
        if not ids:
            return JsonResponse({'status': 'error', 'message': 'Aucun élément sélectionné.'}, status=400)
            
        with tenant_context(institut):
            from t_tresorerie.models import DuePaiements, Paiements, EcheancierSpecial
            from t_crm.models import FicheDeVoeux, FicheVoeuxDouble
            
            deleted_count = 0
            affected_prospects = set()
            
            if model_name == 'due_paiement':
                for dp_id in ids:
                    try:
                        dp = DuePaiements.objects.get(id=dp_id)
                        affected_prospects.add(dp.client)
                        dp.delete()
                        deleted_count += 1
                    except DuePaiements.DoesNotExist:
                        continue
                
                # Update financial status for affected prospects
                for prospect in affected_prospects:
                    if prospect and not DuePaiements.objects.filter(client=prospect).exists():
                        from t_tresorerie.models import ClientPaiementsRequest
                        EcheancierSpecial.objects.filter(prospect=prospect).update(is_validate=False, is_approuved=False)
                        ClientPaiementsRequest.objects.filter(client=prospect).delete()
                        
                        # Reset special echeancier flag on prospect
                        prospect.has_special_echeancier = False
                        prospect.save()
                        
            elif model_name == 'paiement':
                for p_id in ids:
                    try:
                        p = Paiements.objects.get(id=p_id)
                        if p.due_paiements:
                            dp = p.due_paiements
                            amount_to_restore = Decimal(str(p.montant_paye or 0))
                            if dp.montant_restant is not None:
                                dp.montant_restant += amount_to_restore
                            else:
                                dp.montant_restant = amount_to_restore
                            
                            if dp.montant_restant > 0:
                                dp.is_done = False
                            dp.save()
                        
                        p.delete()
                        deleted_count += 1
                    except Paiements.DoesNotExist:
                        continue
                        
            elif model_name == 'operation_bancaire':
                from t_tresorerie.models import OperationsBancaire
                for ob_id in ids:
                    try:
                        ob = OperationsBancaire.objects.get(id=ob_id)
                        ob.delete()
                        deleted_count += 1
                    except OperationsBancaire.DoesNotExist:
                        continue
                        
            elif model_name == 'demande_paiement':
                from t_tresorerie.models import ClientPaiementsRequest
                for dmp_id in ids:
                    try:
                        dmp = ClientPaiementsRequest.objects.get(id=dmp_id)
                        dmp.delete()
                        deleted_count += 1
                    except ClientPaiementsRequest.DoesNotExist:
                        continue
                        
            elif model_name == 'remboursement':
                from t_tresorerie.models import Rembourssements
                for rem_id in ids:
                    try:
                        rem = Rembourssements.objects.get(id=rem_id)
                        rem.delete()
                        deleted_count += 1
                    except Rembourssements.DoesNotExist:
                        continue
                        
            elif model_name == 'autre_produit':
                from t_tresorerie.models import AutreProduit
                for ap_id in ids:
                    try:
                        ap = AutreProduit.objects.get(id=ap_id)
                        ap.delete()
                        deleted_count += 1
                    except AutreProduit.DoesNotExist:
                        continue
            
            return JsonResponse({
                'status': 'success', 
                'message': f'{deleted_count} éléments supprimés avec succès.'
            })
            
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_update_prospect_action_view(request, tenant_id, prospect_id):
    """Permet de modifier l'état et le statut d'un prospect."""
    from django.http import JsonResponse
    from app.models import Institut
    
    institut = get_object_or_404(Institut, id=tenant_id)
    action = request.POST.get('action')
    new_etat = request.POST.get('etat')
    new_statut = request.POST.get('statut')
    
    try:
        with tenant_context(institut):
            from t_crm.models import Prospets
            prospect = get_object_or_404(Prospets, id=prospect_id)
            
            if action == 'delete':
                prospect.delete()
                return JsonResponse({'status': 'success', 'message': 'Prospect supprimé avec succès.'})

            if new_etat:
                prospect.etat = new_etat
            if new_statut:
                prospect.statut = new_statut
                
            prospect.save()
            
            return JsonResponse({
                'status': 'success', 
                'message': 'Prospect mis à jour avec succès.',
                'new_etat_display': prospect.get_etat_display(),
                'new_statut_display': prospect.get_statut_display()
            })
                
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_reset_prospect_action_view(request, tenant_id, prospect_id):
    """
    Réinitialise complètement l'état financier et le statut d'un prospect.
    Supprime les montants dus, les paiements, la demande de paiement,
    et repasse le prospect en statut 'prinscrit'.
    """
    from django.http import JsonResponse
    from app.models import Institut
    from django.shortcuts import get_object_or_404
    
    institut = get_object_or_404(Institut, id=tenant_id)
    
    try:
        with tenant_context(institut):
            from t_crm.models import Prospets
            from t_tresorerie.models import DuePaiements, Paiements, ClientPaiementsRequest, clientPaiementsRequestLine, EcheancierSpecial
            
            prospect = get_object_or_404(Prospets, id=prospect_id)
            
            # 1. Supprimer les paiements associés
            Paiements.objects.filter(prospect=prospect).delete()
            
            # 2. Supprimer les montants dus
            DuePaiements.objects.filter(client=prospect).delete()
            
            # 3. Supprimer la demande de paiement et ses lignes
            cpr_qs = ClientPaiementsRequest.objects.filter(client=prospect)
            for cpr in cpr_qs:
                clientPaiementsRequestLine.objects.filter(paiement_request=cpr).delete()
            cpr_qs.delete()
            
            # 4. Réinitialiser l'échéancier spécial
            EcheancierSpecial.objects.filter(prospect=prospect).update(is_validate=False, is_approuved=False)
            
            # 5. Réinitialiser le prospect
            prospect.statut = 'prinscrit'
            prospect.has_special_echeancier = False
            prospect.instance_date = None
            prospect.convertit_date = None
            prospect.save()
            
            return JsonResponse({
                'status': 'success', 
                'message': f'Le prospect {prospect.nom} {prospect.prenom} a été réinitialisé avec succès.'
            })
                
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_create_tenant_view(request):
    """Crée un nouveau tenant (Institut), son domaine et son superutilisateur."""
    if request.method == 'POST':
        nom = request.POST.get('nom')
        schema_name = request.POST.get('schema_name')
        domain_name = request.POST.get('domain_name')
        admin_username = request.POST.get('admin_username')
        admin_email = request.POST.get('admin_email')
        admin_password = request.POST.get('admin_password')
        tenant_type = request.POST.get('tenant_type', 'second')
        
        if not all([nom, schema_name, domain_name, admin_username, admin_email, admin_password]):
            messages.error(request, "Tous les champs sont obligatoires.")
            return redirect('saas_admin_app:saas_dashboard')
            
        try:
            # 1. Création du Tenant (Institut)
            # schema_name doit être unique et conforme aux règles PostgreSQL
            from app.models import Institut, Domaine
            
            if Institut.objects.filter(schema_name=schema_name).exists():
                messages.error(request, f"Le schéma '{schema_name}' existe déjà.")
                return redirect('saas_admin_app:saas_dashboard')
            
            tenant = Institut(
                nom=nom,
                schema_name=schema_name,
                tenant_type=tenant_type
            )
            tenant.save() # Déclenche la création du schéma via django-tenants
            
            # 2. Création du Domaine
            domain = Domaine(
                domain=domain_name,
                tenant=tenant,
                is_primary=True
            )
            domain.save()
            
            # 3. Création du Super-utilisateur dans le contexte du tenant
            with tenant_context(tenant):
                User = get_user_model()
                user = User.objects.create_superuser(
                    username=admin_username,
                    email=admin_email,
                    password=admin_password
                )
                
                # Initialisation du profil si nécessaire
                try:
                    from institut_app.models import Profile
                    from django.utils import timezone
                    profile, created = Profile.objects.get_or_create(user=user)
                    profile.role = 'admin'
                    profile.last_password_change = timezone.now()
                    profile.save()
                except Exception as e:
                    # On log l'erreur mais on ne bloque pas si le profil échoue (cas rare)
                    print(f"Erreur création profil: {e}")
            
            messages.success(request, f"L'institution '{nom}' a été créée avec succès. Accès : {domain_name}")
            return redirect('saas_admin_app:saas_dashboard')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la création : {str(e)}")
            return redirect('saas_admin_app:saas_dashboard')
            
    return redirect('saas_admin_app:saas_dashboard')

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_tenants_manage_view(request):
    """Page dédiée à la gestion complète des tenants et de leurs domaines."""
    instituts = Institut.objects.all().order_by('-date_creation')
    
    metrics_list = []
    for inst in instituts:
        # Get metrics for each tenant
        user_count = 0
        db_size = "0 KB"
        doc_count = 0
        media_size = "0 KB"
        unknown_count = 0
        
        try:
            with tenant_context(inst):
                user_count = get_user_model().objects.count()
                db_size = format_size(get_schema_size_bytes(inst.schema_name))
                
                # Documents CRM
                try:
                    from t_crm.models import CRMDocument
                    doc_count = CRMDocument.objects.count()
                except: pass
                
                # Media size & Orphans
                media_size_bytes = get_directory_size(os.path.join(settings.MEDIA_ROOT, inst.schema_name))
                unknown_count = get_unknown_files_count(inst)
                media_size = format_size(media_size_bytes)
        except Exception as e:
            print(f"Error metrics for {inst.schema_name}: {e}")

        # Get domain info
        primary_domain = inst.domains.filter(is_primary=True).first()
        other_domains = inst.domains.filter(is_primary=False)

        metrics_list.append({
            'institut': inst,
            'primary_domain': primary_domain,
            'other_domains': other_domains,
            'user_count': user_count,
            'db_size': db_size,
            'doc_count': doc_count,
            'media_size': media_size,
            'unknown_count': unknown_count,
            'date_creation': inst.date_creation
        })

    context = {
        'metrics_list': metrics_list,
        'total_tenants': instituts.count(),
    }
    return render(request, 'saas_admin_app/saas_tenants_manage.html', context)


from django.utils.decorators import method_decorator
from functools import wraps
from django.shortcuts import redirect
from django.contrib import messages

def system_access_required(view_func):
    """Décorateur pour vérifier que l'accès système est déverrouillé par le PIN."""
    @wraps(view_func)
    @user_passes_test(superadmin_only, login_url='/saas-admin/login/')
    def _wrapped_view(request, *args, **kwargs):
        if not request.session.get('saas_system_unlocked', False):
            # Stocker l'URL d'origine pour redirection après déverrouillage
            request.session['next_system_url'] = request.path
            return redirect('saas_admin_app:saas_system_lock')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_system_lock_view(request):
    """Vue pour saisir le code PIN de sécurité SaaS."""
    next_url = request.session.get('next_system_url', reverse('saas_admin_app:saas_dashboard'))
    
    if request.method == 'POST':
        pin = request.POST.get('pin', '').strip()
        # Nettoyage profond du PIN configuré (enlève espaces et guillemets éventuels)
        configured_pin = str(settings.SAAS_SYSTEM_PIN).strip().strip("'").strip('"')
        
        if pin == configured_pin:
            request.session['saas_system_unlocked'] = True
            messages.success(request, "Accès système déverrouillé.")
            return redirect(next_url)
        else:
            messages.error(request, f"Code PIN incorrect.")
            
    return render(request, 'saas_admin_app/saas_system_lock.html', {'next_url': next_url})

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_system_logout_view(request):
    """Verrouille à nouveau l'accès aux outils système."""
    request.session['saas_system_unlocked'] = False
    messages.info(request, "Accès système verrouillé.")
    return redirect('saas_admin_app:saas_dashboard')

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_performance_view(request):
    """Affiche les performances système du serveur."""
    
    # CPU Metrics
    cpu_percent = psutil.cpu_percent(interval=1)
    cpu_count = psutil.cpu_count()
    cpu_freq = psutil.cpu_freq()
    cpu_freq_current = cpu_freq.current if cpu_freq else 0
    cpu_freq_max = cpu_freq.max if cpu_freq else 0
    
    # Memory Metrics
    memory = psutil.virtual_memory()
    memory_total = format_size(memory.total)
    memory_used = format_size(memory.used)
    memory_available = format_size(memory.available)
    memory_percent = memory.percent
    
    # Disk Metrics
    disk = psutil.disk_usage('/')
    disk_total = format_size(disk.total)
    disk_used = format_size(disk.used)
    disk_free = format_size(disk.free)
    disk_percent = disk.percent
    
    # Network Metrics
    net_io = psutil.net_io_counters()
    net_bytes_sent = format_size(net_io.bytes_sent)
    net_bytes_recv = format_size(net_io.bytes_recv)
    net_packets_sent = net_io.packets_sent
    net_packets_recv = net_io.packets_recv
    
    # System Info
    system_info = {
        'os': f"{platform.system()} {platform.release()}",
        'architecture': platform.architecture()[0],
        'processor': platform.processor() or 'N/A',
        'hostname': platform.node(),
        'python_version': platform.python_version(),
        'uptime_seconds': int(time.time() - psutil.boot_time()),
    }
    
    # Process Metrics (Django)
    current_process = psutil.Process()
    process_memory = format_size(current_process.memory_info().rss)
    process_cpu = current_process.cpu_percent(interval=0.1)
    process_threads = current_process.num_threads()
    
    # Database Connections
    db_connections = 0
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT count(*) 
                FROM pg_stat_activity 
                WHERE datname = current_database();
            """)
            row = cursor.fetchone()
            db_connections = row[0] if row else 0
    except Exception:
        db_connections = 0
    
    # Tenant Performance Data
    instituts = Institut.objects.all().order_by('nom')
    tenant_metrics = []
    
    for institut in instituts:
        schema_name = institut.schema_name
        
        # Database size
        db_bytes = get_schema_size_bytes(schema_name)
        db_size_formatted = format_size(db_bytes)
        
        # User count
        try:
            with tenant_context(institut):
                User = get_user_model()
                user_count = User.objects.count()
        except Exception:
            user_count = 0
        
        # Active sessions (approximation)
        active_sessions = 0
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT count(*) 
                    FROM pg_stat_activity 
                    WHERE datname = current_database() 
                    AND state = 'active';
                """)
                row = cursor.fetchone()
                active_sessions = row[0] if row else 0
        except Exception:
            active_sessions = 0
        
        tenant_metrics.append({
            'institut': institut,
            'db_size': db_size_formatted,
            'db_bytes': db_bytes,
            'user_count': user_count,
            'active_sessions': active_sessions,
        })
    
    # Sort tenants by DB size (descending)
    tenant_metrics.sort(key=lambda x: x['db_bytes'], reverse=True)
    
    # Calculate uptime
    uptime = system_info['uptime_seconds']
    uptime_days = uptime // 86400
    uptime_hours = (uptime % 86400) // 3600
    uptime_minutes = (uptime % 3600) // 60
    uptime_str = f"{uptime_days}j {uptime_hours}h {uptime_minutes}m"
    
    context = {
        'cpu_percent': cpu_percent,
        'cpu_count': cpu_count,
        'cpu_freq_current': cpu_freq_current,
        'cpu_freq_max': cpu_freq_max,
        'memory_total': memory_total,
        'memory_used': memory_used,
        'memory_available': memory_available,
        'memory_percent': memory_percent,
        'disk_total': disk_total,
        'disk_used': disk_used,
        'disk_free': disk_free,
        'disk_percent': disk_percent,
        'net_bytes_sent': net_bytes_sent,
        'net_bytes_recv': net_bytes_recv,
        'net_packets_sent': net_packets_sent,
        'net_packets_recv': net_packets_recv,
        'system_info': system_info,
        'uptime_str': uptime_str,
        'process_memory': process_memory,
        'process_cpu': process_cpu,
        'process_threads': process_threads,
        'db_connections': db_connections,
        'tenant_metrics': tenant_metrics,
        'timestamp': datetime.now().isoformat(),
    }
    
    return render(request, 'saas_admin_app/saas_performance.html', context)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_logs_view(request):
    """Affiche les logs système de l'application."""
    
    # Get log file paths
    log_files = []
    log_entries = []
    
    # Common log file locations
    possible_log_paths = [
        os.path.join(settings.BASE_DIR, '*.log'),
        os.path.join(settings.BASE_DIR, 'logs', '*.log'),
        os.path.join(settings.BASE_DIR, 'logs', '**', '*.log'),
    ]
    
    # Add VPS System logs for Linux
    if platform.system() == 'Linux':
        linux_logs = [
            '/var/log/syslog',
            '/var/log/auth.log',
            '/var/log/nginx/access.log',
            '/var/log/nginx/error.log',
            '/var/log/postgresql/postgresql-*.log',
            '/var/log/dpkg.log',
        ]
        possible_log_paths.extend(linux_logs)
    
    # Find all log files
    for pattern in possible_log_paths:
        log_files.extend(glob.glob(pattern, recursive=True))
    
    # Also check for Django log files
    if hasattr(settings, 'LOGGING'):
        log_config = settings.LOGGING
        if 'handlers' in log_config:
            for handler_name, handler_config in log_config['handlers'].items():
                if 'filename' in handler_config:
                    log_file = handler_config['filename']
                    if os.path.exists(log_file) and log_file not in log_files:
                        log_files.append(log_file)
    
    # Get the selected log file (default to first one)
    selected_log_file = request.GET.get('file', log_files[0] if log_files else None)
    log_level_filter = request.GET.get('level', 'ALL')
    search_query = request.GET.get('search', '')
    
    # Read log entries from selected file
    if selected_log_file and os.path.exists(selected_log_file):
        try:
            with open(selected_log_file, 'r', encoding='utf-8', errors='ignore') as f:
                raw_lines = f.readlines()
            
            # Parse log lines (typical Django log format)
            for line in raw_lines:
                line = line.strip()
                if not line:
                    continue
                
                # Determine log level
                level = 'INFO'
                if 'ERROR' in line.upper():
                    level = 'ERROR'
                elif 'WARNING' in line.upper() or 'WARN' in line.upper():
                    level = 'WARNING'
                elif 'DEBUG' in line.upper():
                    level = 'DEBUG'
                elif 'CRITICAL' in line.upper() or 'FATAL' in line.upper():
                    level = 'CRITICAL'
                
                # Apply level filter
                if log_level_filter != 'ALL' and level != log_level_filter:
                    continue
                
                # Apply search filter
                if search_query and search_query.lower() not in line.lower():
                    continue
                
                log_entries.append({
                    'line': line,
                    'level': level,
                    'timestamp': extract_timestamp(line),
                })
        except PermissionError:
            log_entries.append({
                'line': f"ACCÈS REFUSÉ : Le serveur web n'a pas les permissions pour lire ce fichier de log ({os.path.basename(selected_log_file)}).",
                'level': 'CRITICAL',
                'timestamp': None,
            })
        except Exception as e:
            log_entries.append({
                'line': f"Error reading log file: {str(e)}",
                'level': 'ERROR',
                'timestamp': None,
            })
    
    # Get log file statistics
    log_file_stats = []
    for log_file in log_files:
        try:
            file_size = os.path.getsize(log_file)
            file_mod_time = datetime.fromtimestamp(os.path.getmtime(log_file))
            log_file_stats.append({
                'path': log_file,
                'name': os.path.basename(log_file),
                'size': format_size(file_size),
                'modified': file_mod_time,
                'is_selected': log_file == selected_log_file,
            })
        except Exception:
            pass
    
    # Sort log files by modification time (most recent first)
    log_file_stats.sort(key=lambda x: x['modified'], reverse=True)
    
    # Count log levels
    level_counts = {'ERROR': 0, 'WARNING': 0, 'INFO': 0, 'DEBUG': 0, 'CRITICAL': 0}
    for entry in log_entries:
        if entry['level'] in level_counts:
            level_counts[entry['level']] += 1
    
    context = {
        'log_files': log_file_stats,
        'selected_file': selected_log_file,
        'log_entries': log_entries[-500:],  # Limit to last 500 entries for performance
        'level_counts': level_counts,
        'total_entries': len(log_entries),
        'current_level_filter': log_level_filter,
        'search_query': search_query,
    }
    
    return render(request, 'saas_admin_app/saas_logs.html', context)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_processes_view(request):
    """Vue pour afficher les processus système et les services."""
    import subprocess
    import json
    
    # 1. Processus
    processes = []
    # On limite pour éviter de surcharger la page (top CPU/Memory)
    for proc in psutil.process_iter(['pid', 'name', 'username', 'status', 'cpu_percent', 'memory_percent', 'create_time']):
        try:
            pinfo = proc.info
            # Convert timestamp to datetime
            pinfo['uptime'] = datetime.fromtimestamp(pinfo['create_time']).strftime("%Y-%m-%d %H:%M:%S")
            processes.append(pinfo)
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass
            
    # Tri par CPU ou Mémoire (ici par CPU descendant)
    processes.sort(key=lambda x: x['cpu_percent'], reverse=True)
    
    # 2. Services
    services = []
    system_os = platform.system()
    
    if system_os == 'Windows':
        # Utilisation de PowerShell pour obtenir les services
        try:
            cmd = 'powershell -Command "Get-Service | Select-Object Name, DisplayName, Status | ConvertTo-Json"'
            result = subprocess.run(cmd, capture_output=True, shell=True)
            if result.returncode == 0:
                # On décode avec errors='ignore' pour éviter les plantages sur Windows (encoding cp1252 vs powershell output)
                stdout_str = result.stdout.decode('utf-8', errors='ignore')
                raw_services = json.loads(stdout_str)
                # S'il n'y a qu'un service, JSON.loads retourne un dict au lieu d'une liste
                if isinstance(raw_services, dict):
                    raw_services = [raw_services]
                
                for s in raw_services:
                    status_map = {0: 'Stopped', 1: 'StartPending', 2: 'StopPending', 3: 'Running', 4: 'ContinuePending', 5: 'PausePending', 6: 'Paused'}
                    # PowerShell retourne parfois des codes au lieu de strings selon la version
                    status = s.get('Status')
                    if isinstance(status, int):
                        status = status_map.get(status, f"Unknown ({status})")
                    
                    services.append({
                        'name': s.get('Name'),
                        'display_name': s.get('DisplayName'),
                        'status': status,
                        'type': 'Windows Service'
                    })
        except Exception as e:
            services.append({'name': 'Error', 'display_name': f"Could not fetch Windows services: {str(e)}", 'status': 'Error', 'type': 'System'})

    elif system_os == 'Linux':
        # Utilisation de systemctl
        try:
            # On tente de lister les services
            cmd = ['systemctl', 'list-units', '--type=service', '--all', '--no-legend', '--no-pager']
            result = subprocess.run(cmd, capture_output=True, text=True)
            if result.returncode == 0:
                for line in result.stdout.splitlines():
                    parts = line.split(None, 4)
                    if len(parts) >= 4:
                        services.append({
                            'name': parts[0],
                            'display_name': parts[4] if len(parts) > 4 else parts[0],
                            'status': parts[3], # active/inactive/failed
                            'sub_status': parts[2], # running/exited/dead
                            'type': 'Systemd Service'
                        })
        except Exception as e:
            services.append({'name': 'Error', 'display_name': f"Could not fetch Linux services: {str(e)}", 'status': 'Error', 'type': 'System'})

    context = {
        'processes': processes[:100],  # Top 100 processus
        'services': services,
        'system_os': system_os,
        'timestamp': datetime.now().strftime("%H:%M:%S")
    }
    
    return render(request, 'saas_admin_app/saas_processes.html', context)

@system_access_required
def saas_terminal_view(request):
    """Affiche la page du terminal système."""
    # Initialiser le répertoire de travail dans la session s'il n'existe pas ou sur demande
    if 'terminal_cwd' not in request.session or request.GET.get('reset') == '1':
        request.session['terminal_cwd'] = str(settings.BASE_DIR)
        
    context = {
        'current_dir': request.session['terminal_cwd'],
        'system_os': platform.system(),
    }
    return render(request, 'saas_admin_app/saas_terminal.html', context)

@system_access_required
def saas_terminal_exec_view(request):
    """API endpoint pour exécuter des commandes système avec suivi du répertoire."""
    import subprocess
    import os
    from django.http import JsonResponse
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    command = request.POST.get('command')
    if not command:
        return JsonResponse({'success': False, 'error': 'Commande vide'})
    
    # Récupérer le répertoire actuel depuis la session
    cwd = request.session.get('terminal_cwd', str(settings.BASE_DIR))
    if not os.path.exists(cwd):
        cwd = str(settings.BASE_DIR)
        request.session['terminal_cwd'] = cwd

    system_os = platform.system()
    
    try:
        # Pour supporter le changement de répertoire (cd), on exécute la commande 
        # puis on demande le nouveau chemin (pwd).
        # On utilise un marqueur spécial pour isoler le nouveau CWD.
        cwd_marker = "---SAAS_CWD_MARKER---"
        
        if system_os == 'Windows':
            # Sur Windows (cmd/powershell), cd et pwd fonctionnent différemment
            full_command = f"{command} & echo {cwd_marker} & cd"
        else:
            # Sur Linux (bash/sh), on utilise subshell et pwd
            full_command = f"({command}); echo {cwd_marker}; pwd"
        
        result = subprocess.run(
            full_command,
            shell=True,
            capture_output=True,
            text=True,
            cwd=cwd,
            timeout=60,
            encoding='utf-8',
            errors='replace'
        )
        
        stdout = result.stdout
        new_cwd = cwd
        
        # Extraire le nouveau CWD depuis la sortie
        if cwd_marker in stdout:
            parts = stdout.rsplit(cwd_marker, 1)
            stdout = parts[0].strip()
            potential_new_cwd = parts[1].strip()
            if os.path.exists(potential_new_cwd) and os.path.isdir(potential_new_cwd):
                new_cwd = potential_new_cwd
                request.session['terminal_cwd'] = new_cwd
        
        return JsonResponse({
            'success': True,
            'stdout': stdout,
            'stderr': result.stderr,
            'returncode': result.returncode,
            'current_dir': new_cwd
        })
        
    except subprocess.TimeoutExpired:
        return JsonResponse({'success': False, 'error': 'Délai d\'exécution dépassé (60s)'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@system_access_required
def saas_explorer_view(request):
    """Affiche la page principale de l'explorateur de fichiers système."""
    initial_path = request.GET.get('path', str(settings.BASE_DIR))
    context = {
        'initial_path': initial_path,
        'system_os': platform.system(),
    }
    return render(request, 'saas_admin_app/saas_explorer.html', context)

@system_access_required
def api_explorer_browse(request):
    """API pour lister le contenu d'un répertoire système quelconque."""
    import os
    from django.http import JsonResponse
    
    path = request.GET.get('path', str(settings.BASE_DIR))
    
    # Normalisation pour Windows/Linux
    path = os.path.normpath(path)
    
    if not os.path.exists(path):
        return JsonResponse({'success': False, 'error': 'Chemin inexistant'}, status=404)
    
    if not os.path.isdir(path):
        return JsonResponse({'success': False, 'error': 'N\'est pas un répertoire'}, status=400)
    
    items = []
    try:
        for entry in os.scandir(path):
            try:
                info = entry.stat()
                items.append({
                    'name': entry.name,
                    'is_dir': entry.is_dir(),
                    'size': info.st_size,
                    'size_formatted': format_size(info.st_size) if entry.is_file() else '--',
                    'modified': datetime.fromtimestamp(info.st_mtime).strftime('%d/%m/%Y %H:%M'),
                    'permissions': oct(info.st_mode & 0o777),
                })
            except Exception:
                continue
                
        # Tri : dossiers d'abord, puis fichiers
        items.sort(key=lambda x: (not x['is_dir'], x['name'].lower()))
        
        return JsonResponse({
            'success': True,
            'current_path': path,
            'parent_path': os.path.dirname(path),
            'items': items,
            'sep': os.sep
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@system_access_required
def api_explorer_read(request):
    """Lit le contenu d'un fichier texte."""
    import os
    from django.http import JsonResponse
    
    path = request.GET.get('path')
    if not path or not os.path.isfile(path):
        return JsonResponse({'success': False, 'error': 'Fichier invalide'}, status=400)
    
    try:
        # Vérification si binaire (basique)
        with open(path, 'rb') as f:
            chunk = f.read(1024)
            if b'\0' in chunk:
                return JsonResponse({'success': False, 'is_binary': True, 'error': 'Fichier binaire non éditable'})
        
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()
            
        return JsonResponse({
            'success': True,
            'content': content,
            'filename': os.path.basename(path),
            'extension': os.path.splitext(path)[1].lower()
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@system_access_required
def api_explorer_save(request):
    """Enregistre les modifications d'un fichier."""
    import os
    from django.http import JsonResponse
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
        
    path = request.POST.get('path')
    content = request.POST.get('content', '')
    
    if not path or not os.path.isfile(path):
        return JsonResponse({'success': False, 'error': 'Fichier invalide'}, status=400)
        
    try:
        with open(path, 'w', encoding='utf-8') as f:
            f.write(content)
        return JsonResponse({'success': True, 'message': 'Fichier enregistré avec succès'})
    except PermissionError:
        return JsonResponse({'success': False, 'error': 'Permission refusée : l\'utilisateur système du site n\'a pas les droits pour modifier ce fichier.'}, status=403)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@system_access_required
def api_explorer_delete(request):
    """Supprime un fichier ou un répertoire."""
    import os
    import shutil
    from django.http import JsonResponse
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
        
    path = request.POST.get('path')
    if not path or not os.path.exists(path):
        return JsonResponse({'success': False, 'error': 'Chemin invalide'}, status=400)
        
    try:
        if os.path.isdir(path):
            shutil.rmtree(path)
        else:
            os.remove(path)
        return JsonResponse({'success': True, 'message': 'Suppression effectuée avec succès'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def extract_timestamp(log_line):
    """Extract timestamp from a log line."""
    try:
        # Try to find common timestamp patterns
        import re
        # Pattern: 2024-01-15 10:30:45 or 2024-01-15T10:30:45
        match = re.search(r'\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}', log_line)
        if match:
            timestamp_str = match.group(0)
            timestamp_str = timestamp_str.replace('T', ' ')
            return datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
    except Exception:
        pass
    return None

def get_firewall_info():
    """Détecte et retourne le statut et les règles du pare-feu actif."""
    import subprocess
    import platform
    
    info = {
        'active': False,
        'type': 'Aucun détecté',
        'rules': [],
        'error': None,
        'raw_output': ""
    }
    
    if platform.system() != 'Linux':
        info['type'] = platform.system()
        info['error'] = "La détection détaillée du pare-feu n'est supportée que sur Linux (UFW, iptables...)"
        return info

    # 1. Tester UFW
    try:
        result = subprocess.run(['sudo', 'ufw', 'status', 'numbered'], capture_output=True, text=True, timeout=5)
        if result.returncode == 0:
            stdout = result.stdout
            info['raw_output'] = stdout
            if "Status: active" in stdout:
                info['active'] = True
                info['type'] = 'UFW'
                # Parsing basique des règles UFW numérotées
                lines = stdout.splitlines()
                for line in lines:
                    if '[' in line and ']' in line:
                        info['rules'].append({'raw': line.strip()})
                return info
    except Exception:
        pass

    # 2. Tester iptables
    try:
        result = subprocess.run(['sudo', 'iptables', '-L', '-n', '-v'], capture_output=True, text=True, timeout=5)
        if result.returncode == 0:
            info['active'] = True
            info['type'] = 'iptables'
            info['raw_output'] = result.stdout
            return info
    except Exception:
        pass

    # 3. Tester Firewalld
    try:
        result = subprocess.run(['sudo', 'firewall-cmd', '--state'], capture_output=True, text=True, timeout=2)
        if result.returncode == 0 and "running" in result.stdout:
            info['active'] = True
            info['type'] = 'Firewalld'
            rules_res = subprocess.run(['sudo', 'firewall-cmd', '--list-all'], capture_output=True, text=True, timeout=5)
            info['raw_output'] = rules_res.stdout
            return info
    except Exception:
        pass

    return info

def get_fail2ban_info():
    """Récupère le statut de Fail2Ban et les IPs bannies."""
    import subprocess
    info = {
        'active': False,
        'jails': [],
        'error': None,
        'total_banned': 0
    }
    
    try:
        result = subprocess.run(['sudo', 'fail2ban-client', 'status'], capture_output=True, text=True, timeout=5)
        if result.returncode == 0:
            info['active'] = True
            # Parsing des jails
            for line in result.stdout.splitlines():
                if "Jail list:" in line:
                    jails_str = line.split(":", 1)[1].strip()
                    if jails_str:
                        jails = jails_str.split(",")
                        for jail in jails:
                            jail_name = jail.strip()
                            jail_data = {'name': jail_name, 'banned_ips': [], 'count': 0}
                            # Status de chaque jail
                            jail_res = subprocess.run(['sudo', 'fail2ban-client', 'status', jail_name], capture_output=True, text=True, timeout=5)
                            if jail_res.returncode == 0:
                                for jline in jail_res.stdout.splitlines():
                                    if "Banned IP list:" in jline:
                                        ips_part = jline.split(":", 1)[1].strip()
                                        if ips_part:
                                            ips = ips_part.split(" ")
                                            jail_data['banned_ips'] = [ip.strip() for ip in ips if ip.strip()]
                                            jail_data['count'] = len(jail_data['banned_ips'])
                                            info['total_banned'] += jail_data['count']
                            info['jails'].append(jail_data)
    except Exception as e:
        info['error'] = str(e)
        
    return info

@system_access_required
def saas_firewall_view(request):
    """Affiche l'état de la sécurité réseau (Pare-feu & Fail2Ban)."""
    firewall = get_firewall_info()
    fail2ban = get_fail2ban_info()
    
    context = {
        'firewall': firewall,
        'fail2ban': fail2ban,
        'system_os': platform.system(),
        'timestamp': datetime.now().strftime("%H:%M:%S")
    }
    
    return render(request, 'saas_admin_app/saas_firewall.html', context)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_tenant_detail_view(request, tenant_id):
    """Affiche les détails complets d'un tenant/institution."""
    
    # Get the tenant/institut
    try:
        institut = Institut.objects.get(id=tenant_id)
    except Institut.DoesNotExist:
        from django.http import Http404
        raise Http404("Tenant non trouvé")
    
    schema_name = institut.schema_name
    User = get_user_model()
    
    # Basic tenant metrics
    db_bytes = get_schema_size_bytes(schema_name)
    db_size_formatted = format_size(db_bytes)
    
    tenant_media_dir = os.path.join(settings.MEDIA_ROOT, schema_name)
    media_bytes = get_directory_size(tenant_media_dir)
    media_size_formatted = format_size(media_bytes)
    
    # Initialize tenant data
    tenant_data = {
        'institut': institut,
        'db_size': db_size_formatted,
        'db_bytes': db_bytes,
        'media_size': media_size_formatted,
        'media_bytes': media_bytes,
        'users': [],
    }
    
    # Get tenant-specific data using tenant_context
    try:
        with tenant_context(institut):
            # Get all users
            users = User.objects.all().order_by('date_joined')
            for user in users:
                tenant_data['users'].append({
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'is_active': user.is_active,
                    'is_superuser': user.is_superuser,
                    'is_staff': user.is_staff,
                    'date_joined': user.date_joined,
                    'last_login': user.last_login,
                })
            
            tenant_data['user_count'] = users.count()
            
    except Exception as e:
        tenant_data['error'] = str(e)
    
    if request.method == 'POST' and 'update_upload_limit' in request.POST:
        try:
            limit = request.POST.get('max_upload_size')
            is_visible = request.POST.get('is_visible') == 'on'
            
            if limit:
                institut.max_upload_size = int(limit)
            else:
                institut.max_upload_size = None
            
            institut.is_visible = is_visible
            institut.save()
            
            from django.contrib import messages
            messages.success(request, f"Configuration mise à jour pour {institut.nom}")
            return redirect('saas_admin_app:saas_tenant_detail', tenant_id=tenant_id)
        except ValueError:
            pass

    context = {
        'tenant_data': tenant_data,
        'global_config': SaaSGlobalConfiguration.get_solo(),
    }
    
    return render(request, 'saas_admin_app/saas_tenant_detail.html', context)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_tenant_sync_pedagogical_data_view(request, tenant_id):
    """API endpoint pour synchroniser des données pédagogiques du MASTER vers un tenant."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)
    
    try:
        # 1. Identifier le tenant cible
        target_tenant = Institut.objects.get(id=tenant_id)
        
        # 2. Identifier le tenant MASTER (exclure le schéma public)
        master_tenant = Institut.objects.filter(tenant_type='master').exclude(schema_name='public').first()
        if not master_tenant:
            return JsonResponse({'success': False, 'error': 'Tenant MASTER introuvable'})
            
        # 3. Récupérer les formations à synchroniser
        formation_ids = request.POST.getlist('formation_ids[]')
        
        with schema_context(master_tenant.schema_name):
            query = Formation.objects.all()
            if formation_ids:
                query = query.filter(id__in=formation_ids)
            formations_master = list(query)
        
        if not formations_master:
            return JsonResponse({'success': False, 'error': 'Aucune formation à synchroniser'})
            
        # 4. Synchroniser chaque formation
        sync_count = 0
        errors = []
        
        for formation in formations_master:
            try:
                sync_formation_to_tenant(formation, target_tenant, master_tenant)
                sync_count += 1
            except Exception as e:
                import traceback
                print(f"Error syncing formation {formation.nom}: {str(e)}")
                print(traceback.format_exc())
                errors.append(f"{formation.nom}: {str(e)}")
        
        message = f"Synchronisation terminée : {sync_count} formations synchronisées."
        if errors:
            message += f" {len(errors)} erreurs rencontrées."
            
        return JsonResponse({
            'success': True,
            'message': message,
            'sync_count': sync_count,
            'errors': errors
        })
        
    except Institut.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Tenant introuvable'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def api_master_formations_list_view(request):
    """API endpoint pour lister les formations du tenant MASTER."""
    from django.http import JsonResponse
    
    master_tenant = Institut.objects.filter(tenant_type='master').exclude(schema_name='public').first()
    if not master_tenant:
        return JsonResponse({'success': False, 'error': 'Tenant MASTER introuvable'})
        
    with schema_context(master_tenant.schema_name):
        formations = list(Formation.objects.all().values('id', 'nom', 'code'))
        
    return JsonResponse({'success': True, 'formations': formations})

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def api_tenant_formations_list_view(request, tenant_id):
    """API endpoint pour lister les formations d'un tenant spécifique."""
    target_tenant = get_object_or_404(Institut, id=tenant_id)
    
    with schema_context(target_tenant.schema_name):
        formations = []
        for f in Formation.objects.all():
            specs = list(Specialites.objects.filter(formation=f).values('id', 'label', 'code'))
            formations.append({
                'id': f.id,
                'nom': f.nom,
                'code': f.code,
                'specialites': specs
            })
            
    return JsonResponse({'success': True, 'formations': formations})

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
@require_POST
def api_tenant_delete_formation_view(request, tenant_id):
    """API endpoint pour supprimer une formation d'un tenant."""
    target_tenant = get_object_or_404(Institut, id=tenant_id)
    formation_id = request.POST.get('formation_id')
    
    try:
        with schema_context(target_tenant.schema_name):
            formation = Formation.objects.get(id=formation_id)
            formation_nom = formation.nom
            
            # Supprimer les dossiers d'inscription associés (car ils sont en DO_NOTHING dans le modèle)
            DossierInscription.objects.filter(formation=formation).delete()
            
            try:
                formation.delete()
                return JsonResponse({'success': True, 'message': f"Formation '{formation_nom}' supprimée avec succès."})
            except ProtectedError:
                 return JsonResponse({'success': False, 'error': "Cette formation ne peut pas être supprimée car elle est liée à d'autres données (ex: inscriptions)."})
    except Formation.DoesNotExist:
        return JsonResponse({'success': False, 'error': "Formation introuvable."})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
@require_POST
def api_tenant_delete_specialite_view(request, tenant_id):
    """API endpoint pour supprimer une spécialité d'un tenant."""
    target_tenant = get_object_or_404(Institut, id=tenant_id)
    specialite_id = request.POST.get('specialite_id')
    
    try:
        with schema_context(target_tenant.schema_name):
            specialite = Specialites.objects.get(id=specialite_id)
            label = specialite.label
            try:
                specialite.delete()
                return JsonResponse({'success': True, 'message': f"Spécialité '{label}' supprimée avec succès."})
            except ProtectedError:
                 return JsonResponse({'success': False, 'error': "Cette spécialité ne peut pas être supprimée car elle est liée à d'autres données."})
    except Specialites.DoesNotExist:
        return JsonResponse({'success': False, 'error': "Spécialité introuvable."})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_global_config_view(request):
    """Page de configuration globale du SaaS."""
    from django.contrib import messages
    config = SaaSGlobalConfiguration.get_solo()
    
    if request.method == 'POST':
        try:
            config.max_upload_size = int(request.POST.get('max_upload_size', 400))
            config.save()
            messages.success(request, 'Configuration globale sauvegardée avec succès')
            return redirect('saas_admin_app:saas_global_config')
        except ValueError:
            messages.error(request, 'Valeur invalide pour la taille d\'upload')
    
    context = {
        'config': config,
    }
    
    return render(request, 'saas_admin_app/saas_global_config.html', context)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def api_tenants_pedagogical_sync_list_view(request):
    """API endpoint pour lister les tenants secondaires avec leur statut de base."""
    from django.http import JsonResponse
    
    # Récupérer les tenants de type 'second'
    tenants = Institut.objects.filter(tenant_type='second').order_by('nom')
    
    data = []
    for t in tenants:
        data.append({
            'id': t.id,
            'nom': t.nom,
            'schema_name': t.schema_name,
            'is_active': t.is_active,
        })
        
    return JsonResponse({'success': True, 'tenants': data})

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_update_user_view(request, tenant_id, user_id):
    """API endpoint pour modifier les informations d'un utilisateur d'un tenant."""
    from django.http import JsonResponse
    from django.views.decorators.http import require_http_methods
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        institut = Institut.objects.get(id=tenant_id)
    except Institut.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Tenant non trouvé'}, status=404)
    
    try:
        with tenant_context(institut):
            User = get_user_model()
            try:
                user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Utilisateur non trouvé'}, status=404)
            
            # Update user fields
            if 'first_name' in request.POST:
                user.first_name = request.POST['first_name']
            if 'last_name' in request.POST:
                user.last_name = request.POST['last_name']
            if 'email' in request.POST:
                user.email = request.POST['email']
            if 'is_active' in request.POST:
                user.is_active = request.POST['is_active'] == 'true'
            if 'is_superuser' in request.POST:
                user.is_superuser = request.POST['is_superuser'] == 'true'
            if 'is_staff' in request.POST:
                user.is_staff = request.POST['is_staff'] == 'true'
            
            user.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Utilisateur mis à jour avec succès',
                'user': {
                    'id': user.id,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'email': user.email,
                    'is_active': user.is_active,
                }
            })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_tenant_files_view(request, tenant_id):
    """Affiche le navigateur de fichiers pour un tenant."""
    try:
        institut = Institut.objects.get(id=tenant_id)
    except Institut.DoesNotExist:
        from django.http import Http404
        raise Http404("Tenant non trouvé")
    
    schema_name = institut.schema_name
    tenant_media_dir = os.path.join(settings.MEDIA_ROOT, schema_name)
    
    # Nombre de Documents d'Inscription CRM
    try:
        with tenant_context(institut):
            from t_crm.models import DocumentsDemandeInscription
            doc_count = DocumentsDemandeInscription.objects.count()
    except Exception:
        doc_count = 0
    
    # Nombre de fichiers inconnus (orphelins)
    unknown_count = get_unknown_files_count(institut)
    
    context = {
        'institut': institut,
        'tenant_id': tenant_id,
        'schema_name': schema_name,
        'media_root': tenant_media_dir,
        'media_exists': os.path.exists(tenant_media_dir),
        'doc_count': doc_count,
        'unknown_count': unknown_count,
    }
    
    return render(request, 'saas_admin_app/saas_tenant_files.html', context)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def api_tenant_file_status(request, tenant_id):
    """API pour lister les fichiers connus et inconnus d'un tenant."""
    from django.http import JsonResponse
    try:
        institut = Institut.objects.get(id=tenant_id)
    except Institut.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Tenant non trouvé'}, status=404)

    schema_name = institut.schema_name
    tenant_media_dir = os.path.join(settings.MEDIA_ROOT, schema_name)
    
    if not os.path.exists(tenant_media_dir):
        return JsonResponse({'success': True, 'known_files': [], 'unknown_files': []})

    # 1. Lister tous les fichiers physiques
    physical_files = []
    for root, _, files in os.walk(tenant_media_dir):
        for f in files:
            abs_f = os.path.join(root, f)
            rel_f = os.path.relpath(abs_f, tenant_media_dir).replace('\\', '/')
            db_path = f"{schema_name}/{rel_f}"
            physical_files.append({
                'db_path': db_path,
                'rel_path': rel_f,
                'name': f,
                'size': format_size(os.path.getsize(abs_f)),
                'mtime': datetime.fromtimestamp(os.path.getmtime(abs_f)).strftime('%d/%m/%Y %H:%M')
            })

    known_files = []
    unknown_files = []
    
    # 2. Récupérer les correspondances en base
    try:
        with tenant_context(institut):
            from t_crm.models import DocumentsDemandeInscription, Prospets
            
            # Mapping complet des fichiers documentés
            prospect_mapping = {}
            
            # Documents
            docs = DocumentsDemandeInscription.objects.filter(file__isnull=False).select_related('prospect')
            for d in docs:
                p_name = f"{d.prospect.nom} {d.prospect.prenom}" if d.prospect else "Sans prospect"
                prospect_mapping[str(d.file)] = {'owner': p_name, 'type': 'Document Dossier'}
            
            # Photos/Logos
            prospects = Prospets.objects.all()
            for p in prospects:
                p_name = f"{p.nom} {p.prenom}"
                if p.photo: prospect_mapping[str(p.photo)] = {'owner': p_name, 'type': 'Photo Profil'}
                if p.logo_entreprise: prospect_mapping[str(p.logo_entreprise)] = {'owner': p_name, 'type': 'Logo Entreprise'}

            # Classification
            for f_info in physical_files:
                match = prospect_mapping.get(f_info['db_path'])
                if match:
                    f_info.update(match)
                    known_files.append(f_info)
                else:
                    unknown_files.append(f_info)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({
        'success': True,
        'tenant_name': institut.nom,
        'known_files': known_files,
        'unknown_files': unknown_files
    })

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_file_browser_view(request, tenant_id):
    """API endpoint pour naviguer dans les fichiers d'un tenant avec filtres."""
    from django.http import JsonResponse
    
    try:
        institut = Institut.objects.get(id=tenant_id)
    except Institut.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Tenant non trouvé'}, status=404)
    
    schema_name = institut.schema_name
    tenant_media_dir = os.path.join(settings.MEDIA_ROOT, schema_name)
    
    # Paramètres de filtrage et tri
    rel_path = request.GET.get('path', '')
    sort_by = request.GET.get('sort', 'name')  # name, size, modified
    order = request.GET.get('order', 'asc')     # asc, desc
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    
    # Normalisation pour Windows/Linux
    rel_path = rel_path.replace('\\', '/')
    
    # Sécurité : empêcher la traversée de répertoire
    abs_path = os.path.normpath(os.path.join(tenant_media_dir, rel_path))
    if not abs_path.startswith(os.path.normpath(tenant_media_dir)):
        return JsonResponse({'success': False, 'error': 'Accès non autorisé'}, status=403)
    
    if not os.path.exists(abs_path):
        return JsonResponse({'success': False, 'error': 'Chemin non trouvé'}, status=404)
    
    # Fonction de filtrage par date
    def is_in_date_range(mtime):
        dt = datetime.fromtimestamp(mtime).date()
        if date_start:
            try:
                ds = datetime.strptime(date_start, '%Y-%m-%d').date()
                if dt < ds: return False
            except ValueError: pass
        if date_end:
            try:
                de = datetime.strptime(date_end, '%Y-%m-%d').date()
                if dt > de: return False
            except ValueError: pass
        return True

    files_and_dirs = []
    
    # Si c'est un fichier, on retourne ses infos (cas rare pour cette API)
    if os.path.isfile(abs_path):
        stat = os.stat(abs_path)
        if is_in_date_range(stat.st_mtime):
            files_and_dirs.append({
                'name': os.path.basename(abs_path),
                'type': 'file',
                'size': stat.st_size,
                'size_formatted': format_size(stat.st_size),
                'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%d/%m/%Y %H:%M'),
                'modified_raw': stat.st_mtime,
                'path': rel_path,
                'extension': os.path.splitext(abs_path)[1].lower(),
            })
        return JsonResponse({
            'success': True,
            'current_path': rel_path,
            'parent_path': os.path.dirname(rel_path),
            'items': files_and_dirs,
        })
    
    # Si c'est un répertoire, on liste le contenu
    try:
        # Récupération des informations de prospect pour les fichiers de ce tenant
        prospect_mapping = {}
        with tenant_context(institut):
            from t_crm.models import DocumentsDemandeInscription, Prospets
            
            # Préfixe du chemin pour les recherches en base de données
            # Les chemins en base commencent par schema_name/
            db_path_prefix = f"{schema_name}/"
            if rel_path:
                db_path_prefix += f"{rel_path}/"
            
            # 1. Documents liés aux dossiers d'inscription
            docs_qs = DocumentsDemandeInscription.objects.filter(
                file__startswith=db_path_prefix,
                prospect__isnull=False
            ).select_related('prospect').only('file', 'prospect__id', 'prospect__nom', 'prospect__prenom')
            
            for d in docs_qs:
                if d.file:
                    prospect_mapping[str(d.file)] = {
                        'id': d.prospect.id,
                        'nom': d.prospect.nom,
                        'prenom': d.prospect.prenom
                    }
            
            # 2. Photos et logos des prospects
            prospects_qs = Prospets.objects.all().only('id', 'nom', 'prenom', 'photo', 'logo_entreprise')
            for p in prospects_qs:
                if p.photo and str(p.photo).startswith(db_path_prefix):
                    prospect_mapping[str(p.photo)] = {'id': p.id, 'nom': p.nom, 'prenom': p.prenom}
                if p.logo_entreprise and str(p.logo_entreprise).startswith(db_path_prefix):
                    prospect_mapping[str(p.logo_entreprise)] = {'id': p.id, 'nom': p.nom, 'prenom': p.prenom}

        if os.path.isdir(abs_path):
            with os.scandir(abs_path) as entries:
                for entry in entries:
                    try:
                        stat = entry.stat()
                        # Filtrage par date
                        if not is_in_date_range(stat.st_mtime):
                            continue
                            
                        is_dir = entry.is_dir()
                        rel_item_path = os.path.relpath(entry.path, tenant_media_dir).replace('\\', '/')
                        
                        # Recherche du prospect associé
                        full_db_path = f"{schema_name}/{rel_item_path}"
                        p_info = prospect_mapping.get(full_db_path)
                        
                        item_info = {
                            'name': entry.name,
                            'type': 'directory' if is_dir else 'file',
                            'size': stat.st_size if not is_dir else 0,
                            'size_formatted': '—' if is_dir else format_size(stat.st_size),
                            'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%d/%m/%Y %H:%M'),
                            'modified_raw': stat.st_mtime,
                            'path': rel_item_path,
                            'prospect_info': p_info if p_info else "Inconnue",
                        }
                        
                        if not is_dir:
                            item_info['extension'] = os.path.splitext(entry.name)[1].lower()
                        
                        files_and_dirs.append(item_info)
                    except Exception:
                        continue
        
        # Logique de tri
        is_reverse = (order == 'desc')
        
        # Tri principal (selon le critère choisi)
        if sort_by == 'size':
            files_and_dirs.sort(key=lambda x: x['size'], reverse=is_reverse)
        elif sort_by == 'modified':
            files_and_dirs.sort(key=lambda x: x['modified_raw'], reverse=is_reverse)
        else: # name par défaut
            files_and_dirs.sort(key=lambda x: x['name'].lower(), reverse=is_reverse)
            
        # Toujours garder les dossiers en premier (tri stable)
        files_and_dirs.sort(key=lambda x: x['type'] != 'directory')
        
        # Calcul du chemin parent
        parent_path = os.path.dirname(rel_path).replace('\\', '/') if rel_path else ''
        
        return JsonResponse({
            'success': True,
            'current_path': rel_path,
            'parent_path': parent_path,
            'items': files_and_dirs,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@system_access_required
def saas_reset_user_sessions_view(request, tenant_id, user_id):
    """Réinitialise toutes les sessions actives pour un utilisateur spécifique."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
        
    try:
        from django_tenants.utils import tenant_context
        from app.models import Institut
        from django.contrib.auth import get_user_model
        from django.contrib.sessions.models import Session
        from django.utils import timezone
        
        institut = Institut.objects.get(id=tenant_id)
        with tenant_context(institut):
            User = get_user_model()
            user = User.objects.get(id=user_id)
            
            # Pour vider les sessions en DB
            sessions = Session.objects.filter(expire_date__gte=timezone.now())
            count = 0
            for session in sessions:
                data = session.get_decoded()
                if str(user.id) == data.get('_auth_user_id'):
                    session.delete()
                    count += 1
            
            # Note: Si Redis est utilisé en prod, il faudrait une logique différente
            # pour scanner les clés redis, mais ici on gère le cas standard DB.
            
            return JsonResponse({
                'success': True, 
                'message': f"Toutes les sessions ({count}) de {user.username} ont été réinitialisées."
            })
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@system_access_required
def saas_file_serve_view(request, tenant_id, file_path):
    """Sert un fichier spécifique d'un tenant."""
    from django.http import FileResponse, Http404, HttpResponseForbidden
    import mimetypes
    
    try:
        institut = Institut.objects.get(id=tenant_id)
    except Institut.DoesNotExist:
        raise Http404("Tenant non trouvé")
    
    schema_name = institut.schema_name
    tenant_media_dir = os.path.join(settings.MEDIA_ROOT, schema_name)
    
    # Build the absolute file path
    abs_file_path = os.path.normpath(os.path.join(tenant_media_dir, file_path))
    
    # Security check: prevent path traversal
    if not abs_file_path.startswith(os.path.normpath(tenant_media_dir)):
        return HttpResponseForbidden("Accès non autorisé")
    
    if not os.path.exists(abs_file_path):
        raise Http404("Fichier non trouvé")
    
    if not os.path.isfile(abs_file_path):
        raise Http404("Ce n'est pas un fichier")
    
    # Determine content type
    content_type, _ = mimetypes.guess_type(abs_file_path)
    if content_type is None:
        content_type = 'application/octet-stream'
    
    # Serve the file
    response = FileResponse(
        open(abs_file_path, 'rb'),
        content_type=content_type
    )
    response['Content-Disposition'] = f'inline; filename="{os.path.basename(abs_file_path)}"'
    response['Content-Length'] = os.path.getsize(abs_file_path)
    
    # Allow embedding in iframes for preview
    response['X-Frame-Options'] = 'SAMEORIGIN'
    response['Content-Security-Policy'] = "frame-ancestors 'self'"
    
    return response

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_reset_user_password_view(request, tenant_id, user_id):
    """API endpoint pour réinitialiser le mot de passe d'un utilisateur."""
    from django.http import JsonResponse
    import secrets
    from .models import SaaSEmailConfiguration
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    # Check if we should send email
    send_email = request.POST.get('send_email', 'false').lower() == 'true'
    
    try:
        institut = Institut.objects.get(id=tenant_id)
    except Institut.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Tenant non trouvé'}, status=404)
    
    try:
        with tenant_context(institut):
            User = get_user_model()
            try:
                user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Utilisateur non trouvé'}, status=404)
            
            # Generate a random password
            new_password = secrets.token_urlsafe(12)
            
            # Set the new password
            user.set_password(new_password)
            user.save()
            
            # Send email if requested and email is enabled
            email_sent = False
            email_error = None
            
            if send_email and user.email:
                try:
                    # Get email configuration from SaaS admin
                    config = SaaSEmailConfiguration.get_solo()
                    
                    if config.email_enabled:
                        # Apply email settings
                        config.apply_email_settings()
                        
                        # Format email content
                        subject = config.email_reset_password_subject
                        message = config.email_reset_password_template.format(
                            user_name=user.get_full_name() or user.username,
                            password=new_password
                        )
                        
                        # Send email
                        send_mail(
                            subject=subject,
                            message=message,
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            recipient_list=[user.email],
                            fail_silently=False,
                        )
                        email_sent = True
                    else:
                        email_error = "L'envoi d'emails n'est pas activé dans la configuration"
                except Exception as email_err:
                    email_error = f"Erreur lors de l'envoi: {str(email_err)}"
                    import logging
                    logging.getLogger(__name__).error(f"Email send error: {email_err}")
            
            response_data = {
                'success': True,
                'message': f'Mot de passe réinitialisé pour {user.get_full_name() or user.email}',
                'new_password': new_password,
                'email_sent': email_sent,
                'email_error': email_error,
                'user': {
                    'id': user.id,
                    'email': user.email,
                    'full_name': user.get_full_name() or user.username,
                }
            }
            
            return JsonResponse(response_data)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_email_config_view(request):
    """Page de configuration des emails."""
    from .models import SaaSEmailConfiguration
    from django.shortcuts import redirect
    from django.contrib import messages
    
    config = SaaSEmailConfiguration.get_solo()
    
    if request.method == 'POST':
        config.email_enabled = request.POST.get('email_enabled', '').lower() in ['true', 'on', '1']
        config.email_host = request.POST.get('email_host', 'smtp.gmail.com')
        config.email_port = int(request.POST.get('email_port', 587))
        config.email_use_tls = request.POST.get('email_use_tls', '').lower() in ['true', 'on', '1']
        config.email_host_user = request.POST.get('email_host_user', '')
        config.email_host_password = request.POST.get('email_host_password', '')
        config.default_from_email = request.POST.get('default_from_email', 'noreply@school-saas.com')
        config.email_reset_password_subject = request.POST.get('email_reset_password_subject', config.email_reset_password_subject)
        config.email_reset_password_template = request.POST.get('email_reset_password_template', config.email_reset_password_template)
        config.save()
        
        messages.success(request, 'Configuration email sauvegardée avec succès')
        return redirect('saas_admin_app:saas_email_config')
    
    context = {
        'config': config,
    }
    
    return render(request, 'saas_admin_app/saas_email_config.html', context)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_test_email_view(request):
    """API pour tester l'envoi d'email."""
    from .models import SaaSEmailConfiguration
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        config = SaaSEmailConfiguration.get_solo()
        
        if not config.email_enabled:
            return JsonResponse({'success': False, 'error': "L'envoi d'emails n'est pas activé"}, status=400)
        
        test_email = request.POST.get('email', config.email_host_user)
        if not test_email:
            return JsonResponse({'success': False, 'error': 'Aucune email de test fournie'}, status=400)
        
        # Send test email via notre mécanisme central
        from saas_admin_app.email_utils import send_platform_email
        success = send_platform_email(
            subject='Test Email - School SaaS',
            message='Ceci est un email de test pour vérifier la configuration SMTP de School SaaS.\n\nSi vous avez reçu cet email, la configuration fonctionne correctement.',
            recipient_list=[test_email]
        )
        
        if success:
            return JsonResponse({
                'success': True,
                'message': f'Email de test envoyé avec succès à {test_email}'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': "Échec de l'envoi. Vérifiez vos identifiants SMTP (Serveur, Port, Utilisateur, Mot de passe) dans la configuration."
            }, status=500)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_toggle_maintenance_view(request):
    """API endpoint pour basculer le mode de développement/maintenance."""
    from .models import SaaSMaintenanceConfiguration
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        config = SaaSMaintenanceConfiguration.get_solo()
        
        # Le frontend nous envoie 'true' ou 'false' sous forme de chaine
        is_maintenance = request.POST.get('is_maintenance_mode', 'false').lower() == 'true'
        config.is_maintenance_mode = is_maintenance
        
        # Récupération optionnelle du message et de l'heure de fin
        message = request.POST.get('maintenance_message')
        if message:
            config.maintenance_message = message
            
        end_time_raw = request.POST.get('maintenance_end_time')
        if end_time_raw:
            try:
                # Le format envoyé par datetime-local est souvent YYYY-MM-DDTHH:MM
                from django.utils.dateparse import parse_datetime
                config.maintenance_end_time = parse_datetime(end_time_raw)
            except Exception:
                config.maintenance_end_time = None
        else:
            config.maintenance_end_time = None
            
        config.save()
        
        status_text = "activé" if is_maintenance else "désactivé"
        return JsonResponse({
            'success': True,
            'message': f'Mode de développement (maintenance) {status_text} avec succès.',
            'is_maintenance_mode': config.is_maintenance_mode,
            'maintenance_end_time': config.maintenance_end_time.isoformat() if config.maintenance_end_time else None
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_backups_view(request):
    """Affiche la liste des sauvegardes effectuées."""
    backups = DatabaseBackup.objects.all()
    
    # Statistiques simples
    total_backups = backups.count()
    global_backups = backups.filter(backup_type='GLOBAL').count()
    tenant_backups = backups.filter(backup_type='TENANT').count()
    total_size = sum(b.size for b in backups)
    
    context = {
        'backups': backups,
        'instituts': Institut.objects.all().order_by('nom'),
        'total_backups': total_backups,
        'global_backups': global_backups,
        'tenant_backups': tenant_backups,
        'total_size_formatted': format_size(total_size),
    }
    return render(request, 'saas_admin_app/saas_backups.html', context)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_create_backup_view(request):
    """Déclenche une nouvelle sauvegarde."""
    tenant_id = request.GET.get('tenant_id')
    tenant = None
    
    if tenant_id:
        try:
            tenant = Institut.objects.get(id=tenant_id)
        except Institut.DoesNotExist:
            messages.error(request, "Tenant non trouvé.")
            return redirect('saas_admin_app:saas_backups')
    
    success, result = perform_backup(tenant)
    
    if success:
        messages.success(request, f"Sauvegarde {'globale' if not tenant else f'du tenant {tenant.nom}'} réussie.")
    else:
        messages.error(request, f"Échec de la sauvegarde : {result}")
        
    return redirect('saas_admin_app:saas_backups')

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_delete_backup_view(request, backup_id):
    """Supprime une sauvegarde spécifique."""
    try:
        backup = DatabaseBackup.objects.get(id=backup_id)
        filename = backup.filename
        backup.delete()  # Ceci supprimera aussi le fichier physique via la méthode delete du modèle
        messages.success(request, f"Sauvegarde {filename} supprimée.")
    except DatabaseBackup.DoesNotExist:
        messages.error(request, "Sauvegarde non trouvée.")
        
    return redirect('saas_admin_app:saas_backups')

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_download_backup_view(request, backup_id):
    """Permet de télécharger un fichier de sauvegarde."""
    from django.http import FileResponse, Http404
    try:
        backup = DatabaseBackup.objects.get(id=backup_id)
        if backup.file and os.path.exists(backup.file.path):
            return FileResponse(open(backup.file.path, 'rb'), as_attachment=True, filename=backup.filename)
        else:
            messages.error(request, "Fichier physique introuvable.")
            return redirect('saas_admin_app:saas_backups')
    except DatabaseBackup.DoesNotExist:
        raise Http404("Sauvegarde non trouvée.")


@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_restore_backup_view(request, backup_id):
    """Effectue la restauration d'une sauvegarde de tenant."""
    from django.shortcuts import get_object_or_404, redirect
    from django.contrib import messages
    
    if request.method != 'POST':
        messages.error(request, "Méthode non autorisée.")
        return redirect('saas_admin_app:saas_backups')
        
    backup_obj = get_object_or_404(DatabaseBackup, id=backup_id)
    
    # Sécurité supplémentaire : vérification du code de confirmation
    confirmation_code = request.POST.get('confirmation_code')
    if confirmation_code != backup_obj.filename:
        messages.error(request, "Le code de confirmation est incorrect. Veuillez saisir le nom complet du fichier.")
        return redirect('saas_admin_app:saas_backups')

    # 1. Activer le mode maintenance
    maintenance_config = SaaSMaintenanceConfiguration.get_solo()
    old_maintenance_state = maintenance_config.is_maintenance_mode
    maintenance_config.is_maintenance_mode = True
    maintenance_config.maintenance_message = f"Restauration en cours pour {backup_obj.tenant.nom if backup_obj.tenant else 'un tenant'}. Veuillez patienter..."
    maintenance_config.save()
    
    try:
        # 2. Effectuer la restauration
        success, message = perform_restore(backup_obj)
        
        if success:
            messages.success(request, f"Succès : {message}")
        else:
            messages.error(request, f"Échec de la restauration : {message}")
            
    finally:
        # 3. Restaurer l'état précédent du mode maintenance
        maintenance_config.is_maintenance_mode = old_maintenance_state
        maintenance_config.save()
        
    return redirect('saas_admin_app:saas_backups')

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_force_tenant_password_change_view(request, tenant_id):
    """
    Active le forçage de changement de mot de passe pour un tenant et déconnecte tous ses utilisateurs.
    """
    from django.contrib.sessions.models import Session
    from django.utils import timezone
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
        
    try:
        institut = Institut.objects.get(id=tenant_id)
        
        # Basculer l'état
        institut.force_password_change = not institut.force_password_change
            
        if institut.force_password_change:
            institut.password_reset_date = timezone.now()
            
            # Déconnecter tous les utilisateurs de ce tenant
            with tenant_context(institut):
                from institut_app.models import UserSession
                user_sessions = UserSession.objects.all()
                session_keys = [us.last_session_key for us in user_sessions if us.last_session_key]
                
                if session_keys:
                    Session.objects.filter(session_key__in=session_keys).delete()
                    # On vide les clés de session pour forcer le re-login
                    user_sessions.update(last_session_key=None)
                    
            message = f"Le forçage de changement de mot de passe a été activé pour {institut.nom}. Tous les utilisateurs ont été déconnectés."
        else:
            message = f"Le forçage de changement de mot de passe a été désactivé pour {institut.nom}."
            
        institut.save()
        
        return JsonResponse({
            'success': True, 
            'message': message,
            'force_password_change': institut.force_password_change
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_force_user_password_change_view(request, tenant_id, user_id):
    """
    Force le changement de mot de passe pour un utilisateur spécifique d'un tenant et le déconnecte.
    """
    from django.contrib.sessions.models import Session
    from django.contrib.auth.models import User
    from app.models import Institut
    from django_tenants.utils import tenant_context
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
        
    try:
        institut = Institut.objects.get(id=tenant_id)
        
        with tenant_context(institut):
            from institut_app.models import Profile, UserSession
            
            user = User.objects.get(id=user_id)
            profile, _ = Profile.objects.get_or_create(user=user)
            
            profile.force_password_change = True
            profile.save()
            
            # Déconnecter l'utilisateur
            try:
                user_session = UserSession.objects.get(user=user)
                if user_session.last_session_key:
                    Session.objects.filter(session_key=user_session.last_session_key).delete()
                    user_session.last_session_key = None
                    user_session.save(update_fields=['last_session_key'])
            except UserSession.DoesNotExist:
                pass
                
        message = f"Le forçage de changement de mot de passe a été activé pour l'utilisateur {user.username}."
        
        return JsonResponse({
            'success': True, 
            'message': message,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

from django.shortcuts import get_object_or_404

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_knowledge_center_view(request):
    from .models import KnowledgeCategory, KnowledgeResource
    
    categories = KnowledgeCategory.objects.all().order_by('order', 'name')
    resources = KnowledgeResource.objects.all().order_by('category__order', 'order', '-created_at')
    
    context = {
        'categories': categories,
        'resources': resources,
    }
    return render(request, 'saas_admin_app/saas_knowledge_list.html', context)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_knowledge_category_action_view(request):
    from .models import KnowledgeCategory
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            name = request.POST.get('name')
            description = request.POST.get('description')
            KnowledgeCategory.objects.create(name=name, description=description)
            messages.success(request, "Catégorie ajoutée avec succès.")
        elif action == 'edit':
            cat_id = request.POST.get('id')
            cat = get_object_or_404(KnowledgeCategory, id=cat_id)
            cat.name = request.POST.get('name')
            cat.description = request.POST.get('description')
            cat.save()
            messages.success(request, "Catégorie modifiée avec succès.")
        elif action == 'delete':
            cat_id = request.POST.get('id')
            cat = get_object_or_404(KnowledgeCategory, id=cat_id)
            cat.delete()
            messages.success(request, "Catégorie supprimée.")
    return redirect('saas_admin_app:saas_knowledge_center')

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_knowledge_resource_action_view(request):
    from .models import KnowledgeCategory, KnowledgeResource
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            title = request.POST.get('title')
            cat_id = request.POST.get('category')
            resource_type = request.POST.get('resource_type')
            description = request.POST.get('description')
            video_url = request.POST.get('video_url')
            is_published = request.POST.get('is_published') == 'on'
            file_attachment = request.FILES.get('file_attachment')
            
            category = get_object_or_404(KnowledgeCategory, id=cat_id)
            KnowledgeResource.objects.create(
                title=title, category=category, resource_type=resource_type,
                description=description, video_url=video_url, is_published=is_published,
                file_attachment=file_attachment
            )
            messages.success(request, "Ressource ajoutée avec succès.")
        elif action == 'edit':
            res_id = request.POST.get('id')
            res = get_object_or_404(KnowledgeResource, id=res_id)
            res.title = request.POST.get('title')
            cat_id = request.POST.get('category')
            res.category = get_object_or_404(KnowledgeCategory, id=cat_id)
            res.resource_type = request.POST.get('resource_type')
            res.description = request.POST.get('description')
            res.video_url = request.POST.get('video_url')
            res.is_published = request.POST.get('is_published') == 'on'
            if 'file_attachment' in request.FILES:
                res.file_attachment = request.FILES.get('file_attachment')
            res.save()
            messages.success(request, "Ressource modifiée avec succès.")
        elif action == 'delete':
            res_id = request.POST.get('id')
            res = get_object_or_404(KnowledgeResource, id=res_id)
            res.delete()
            messages.success(request, "Ressource supprimée.")
    return redirect('saas_admin_app:saas_knowledge_center')

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_toggle_tenant_visibility_view(request, tenant_id):
    """Bascule la visibilité d'un tenant dans la sélection d'organisation."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
        
    try:
        institut = Institut.objects.get(id=tenant_id)
        institut.is_visible = not institut.is_visible
        institut.save()
        
        status_text = "visible" if institut.is_visible else "invisible"
        message = f"L'institution '{institut.nom}' est désormais {status_text} dans la sélection."
        
        return JsonResponse({
            'success': True, 
            'message': message,
            'is_visible': institut.is_visible
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_toggle_tenant_active_view(request, tenant_id):
    """Bascule l'état actif/désactivé d'un tenant."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
        
    try:
        institut = Institut.objects.get(id=tenant_id)
        institut.is_active = not institut.is_active
        institut.save()
        
        status_text = "activé" if institut.is_active else "désactivé"
        message = f"L'institution '{institut.nom}' est désormais {status_text}. L'accès via URL est {'autorisé' if institut.is_active else 'bloqué'}."
        
        return JsonResponse({
            'success': True, 
            'message': message,
            'is_active': institut.is_active
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_audit_logs_view(request):
    """Affiche les logs d'audit (UserActionLog) de tous les tenants."""
    instituts = Institut.objects.all().order_by('nom')
    all_logs = []
    
    selected_tenant_id = request.GET.get('tenant')
    action_type = request.GET.get('action_type')
    
    for inst in instituts:
        if selected_tenant_id and str(inst.id) != selected_tenant_id:
            continue
            
        try:
            with tenant_context(inst):
                from t_crm.models import UserActionLog
                qs = UserActionLog.objects.all()
                if action_type:
                    qs = qs.filter(action_type=action_type)
                logs = qs.select_related('user').order_by('-created_at')[:50]
                for log in logs:
                    all_logs.append({
                        'log': log,
                        'tenant': inst,
                    })
        except Exception:
            pass
            
    # Sort all logs by date
    all_logs.sort(key=lambda x: x['log'].created_at, reverse=True)
    all_logs = all_logs[:200] # Limit to 200 for performance
    
    from t_crm.models import UserActionLog
    action_choices = UserActionLog.ACTION_CHOICES
    
    context = {
        'logs': all_logs,
        'instituts': instituts,
        'action_choices': action_choices,
        'selected_tenant_id': int(selected_tenant_id) if selected_tenant_id else None,
        'selected_action': action_type,
    }
    return render(request, 'saas_admin_app/saas_audit_logs.html', context)

@user_passes_test(superadmin_only, login_url='/saas-admin/login/')
def saas_clear_tenant_logs_view(request):
    """Vide les logs d'audit pour un ou tous les tenants."""
    if request.method == 'POST':
        tenant_id = request.POST.get('tenant_id')
        delete_mode = request.POST.get('delete_mode')
        date_threshold = request.POST.get('date_threshold')
        
        instituts = Institut.objects.all()
        if tenant_id and tenant_id != 'all':
            instituts = instituts.filter(id=tenant_id)
            
        total_deleted = 0
        for inst in instituts:
            try:
                with tenant_context(inst):
                    from t_crm.models import UserActionLog
                    qs = UserActionLog.objects.all()
                    
                    if delete_mode == 'older_than' and date_threshold:
                        qs = qs.filter(created_at__lt=date_threshold)
                    
                    count, _ = qs.delete()
                    total_deleted += count
            except Exception:
                pass
                
        messages.success(request, f"Nettoyage terminé : {total_deleted} entrées supprimées au total.")
        
    return redirect('saas_admin_app:saas_audit_logs')


from django.views.decorators.http import require_POST

@require_POST
def saas_reset_tresorerie_view(request, tenant_id):
    """
    Rinitialise les donnes financires d'un locataire (tenant).
    Purge les tables: Rembourssements, PaiementRemboursement, Depenses, OperationsBancaire, DepotBanque
    """
    from django.http import JsonResponse
    from app.models import Institut
    from django.shortcuts import get_object_or_404
    from django_tenants.utils import tenant_context
    
    institut = get_object_or_404(Institut, id=tenant_id)
    
    try:
        with tenant_context(institut):
            from t_tresorerie.models import (
                Rembourssements, PaiementRemboursement, 
                Depenses, OperationsBancaire, DepotBanque
            )
            from t_conseil.models import Facture
            
            # Delete in reverse dependency order if any, though Django's CASCADE handles most.
            # We explicitly delete from all 5 models to clear them out.
            PaiementRemboursement.objects.all().delete()
            Rembourssements.objects.all().delete()
            Depenses.objects.all().delete()
            OperationsBancaire.objects.all().delete()
            DepotBanque.objects.all().delete()
            Facture.objects.filter(module_source='tresorerie').delete()
            
            return JsonResponse({'status': 'success', 'message': 'La trsorerie a t rinitialise avec succs.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
