from django.shortcuts import render, redirect
from django.contrib import messages
from ..models import *
from ..forms import *
from django.contrib.auth.decorators import login_required
from institut_app.decorators import module_permission_required
from django.db import transaction, IntegrityError
from django_tenants.utils import get_tenant_model, schema_context
from django.http import JsonResponse
from django.db.models import Q
from datetime import datetime
import json
import csv
import openpyxl
from django.http import HttpResponse
from django.core.paginator import Paginator

def format_dispo(dispo):
    if not dispo or 'disponibilites' not in dispo:
        return ""
    items = []
    for d in dispo['disponibilites']:
        jour = d.get('jour', '').capitalize()
        debut = d.get('heure_debut', '')
        fin = d.get('heure_fin', '')
        items.append(f"{jour}:{debut}-{fin}")
    return ", ".join(items)

@login_required(login_url="institut_app:login")
def export_formateurs(request):
    formateurs = Formateurs.objects.all()

    headers = ['Email', 'Nom', 'Prénom', 'Téléphone', 'Diplôme', 'NIN', 'Disponibilité']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="formateurs_export.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Formateurs'
    sheet.append(headers)

    for f in formateurs:
        sheet.append([
            f.email,
            f.nom,
            f.prenom,
            f.telephone,
            f.diplome,
            f.nin,
            format_dispo(f.dispo)
        ])

    workbook.save(response)
    return response

@login_required(login_url="institut_app:login")
@module_permission_required('int', 'add')
def import_formateurs(request):
    if request.method == "POST":
        if 'file' not in request.FILES:
            messages.error(request, 'Aucun fichier sélectionné.')
            return redirect('t_formations:PageFormateurs')
            
        excel_file = request.FILES['file']
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'Veuillez utiliser un fichier au format Excel (.xlsx)')
            return redirect('t_formations:PageFormateurs')

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            
            headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
            
            email_idx = headers.index('Email') if 'Email' in headers else -1
            nom_idx = headers.index('Nom') if 'Nom' in headers else -1
            prenom_idx = headers.index('Prénom') if 'Prénom' in headers else -1
            tel_idx = headers.index('Téléphone') if 'Téléphone' in headers else -1
            diplome_idx = headers.index('Diplôme') if 'Diplôme' in headers else -1
            nin_idx = headers.index('NIN') if 'NIN' in headers else -1
            
            if nom_idx == -1 or prenom_idx == -1 or tel_idx == -1 or email_idx == -1:
                messages.error(request, 'Format invalide. Les colonnes Email, Nom, Prénom, Téléphone sont obligatoires.')
                return redirect('t_formations:PageFormateurs')
                
            count_new = 0
            count_updated = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                email = str(row[email_idx]).strip() if email_idx != -1 and row[email_idx] else None
                nom = str(row[nom_idx]).strip() if nom_idx != -1 and row[nom_idx] else None
                prenom = str(row[prenom_idx]).strip() if prenom_idx != -1 and row[prenom_idx] else None
                telephone = str(row[tel_idx]).strip() if tel_idx != -1 and row[tel_idx] else None
                diplome = str(row[diplome_idx]).strip() if diplome_idx != -1 and row[diplome_idx] else ""
                nin = str(row[nin_idx]).strip() if nin_idx != -1 and row[nin_idx] else ""
                
                # Ignorer la disponibilité car ce n'est pas requis
                
                if not nom or not prenom or not email or not telephone or email.lower() == 'none':
                    continue
                    
                formateur, created = Formateurs.objects.update_or_create(
                    email=email,
                    defaults={
                        'nom': nom,
                        'prenom': prenom,
                        'telephone': telephone,
                        'diplome': diplome if diplome and diplome.lower() != 'none' else "",
                        'nin': nin if nin and nin.lower() != 'none' else "",
                    }
                )
                if created:
                    count_new += 1
                else:
                    count_updated += 1
                    
            messages.success(request, f'Import terminé avec succès. {count_new} créés, {count_updated} mis à jour.')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation: {str(e)}")
            
    return redirect('t_formations:PageFormateurs')


@login_required(login_url="institut_app:login")
@module_permission_required('int', 'view')
def PageFormateurs(request):
    search_query = request.GET.get('search', '')
    
    queryset = Formateurs.objects.all().order_by('-nom')

    if search_query:
        for term in search_query.split():
            queryset = queryset.filter(
                Q(nom__icontains=term) |
                Q(prenom__icontains=term) |
                Q(email__icontains=term) |
                Q(telephone__icontains=term) |
                Q(nin__icontains=term)
            )

    # Stats (full dataset for total counts)
    total_count = Formateurs.objects.count()
    
    # Pagination
    paginator = Paginator(queryset, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        "formateurs": page_obj, # Template expects 'formateurs' to iterate
        "page_obj": page_obj,
        "total_count": total_count,
        "filters": {
            "search": search_query,
        }
    }
    return render(request, 'tenant_folder/formateur/liste_des_formateur.html', context)

@login_required(login_url="institut_app:login")
@module_permission_required('int', 'view')
def ChargeHoraireFormateur(request):
    from t_timetable.models import TimetableEntry
    from datetime import date

    formateur_id = request.GET.get('formateur_id')
    mode_calcul = request.GET.get('mode', 'standard')

    # Get list of all formateurs for the dropdown select list
    formateurs = Formateurs.objects.all().order_by('nom', 'prenom')

    selected_formateur = None
    workload_data = []
    totaux = {
        'hebdo': 0,
        'mensuel_standard': 0,
        'mensuel_reel': 0,
        'semestre': 0,
    }

    if formateur_id:
        try:
            selected_formateur = Formateurs.objects.get(id=formateur_id)
            # Find timetable entries for this formateur, but only in active/en cours timetables
            # and order them to get consistent output
            entries = TimetableEntry.objects.filter(
                formateur=selected_formateur,
                timetable__status__in=['enc', 'val']
            ).select_related('timetable', 'timetable__groupe')

            workload_by_group_sem = {}

            for entry in entries:
                if not entry.heure_debut or not entry.heure_fin or not entry.jour:
                    continue

                # Calculate duration in hours
                t1 = datetime.combine(date(2000, 1, 1), entry.heure_debut)
                t2 = datetime.combine(date(2000, 1, 1), entry.heure_fin)
                duration_hours = (t2 - t1).total_seconds() / 3600.0

                timetable = entry.timetable
                groupe_name = timetable.groupe.nom if (timetable and timetable.groupe) else "N/A"
                semestre_name = timetable.get_semestre_display() if hasattr(timetable, 'get_semestre_display') else f"Semestre {timetable.semestre}"

                key = (groupe_name, semestre_name)
                if key not in workload_by_group_sem:
                    workload_by_group_sem[key] = {
                        'groupe': groupe_name,
                        'semestre': semestre_name,
                        'jour_stats': {
                            'lundi': 0.0,
                            'mardi': 0.0,
                            'mercredi': 0.0,
                            'jeudi': 0.0,
                            'vendredi': 0.0,
                            'samedi': 0.0,
                            'dimanche': 0.0,
                        },
                        'total_hebdomadaire': 0.0,
                    }

                day = entry.jour.strip().lower()
                if day in workload_by_group_sem[key]['jour_stats']:
                    workload_by_group_sem[key]['jour_stats'][day] += duration_hours
                    workload_by_group_sem[key]['total_hebdomadaire'] += duration_hours

            totaux_hebdo = 0.0
            for key, data in workload_by_group_sem.items():
                # Format each day's hour values
                for d in data['jour_stats']:
                    val = data['jour_stats'][d]
                    data['jour_stats'][d] = int(val) if val.is_integer() else round(val, 2)

                tot_hebdo = data['total_hebdomadaire']
                data['total_hebdomadaire'] = int(tot_hebdo) if tot_hebdo.is_integer() else round(tot_hebdo, 2)
                totaux_hebdo += tot_hebdo

                # Mode-specific calculations for row totals
                m_std = tot_hebdo * 4
                m_reel = tot_hebdo * 4.33
                sem_tot = tot_hebdo * 14

                data['mensuel_standard'] = int(m_std) if m_std.is_integer() else round(m_std, 2)
                data['mensuel_reel'] = int(m_reel) if m_reel.is_integer() else round(m_reel, 2)
                data['semestre_total'] = int(sem_tot) if sem_tot.is_integer() else round(sem_tot, 2)

                workload_data.append(data)

            # Global totals
            totaux = {
                'hebdo': int(totaux_hebdo) if totaux_hebdo.is_integer() else round(totaux_hebdo, 2),
                'mensuel_standard': int(totaux_hebdo * 4) if (totaux_hebdo * 4).is_integer() else round(totaux_hebdo * 4, 2),
                'mensuel_reel': int(totaux_hebdo * 4.33) if (totaux_hebdo * 4.33).is_integer() else round(totaux_hebdo * 4.33, 2),
                'semestre': int(totaux_hebdo * 14) if (totaux_hebdo * 14).is_integer() else round(totaux_hebdo * 14, 2),
            }

        except Formateurs.DoesNotExist:
            messages.error(request, "Formateur introuvable.")

    context = {
        'formateurs': formateurs,
        'selected_formateur': selected_formateur,
        'mode_calcul': mode_calcul,
        'workload_data': workload_data,
        'totaux': totaux,
    }
    return render(request, 'tenant_folder/formateur/charge_horaire.html', context)

@login_required(login_url="institut_app:login")
@module_permission_required('int', 'change')
def request_formateur_dispo(request):
    if request.method == "POST":
        formateur_id = request.POST.get("id")
        try:
            formateur = Formateurs.objects.get(id=formateur_id)
            if not isinstance(formateur.dispo, dict):
                formateur.dispo = {}
            
            formateur.dispo["demande_dispo"] = True
            formateur.save()
            
            return JsonResponse({"status": "success", "message": f"Demande envoyée à {formateur.nom} {formateur.prenom}."})
        except Formateurs.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Formateur introuvable."}, status=404)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
    return JsonResponse({"status": "error", "message": "Méthode non autorisée."}, status=405)

@login_required(login_url="institut_app:login")
@module_permission_required('int', 'add')
@transaction.atomic
def create_formateur(request):
    if request.method == "POST":
        nom = request.POST.get('nom')
        prenom = request.POST.get('prenom')
        telephone = request.POST.get('telephone')
        email = request.POST.get('email')
        diplome = request.POST.get('diplome')
        nin = request.POST.get('nin')
        password = request.POST.get('password')
        is_particular_irg = request.POST.get('is_particular_irg') in ['true', 'on', '1']

        Formateurs.objects.create(
            nom = nom,
            prenom = prenom,
            telephone = telephone,
            email = email,
            diplome = diplome,
            nin = nin,
            password = password,
            is_particular_irg = is_particular_irg,
        )
        messages.success(request,'Les données du formateur ont été sauvegarder avec succès.')
        return JsonResponse({'status': 'success'})
    else:
        return JsonResponse({"status" : "error","message":"Method non autoriser"})


@login_required(login_url="institut_app:login")
@module_permission_required('int', 'change')
@transaction.atomic
def update_formateur(request):
    if request.method == "POST":
        formateur_id = request.POST.get('id')
        nom = request.POST.get('nom')
        prenom = request.POST.get('prenom')
        telephone = request.POST.get('telephone')
        email = request.POST.get('email')
        diplome = request.POST.get('diplome')
        nin = request.POST.get('nin')
        password = request.POST.get('password')
        is_particular_irg = request.POST.get('is_particular_irg') in ['true', 'on', '1']

        try:
            formateur = Formateurs.objects.get(id=formateur_id)
            formateur.nom = nom
            formateur.prenom = prenom
            formateur.telephone = telephone
            formateur.email = email
            formateur.diplome = diplome
            formateur.nin = nin
            formateur.is_particular_irg = is_particular_irg
            if password:
                formateur.password = password

            formateur.save()
            messages.success(request,'Les données du formateur ont été mises à jour avec succès.')
            return JsonResponse({'status': 'success'})
        except Formateurs.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Le formateur spécifié n\'existe pas.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    else:
        return JsonResponse({"status": "error", "message": "Méthode non autorisée"})


@login_required(login_url="institut_app:login")
@module_permission_required('int', 'delete')
def delete_formateur(request):
    if request.method == "POST":
        formateur_id = request.POST.get('id')

        try:
            formateur = Formateurs.objects.get(id=formateur_id)
            formateur.delete()
            messages.success(request, 'Le formateur a été supprimé avec succès.')
            return JsonResponse({'status': 'success'})
        except Formateurs.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Le formateur spécifié n\'existe pas.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    else:
        return JsonResponse({"status": "error", "message": "Méthode non autorisée"})


@login_required(login_url="institut_app:login")
@module_permission_required('int', 'view')
def ApiGetFormateurs(request):
    try:
        formateurs = Formateurs.objects.all()
        formateurs_data = []
        
        for formateur in formateurs:
            formateurs_data.append({
                'id': formateur.id,
                'nom': formateur.nom,
                'prenom': formateur.prenom,
                'telephone': formateur.telephone,
                'email': formateur.email,
                'diplome': formateur.diplome,
                'nin': formateur.nin,
                'is_particular_irg': formateur.is_particular_irg,
            })
        
        return JsonResponse(formateurs_data, safe=False)
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})


@login_required(login_url="institut_app:login")
@module_permission_required('int', 'view')
def load_module_teachers(request):
    module_id = request.GET.get('module_id')
    if module_id:
        try:
            # Get all trainers assigned to this module
            teacher_modules = EnseignantModule.objects.filter(module_id=module_id).select_related('formateur')
            teachers_data = []
            
            for tm in teacher_modules:
                formateur = tm.formateur
                teachers_data.append({
                    'id': formateur.id,
                    'nom': formateur.nom,
                    'prenom': formateur.prenom,
                    'telephone': formateur.telephone,
                    'email': formateur.email,
                    'diplome': formateur.diplome,
                })
            
            return JsonResponse(teachers_data, safe=False)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    else:
        return JsonResponse({'status': 'error', 'message': 'ID du module requis'})


@login_required(login_url="institut_app:login")
@module_permission_required('int', 'change')
@transaction.atomic
def assign_trainers_to_module(request):
    if request.method == "POST":
        try:
            module_id = request.POST.get('module_id')
            trainer_ids = request.POST.getlist('trainer_ids[]')  # For multiple trainer IDs
        
            if not module_id:
                return JsonResponse({'status': 'error', 'message': 'ID du module est requis'})
            
            if not trainer_ids:
                return JsonResponse({'status': 'error', 'message': 'Au moins un ID d\'enseignant est requis'})
               
            valid_trainer_ids = []
            for id_str in trainer_ids:
                try:
                    valid_trainer_ids.append(int(id_str))
                except ValueError:
                    continue 
            
            if not valid_trainer_ids:
                return JsonResponse({'status': 'error', 'message': 'Aucun ID d\'enseignant valide'})
            
            try:
                module = Modules.objects.get(id=module_id)
            except Modules.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Le module spécifié n\'existe pas'})
            
            formateurs = Formateurs.objects.filter(id__in=valid_trainer_ids)
            found_ids = [f.id for f in formateurs]
            missing_ids = set(valid_trainer_ids) - set(found_ids)
            
            if missing_ids:
                return JsonResponse({'status': 'error', 'message': f'Les formateurs avec les IDs {list(missing_ids)} n\'existent pas'})
            
            EnseignantModule.objects.filter(module=module).delete()
            
            assigned_count = 0
            for formateur in formateurs:
                EnseignantModule.objects.create(
                    module=module,
                    formateur=formateur
                )
                assigned_count += 1
                
            return JsonResponse({
                'status': 'success',
                'message': f'{assigned_count} enseignant(s) affecté(s) au module avec succès'
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    else:
        return JsonResponse({"status": "error", "message": "Méthode non autorisée"})

@login_required(login_url="institut_app:login")
@module_permission_required('int', 'change')
def create_availability(request):
    if request.method == "POST":
        try:
            formateur_id = request.POST.get('formateur_id')
            availabilities_json = request.POST.get('availabilities')

            if not formateur_id or not availabilities_json:
                return JsonResponse({'status': 'error', 'message': 'Tous les champs sont obligatoires'})

            # Convertir le JSON en Python
            try:
                availabilities = json.loads(availabilities_json)
            except json.JSONDecodeError:
                return JsonResponse({'status': 'error', 'message': 'Données de disponibilité invalides'})

            try:
                formateur = Formateurs.objects.get(id=formateur_id)
            except Formateurs.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Le formateur spécifié n\'existe pas'})

            # Initialiser la structure JSON si elle est vide
            if not formateur.dispo:
                formateur.dispo = {}

            if 'disponibilites' not in formateur.dispo:
                formateur.dispo['disponibilites'] = []

            # Ajouter les nouvelles disponibilités sans doublons
            for availability in availabilities:
                jour = availability.get('jour')
                heure_debut = availability.get('heure_debut')
                heure_fin = availability.get('heure_fin')

                if not all([jour, heure_debut, heure_fin]):
                    return JsonResponse({'status': 'error', 'message': 'Tous les champs de disponibilité sont obligatoires'})

                # Normaliser le jour (ex: "Lundi" -> "lundi")
                jour = jour.strip().lower()

                # Vérifier si ce créneau existe déjà
                exists = any(
                    d['jour'].lower() == jour and
                    d['heure_debut'] == heure_debut and
                    d['heure_fin'] == heure_fin
                    for d in formateur.dispo['disponibilites']
                )

                if not exists:
                    formateur.dispo['disponibilites'].append({
                        'jour': jour,
                        'heure_debut': heure_debut,
                        'heure_fin': heure_fin
                    })

            # 🔹 Ordre logique des jours
            ordre_jours = [
                "dimanche", "lundi", "mardi", "mercredi", "jeudi", "vendredi", "samedi"
            ]

            # 🔹 Tri par jour puis par heure_debut
            def sort_key(d):
                jour_index = ordre_jours.index(d['jour']) if d['jour'] in ordre_jours else 999
                try:
                    heure_obj = datetime.strptime(d['heure_debut'], "%H:%M")
                except ValueError:
                    heure_obj = datetime.strptime("00:00", "%H:%M")
                return (jour_index, heure_obj)

            formateur.dispo['disponibilites'].sort(key=sort_key)

            formateur.save()

            return JsonResponse({'status': 'success', 'message': 'Disponibilités enregistrées avec succès'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({"status": "error", "message": "Méthode non autorisée"})

@login_required(login_url="institut_app:login")
@module_permission_required('int', 'delete')
@transaction.atomic
def remove_trainer_from_module(request):
    if request.method == "POST":
        try:
            # Get module ID and trainer ID from the request
            trainer_id = request.POST.get('trainer_id')
            module_id = request.POST.get('module_id')
            
            if not module_id:
                return JsonResponse({'status': 'error', 'message': 'ID du module est requis'})
            
            if not trainer_id:
                return JsonResponse({'status': 'error', 'message': 'ID de l\'enseignant est requis'})
            
            # Validate module exists
            try:
                module = Modules.objects.get(id=module_id)
            except Modules.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Le module spécifié n\'existe pas'})
            
            # Validate trainer exists
            try:
                formateur = Formateurs.objects.get(id=trainer_id)
            except Formateurs.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'L\'enseignant spécifié n\'existe pas'})
            
            # Remove the specific association
            association = EnseignantModule.objects.filter(module=module, formateur=formateur).first()
            if association:
                association.delete()
                messages.success(request, f'L\'enseignant {formateur.nom} {formateur.prenom} a été retiré du module avec succès.')
                return JsonResponse({
                    'status': 'success',
                    'message': f'Enseignant retiré du module avec succès'
                })
            else:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Aucune association trouvée entre cet enseignant et ce module'
                })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    else:
        return JsonResponse({"status": "error", "message": "Méthode non autorisée"})


@login_required(login_url="institut_app:login")
@module_permission_required('int', 'change')
@transaction.atomic
def update_module_details(request):
    if request.method == "POST":
        try:
            # Get module ID and updated details from the request
            module_id = request.POST.get('id')
            code = request.POST.get('code')
            label = request.POST.get('label')
            duree = request.POST.get('duree')
            coef = request.POST.get('coef')
            n_elimate = request.POST.get('n_elimate')
            systeme_eval = request.POST.get('systeme_eval')
            
            if not module_id:
                return JsonResponse({'status': 'error', 'message': 'ID du module est requis'})
            
            # Validate module exists
            try:
                module = Modules.objects.get(id=module_id)
            except Modules.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Le module spécifié n\'existe pas'})
            
            # Update the module details
            if code is not None:
                module.code = code
            if label is not None:
                module.label = label
            if duree is not None:
                module.duree = duree
            if coef is not None:
                module.coef = coef
            if n_elimate is not None:
                module.n_elimate = n_elimate
            if systeme_eval is not None:
                module.systeme_eval = systeme_eval
            
            # Update the updated_by field if user is available
            if hasattr(request, 'user') and request.user.is_authenticated:
                module.updated_by = request.user
            
            module.save()
            
            messages.success(request, f'Les détails du module {module.label} ont été mis à jour avec succès.')
            return JsonResponse({
                'status': 'success',
                'message': 'Détails du module mis à jour avec succès'
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    else:
        return JsonResponse({"status": "error", "message": "Méthode non autorisée"})


@login_required(login_url="institut_app:login")
@module_permission_required('int', 'approuv')
def validate_module(request):
    if request.method == "POST":
        try:
            module_id = request.POST.get('module_id')
            
            if not module_id:
                return JsonResponse({'status': 'error', 'message': 'ID du module est requis'})
            
            # Validate module exists
            try:
                module = Modules.objects.get(id=module_id)
            except Modules.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Le module spécifié n\'existe pas'})
            
            # Update the validation status
            module.est_valider = True
            module.save()
            
            return JsonResponse({
                'status': 'success',
                'message': 'Module validé avec succès'
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    else:
        return JsonResponse({"status": "error", "message": "Méthode non autorisée"})


@login_required(login_url="institut_app:login")
@module_permission_required('int', 'view')
def get_availability(request):
    if request.method == "GET":
        try:
            formateur_id = request.GET.get('formateur_id')
            
            # Validate required field
            if not formateur_id:
                return JsonResponse({'status': 'error', 'message': 'ID du formateur est requis'})

            # Validate formateur exists
            try:
                formateur = Formateurs.objects.get(id=formateur_id)
            except Formateurs.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Le formateur spécifié n\'existe pas'})

            # Get availability data
            disponibilites = formateur.dispo.get('disponibilites', []) if formateur.dispo else []
            
            return JsonResponse({
                'status': 'success', 
                'disponibilites': disponibilites,
                'formateur_nom': formateur.nom,
                'formateur_prenom': formateur.prenom
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    else:
        return JsonResponse({"status": "error", "message": "Méthode non autorisée"})


@login_required(login_url="institut_app:login")
@module_permission_required('int', 'delete')
def delete_availability(request):
    if request.method == "POST":
        try:
            formateur_id = request.POST.get('formateur_id')
            availability_index = request.POST.get('availability_index')
            
            # Validate required fields
            if not formateur_id or availability_index is None:
                return JsonResponse({'status': 'error', 'message': 'ID du formateur et index de disponibilité sont requis'})

            # Validate formateur exists
            try:
                formateur = Formateurs.objects.get(id=formateur_id)
            except Formateurs.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Le formateur spécifié n\'existe pas'})

            # Validate index is a number
            try:
                availability_index = int(availability_index)
            except ValueError:
                return JsonResponse({'status': 'error', 'message': 'Index de disponibilité invalide'})

            # Get availability data
            if not formateur.dispo:
                formateur.dispo = {}
            
            disponibilites = formateur.dispo.get('disponibilites', [])
            
            # Check if index is valid
            if availability_index < 0 or availability_index >= len(disponibilites):
                return JsonResponse({'status': 'error', 'message': 'Index de disponibilité invalide'})

            # Remove the availability at the specified index
            disponibilites.pop(availability_index)
            
            # Update the formateur's dispo with the modified list
            formateur.dispo['disponibilites'] = disponibilites
            
            # Save the formateur with updated dispo
            formateur.save()

            return JsonResponse({'status': 'success', 'message': 'Disponibilité supprimée avec succès'})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    else:
        return JsonResponse({"status": "error", "message": "Méthode non autorisée"})


@login_required(login_url="institut_app:login")
def get_formateur_assignments(request):
    formateur_id = request.GET.get('formateur_id')
    if not formateur_id:
        return JsonResponse({'status': 'error', 'message': 'ID formateur requis'})
    try:
        formateur = Formateurs.objects.get(id=formateur_id)
        assignments = EnseignantModule.objects.filter(formateur=formateur).select_related('module', 'module__specialite')
        
        assignments_data = []
        for ass in assignments:
            has_plan = PlansCours.objects.filter(assignment=ass).exists()
            assignments_data.append({
                'assignment_id': ass.id,
                'module_code': ass.module.code if ass.module else 'N/A',
                'module_label': ass.module.label if ass.module else 'N/A',
                'specialite': ass.module.specialite.label if ass.module and ass.module.specialite else 'N/A',
                'has_plan': has_plan,
            })
            
        return JsonResponse({
            'status': 'success',
            'formateur_name': f"{formateur.nom} {formateur.prenom}",
            'assignments': assignments_data
        })
    except Formateurs.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Formateur introuvable'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@login_required(login_url="institut_app:login")
def demand_plan_cours(request):
    if request.method == "POST":
        assignment_id = request.POST.get('assignment_id')
        if not assignment_id:
            return JsonResponse({'status': 'error', 'message': 'ID d\'affectation requis'})
        try:
            assignment = EnseignantModule.objects.get(id=assignment_id)
            assignment.demande_plan_cours = True
            assignment.save()
            
            formateur = assignment.formateur
            return JsonResponse({'status': 'success', 'message': f'Demande envoyée à {formateur.nom} {formateur.prenom}.'})
        except EnseignantModule.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Affectation introuvable'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)


@login_required(login_url="institut_app:login")
def get_plan_cours_details(request):
    assignment_id = request.GET.get('assignment_id')
    if not assignment_id:
        return JsonResponse({'status': 'error', 'message': 'ID d\'affectation requis'})
    try:
        assignment = EnseignantModule.objects.get(id=assignment_id)
        plan_cours = PlansCours.objects.filter(assignment=assignment).first()
        if not plan_cours:
            return JsonResponse({'status': 'error', 'message': 'Plan de cours non rédigé.'})
        
        return JsonResponse({
            'status': 'success',
            'module_label': assignment.module.label if assignment.module else '',
            'module_code': assignment.module.code if assignment.module else '',
            'formateur_name': f"{assignment.formateur.nom} {assignment.formateur.prenom}" if assignment.formateur else '',
            'general': plan_cours.general or {},
            'deroulment': plan_cours.deroulment or [],
            'evaluation': plan_cours.evaluation or [],
            'autres': plan_cours.autres or {}
        })
    except EnseignantModule.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Affectation introuvable'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
