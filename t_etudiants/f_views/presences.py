from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from ..models import *
from django.contrib.auth.decorators import login_required
from ..forms import *
from t_crm.models import NotesProcpects, RendezVous
from django.db import transaction
from t_formations.models import *
from t_groupe.models import GroupeLine, Group
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime
import json
from django.shortcuts import get_object_or_404



@login_required(login_url="institut_app:login")
def RegistrePage(request):

    groupes = Groupe.objects.all()
    registres = RegistrePresence.objects.all()

    # Calcul des statistiques
    total_count = registres.count()
    in_progress_count = registres.filter(status='enc').count()
    archived_count = registres.filter(status='ter').count()

    # Calcul du taux moyen de présence
    # Le calcul exact dépend de la structure du modèle, mais nous pouvons estimer en fonction
    # des séances effectuées et du nombre d'absences
    total_sessions = 0
    total_absences = 0

    for registre in registres:
        # Pour chaque registre, on va compter les séances et les absences
        lignes_registre = LigneRegistrePresence.objects.filter(registre=registre)
        for ligne in lignes_registre:
            # Compter les séances effectuées pour cette ligne
            suivi_cours_list = SuiviCours.objects.filter(ligne_presence=ligne, is_done=True)
            for seance in suivi_cours_list:
                # Utiliser la méthode nombre_absents du modèle SuiviCours
                absences_dans_seance = seance.nombre_absents()
                total_absences += absences_dans_seance
                total_sessions += 1  # Une séance peut impliquer plusieurs étudiants

    # Pour calculer le taux de présence, nous avons besoin d'avoir plus d'informations
    # Pour l'instant, nous allons estimer basé sur les données disponibles
    # Si nous avons des données d'historique d'absences, comptons le total
    historiques = HistoriqueAbsence.objects.all()
    total_records = 0
    total_presences = 0

    for historique in historiques:
        for entry in historique.historique:
            for data in entry.get("data", []):
                total_records += 1
                etat = data.get("etat", "")
                if etat.upper() == "P":  # Présent
                    total_presences += 1
                # "A" pour absent, "J" pour justifié, etc.

    # Calcul du taux moyen de présence si nous avons des données d'historique
    if total_records > 0:
        average_attendance_rate = (total_presences / total_records) * 100
    else:
        # Si aucune donnée d'historique n'est disponible, nous ne pouvons pas calculer avec précision
        # Utilisons une estimation basée sur les séances effectuées
        total_students_in_registres = 0
        for registre in registres:
            groupe_lines = GroupeLine.objects.filter(groupe=registre.groupe)
            total_students_in_registres += groupe_lines.count()

        if total_sessions > 0 and total_students_in_registres > 0:
            # Estimation: nombre total d'étudiants * nombre de séances - absences
            total_possible_attendances = total_sessions * total_students_in_registres
            if total_possible_attendances > total_absences:
                average_attendance_rate = ((total_possible_attendances - total_absences) / total_possible_attendances) * 100
            else:
                average_attendance_rate = 0
        else:
            average_attendance_rate = 0

    context = {
        'groupes': groupes,
        'registres': registres,
        'total_count': total_count,
        'in_progress_count': in_progress_count,
        'archived_count': archived_count,
        'average_attendance_rate': average_attendance_rate,
    }
    return render(request, 'tenant_folder/presences/registres.html', context)


@login_required(login_url="institut_app:login")
@transaction.atomic
def ApiSaveRegistreGroupe(request):
    registerName = request.POST.get('registerName')
    semester = request.POST.get('semester')
    group = request.POST.get('group')


    try:
        RegistrePresence.objects.create(
            label = registerName,
            semestre = semester,
            groupe_id = group,
        )

        return JsonResponse({"status" : "success"})
    except:
        return JsonResponse({"status" : "error"})
    

@login_required(login_url="institut_app:login")
def DetailsRegistrePresence(request, pk):
    registre = RegistrePresence.objects.get(id= pk)

    modules = ProgrammeFormation.objects.filter(specialite = registre.groupe.specialite, semestre = registre.semestre)
    enseignants = Employees.objects.filter(is_teacher = True)
    listes = LigneRegistrePresence.objects.filter(registre = registre)

    context = {
        'registre' : registre,
        'modules' : modules,
        'enseignants' : enseignants,
        'listes' : listes,
    }

    return render(request, 'tenant_folder/presences/details_registre.html', context)

@login_required(login_url="institut_app:login")
def liste_registres(request):
    if request.method == "POST":
        module_id = request.POST.get('module_id')
        teacher_id = request.POST.get('teacher_id')
        hours = request.POST.get('hours')
        type = request.POST.get('type')
        room = request.POST.get('room')
        registre_id = request.POST.get('registre_id')

        try:
            LigneRegistrePresence.objects.create(
                module_id = module_id,
                teacher_id = teacher_id,
                hours = hours,
                type = type,
                room = room,
                registre_id = registre_id
            )
            return JsonResponse({"status" : "success"})
        
        except Exception as e:
            error_message = str(e)
        
            return JsonResponse({ "status": "error","message": "Une erreur s'est produite lors du traitement de la requête","error": error_message})
    else:
        return JsonResponse({"status" : "error",'message' : "Méthode non autorisée"})
    

@login_required(login_url="institut_app:login")
def DetailsListePresence(request, pk):
    object = LigneRegistrePresence.objects.get(id = pk)

    student = GroupeLine.objects.filter(groupe = object.registre.groupe)

    context =  {
        'etudiants' : student,
        'ligne_presence' : object,
        'pk' : pk,
    }
    return render(request, 'tenant_folder/presences/details_ligne_presence.html', context)

@login_required(login_url="institut_app:login")
def ApiLoadDatas(request):
    if request.method == "GET":
        id_ligne_presence = request.GET.get('id_ligne_presence')

        if not id_ligne_presence:
            return JsonResponse({"status" : "error","message":"Informations manquantes"})
        
        obj = LigneRegistrePresence.objects.get(id = id_ligne_presence)
        student = GroupeLine.objects.filter(groupe = obj.registre.groupe)

        liste = []
        for i in student:
            liste.append({
                "nom" : i.student.nom,
                "prenom" : i.student.prenom,
                "id_student" : i.student.id,
                "matricule_interne" : i.student.matricule_interne,
                "id" : i.id,
            })

        infos = []
        infos.append({
            "module" : obj.module.label,
            "code_module" : obj.module.code,
            "nom_formateur" : obj.teacher.nom,
            "prenom_formateur" : obj.teacher.prenom,
            
        })

        data = {
            'liste' : liste,
            'infos' : infos,
        }

        return JsonResponse(data, safe=False)


    else:
        return JsonResponse({"status" : "error"})

@login_required(login_url="institut_app:login")
@csrf_exempt
def ApiAjouterHistoriqueAbsence(request):
    if request.method == "POST":
        try:
            payload = json.loads(request.POST.get("data"))
            ligne_id = payload.get("ligne_presence")
            date_str = payload.get("date")
            is_done = payload.get("is_done", True)
            observation = payload.get("observation", "")
            records = payload.get("records", [])

            ligne = LigneRegistrePresence.objects.get(id=ligne_id)
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()

            SuiviCours.objects.create(
                is_done = is_done,
                ligne_presence_id = ligne_id,
                module =ligne.module,
                date_seance = date_obj,
                observation = observation,
            )

            if is_done:
                for record in records:
                    student_id = record.get("student_id")
                    status = record.get("status", "P")

                    etudiant = Prospets.objects.get(id=student_id)
                    historique, _ = HistoriqueAbsence.objects.get_or_create(
                        etudiant=etudiant,
                        ligne_presence=ligne
                    )

                    module_label = ligne.module.label if ligne.module else "N/A"
                    module_code  = ligne.module.code if ligne.module.code else "N/A"

                    # Appel de la méthode du modèle
                    historique.ajouter_entree(date_obj, module_label, module_code, status)

            return JsonResponse({"status": "success", "message": "Historique mis à jour avec succès"})
        
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})

    return JsonResponse({"status": "error", "message": "Méthode non autorisée"})

@login_required(login_url="institut_app:login")
def ApiGetHistoriqueEtudiant(request, pk, id_ligne):
    try:
        historique_obj = get_object_or_404(HistoriqueAbsence, etudiant_id=pk, ligne_presence_id=id_ligne)
        data = historique_obj.historique or []
        return JsonResponse({"status": "success", "historique": data})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})

@login_required(login_url="institut_app:login")
@csrf_exempt
def ApiUpdateAbsenceReason(request):
    if request.method == "POST":
        try:
            student_id = request.POST.get('student_id')
            ligne_id = request.POST.get('ligne_id')
            date_str = request.POST.get('date')
            module_name = request.POST.get('module')
            index_str = request.POST.get('index')  # The specific index in the daily data array
            new_reason = request.POST.get('reason')
            new_status = request.POST.get('status')
            
            if not new_reason or not new_reason.strip():
                return JsonResponse({"status": "error", "message": "Le motif d'absence est requis."})

            historique_obj = get_object_or_404(HistoriqueAbsence, etudiant_id=student_id, ligne_presence_id=ligne_id)
            
            # Normalize module name for comparison fallback
            module_name_clean = module_name.strip().lower() if module_name else ""
            
            updated = False
            error_details = "Date non trouvée dans l'historique."
            
            if historique_obj.historique:
                # We work on a copy to ensure mutation detection
                new_historique = list(historique_obj.historique)
                
                for entry in new_historique:
                    # Date comparison
                    if entry.get('date') == date_str:
                        error_details = f"Session non trouvée pour la date {date_str}."
                        data_array = entry.get('data', [])
                        
                        # 1. Targeted update by index (most reliable)
                        if index_str is not None:
                            try:
                                idx = int(index_str)
                                if 0 <= idx < len(data_array):
                                    item = data_array[idx]
                                    # Verify it's the right module as a safety check
                                    if str(item.get('module', '')).strip().lower() == module_name_clean:
                                        item['reason'] = new_reason
                                        item['motif'] = new_reason
                                        if new_status:
                                            item['etat'] = new_status
                                        updated = True
                            except (ValueError, IndexError):
                                pass
                        
                        # 2. Fallback to name-based search if index failed or wasn't provided
                        if not updated:
                            for item in data_array:
                                stored_module = str(item.get('module', '')).strip().lower()
                                if stored_module == module_name_clean:
                                    item['reason'] = new_reason
                                    item['motif'] = new_reason
                                    if new_status:
                                        item['etat'] = new_status
                                    updated = True
                                    break
                                    
                        if updated:
                            break
                
                if updated:
                    historique_obj.historique = new_historique
                    historique_obj.save()
                    return JsonResponse({"status": "success", "message": "Enregistrement réussi."})
            
            return JsonResponse({"status": "error", "message": f"Erreur : {error_details}"})
                
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})
            
    return JsonResponse({"status": "error", "message": "Méthode non autorisée"})


@login_required(login_url="institut_app:login")
def ListeDesEtudiants(request):
    liste = Prospets.objects.filter(statut = "convertit")
    context = {
        'etudiants' : liste
    }
    return render(request, 'tenant_folder/presences/liste_des_etudiants.html', context)


def get_attendance_data(request):
    """
    Helper pour récupérer les données d'assiduité filtrées.
    Partagé entre la vue HTML, l'export Excel et l'export PDF.
    """
    search_query = request.GET.get('search', '').lower()
    selected_promo = request.GET.get('promo', '')
    selected_semester = request.GET.get('semester', '')
    selected_group = request.GET.get('group', '')

    # Récupérer tous les étudiants convertis
    students = Prospets.objects.filter(statut="convertit")
    
    # Récupérer tous les historiques d'absence avec les relations nécessaires
    all_histories = HistoriqueAbsence.objects.filter(etudiant__in=students).select_related(
        'etudiant',
        'ligne_presence__registre__groupe__promotion'
    )
    
    # Clé : (student_id, promo_id, semester_code)
    student_period_stats = {}
    
    for history in all_histories:
        registre = history.ligne_presence.registre if history.ligne_presence else None
        if not registre: continue
            
        student = history.etudiant
        groupe = registre.groupe
        promo = groupe.promotion if groupe else None
        semestre_code = registre.semestre or '?'
        semestre_display = registre.get_semestre_display() if semestre_code != '?' else "N/A"
        promo_label = promo.label if promo else "N/A"
        
        # Filtres côte serveur (pour les exports)
        if selected_promo and promo_label != selected_promo: continue
        if selected_semester and semestre_display != selected_semester: continue
        if selected_group and (not groupe or groupe.nom != selected_group): continue
        if search_query:
            full_name = f"{student.nom} {student.prenom}".lower()
            matricule = (student.matricule_interne or "").lower()
            if search_query not in full_name and search_query not in matricule:
                continue

        promo_id = promo.id if promo else 0
        key = (student.id, promo_id, semestre_code)
        
        if key not in student_period_stats:
            student_period_stats[key] = {
                'student': student,
                'promo': promo_label,
                'semester': semestre_code,
                'semester_display': semestre_display,
                'presence': 0, 'absence': 0, 'justified': 0, 'total': 0,
                'groups': set(), 'details': []
            }
            
        data = student_period_stats[key]
        if groupe: data['groups'].add(groupe.nom)
            
        teacher_name = history.ligne_presence.teacher.nom if (history.ligne_presence and history.ligne_presence.teacher) else "N/A"
        historique_data = history.historique or []
        for entry in historique_data:
            session_date = entry.get('date')
            for item in entry.get('data', []):
                etat = str(item.get('etat', 'P')).upper()
                reason = item.get('reason') or item.get('motif')
                data['total'] += 1
                if etat == 'P': data['presence'] += 1
                elif etat == 'J' or (etat == 'A' and reason): data['justified'] += 1
                elif etat == 'A': data['absence'] += 1
                else: data['presence'] += 1
                    
                data['details'].append({
                    'date': session_date,
                    'module': item.get('module') or "Module",
                    'teacher': teacher_name,
                    'status': 'Présent' if etat == 'P' else ('Justifié' if (etat == 'J' or reason) else 'Absent'),
                    'badge': 'success' if etat == 'P' else ('info' if (etat == 'J' or reason) else 'danger')
                })
                        
    final_stats = []
    for key, data in student_period_stats.items():
        if data['total'] > 0:
            data['rate'] = round((data['presence'] / data['total']) * 100, 1)
        else: data['rate'] = 0.0
        
        try:
            data['details'].sort(key=lambda x: datetime.strptime(x['date'], "%d/%m/%Y"), reverse=True)
        except: pass
            
        data['groups'] = sorted(list(data['groups']))
        data['details_json'] = json.dumps(data['details'])
        final_stats.append(data)
        
    final_stats.sort(key=lambda x: (x['student'].nom or '', x['student'].prenom or '', x['promo'], x['semester']))
    return final_stats


@login_required(login_url="institut_app:login")
def EtatPresences(request):
    """Vue HTML principale"""
    final_stats = get_attendance_data(request)
    context = {'stats': final_stats}
    return render(request, 'tenant_folder/presences/etat_presences.html', context)


@login_required(login_url="institut_app:login")
def ExportPresencesExcel(request):
    """Génération du fichier Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    stats = get_attendance_data(request)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "État des Présences"
    
    # En-têtes
    headers = [
        "Étudiant", "Matricule", "Promotion", "Semestre", "Groupes",
        "Total Séances", "Présences", "Absences", "Justifiées", "Taux (%)"
    ]
    ws.append(headers)
    
    # Style en-tête
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Données
    for idx, item in enumerate(stats, 2):
        ws.append([
            f"{item['student'].nom} {item['student'].prenom}",
            item['student'].matricule_interne or "N/A",
            item['promo'],
            item['semester_display'],
            ", ".join(item['groups']),
            item['total'],
            item['presence'],
            item['absence'],
            item['justified'],
            f"{item['rate']}%"
        ])
    
    # Ajustement largeur colonnes
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Etat_Presences.xlsx'
    wb.save(response)
    return response


@login_required(login_url="institut_app:login")
def ExportPresencesPDF(request):
    """Génération du fichier PDF via WeasyPrint"""
    from weasyprint import HTML
    from django.template.loader import render_to_string
    from django.utils import timezone
    
    stats = get_attendance_data(request)
    
    html_string = render_to_string('tenant_folder/presences/etat_presences_pdf.html', {
        'stats': stats,
        'now': timezone.now(),
        'tenant': request.tenant
    }, request=request)
    
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    pdf = html.write_pdf()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="Etat_Presences.pdf"'
    return response