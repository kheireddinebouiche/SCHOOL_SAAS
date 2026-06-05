from institut_app.decorators import module_permission_required
from django.shortcuts import render,redirect
from django.http import JsonResponse, HttpResponse
import openpyxl
from ..models import *
from ..forms import *
from django.contrib import messages
from t_tresorerie.models import *
from t_formations.models import *
from django.db import transaction
from django.db.models import Count, Q
from django.core.exceptions import PermissionDenied
from functools import wraps
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.utils.dateformat import format


@login_required(login_url='institut_app:login')
def ApiLoadProspectPerosnalInfos(request):
    id_prospect = request.GET.get('id_prospect')
    prospect = Prospets.objects.filter(id=id_prospect).values('created_at','id','nin','nom','prenom','email','indic','telephone','type_prospect','canal','contact_situation','statut','etat','entreprise','poste_dans_entreprise','observation','has_second_wish', 'motif_annulation', 'created_by__username', 'is_double').first()
    
    if prospect:
        obj = Prospets.objects.get(id= prospect['id'])
        prospect['created_at'] = prospect['created_at'].strftime("%Y-%m-%d %H:%M")
        prospect['statut_key'] = prospect['statut'] # The internal code (e.g. 'annuler')
        prospect['statut_label'] = obj.get_statut_display()
        prospect['statut'] = obj.get_statut_display() # Consistent with ApiLoadPreinscrisPerosnalInfos
        prospect['created_by'] = prospect['created_by__username'] if prospect['created_by__username'] else "Système"
        prospect['canal_label'] = obj.get_canal_display()
        prospect['contact_situation_label'] = obj.get_contact_situation_display()

    return JsonResponse(prospect, safe=False)

@login_required(login_url='institut_app:login')
def ApiLoadProspectRendezVous(request):
   id_prospect = request.GET.get('id_prospect')
   rendez_vous = list(RendezVous.objects.filter(prospect__id=id_prospect, archived=False).values('id','date_rendez_vous','heure_rendez_vous','type','object','created_at','statut'))
   for l in rendez_vous:
       l_obj = RendezVous.objects.get(id = l['id'])
       l['status_label'] = l_obj.get_statut_display()
       l['type_label'] = l_obj.get_type_display()
       l['type_label'] = l_obj.get_type_display()
       l['created_at'] = format(l_obj.created_at, "Y-m-d H:i")
   return JsonResponse(rendez_vous, safe=False)

################################### Gestion des notes ##################################################
@login_required(login_url='institut_app:login')
def ApiLoadNote(request):
    prospect_id = request.GET.get('id_prospect')
    notes = list(NotesProcpects.objects.filter(prospect__id = prospect_id).values('id','prospect','created_by__username','created_at','note','tage'))
    for l in notes:
        l_obj = NotesProcpects.objects.get(id = l['id'])
        l['tage_label'] = l_obj.get_tage_display()
        l['created_at'] = format(l_obj.created_at, "Y-m-d H:i")
    return JsonResponse(notes, safe=False)
    
@login_required(login_url='institut_app:login')
@transaction.atomic
def ApiStoreNote(request):
    id_prospect = request.POST.get('id_prospect')
    content = request.POST.get('content')
    tags = request.POST.get('tags')

    if not content:
        return JsonResponse({'status': 'error', 'message': 'Le contenu de la note est requis.'}, status=400)

    note = NotesProcpects.objects.create(
        prospect=Prospets.objects.get(id=id_prospect),
        created_by=request.user,
        note=content,
        tage = tags,
        context = "prospect",
    )
    note.save()
    return JsonResponse({'status': 'success', 'message': 'Note enregistrée avec succès.'})

@login_required(login_url='institut_app:login')
def ApiDeleteNote(request):
    id_note = request.POST.get('id_note')
    try:
        note = NotesProcpects.objects.get(id=id_note)
        note.delete()
        return JsonResponse({'status': 'success', 'message': 'Note supprimée avec succès.'})
    except NotesProcpects.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Note non trouvée.'}, status=404)

@login_required(login_url='institut_app:login')
@transaction.atomic
def ApiUpdateNote(request):
    note_id = request.POST.get('id')
    content = request.POST.get('note')
    tage = request.POST.get('tage')

    try:
        note = NotesProcpects.objects.get(id=note_id)
        note.note = content
        note.tage = tage
        note.save()
        return JsonResponse({'status': 'success', 'message': 'Note mise à jour avec succès.'})
    except NotesProcpects.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Note introuvable.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


################################### !Gestion des notes##################################################



###################################Fiche de voeux prospect #############################################
@login_required(login_url='institut_app:login')
def ApiLoadFicheVoeuxProspect(request):
    id_prospect = request.GET.get('id_prospect')
    prospect = Prospets.objects.get(id = id_prospect)
    fiche_voeux = FicheDeVoeux.objects.filter(prospect = prospect)

    fiche_voeux_list = []
    for fiche in fiche_voeux:
        fiche_voeux_list.append({
            'id': fiche.id,
            'formation_label': fiche.specialite.formation.nom,
            'specialite_code': fiche.specialite.code,
            'specialite_label': f"{fiche.specialite.label} ({fiche.specialite.version})" if fiche.specialite.version else fiche.specialite.label,
            'specialite_id' : fiche.specialite.id,
            'specialite_id_formation': fiche.specialite.formation.id,
            'promo' : fiche.promo.get_session_display()+'-'+fiche.promo.begin_year+'/'+fiche.promo.end_year,
            'created_at' : format(fiche.created_at,"Y-m-d H:i"),
            'updated_at' : format(fiche.updated_at,"Y-m-d H:i"),
            
        })
    return JsonResponse({'fiche_voeux': fiche_voeux_list})

@login_required(login_url="institut_app:login")
def ApiLoadFicheVoeuxDoubleProspect(request):
    if request.method == "GET":
        id_prospect = request.GET.get('id_prospect')
        prospect = Prospets.objects.get(id = id_prospect)
        fiche_voeux = FicheVoeuxDouble.objects.filter(prospect = prospect)

        fiche_voeux_list = []
        for fiche in fiche_voeux:
            fiche_voeux_list.append({
                'id': fiche.id,
                'specialite1_code': fiche.specialite.specialite1.code,
                'specialite1_label': f"{fiche.specialite.specialite1.label} ({fiche.specialite.specialite1.version})" if fiche.specialite.specialite1.version else fiche.specialite.specialite1.label,
                'specialite1_version': fiche.specialite.specialite1.version,
                'formation_1' : fiche.specialite.specialite1.formation.nom,
                'formation_2' : fiche.specialite.specialite2.formation.nom,
                'specialite2_code': fiche.specialite.specialite2.code,
                'specialite2_label': f"{fiche.specialite.specialite2.label} ({fiche.specialite.specialite2.version})" if fiche.specialite.specialite2.version else fiche.specialite.specialite2.label,
                'specialite2_version': fiche.specialite.specialite2.version,
                'promo' : fiche.promo.get_session_display()+'-'+fiche.promo.begin_year+'/'+fiche.promo.end_year,
                'created_at' : fiche.created_at,
                'updated_at' : fiche.updated_at
                
            })
        return JsonResponse({'fiche_voeux': fiche_voeux_list})
    else:
        return JsonResponse({"status":"error"})


@login_required(login_url="institut_app:login")
def ApiLoadDoubleDiplomations(request):
    if request.method == "GET":
        queryset = DoubleDiplomation.objects.all()
        if request.tenant.tenant_type != 'master':
            queryset = queryset.filter(specialite1__is_visible=True, specialite2__is_visible=True)
        
        liste = queryset.values('id', 'label')
        return JsonResponse(list(liste), safe=False)
    else:
        return JsonResponse({"status":"error"})
    
@login_required(login_url="institut_app:login")
def ApiLoadDoubleSpecialite(request):
    if request.method == "GET":
        id_formation = request.GET.get('id_formation')

        if not id_formation:
            return JsonResponse({"status":"error","message":"Informations manquantes"})

        obj = DoubleDiplomation.objects.get(id = id_formation)

        data = {
            'specialite1' : f"{obj.specialite1.label} ({obj.specialite1.version})" if obj.specialite1.version else obj.specialite1.label,
            'specialite1_version' : obj.specialite1.version,
            'specialite1_formation' : obj.specialite1.formation.nom,

            'specialite2' : f"{obj.specialite2.label} ({obj.specialite2.version})" if obj.specialite2.version else obj.specialite2.label,
            'specialite2_version' : obj.specialite2.version,
            'specialite2_formation' : obj.specialite2.formation.nom,

        }

        return JsonResponse(data, safe=False)
    else:
        return JsonResponse({"status":"error"})

@login_required(login_url="institut_app:login")
def ApiLoadPromos(request):
    if request.method == "GET":
        liste = Promos.objects.all()
        data = []
        for i in liste:
            data.append({
                "id" : i.id,
                "code" : i.code,
                "start_year" : i.begin_year,
                "end_year" : i.end_year,
                "session" : i.get_session_display()
            })
        
        return JsonResponse({"data" : data},  safe=False)
        
    else:
        return JsonResponse({"status":"error"})

@login_required(login_url="institut_app")
def ApiCreateDoubleDiplomation(request):
    if request.method == "POST":
        id_prospect = request.POST.get('id_prospect')
        promo = request.POST.get('promo')
        formation = request.POST.get('formation')
        commentaires = request.POST.get('commentaires')

        if not id_prospect or not promo or not formation:
            return JsonResponse({"status":"error",'message':"Informations manquantes"})
        
        try:
            FicheVoeuxDouble.objects.create(
                prospect_id = id_prospect,
                specialite_id = formation,
                promo = Promos.objects.get(code = promo),
                commentaire = commentaires,
            )
            prospect = Prospets.objects.get(id = id_prospect)
            prospect.is_double =True
            prospect.save()

            ## Rajouter la suppression de la fiche de voeux standard dans le cas ou elle existe
            voeux_standard = FicheDeVoeux.objects.filter(prospect_id=id_prospect, promo = Promos.objects.get(code=promo), is_confirmed=False)

            if voeux_standard.exists():
                voeux_standard.delete()

            messages.success(request, "Inscription du prospect en double diplômation confirmée.")
            return JsonResponse({"status" : "success"})
        
        except Exception as e:
            return JsonResponse({"status" : "error", 'message' : str(e)})
        
    else:
        return JsonResponse({"status":"error"})

@login_required(login_url="institut_app:login")
def ApiUpdateDoubleVoeux(request):
    if request.method == "POST":
        id_formation = request.POST.get('id_formation')
        id_prospect = request.POST.get('id_prospect')
        id_voeux = request.POST.get('id_voeux')
        promo = request.POST.get('promo')
        comment = request.POST.get('comment')

        if not id_formation or not promo or not id_voeux or not id_prospect:
            return JsonResponse({"status":"error","message":"Informations manquante"})
        
        obj = FicheVoeuxDouble.objects.get(id = id_voeux, prospect_id = id_prospect)

        obj.specialite = DoubleDiplomation.objects.get(id = id_formation)
        obj.promo = Promos.objects.get(code = promo)
        obj.commentaire = comment
        obj.save()

        return JsonResponse({"status" : "success", "message":"Les modifications ont été effectuées avec succès"})



    else:
        return JsonResponse({"status":"error"})

@login_required(login_url="insitut_app:login")
@transaction.atomic
def ApiChangeToStandardCursus(request):
    if request.method == "POST":
        formation = request.POST.get('formation')
        promo = request.POST.get('promo')
        specialite = request.POST.get('specialite')
        commentaires = request.POST.get('commentaires')
        id_prospect = request.POST.get('id_prospect')

        if not formation or not promo or not specialite or not id_prospect:
            return JsonResponse({"status":"error",'message':"Informations manquante"})
        
        try:
            FicheDeVoeux.objects.create(
                prospect_id = id_prospect,
                specialite_id = specialite,
                promo = Promos.objects.get(code = promo),
                commentaire = commentaires
            )

            fiche_double = FicheVoeuxDouble.objects.get(prospect_id = id_prospect, is_confirmed=False, promo__code = promo)
            fiche_double.delete()

            prosepect = Prospets.objects.get(id = id_prospect)
            prosepect.is_double = False
            prosepect.save()

            messages.success(request, "L'inscription de l'étudiant à un cursus standard a été effectuée avec succès")
            return JsonResponse({"status":"success"})

        except Exception as e:
            return JsonResponse({"status" : "error", "message" : str(e)})

    else:
        return JsonResponse({"status":"error"})

@login_required(login_url='institut_app:login')
def ApiUpdateFicheVoeuxProspect(request):
    pass

@login_required(login_url='institut_app:login')
def ApiDeleteFicheVoeuxProspect(request):
    pass


@login_required(login_url='institut_app:login')
def ApiStoreSecondVoeuxProspect(request):
    pass

@login_required(login_url='institut_app:login')
def ApiLoadSecondVoeuxProspect(request):
    pass


###################################Fiche de voeux prospect #############################################

###################################Gestion des rappels #############################################

@login_required(login_url='institut_app:login')
@transaction.atomic
def ApiStoreRappel(request):
    type = request.POST.get('type')
    subject = request.POST.get('subject')
    date = request.POST.get('date')
    time = request.POST.get('time')
    description = request.POST.get('description')
    id_prospect = request.POST.get('id_prospect')

    if not all([type, subject, date, time, description, id_prospect]):
        return JsonResponse({'status': 'error', 'message': 'Tous les champs sont requis.'})

    rappel = RendezVous.objects.create(
        type=type,
        object=subject,
        date_rendez_vous=date,
        heure_rendez_vous=time,
        description=description,
        prospect=Prospets.objects.get(id=id_prospect),
        created_by=request.user,
        context = "prospect",
    )
    rappel.save()
    return JsonResponse({'status': 'success', 'message': 'Rappel enregistré avec succès.'})

@login_required(login_url='institut_app:login')
def ApiLoadRappel(request):
    id_prospect = request.GET.get('id_prospect')
    rappel = RendezVous.objects.filter(prospect__id=id_prospect, context="prospect", archived=False).values(
        'id', 'type','objet', 'date_rendez_vous', 'heure_rendez_vous', 'description', 'created_at','created_by'
    )
    for l in rappel:
        l_obj = RendezVous.objects.get(id = l['id'])
        l['type_label'] = l_obj.get_type_display()
        
    return JsonResponse(list(rappel), safe=False)

@login_required(login_url='institut_app:login')
@transaction.atomic
def ApiDeleteRappel(request):
    id_rappel = request.POST.get('id_rappel')
    try:
        rappel = RendezVous.objects.get(id=id_rappel)
        rappel.delete()
        return JsonResponse({'status': 'success', 'message': 'Rappel supprimé avec succès.'})
    except RendezVous.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Rappel non trouvé.'}, status=404)

@login_required(login_url='institut_app:login')
@transaction.atomic
def ApiUpdateRappel(request):
    id_rappel = request.POST.get('id_rappel')
    type = request.POST.get('type')
    subject = request.POST.get('subject')
    date = request.POST.get('date')
    time = request.POST.get('time')
    description = request.POST.get('description')

    if not all([id_rappel, type, subject, date, time, description]):
        return JsonResponse({'status': 'error', 'message': 'Tous les champs sont requis.'}, status=400)

    try:
        rappel = RendezVous.objects.get(id=id_rappel)
        rappel.type = type
        rappel.object = subject
        rappel.date_rendez_vous = date
        rappel.heure_rendez_vous = time
        rappel.description = description
        rappel.save()
        return JsonResponse({'status': 'success', 'message': 'Rappel mis à jour avec succès.'})
    except RendezVous.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Rappel non trouvé.'}, status=404)
    
@login_required(login_url='institut_app:login')
@transaction.atomic
def ApiChangeStateRappel(request):
    pass

@login_required(login_url='institut_app:login')
def ApiLoadRendezVousDetails(request):
    id_rendez_vous = request.GET.get('id_rendez_vous')
    object = RendezVous.objects.filter(id=id_rendez_vous).values('id','type','object','description','statut','date_rendez_vous','heure_rendez_vous')

    for i in object:
        i_obj = RendezVous.objects.get(id = i['id'])
        i['type_label'] = i_obj.get_type_display()
        i['statut_label'] = i_obj.get_statut_display()

    return JsonResponse(list(object), safe=False)

###################################Fiche de voeux prospect #############################################
from datetime import datetime

@login_required(login_url='institut_app:login')
@transaction.atomic
@module_permission_required('crm', 'approuv')
def ApiValidateProspect(request):
    if request.method == 'POST':
        id_prospect = request.POST.get('id_prospect')
        id_fiche_voeux = request.POST.get("id_fiche_voeux")
        try:
            prospect = Prospets.objects.get(id=id_prospect)

            # Vérification de la fiche de vœux (Uniquement si on n'est pas dans le contexte Conseil)
            if prospect.context != 'con' and not id_fiche_voeux:
                fiche_existe = FicheDeVoeux.objects.filter(prospect=prospect, is_confirmed=True).exists()
                if not fiche_existe:
                    # On vérifie aussi s'il y a une fiche non confirmée qu'on pourrait confirmer automatiquement ? 
                    # Non, on suit la règle : il faut une fiche.
                    return JsonResponse({
                        'status': 'no_voeux', 
                        'message': "Impossible de valider : ce prospect n'a pas de fiche de vœux confirmée. Veuillez en créer une avant de valider."
                    })

            prospect.etat = "accepte"
            if prospect.context == "con":
                prospect.statut = "convertit"
                prospect.is_client = True
                prospect.convertit_date = datetime.now()
            else:
                prospect.statut = "prinscrit"
                prospect.preinscri_date = datetime.now()
            prospect.motif_annulation = ""
            prospect.save()

            if id_fiche_voeux:
                FicheDeVoeux.objects.filter(id=id_fiche_voeux).update(is_confirmed=True)

            return JsonResponse({'status': 'success', 'message': 'Prospect validé avec succès.'})
        except Prospets.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Prospect non trouvé.'})
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'})

@login_required(login_url="institut_app:login")
@module_permission_required('crm', 'approuv')
def ApiValidateProspectDouble(request):
    if request.method == "POST":
        id_prospect = request.POST.get('id_prospect')
        id_fiche_voeux = request.POST.get("id_fiche_voeux")
        try:
            prospect = Prospets.objects.get(id=id_prospect)

            # Vérification de la fiche de vœux double
            if not id_fiche_voeux:
                fiche_existe = FicheVoeuxDouble.objects.filter(prospect=prospect, is_confirmed=True).exists()
                if not fiche_existe:
                    return JsonResponse({
                        'status': 'no_voeux', 
                        'message': "Impossible de valider : ce prospect n'a pas de fiche de vœux (double diplomation) confirmée. Veuillez en créer une avant de valider."
                    })

            prospect.etat = "accepte"
            prospect.statut = "prinscrit"
            prospect.motif_annulation = ""
            prospect.preinscri_date = datetime.now()
            prospect.save()


            if id_fiche_voeux:
                FicheVoeuxDouble.objects.filter(id=id_fiche_voeux).update(is_confirmed=True)

            return JsonResponse({'status': 'success', 'message': 'Prospect validé avec succès.'})
        except Prospets.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Prospect non trouvé.'})
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'})

@login_required(login_url='institut_app:login')
def ApiCheckStatutProspect(request):
    
    prospect = Prospets.objects.get(id=request.GET.get('id_prospect'))

    data = {
        'id': prospect.id,
        'etat': prospect.etat,
        'statut': prospect.statut,
        'motif_annulation': prospect.motif_annulation
    }
    return JsonResponse({'data': data})

@login_required(login_url="institut_app:login")
def ApiLoadFormationAndSpecialite(request):
    formation = Formation.objects.all()
    spec_queryset = Specialites.objects.all()
    
    if request.tenant.tenant_type != 'master':
        spec_queryset = spec_queryset.filter(is_visible=True)

    formation_liste = []
    for i in formation:
        formation_liste.append({
            'id' : i.id,
            'code' : i.code,
            'nom' : i.nom,
        })
    
    specialite_liste = []
    for i in spec_queryset:
        specialite_liste.append({
            'id' : i.id,
            'code' : i.code,
            'label' : f"{i.label} ({i.version})" if i.version else i.label,
            'version': i.version,
        })

    return JsonResponse({'formation' : formation_liste, 'specialite' : specialite_liste})

@login_required(login_url="institut_app:login")
def ApiLoadFormation(request):
    formation = Formation.objects.all()
    promos = Promos.objects.all()
    
    formation_liste = []
    for i in formation:
        formation_liste.append({
            'id' : i.id,
            'code' : i.code,
            'nom' : i.nom,
        })

    promo_liste = []
    for i in promos:
        promo_liste.append({
            "id" : i.id,
            "code" : i.code,
            'begin_year' : i.begin_year,
            'end_year' : i.end_year,
            'session' : i.get_session_display(),
        })

    return JsonResponse({'formation':formation_liste,"promo" : promo_liste})

@login_required(login_url="institut_app:login")
def ApiLoadSpecialiteProspect(request):
    id_formation = request.GET.get('id_formation')
    formation = Formation.objects.get(id = id_formation)
    queryset = Specialites.objects.filter(formation__code = formation.code)
    
    if request.tenant.tenant_type != 'master':
        queryset = queryset.filter(is_visible=True)

    data = []
    for i in queryset:
        data.append({
            "id" : i.id,
            "label" : f"{i.label} ({i.version})" if i.version else i.label,
            "version": i.version,
            "code" : i.code,
        })

    return JsonResponse({'specialite' : data}, safe=False)

@login_required(login_url="intitut_app:login")
@transaction.atomic
def ApiUpdateVoeux(request):
    id_prospect = request.POST.get('id_prospect')
    id_specialite = request.POST.get('specialite')
    id_fiche = request.POST.get('id_voeux')
    promo = request.POST.get('promo')
    
    if not id_prospect or not id_specialite or not id_fiche:
        return JsonResponse({"status" : "error", 'message': "Veuillez remplir tous les champs"})
    
    fiche = FicheDeVoeux.objects.get(id = id_fiche, prospect = Prospets.objects.get(id=id_prospect))    
    fiche.specialite = Specialites.objects.get(id = id_specialite)
    fiche.promo = Promos.objects.get(code = promo)
    fiche.save()
    return JsonResponse({'status' : "success", 'message' : "Fiche de voeux mises à jour avec succès"})


@login_required(login_url='institut_app:login')
@transaction.atomic
def ApiUpdateProspectData(request):
    nom = request.POST.get('nom')
    prenom = request.POST.get('prenom')
    email = request.POST.get('email')
    telephone = request.POST.get('telephone')
    observation = request.POST.get('observation')
    id_prospect = request.POST.get('id_prospect')
    indic = request.POST.get('indic')
    contact_situation = request.POST.get('contact_situation')
    canal = request.POST.get('canal')
    nin = request.POST.get('nin')

    prospect = Prospets.objects.get(id = id_prospect)

    prospect.nom = nom
    prospect.prenom = prenom
    prospect.email = email
    prospect.observation = observation
    prospect.telephone = telephone
    prospect.indic = indic
    prospect.contact_situation = contact_situation
    prospect.canal = canal
    prospect.nin = nin

    prospect.save()

    return JsonResponse({'status' : 'success', 'message' : "Les informations du prospect ont été mises à jour"})

@login_required(login_url="institut_app:login")
@transaction.atomic
def ApiCreateVoeux(request):
    specialite = request.POST.get('specialite')
    id_prospect = request.POST.get('id_prospect')
    promo = request.POST.get('promo')
    comment = request.POST.get('comment')

    FicheDeVoeux.objects.create(
        prospect = Prospets.objects.get(id = id_prospect),
        specialite = Specialites.objects.get(id = specialite),
        promo = Promos.objects.get(code = promo),
        commentaire = comment
    )

    return JsonResponse({"status" : "success", "message" : "La fiche de voeux a été enregistrée avec succès"})

@login_required(login_url="institut_app:login")
@module_permission_required('crm', 'approuv')
def ApiConfirmeDoubleDiplome(request):
    if request.method == "POST":
        pass
    else:
        return JsonResponse({"status":"error"})

@login_required(login_url='institut_app:login')
def ApiLoadProspectFinancialHistory(request):
    id_prospect = request.GET.get('id_prospect')
    if not id_prospect:
        return JsonResponse({'status': 'error', 'message': 'ID prospect manquant'}, status=400)
        
    try:
        prospect = Prospets.objects.get(id=id_prospect)
    except Prospets.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Prospect introuvable'}, status=404)

    # 1. Query DuePaiements (Echéances prévues)
    due_paiements = DuePaiements.objects.filter(client=prospect).order_by('date_echeance', 'id')
    due_list = []
    for dp in due_paiements:
        due_list.append({
            'id': dp.id,
            'label': dp.label or 'N/A',
            'montant_due': float(dp.montant_due) if dp.montant_due is not None else 0.0,
            'montant_restant': float(dp.montant_restant) if dp.montant_restant is not None else 0.0,
            'date_echeance': dp.date_echeance.strftime('%d/%m/%Y') if dp.date_echeance else 'N/A',
            'is_done': dp.is_done,
            'is_annulated': dp.is_annulated,
            'type': dp.get_type_display() if hasattr(dp, 'get_type_display') and dp.type else (dp.type or 'N/A'),
            'observation': dp.observation or '',
        })

    # 2. Query Paiements (Paiements effectués)
    paiements = Paiements.objects.filter(prospect=prospect).order_by('-date_paiement', '-id')
    paiements_list = []
    for p in paiements:
        paiements_list.append({
            'id': p.id,
            'num': p.num or 'N/A',
            'paiement_label': p.paiement_label or p.observation or 'Paiement',
            'montant_paye': float(p.montant_paye) if p.montant_paye is not None else 0.0,
            'date_paiement': p.date_paiement.strftime('%d/%m/%Y') if p.date_paiement else 'N/A',
            'mode_paiement': p.get_mode_paiement_display() if hasattr(p, 'get_mode_paiement_display') and p.mode_paiement else (p.mode_paiement or 'N/A'),
            'is_done': p.is_done,
            'is_refund': p.is_refund,
            'refund_id': p.refund_id.id if p.refund_id else None,
        })

    # 3. Query Remboursements
    refunds = Rembourssements.objects.filter(client=prospect).order_by('-created_at')
    refunds_list = []
    for r in refunds:
        refunds_list.append({
            'id': r.id,
            'motif_rembourssement': r.motif_rembourssement or 'N/A',
            'allowed_amount': float(r.allowed_amount) if r.allowed_amount is not None else 0.0,
            'created_at': r.created_at.strftime('%d/%m/%Y %H:%M') if r.created_at else 'N/A',
            'mode_rembourssement': r.get_mode_rembourssement_display() if hasattr(r, 'get_mode_rembourssement_display') and r.mode_rembourssement else (r.mode_rembourssement or 'N/A'),
            'etat': r.get_etat_display() if hasattr(r, 'get_etat_display') and r.etat else (r.etat or 'N/A'),
            'etat_key': r.etat,
            'is_done': r.is_done,
            'observation': r.observation or '',
        })

    return JsonResponse({
        'status': 'success',
        'due_paiements': due_list,
        'paiements': paiements_list,
        'remboursements': refunds_list
    })

@login_required(login_url='institut_app:login')
@module_permission_required('crm', 'add')
def DownloadProspectsTemplate(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modèle Prospects"

    # En-têtes (Nom, Prénom, Numéro téléphone, Email)
    headers = ["Nom", "Prénom", "Téléphone", "Email"]
    ws.append(headers)

    # Style des en-têtes
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Ajouter un exemple pour guider l'utilisateur
    ws.append(["DUPONT", "Jean", "0612345678", "jean.dupont@example.com"])

    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=modele_import_prospects.xlsx'
    wb.save(response)
    return response

@login_required(login_url='institut_app:login')
@module_permission_required('crm', 'add')
@transaction.atomic
def ApiImportProspectsParticuliers(request):
    if request.method == "POST" and request.FILES.get('file'):
        excel_file = request.FILES['file']
        
        # Validation d'extension
        if not excel_file.name.endswith('.xlsx'):
            return JsonResponse({'status': 'error', 'message': 'Seuls les fichiers .xlsx sont autorisés.'})
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return JsonResponse({'status': 'error', 'message': 'Le fichier semble vide ou ne contient que des en-têtes.'})
            
            headers = [str(h).strip().lower() for h in rows[0] if h]
            
            # Vérification des colonnes requises
            if 'nom' not in headers or 'prénom' not in headers and 'prenom' not in headers:
                return JsonResponse({'status': 'error', 'message': 'Le fichier doit contenir au moins les colonnes "Nom" et "Prénom".'})
            
            nom_idx = headers.index('nom')
            prenom_idx = headers.index('prénom') if 'prénom' in headers else headers.index('prenom')
            tel_idx = headers.index('téléphone') if 'téléphone' in headers else (headers.index('telephone') if 'telephone' in headers else -1)
            email_idx = headers.index('email') if 'email' in headers else -1
            
            imported_count = 0
            for row in rows[1:]:
                if len(row) <= max(nom_idx, prenom_idx):
                    continue
                
                nom = str(row[nom_idx] or '').strip().upper()
                prenom = str(row[prenom_idx] or '').strip().capitalize()
                
                if not nom or not prenom:
                    continue
                
                tel = str(row[tel_idx]).strip() if tel_idx != -1 and row[tel_idx] else ''
                if tel.endswith('.0'): tel = tel[:-2] # Corriger le formatage float
                
                email = str(row[email_idx]).strip() if email_idx != -1 and row[email_idx] else ''
                
                Prospets.objects.create(
                    nom=nom,
                    prenom=prenom,
                    telephone=tel,
                    email=email,
                    type_prospect='particulier',
                    context='acc',  # Sans voeux formulés pour le moment
                    created_by=request.user
                )
                imported_count += 1
                
            return JsonResponse({'status': 'success', 'message': f'{imported_count} prospects ont été importés avec succès.'})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Erreur lors de l\'importation : {str(e)}'})
            
    return JsonResponse({'status': 'error', 'message': 'Aucun fichier fourni.'})