from institut_app.decorators import module_permission_required
from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from .forms import *
from .models import *
from t_tresorerie.models import PaymentCategory, OperationsBancaire
from t_crm.models import Opportunite, Prospets, ProspectBankAccount
from django.template.loader import render_to_string
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Q
from weasyprint import HTML
import os
from django.conf import settings
from .utils import num_to_words_fr
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from institut_app.decorators import ajax_required

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'view')
def ListeThematique(request):
    context = {
        'tenant' : request.tenant,
        'tvas': TvaConseil.objects.all().order_by('valeur'),
    }
    return render(request, 'tenant_folder/conseil/liste-des-thematiques.html', context)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'view')
def ApiLoadThematique(request):
    thematique = Thematiques.objects.filter(etat  = "active").values('id', 'label', 'description', 'prix', 'created_at', 'billing_type', 'default_tva', 'categorie')
    return JsonResponse(list(thematique), safe=False)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'add')
@transaction.atomic
def ApiSaveThematique(request):
    label = request.POST.get('label')
    prix = request.POST.get('prix')
    description = request.POST.get('description')
    billing_type = request.POST.get('billing_type', 'heure')
    default_tva = request.POST.get('default_tva', 19.00)
    categorie = request.POST.get('categorie', 'service')

    thematique = Thematiques.objects.create(
        label = label,
        description = description,
        prix = prix,
        billing_type = billing_type,
        default_tva = default_tva,
        categorie = categorie
    )

    from t_crm.models import UserActionLog
    UserActionLog.objects.create(
        user=request.user,
        action_type='CREATE',
        target_model='Thematiques',
        target_id=str(thematique.id),
        details=f"Création de la thématique: {label}",
        ip_address=request.META.get('REMOTE_ADDR')
    )

    return JsonResponse({'status': 'success', 'message': 'Thématique ajoutée avec succès.'})

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'view')
def ApiLoadThematiqueDetails(request):
    id_thematique = request.GET.get('id_thematique')
    obj = Thematiques.objects.filter(id = id_thematique)
    data =[]
    for i in obj:
        data = {
            'id': i.id,
            'label': i.label,
            'description': i.description,
            'prix': i.prix,
            'default_tva': i.default_tva,
            'categorie': i.categorie
        }
    return JsonResponse(data, safe=False)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'view')
def DetailsProspectConseil(request, slug):
    prospect = get_object_or_404(Prospets, slug=slug)
    from t_crm.models import RendezVous
    rendez_vous = RendezVous.objects.filter(prospect=prospect).order_by('-date_rendez_vous', '-heure_rendez_vous')
    
    context = {
        'tenant': request.tenant,
        'prospect': prospect,
        'pk': prospect.id,
        'slug': prospect.slug,
        'rendez_vous': rendez_vous,
    }
    
    # Use dedicated template for Conseil/Executive Education
    return render(request, 'tenant_folder/conseil/prospect/details_prospect.html', context)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'view')
def ApiLoadProspectDevis(request):
    id_prospect = request.GET.get('id_prospect')
    devis = Devis.objects.filter(client_id=id_prospect).values('id', 'num_devis', 'montant', 'date_emission', 'etat')
    return JsonResponse(list(devis), safe=False)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'view')
def ApiLoadProspectFactures(request):
    id_prospect = request.GET.get('id_prospect')
    factures = Facture.objects.filter(client_id=id_prospect)
    
    data = []
    for f in factures:
        data.append({
            'id': f.id,
            'num_facture': f.num_facture,
            'date_emission': f.date_emission,
            'etat': f.etat,
            'montant': f.total_ttc()
        })
    return JsonResponse(data, safe=False)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'view')
def ApiLoadProspectFinancials(request):
    id_prospect = request.GET.get('id_prospect')
    factures = Facture.objects.filter(client_id=id_prospect)
    
    total_facture = 0
    total_paye = 0
    
    for f in factures:
        if getattr(f, 'type_facture', 'standard') == 'avoir':
            total_facture -= f.total_ttc()
            payments_sum = f.paiements.filter(is_done=True).aggregate(Sum('montant'))['montant__sum'] or 0
            total_paye -= payments_sum
        else:
            total_facture += f.total_ttc()
            # Sum payments for this facture
            payments_sum = f.paiements.filter(is_done=True).aggregate(Sum('montant'))['montant__sum'] or 0
            total_paye += payments_sum
        
    return JsonResponse({
        'total_invoiced': float(total_facture),
        'total_paid': float(total_paye),
        'balance': float(total_facture - total_paye)
    })

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'add')
def AddNewDevis(request):
    form = NewDevisForms()
    if request.method == "POST":
        form = NewDevisForms(request.POST)
        if form.is_valid():
            devis = form.save(commit=False)
            
            config = None
            if devis.entreprise:
                config = ConseilConfiguration.objects.filter(entreprise=devis.entreprise).first()
            if not config:
                config = ConseilConfiguration.objects.filter(entreprise=None).first()
            
            if config and config.default_conditions_commerciales:
                devis.conditions_commerciales = config.default_conditions_commerciales
                
            devis.save()
            
            # Create a new Opportunity for this manual Devis creation
            # Since the user explicitly creates a NEW quote, we assume a NEW opportunity unless specified otherwise.
            # (Ideally we'd let them pick an existing opportunity, but for now this fixes the "doesn't add to pipeline" issue)
            try:
                opp = Opportunite.objects.create(
                    prospect=devis.client,
                    nom=f"Devis #{devis.num_devis}",
                    stage='entrant', # Start at entrant or devis_envoye? If it's a draft, maybe just entrant or nego?
                    # Actually if it is just "Nouveau Devis" it is a draft. Let's say 'negociation' or keep 'entrant'.
                    # User complaint: "l'opportunité ne se rajoute pas".
                    budget=devis.montant or 0,
                    commercial=request.user # Assign current user as commercial?
                )
                devis.opportunite = opp
                devis.save()
            except Exception as e:
                # Log error but don't crash
                print(f"Error creating opportunity: {e}")
                
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='CREATE',
                target_model='Devis',
                target_id=str(devis.num_devis),
                details=f"Création d'un devis: {devis.num_devis} (Montant: {devis.montant})",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            messages.success(request, "Devis ajouté avec succès.")
            return redirect('t_conseil:configure-devis', pk=form.instance.num_devis)
        else:
            messages.error(request, "Erreur lors de l'ajout du devis.")
            return redirect('t_conseil:AddNewDevis')

    context = {
        'form' : form,
        'tenant' : request.tenant,
    }

    return render(request, 'tenant_folder/conseil/nouveau-devis.html', context)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'add')
def AddNewFacture(request):
    form = NewFactureForms()
    if request.method == "POST":
        form = NewFactureForms(request.POST)
        if form.is_valid():
            facture = form.save(commit=False)
            
            config = None
            if facture.entreprise:
                config = ConseilConfiguration.objects.filter(entreprise=facture.entreprise).first()
            if not config:
                config = ConseilConfiguration.objects.filter(entreprise=None).first()
            
            if config and config.default_conditions_commerciales:
                facture.conditions_commerciales = config.default_conditions_commerciales
                
            facture.save()
            
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='CREATE',
                target_model='Facture',
                target_id=str(facture.num_facture),
                details=f"Création d'une nouvelle facture: {facture.num_facture}",
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, "Facture ajoutée avec succès.")
            return redirect('t_conseil:configure-facture', pk=form.instance.num_facture)
        else:
            messages.error(request, "Erreur lors de l'ajout de la facture.")
            return redirect('t_conseil:AddNewFacture')

    context = {
        'form' : form,
        'tenant' : request.tenant,
    }

    return render(request, 'tenant_folder/conseil/nouveau-facture.html', context)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'view')
def ApiFetchEnterpriseTvas(request):
    ent_id = request.GET.get('enterprise_id')
    if not ent_id:
        return JsonResponse({'status': 'error', 'message': 'ID Entreprise manquant'}, status=400)
    
    tvas = TvaConseil.objects.all().order_by('valeur')
    data = [{'valeur': float(t.valeur), 'label': t.label, 'is_default': t.is_default} for t in tvas]
    return JsonResponse({'status': 'success', 'tvas': data})

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'change')
def configure_devis(request, pk):
    if pk is None or pk == '0':
        return redirect('t_conseil:AddNewDevis')
    else:
        try:
            devis = Devis.objects.get(num_devis=pk) 
        except Devis.DoesNotExist:
            messages.error(request, 'Devis introuvable.')
            return redirect(request.META.get('HTTP_REFERER', '/'))
        lignes_devis = devis.lignes_devis.all()

        config = None
        if devis.entreprise:
            config = ConseilConfiguration.objects.filter(entreprise=devis.entreprise).first()
        
        if not config:
            config, _ = ConseilConfiguration.objects.get_or_create(entreprise=None)
        
        tvas = TvaConseil.objects.all().order_by('valeur')
        context = {
            'tenant' : request.tenant,
            'devis' : devis,
            'lignes_devis' : lignes_devis,
            'config': config,
            'tvas': tvas,
            'enterprises': Entreprise.objects.all(),
            'can_edit': devis.etat == 'brouillon'
        }

        return render(request, 'tenant_folder/conseil/configure-devis.html', context)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'change')
def configure_facture(request, pk):
    if pk is None or pk == '0':
        return redirect('t_conseil:AddNewFacture')
    else:
        try:
            facture = Facture.objects.get(num_facture=pk) 
        except Facture.DoesNotExist:
            messages.error(request, 'Facture introuvable.')
            return redirect(request.META.get('HTTP_REFERER', '/'))
        lignes_facture = facture.lignes_facture.all()

        config = None
        if facture.entreprise:
            config = ConseilConfiguration.objects.filter(entreprise=facture.entreprise).first()
        
        if not config:
            config, _ = ConseilConfiguration.objects.get_or_create(entreprise=None)
        
        tvas = TvaConseil.objects.all().order_by('valeur')
        context = {
            'tenant' : request.tenant,
            'facture' : facture,
            'lignes_facture' : lignes_facture,
            'config': config,
            'tvas': tvas,
            'enterprises': Entreprise.objects.all(),
            'can_edit': facture.etat == 'brouillon'
        }

        return render(request, 'tenant_folder/conseil/configure-facture.html', context)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def ListeDesDevis(request):
    devis = Devis.objects.all().order_by('-created_at')
    
    # Calculate KPIs
    stats = {
        'total': devis.count(),
        'accepte': devis.filter(etat='accepte').count(),
        'attente': devis.filter(etat__in=['brouillon', 'envoye']).count(),
        'rejete': devis.filter(etat='rejete').count(),
    }
    
    # Get distinct issuing entities (entreprises)
    from institut_app.models import Entreprise
    entites = Entreprise.objects.filter(devis_entreprise__isnull=False).distinct()
    
    context = {
        "devis" : devis,
        "stats": stats,
        "entites": entites,
    }
    return render(request,'tenant_folder/conseil/liste_des_devis.html', context)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'delete')
def ArchiveThematique(request):
    context = {
        'tenant' : request.tenant
    }

    return render(request, 'tenant_folder/conseil/archive_thematique.html', context)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'view')
def ApiLoadArchivedThematique(request):
    thematique = Thematiques.objects.filter(etat = "archive").values('id', 'label', 'prix', 'created_at', 'categorie', 'billing_type')
    return JsonResponse(list(thematique), safe=False)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'delete')
@transaction.atomic
def ApiArchiveThematique(request):
    id_thematique = request.POST.get('id_thematique')
    try:
        thematique = Thematiques.objects.get(id=id_thematique) 
    except Thematiques.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Thematiques introuvable.'})
    thematique.etat = "archive"
    thematique.save()
    
    from t_crm.models import UserActionLog
    UserActionLog.objects.create(
        user=request.user,
        action_type='UPDATE',
        target_model='Thematiques',
        target_id=str(id_thematique),
        details=f"Archivage de la thématique: {thematique.label}",
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    return JsonResponse({'status': 'success', 'message': 'Thématique archivée avec succès.'})   
    
@login_required(login_url='institut_app:login')
@module_permission_required('con', 'change')
@transaction.atomic
def ApiActivateThematique(request):
    id_thematique = request.POST.get('id_thematique')
    try:
        thematique = Thematiques.objects.get(id=id_thematique) 
    except Thematiques.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Thematiques introuvable.'})
    thematique.etat = "active"
    thematique.save()
    
    from t_crm.models import UserActionLog
    UserActionLog.objects.create(
        user=request.user,
        action_type='UPDATE',
        target_model='Thematiques',
        target_id=str(id_thematique),
        details=f"Activation de la thématique: {thematique.label}",
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    return JsonResponse({'status': 'success', 'message': 'Thématique activée avec succès.'})

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'delete')
@transaction.atomic
def ApiDeleteFinalThematique(request):
    id_thematique = request.POST.get('id_thematique')
    try:
        thematique = Thematiques.objects.get(id=id_thematique) 
    except Thematiques.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Thematiques introuvable.'})
    label = thematique.label
    thematique.delete()
    
    from t_crm.models import UserActionLog
    UserActionLog.objects.create(
        user=request.user,
        action_type='DELETE',
        target_model='Thematiques',
        target_id=str(id_thematique),
        details=f"Suppression définitive de la thématique: {label}",
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    return JsonResponse({'status': 'success', 'message': 'Thématique supprimée définitivement.'})

@module_permission_required('con', 'change')
def make_prospect_client(request):
    pass

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'change')
@transaction.atomic
def ApiUpdateThematique(request):
    id_thematique = request.POST.get('id_thematique')
    label = request.POST.get('label')
    prix = request.POST.get('prix')
    description = request.POST.get('description')
    default_tva = request.POST.get('default_tva')
    categorie = request.POST.get('categorie')

    try:
        thematique = Thematiques.objects.get(id=id_thematique)
        thematique.label = label
        thematique.categorie = categorie
        
        # Handle numeric fields safely
        if prix:
            thematique.prix = str(prix).replace(',', '.')
        else:
            thematique.prix = None
            
        if default_tva:
            thematique.default_tva = str(default_tva).replace(',', '.')
        else:
            thematique.default_tva = 19.00
            
        thematique.description = description
        thematique.save()
        
        from t_crm.models import UserActionLog
        UserActionLog.objects.create(
            user=request.user,
            action_type='UPDATE',
            target_model='Thematiques',
            target_id=str(id_thematique),
            details=f"Mise à jour de la thématique: {label}",
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({'status': 'success', 'message': 'Thématique mise Ã  jour avec succès.'})
    except Thematiques.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': f'Thématique introuvable (ID: {id_thematique})'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Erreur lors de la mise Ã  jour : {str(e)}'})

import openpyxl
from django.http import HttpResponse

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def ApiDownloadThematiqueTemplate(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Thematiques"
    
    headers = ['Label', 'Description']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=thematiques_template.xlsx'
    wb.save(response)
    return response

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def ApiExportThematique(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Thematiques"
    
    headers = ['Label', 'Description']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        
    thematiques = Thematiques.objects.all()
    for t in thematiques:
        ws.append([
            t.label, 
            t.description
        ])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=export_thematiques.xlsx'
    wb.save(response)
    return response

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'add')
def ApiImportThematique(request):
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        if not excel_file.name.endswith('.xlsx'):
            return JsonResponse({'status': 'error', 'message': 'Le fichier doit être au format Excel (.xlsx)'})
            
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            imported_count = 0
            skipped_count = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                label = row[0]
                description = row[1] if len(row) > 1 else None
                
                if not label:
                    skipped_count += 1
                    continue
                    
                if Thematiques.objects.filter(label=label).exists():
                    skipped_count += 1
                    continue
                    
                Thematiques.objects.create(
                    label=str(label),
                    description=str(description) if description else None,
                    etat='active'
                )
                imported_count += 1
                
            return JsonResponse({
                'status': 'success', 
                'message': f'Importation terminée : {imported_count} thématiques ajoutées, {skipped_count} ignorées.'
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Erreur lors de l\'importation : {str(e)}'})
            
    return JsonResponse({'status': 'error', 'message': 'Requête invalide ou fichier manquant'})

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def ListeProspectConseil(request):
    context ={
        'tenant' : request.tenant,
    }
    return render(request, "tenant_folder/conseil/prospect/liste_des_prospects.html",context)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def ApiLoadProspect(request):
    pass

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def ListeDesClients(request):
    context = {
        'tenant': request.tenant,
    }
    return render(request, "tenant_folder/conseil/clients/liste_des_clients.html", context)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def DetailsClient(request, slug):
    client = get_object_or_404(Prospets, slug=slug)
    from t_crm.models import RendezVous
    rendez_vous = RendezVous.objects.filter(prospect=client).order_by('-date_rendez_vous', '-heure_rendez_vous')
    
    context = {
        'tenant': request.tenant,
        'client': client,
        'pk': client.id,
        'rendez_vous': rendez_vous,
    }
    return render(request, "tenant_folder/conseil/clients/details_client.html", context)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def ApiListeClients(request):
    liste = Prospets.objects.filter(context='con', is_client=True).values(
        'id', 'slug', 'nom', 'prenom', 'entreprise', 'email', 'telephone', 'created_at', 'convertit_date'
    )
    return JsonResponse(list(liste), safe=False)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'change')
@transaction.atomic
def ApiTransformeToClient(request):
    id_prospect = request.POST.get('id_prospect')
    try:
        prospect = Prospets.objects.get(id=id_prospect)
        prospect.is_client = True
        prospect.statut = 'convertit'
        prospect.convertit_date = timezone.now().date()
        prospect.save()
        
        from t_crm.models import UserActionLog
        UserActionLog.objects.create(
            user=request.user,
            action_type='UPDATE',
            target_model='Prospets',
            target_id=str(prospect.id),
            details=f"Conversion du prospect en client: {prospect.nom} {prospect.prenom}",
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({'status': 'success', 'message': 'Prospect converti en client.'})
    except Prospets.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Prospect non trouvé.'})

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'add')
@transaction.atomic
def ApiSaveLigneDevis(request):
    import decimal
    devis_id = request.POST.get('devis_id')
    thematique_id = request.POST.get('thematique_id')
    quantite = request.POST.get('quantite')
    description = request.POST.get('description')
    
    try:
        devis = Devis.objects.get(num_devis=devis_id) 
    except Devis.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Devis introuvable.'})
    try:
        thematique = Thematiques.objects.get(id=thematique_id) 
    except Thematiques.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Thematiques introuvable.'})
    
    try:
        qty = decimal.Decimal(quantite)
    except:
        qty = 0
        
    montant_ligne = (thematique.prix or 0) * qty

    LignesDevis.objects.create(
        devis=devis,
        thematique=thematique,
        description=description or thematique.label,
        quantite=qty,
        montant=montant_ligne 
    )
    
    # Recalculate devis total
    total = sum(l.montant for l in devis.lignes_devis.all() if l.montant)
    devis.montant = total
    devis.save()

    return JsonResponse({'status': 'success', 'message': 'Ligne ajoutée avec succès.'})

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'change')
@transaction.atomic
def ApiStartTransformationDevisToFacture(request):
    import decimal
    devis_id = request.POST.get('devis_id')
    tva_input = request.POST.get('tva', 19)
    try:
        tva = decimal.Decimal(str(tva_input).replace(',', '.'))
    except:
        tva = decimal.Decimal('19')
        
    try:
        devis = Devis.objects.get(num_devis=devis_id)
    except Devis.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Devis introuvable.'}, status=404)
    
    # Create Facture copying settings from Devis
    # Automatically accept the devis when converting to invoice
    devis.etat = 'accepte'
    devis.save()

    # Convert the associated prospect to a client if not already
    if devis.client and not devis.client.is_client:
        devis.client.is_client = True
        devis.client.statut = 'convertit'
        devis.client.convertit_date = timezone.now().date()
        devis.client.save()

    facture = Facture.objects.create(
        client=devis.client,
        entreprise=devis.entreprise,
        devis_source=devis,
        date_emission=timezone.now().date(),
        tva=tva, # Global TVA (kept for legacy/defaults, though lines manage their own)
        show_tva=devis.show_tva,
        show_remise=devis.show_remise,
        etat='brouillon',
        module_source='conseil',
        conditions_commerciales=devis.conditions_commerciales
    )
    
    for ligne in devis.lignes_devis.all():
        # Recalculate Unit Price: Montant = (UP * Qty) * (1 - Remise/100)
        # => UP = Montant / (Qty * (1 - Remise/100))
        qty = ligne.quantite if ligne.quantite and ligne.quantite > 0 else 1
        remise_val = ligne.remise_percent if ligne.remise_percent else 0
        remise_factor = decimal.Decimal(1) - (remise_val / decimal.Decimal(100))
        
        if remise_factor <= 0: remise_factor = 1 # Safety
        
        if ligne.prix_unitaire and ligne.prix_unitaire > 0:
            up = ligne.prix_unitaire
        elif ligne.montant:
            try:
                up = (ligne.montant / decimal.Decimal(qty)) / remise_factor
            except:
                up = 0
        else:
            up = 0

        LignesFacture.objects.create(
            facture=facture,
            thematique=ligne.thematique,
            specialite=ligne.specialite, # Support degree formations
            description=ligne.description,
            long_description=ligne.long_description,
            quantite=ligne.quantite,
            prix_unitaire=up,
            montant_ht=ligne.montant,
            remise_percent=ligne.remise_percent,
            tva_percent=ligne.tva_percent
        )
        
    # Update CRM Pipeline Stage
    if devis.opportunite:
        devis.opportunite.stage = 'facture'
        devis.opportunite.save()
    elif facture.client:
        # Check for related opportunity (linked to prospect) - Fallback
        opp = Opportunite.objects.filter(prospect=facture.client).order_by('-created_at').first()
        if opp:
            opp.stage = 'facture'
            opp.save()
        else:
             # Create new won opportunity?
             Opportunite.objects.create(
                 prospect=facture.client,
                 nom=f"Facture {facture.num_facture}",
                 stage='facture',
                 budget=facture.total_ttc or 0
             )
        
    # Carry over participants if any
    for part in Participant.objects.filter(devis=devis):
        Participant.objects.create(
            facture=facture,
            nom=part.nom,
            prenom=part.prenom,
            email=part.email,
            telephone=part.telephone,
            date_naissance=part.date_naissance,
            lieu_naissance=part.lieu_naissance,
            poste=part.poste,
            nin=part.nin
        )
        
    from t_crm.models import UserActionLog
    UserActionLog.objects.create(
        user=request.user,
        action_type='CREATE',
        target_model='Facture',
        target_id=str(facture.num_facture),
        details=f"Transformation du devis {devis.num_devis} en facture: {facture.num_facture}",
        ip_address=request.META.get('REMOTE_ADDR')
    )
        
    return JsonResponse({'status': 'success', 'message': 'Devis transformé en facture avec succès.', 'facture_num': facture.num_facture})

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'add')
@transaction.atomic
def ApiSaveDevisItems(request):
    import json
    import decimal
    
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
        
    try:
        data = json.loads(request.body)
        devis_id = data.get('devis_id')
        items = data.get('items', [])
        show_tva = data.get('show_tva', True)
        show_remise = data.get('show_remise', False)
        
        try:
            devis = Devis.objects.get(num_devis=devis_id) 
        except Devis.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Devis introuvable.'})
        devis.show_tva = show_tva
        devis.show_remise = show_remise
        
        # Associated enterprise if provided
        entreprise_id = data.get('entreprise_id')
        if entreprise_id:
            try:
                devis.entreprise = Entreprise.objects.get(id=entreprise_id)
            except Entreprise.DoesNotExist:
                pass
        
        # Conditions commerciales
        conditions = data.get('conditions_commerciales')
        if conditions is not None:
            devis.conditions_commerciales = conditions
        
        # Clear existing lines to replace with new ones (full sync) or append?
        # Usually full sync is safer for "Save" button unless we track IDs.
        # For simplicity, let's clear and re-create.
        from django.db import transaction
        
        with transaction.atomic():
            # Clear existing lines to replace with new ones (full sync)
            devis.lignes_devis.all().delete()
            
            total_montant = 0
            
            for item in items:
                thematique_id = item.get('thematique_id')
                specialite_id = item.get('specialite_id') # Support degree formations
                description = item.get('description')
                long_description = item.get('long_description', '')
                quantite = item.get('quantity')
                price = item.get('unitPrice')
                remise = item.get('remise_percent', 0)
                tva_l = item.get('tva_percent', 19.00)

                
                thematique = None
                if thematique_id:
                    try:
                        thematique = Thematiques.objects.get(id=thematique_id)
                    except Thematiques.DoesNotExist:
                        pass
                
                specialite = None
                if specialite_id:
                    try:
                        from t_formations.models import Specialites
                        specialite = Specialites.objects.get(id=specialite_id)
                    except Specialites.DoesNotExist:
                        pass
                
                try:
                    qty = decimal.Decimal(str(quantite))
                    unit_price = decimal.Decimal(str(price))
                    remise_percent = decimal.Decimal(str(remise))
                    tva_percent = decimal.Decimal(str(tva_l))
                except:
                    qty = 0
                    unit_price = 0
                    remise_percent = 0
                    tva_percent = 19
                    
                montant_ligne = (unit_price * qty) * (1 - (remise_percent / 100))
                total_montant += montant_ligne
                
                LignesDevis.objects.create(
                    devis=devis,
                    thematique=thematique,
                    specialite=specialite, # Support degree formations
                    description=description,
                    long_description=long_description,
                    quantite=qty,
                    prix_unitaire=unit_price,
                    montant=montant_ligne,
                    remise_percent=remise_percent,
                    tva_percent=tva_percent
                )
            
            devis.montant = total_montant
            devis.save()
            
            # Update linked Opportunity Budget
            if devis.opportunite:
                devis.opportunite.budget = total_montant
                devis.opportunite.save()
                
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='UPDATE',
                target_model='Devis',
                target_id=str(devis.num_devis),
                details=f"Mise à jour des éléments du devis: {devis.num_devis}",
                ip_address=request.META.get('REMOTE_ADDR')
            )

        return JsonResponse({'status': 'success', 'message': 'Devis enregistré avec succès.'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def DetailsDevis(request, pk):
    try:
        devis = Devis.objects.get(num_devis=pk) 
    except Devis.DoesNotExist:
        messages.error(request, 'Devis introuvable.')
        return redirect(request.META.get('HTTP_REFERER', '/'))
        
    if devis.etat == 'brouillon':
        messages.warning(request, "Veuillez valider le devis avant de pouvoir consulter ses détails.")
        return redirect('t_conseil:configure-devis', pk=devis.num_devis)
        
    lignes_devis = devis.lignes_devis.all()
    
    # Calculate totals for summary (since we have per-line TVA)
    total_ht = 0
    tva_breakdown = {}
    
    for ligne in lignes_devis:
        total_ht += ligne.montant
        if devis.show_tva:
            rate = float(ligne.tva_percent)
            amount = float(ligne.montant) * (rate / 100)
            if rate > 0:
                tva_breakdown[rate] = tva_breakdown.get(rate, 0) + amount
    
    total_tva = sum(tva_breakdown.values())
    total_ttc = float(total_ht) + total_tva
    
    # Sort breakdown for consistent display
    sorted_tva = sorted([{'rate': r, 'amount': a} for r, a in tva_breakdown.items()], key=lambda x: x['rate'], reverse=True)
    
    context = {
        'tenant' : request.tenant,
        'devis' : devis,
        'lignes_devis' : lignes_devis,
        'total_ht': total_ht,
        'total_tva': total_tva,
        'total_ttc': total_ttc,
        'tva_breakdown': sorted_tva,
        'has_facture': Facture.objects.filter(devis_source=devis).exists(),
        'participants': Participant.objects.filter(devis=devis)
    }
    return render(request, 'tenant_folder/conseil/details_devis.html', context)


@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def PrintDevisConseil(request, pk):
    from .models import Devis
    from pdf_editor.models import DocumentTemplate, DocumentGeneration
    from pdf_editor.utils import render_template_with_context
    from django.utils import timezone
    from django.contrib import messages

    try:
        devis = Devis.objects.get(num_devis=pk)
    except Devis.DoesNotExist:
        messages.error(request, 'Devis introuvable.')
        return redirect(request.META.get('HTTP_REFERER', '/'))

    # Retrieve the dolibare template
    try:
        template_obj = DocumentTemplate.objects.get(slug='dolibare', is_active=True)
    except DocumentTemplate.DoesNotExist:
        messages.error(request, "Template 'dolibare' introuvable dans l'éditeur de documents.")
        return redirect('t_conseil:DetailsDevis', pk=pk)

    lignes_devis = devis.lignes_devis.all()
    
    total_ht = 0
    total_tva = 0
    tva_breakdown = {}
    
    for ligne in lignes_devis:
        total_ht += float(ligne.montant)
        if devis.show_tva:
            rate = float(getattr(ligne, 'tva_percent', 0))
            amount = float(ligne.montant) * (rate / 100)
            if rate > 0:
                tva_breakdown[rate] = tva_breakdown.get(rate, 0) + amount

    total_tva = sum(tva_breakdown.values())
    total_ttc = total_ht + total_tva
    total_remise = 0 # Not present in Devis model

    emetteur = getattr(devis, 'entreprise', None)

    # Context mapping
    context_data = {
        'entreprise_nom': emetteur.designation if emetteur else 'SALDAE',
        'entreprise_adresse': getattr(emetteur, 'adresse', ''),
        'entreprise_telephone': getattr(emetteur, 'telephone', ''),
        'entreprise_email': getattr(emetteur, 'email', ''),
        'entreprise_rc': getattr(emetteur, 'rc', ''),
        'entreprise_nif': getattr(emetteur, 'nif', ''),
        'entreprise_nis': getattr(emetteur, 'nis', ''),
        'entreprise_art_imp': getattr(emetteur, 'art', ''), # the model uses 'art'
        'entreprise_logo': request.build_absolute_uri(emetteur.logo.url) if emetteur and hasattr(emetteur, 'logo') and emetteur.logo else '',
        
        'devis_numero': devis.num_devis,
        'date_emission': devis.date_emission.strftime("%d/%m/%Y") if devis.date_emission else "",
        'date_echeance': devis.date_echeance.strftime("%d/%m/%Y") if devis.date_echeance else "",
        'conditions_commerciales': devis.conditions_commerciales,
        
        'client_nom': str(devis.client.entreprise) if devis.client.entreprise else f"{devis.client.nom} {devis.client.prenom}",
        'client_adresse': devis.client.adresse,
        'client_telephone': devis.client.telephone,
        'client_email': devis.client.email,
        'client_rc': getattr(devis.client, 'rc', ''),
        'client_nif': getattr(devis.client, 'nif', ''),
        'client_nis': getattr(devis.client, 'nis', ''),
        'client_art_imp': getattr(devis.client, 'art_imp', ''),
        'client_logo': request.build_absolute_uri(devis.client.logo_entreprise.url) if hasattr(devis.client, 'logo_entreprise') and devis.client.logo_entreprise else '',
        
        'lignes': [
            {
                'designation': getattr(ligne.thematique, 'label', '') if hasattr(ligne, 'thematique') and ligne.thematique else getattr(ligne, 'description', ''),
                'description': getattr(ligne, 'long_description', getattr(ligne, 'description', '')),
                'quantite': float(ligne.quantite),
                'prix_unitaire': float(ligne.prix_unitaire),
                'tva_percent': float(getattr(ligne, 'tva_percent', 0)) if devis.show_tva else 0,
                'remise_percent': float(getattr(ligne, 'remise_percent', 0)) if devis.show_remise else 0,
                'montant': float(ligne.montant)
            } for ligne in lignes_devis
        ],
        'total_ht': round(total_ht, 2),
        'total_tva': round(total_tva, 2),
        'total_ttc': round(total_ttc, 2),
        'total_remise': round(total_remise, 2),
        'show_tva': devis.show_tva,
        'show_remise': devis.show_remise,
    }

    try:
        rendered_content, error = render_template_with_context(template_obj.content, context_data)
        if error:
            messages.error(request, f"Erreur de génération : {error}")
            return redirect('t_conseil:DetailsDevis', pk=pk)

        doc_gen = DocumentGeneration.objects.create(
            template=template_obj,
            context_data=context_data,
            rendered_content=rendered_content,
            generated_by=request.user
        )

        return redirect('pdf_editor:document-export', pk=doc_gen.pk)
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération PDF : {str(e)}")
        return redirect('t_conseil:DetailsDevis', pk=pk)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'approuv')
@transaction.atomic
def ApiValidateDevis(request):
    if request.method == 'POST':
        devis_id = request.POST.get('devis_id')
        try:
            devis = Devis.objects.get(num_devis=devis_id)
            devis.etat = 'envoye'
            devis.save()
            
            # Update CRM Pipeline Stage
            if devis.opportunite:
                devis.opportunite.stage = 'devis_envoye'
                devis.opportunite.save()
            elif devis.client:
                 # Fallback for legacy devis without direct link
                opp = Opportunite.objects.filter(prospect=devis.client).order_by('-created_at').first()
                if opp:
                    opp.stage = 'devis_envoye'
                    opp.save()
                    # Link it for future
                    devis.opportunite = opp
                    devis.save()
                else:
                    new_opp = Opportunite.objects.create(
                        prospect=devis.client,
                        nom=f"Devis #{devis.num_devis}",
                        stage='devis_envoye',
                        budget=devis.montant or 0
                    )
                    devis.opportunite = new_opp
                    devis.save()
                    
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='UPDATE',
                target_model='Devis',
                target_id=str(devis.num_devis),
                details=f"Validation et envoi du devis: {devis.num_devis}",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            return JsonResponse({'status': 'success', 'message': 'Devis validé avec succès.'})
        except Devis.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Devis non trouvé.'}, status=404)
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'change')
@transaction.atomic
def ApiRevertDevisToDraft(request):
    if request.method == 'POST':
        devis_id = request.POST.get('devis_id')
        try:
            devis = Devis.objects.get(num_devis=devis_id)
            # Check if has facture
            if Facture.objects.filter(devis_source=devis).exists():
                return JsonResponse({'status': 'error', 'message': 'Impossible de repasser en brouillon : une facture est déjÃ  associée Ã  ce devis.'})
            
            devis.etat = 'brouillon'
            devis.save()
            
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='UPDATE',
                target_model='Devis',
                target_id=str(devis.num_devis),
                details=f"Repassage en brouillon du devis: {devis.num_devis}",
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            return JsonResponse({'status': 'success', 'message': 'Devis repassé en brouillon.'})
        except Devis.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Devis non trouvé.'}, status=404)
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def ListeDesFactures(request):
    factures = Facture.objects.filter(module_source='conseil').exclude(type_facture='avoir').order_by('-created_at')
    
    # Calculate stats
    stats = {
        'total': factures.count(),
        'payee': factures.filter(etat='payee').count(),
        'envoye': factures.filter(etat='envoye').count(),
        'brouillon': factures.filter(etat='brouillon').count(),
    }
    
    config, _ = ConseilConfiguration.objects.get_or_create(id=1)
    tvas = TvaConseil.objects.all().order_by('valeur')
    
    # Get all enterprises for tabs
    enterprises = Entreprise.objects.all()

    context = {
        "tenant": request.tenant,
        "factures": factures,
        "stats": stats,
        "config": config,
        "tvas": tvas,
        "enterprises": enterprises,
    }
    return render(request, 'tenant_folder/conseil/liste_des_factures.html', context)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def ListeDesAvoirs(request):
    factures = Facture.objects.filter(module_source='conseil', type_facture='avoir').order_by('-created_at')
    
    # Calculate stats
    stats = {
        'total': factures.count(),
        'payee': factures.filter(etat='payee').count(),
        'envoye': factures.filter(etat='envoye').count(),
        'brouillon': factures.filter(etat='brouillon').count(),
    }
    
    config, _ = ConseilConfiguration.objects.get_or_create(id=1)
    tvas = TvaConseil.objects.all().order_by('valeur')
    
    enterprises = Entreprise.objects.all()

    context = {
        "tenant": request.tenant,
        "factures": factures,
        "stats": stats,
        "config": config,
        "tvas": tvas,
        "enterprises": enterprises,
        "is_avoir_list": True,
    }
    return render(request, 'tenant_folder/conseil/liste_des_factures.html', context)


@login_required(login_url="institut_app:login")
@module_permission_required('con', 'add')
@transaction.atomic
def ApiQuickCreateProspect(request):
    if request.method != "POST":
         return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    nom = request.POST.get('nom')
    prenom = request.POST.get('prenom')
    email = request.POST.get('email')
    telephone = request.POST.get('telephone')
    type_prospect = 'entreprise'
    entreprise_nom = request.POST.get('entreprise')
    poste = request.POST.get('poste')
    
    if not nom or not telephone or not entreprise_nom:
         return JsonResponse({'status': 'error', 'message': "Le nom, le téléphone et le nom de l'entreprise sont obligatoires."})
         
    try:
        # Check for duplicates? For now, we assume standard creation.
        
        prospect = Prospets.objects.create(
            nom=nom,
            prenom=prenom,  # Keep prenom for both types - it's the contact person name
            email=email,
            telephone=telephone,
            type_prospect=type_prospect,
            context='con', # Conseil
            indic='+213', # Default
            # Enterprise specific fields - always populated as type_prospect is entreprise
            entreprise=entreprise_nom,
            poste_dans_entreprise=poste,
        )
        
        from t_crm.models import UserActionLog
        UserActionLog.objects.create(
            user=request.user,
            action_type='CREATE',
            target_model='Prospets',
            target_id=str(prospect.id),
            details=f"Création rapide d'un prospect (Conseil): {entreprise_nom} - {nom} {prenom}",
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({
            'status': 'success', 
            'message': 'Prospect créé avec succès.',
            'prospect': {
                'id': prospect.id,
                'nom': prospect.nom,
                'prenom': prospect.prenom,
                'type': prospect.type_prospect,
                'entreprise': prospect.entreprise
            }
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@login_required(login_url="institut_app:login")
@ajax_required
@module_permission_required('con', 'delete')
@transaction.atomic
def ApiDeleteOpportunite(request):
    if request.method == "POST":
        id_opp = request.POST.get('id')
        try:
            opp = Opportunite.objects.get(id=id_opp)
            nom_opp = opp.nom
            opp.delete()
            
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='DELETE',
                target_model='Opportunite',
                target_id=str(id_opp),
                details=f"Suppression de l'opportunité: {nom_opp}",
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            return JsonResponse({'status': 'success', 'message': 'Opportunité supprimée.'})
        except Opportunite.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Opportunité non trouvée.'})
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'})

@login_required(login_url="institut_app:login")
@ajax_required
@module_permission_required('con', 'add')
@transaction.atomic
def ApiAddPaiement(request):
    if request.method == "POST":
        facture_id = request.POST.get('facture_id')
        montant = request.POST.get('montant')
        date_p = request.POST.get('date_paiement')
        mode = request.POST.get('mode_paiement')
        ref = request.POST.get('reference', '')
        note = request.POST.get('note', '')

        try:
            facture = Facture.objects.get(num_facture=facture_id)
            
            paiement = Paiement.objects.create(
                facture=facture,
                montant=montant,
                date_paiement=date_p,
                mode_paiement=mode,
                reference=ref,
                note=note,
                is_done=(mode == 'esp' or mode == 'espece')
            )

            if mode in ['vir', 'che']:
                OperationsBancaire.objects.create(
                    operation_type='entree',
                    conseil_paiement=paiement,
                    montant=montant,
                    reference_bancaire=ref,
                )

            # Update Facture Status
            total_ttc = facture.total_ttc()
            total_paye = sum(p.montant for p in facture.paiements.all())

            if total_paye >= total_ttc:
                facture.etat = 'paye'
            elif total_paye > 0:
                facture.etat = 'battente'
            
            facture.save()

            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='CREATE',
                target_model='Paiement',
                target_id=str(paiement.id),
                details=f"Ajout d'un paiement de {montant} pour la facture: {facture.num_facture}",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            return JsonResponse({
                'status': 'success', 
                'message': 'Paiement enregistré avec succès.',
                'new_status': facture.get_etat_display(),
                'total_paye': float(total_paye)
            })

        except Facture.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Facture non trouvée.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
            
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'})

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def DetailsFacture(request, pk):
    try:
        facture = Facture.objects.get(num_facture=pk) 
    except Facture.DoesNotExist:
        messages.error(request, 'Facture introuvable.')
        return redirect(request.META.get('HTTP_REFERER', '/'))
        
    if facture.etat == 'brouillon':
        messages.warning(request, "Veuillez valider la facture avant de pouvoir consulter ses détails.")
        return redirect('t_conseil:configure-facture', pk=facture.num_facture)
        
    lignes_facture = facture.lignes_facture.all()
    
    config = None
    if facture.entreprise:
        config = ConseilConfiguration.objects.filter(entreprise=facture.entreprise).first()
    
    if not config:
        config, _ = ConseilConfiguration.objects.get_or_create(entreprise=None)
        
    tvas = TvaConseil.objects.all().order_by('valeur')
    
    # Calculate totals and breakdown
    total_ht = 0
    tva_breakdown = {}
    
    for ligne in lignes_facture:
        total_ht += ligne.montant_ht
        if facture.show_tva:
            rate = float(ligne.tva_percent)
            amount = float(ligne.montant_ht) * (rate / 100)
            if rate > 0:
                tva_breakdown[rate] = tva_breakdown.get(rate, 0) + amount
            
    total_tva = sum(tva_breakdown.values())
    total_ttc = float(total_ht) + total_tva
    sorted_tva = sorted([{'rate': r, 'amount': a} for r, a in tva_breakdown.items()], key=lambda x: x['rate'], reverse=True)

    context = {
        "tenant": request.tenant,
        "facture": facture,
        "lignes_facture": lignes_facture,
        "config": config,
        "tvas": tvas,
        "total_ht": total_ht,
        "total_tva": total_tva,
        "total_ttc": total_ttc,
        "tva_breakdown": sorted_tva
    }
    return render(request, 'tenant_folder/conseil/details_facture.html', context)


@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def PrintFactureConseil(request, pk):
    from t_conseil.models import Facture
    from pdf_editor.models import DocumentTemplate, DocumentGeneration
    from pdf_editor.utils import render_template_with_context
    from django.utils import timezone
    from django.contrib import messages

    try:
        facture = Facture.objects.get(num_facture=pk)
    except Facture.DoesNotExist:
        messages.error(request, 'Facture introuvable.')
        return redirect(request.META.get('HTTP_REFERER', '/'))

    # Retrieve the dolibare_facture template
    try:
        template_obj = DocumentTemplate.objects.get(slug='dolibare_facture', is_active=True)
    except DocumentTemplate.DoesNotExist:
        messages.error(request, "Template 'dolibare_facture' introuvable dans l'éditeur de documents.")
        return redirect('t_conseil:DetailsFacture', pk=pk)

    lignes_facture = facture.lignes_facture.all()
    
    total_ht = 0
    total_tva = 0
    tva_breakdown = {}
    
    for ligne in lignes_facture:
        total_ht += float(ligne.montant_ht)
        if facture.show_tva:
            rate = float(getattr(ligne, 'tva_percent', 0))
            amount = float(ligne.montant_ht) * (rate / 100)
            if rate > 0:
                tva_breakdown[rate] = tva_breakdown.get(rate, 0) + amount

    total_tva = sum(tva_breakdown.values())
    timbre = float(facture.get_timbre()) if hasattr(facture, 'get_timbre') else 0
    total_ttc = total_ht + total_tva + timbre
    total_remise = getattr(facture, 'montant_remise_globale', 0)

    emetteur = getattr(facture, 'entreprise', None)
    
    # Retrieve configuration for bank account
    from t_conseil.models import ConseilConfiguration
    config_fin = ConseilConfiguration.objects.filter(entreprise=emetteur).first() if emetteur else ConseilConfiguration.objects.filter(entreprise=None).first()
    
    banque_nom = ''
    banque_iban = ''
    if config_fin and config_fin.compte_bancaire_defaut:
        banque_nom = config_fin.compte_bancaire_defaut.bank_name
        banque_iban = config_fin.compte_bancaire_defaut.bank_iban

    # Context mapping
    context_data = {
        'entreprise_nom': emetteur.designation if emetteur else 'SALDAE',
        'entreprise_adresse': getattr(emetteur, 'adresse', ''),
        'entreprise_telephone': getattr(emetteur, 'telephone', ''),
        'entreprise_email': getattr(emetteur, 'email', ''),
        'entreprise_rc': getattr(emetteur, 'rc', ''),
        'entreprise_nif': getattr(emetteur, 'nif', ''),
        'entreprise_nis': getattr(emetteur, 'nis', ''),
        'entreprise_art_imp': getattr(emetteur, 'art', ''), # the model uses 'art'
        'entreprise_logo': request.build_absolute_uri(emetteur.logo.url) if emetteur and hasattr(emetteur, 'logo') and emetteur.logo else '',
        
        'banque_nom': banque_nom,
        'banque_iban': banque_iban,
        
        'facture_numero': facture.num_facture,
        'date_emission': facture.date_facturation.strftime("%d/%m/%Y") if hasattr(facture, 'date_facturation') and facture.date_facturation else "",
        'date_echeance': facture.date_echeance.strftime("%d/%m/%Y") if facture.date_echeance else "",
        'conditions_commerciales': getattr(facture, 'conditions_commerciales', ''),
        'mode_paiement': facture.get_mode_paiement_display() if hasattr(facture, 'get_mode_paiement_display') else getattr(facture, 'mode_paiement', ''),
        
        'client_nom': str(facture.client.entreprise) if getattr(facture.client, 'entreprise', None) else f"{getattr(facture.client, 'nom', '')} {getattr(facture.client, 'prenom', '')}",
        'client_adresse': getattr(facture.client, 'adresse', ''),
        'client_telephone': getattr(facture.client, 'telephone', ''),
        'client_email': getattr(facture.client, 'email', ''),
        'client_rc': getattr(facture.client, 'rc', ''),
        'client_nif': getattr(facture.client, 'nif', ''),
        'client_nis': getattr(facture.client, 'nis', ''),
        'client_art_imp': getattr(facture.client, 'art_imp', ''),
        'client_logo': request.build_absolute_uri(facture.client.logo_entreprise.url) if hasattr(facture.client, 'logo_entreprise') and getattr(facture.client, 'logo_entreprise') else '',
        
        'lignes': [
            {
                'designation': getattr(ligne.thematique, 'label', '') if hasattr(ligne, 'thematique') and ligne.thematique else getattr(ligne, 'description', ''),
                'description': getattr(ligne, 'long_description', getattr(ligne, 'description', '')),
                'quantite': float(ligne.quantite),
                'prix_unitaire': float(ligne.prix_unitaire_ht) if hasattr(ligne, 'prix_unitaire_ht') else float(getattr(ligne, 'prix_unitaire', 0)),
                'tva_percent': float(getattr(ligne, 'tva_percent', 0)) if facture.show_tva else 0,
                'remise_percent': float(getattr(ligne, 'remise_percent', 0)) if getattr(facture, 'show_remise', False) else 0,
                'montant': float(ligne.montant_ht) if hasattr(ligne, 'montant_ht') else float(getattr(ligne, 'montant', 0))
            } for ligne in lignes_facture
        ],
        'total_ht': round(total_ht, 2),
        'total_tva': round(total_tva, 2),
        'total_ttc': round(total_ttc, 2),
        'total_remise': round(total_remise, 2),
        'timbre': round(timbre, 2),
        'show_tva': facture.show_tva,
        'show_remise': getattr(facture, 'show_remise', False),
    }

    try:
        rendered_content, error = render_template_with_context(template_obj.content, context_data)
        if error:
            messages.error(request, f"Erreur de génération : {error}")
            return redirect('t_conseil:DetailsFacture', pk=pk)

        doc_gen = DocumentGeneration.objects.create(
            template=template_obj,
            context_data=context_data,
            rendered_content=rendered_content,
            generated_by=request.user
        )

        return redirect('pdf_editor:document-export', pk=doc_gen.pk)
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération PDF : {str(e)}")
        return redirect('t_conseil:DetailsFacture', pk=pk)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def ConseilDashboard(request):
    from django.db.models import Count, Sum, Q
    from t_crm.models import Prospets, RendezVous
    from t_conseil.models import Opportunite, Devis, Facture, GroupeConseil, Participant
    from django.utils import timezone
    import json
    
    now = timezone.now()
    first_day_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # 1. CRM & Pipeline Stats
    prospects = Prospets.objects.filter(context='con')
    total_prospects = prospects.filter(is_client=False).count()
    new_prospects_this_month = prospects.filter(is_client=False, created_at__gte=first_day_of_month).count()
    
    pipeline_stats = Opportunite.objects.filter(prospect__context='con').values('stage').annotate(count=Count('id'))
    pipeline_dict = {item['stage']: item['count'] for item in pipeline_stats}
    
    # 2. Ventes (Devis & Factures)
    total_clients = prospects.filter(is_client=True).count()
    devis_stats = Devis.objects.aggregate(
        total_attente=Count('id', filter=Q(etat='attente')),
        montant_attente=Sum('montant', filter=Q(etat='attente')),
        total_accepte=Count('id', filter=Q(etat='accepte')),
        total_refuse=Count('id', filter=Q(etat='refuse')),
    )
    
    # Chiffre d'affaires (Factures validées)
    factures = Facture.objects.filter(type_facture='standard', etat__in=['brouillon', 'valide', 'annule']) 
    ca_global = Facture.objects.filter(type_facture='standard', etat='valide').aggregate(ca=Sum('lignes_facture__montant_ht'))['ca'] or 0
    factures_attente = Facture.objects.filter(type_facture='standard', etat='brouillon').count()
    
    # 3. Formation (Groupes & Participants)
    groupes_stats = GroupeConseil.objects.aggregate(
        en_cours=Count('id', filter=Q(etat='enc')),
        brouillon=Count('id', filter=Q(etat='brouillon')),
        cloture=Count('id', filter=Q(etat='cloture')),
    )
    
    # Total participants uniques
    total_participants = Participant.objects.count()
    
    # JSON for Charts
    stages = ['entrant', 'contacte', 'negociation', 'devis_envoye', 'facture', 'recouvrement']
    stages_labels = ['Entrant', 'Contacté', 'Négociation', 'Devis envoyé', 'Facturé', 'Recouvrement']
    pipeline_data = [pipeline_dict.get(stage, 0) for stage in stages]

    # Prochains Rendez-vous
    upcoming_rendez_vous = RendezVous.objects.filter(context='con', date_rendez_vous__gte=now.date()).order_by('date_rendez_vous', 'heure_rendez_vous')[:5]
    
    context = {
        'tenant': request.tenant,
        'page_title': 'Tableau de Bord - Executive Education',
        'crm': {
            'total_prospects': total_prospects,
            'new_prospects': new_prospects_this_month,
            'total_clients': total_clients,
        },
        'ventes': {
            'devis_attente': devis_stats['total_attente'] or 0,
            'devis_attente_montant': devis_stats['montant_attente'] or 0,
            'devis_accepte': devis_stats['total_accepte'] or 0,
            'devis_refuse': devis_stats['total_refuse'] or 0,
            'ca_global': ca_global,
            'factures_attente': factures_attente,
        },
        'formation': {
            'groupes_en_cours': groupes_stats['en_cours'] or 0,
            'groupes_brouillon': groupes_stats['brouillon'] or 0,
            'groupes_cloture': groupes_stats['cloture'] or 0,
            'total_participants': total_participants,
        },
        'charts': {
            'pipeline_labels': json.dumps(stages_labels),
            'pipeline_data': json.dumps(pipeline_data),
        },
        'upcoming_rendez_vous': upcoming_rendez_vous,
    }
    return render(request, 'tenant_folder/conseil/dashboard.html', context)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def PipelineConseil(request):
    from django.db.models import Sum, Q
    from django.db.models.functions import TruncMonth
    
    stages = [
        ('entrant', 'Entrant'),
        ('contacte', 'Contacté'),
        ('negociation', 'Négociation'),
        ('devis_envoye', 'Devis envoyé'),
        ('facture', 'Facturé'),
        ('recouvrement', 'Recouvrement'),
    ]
    
    view_type = request.GET.get('view', 'stage')  # 'stage' or 'closing'
    search_query = request.GET.get('search', '')
    commercial_id = request.GET.get('commercial', '')
    state_filter = request.GET.get('state', 'active') # 'active', 'inactive', 'all'
    
    prospects = Opportunite.objects.filter(prospect__context='con')
    
    # Apply Filters
    if search_query:
        prospects = prospects.filter(Q(prospect__nom__icontains=search_query) | Q(prospect__prenom__icontains=search_query) | Q(prospect__entreprise__icontains=search_query) | Q(nom__icontains=search_query))
    if commercial_id:
        prospects = prospects.filter(commercial_id=commercial_id)
    if state_filter == 'active':
        prospects = prospects.filter(is_active=True)
    elif state_filter == 'inactive':
        prospects = prospects.filter(is_active=False)

    pipeline_data = []
    
    if view_type == 'closing':
        # Group by closing month
        closing_months = prospects.filter(closing_date__isnull=False).annotate(month=TruncMonth('closing_date')).values('month').distinct().order_by('month')
        for m in closing_months:
            month_date = m['month']
            items = prospects.filter(closing_date__year=month_date.year, closing_date__month=month_date.month)
            total_budget = items.aggregate(total=Sum('budget'))['total'] or 0
            pipeline_data.append({
                'code': month_date.strftime('%Y-%m'),
                'label': month_date.strftime('%B %Y'),
                'items': items.order_by('closing_date'),
                'total_budget': total_budget,
                'count': items.count()
            })
        # Add a "Sans date" column if exists
        no_date_items = prospects.filter(closing_date__isnull=True)
        if no_date_items.exists():
             pipeline_data.append({
                'code': 'no-date',
                'label': 'Sans Date',
                'items': no_date_items.order_by('-updated_at'),
                'total_budget': no_date_items.aggregate(total=Sum('budget'))['total'] or 0,
                'count': no_date_items.count()
            })
    else:
        # Default Stage View
        for stage_code, stage_label in stages:
            stage_items = prospects.filter(stage=stage_code).order_by('-updated_at')
            total_budget = stage_items.aggregate(total=Sum('budget'))['total'] or 0
            pipeline_data.append({
                'code': stage_code,
                'label': stage_label,
                'items': stage_items,
                'total_budget': total_budget,
                'count': stage_items.count()
            })
    
    # Get all commercials for filtered list
    commercials = User.objects.filter(opportunites_conseil__isnull=False).distinct()
    
    # Get all active prospects for "New Opportunity" dropdown
    all_prospects = Prospets.objects.filter(context='con', type_prospect='entreprise', conseil_is_active=True).order_by('nom')

    context = {
        'tenant': request.tenant,
        'pipeline_stages': pipeline_data,
        'view_type': view_type,
        'commercials': commercials,
        'all_prospects': all_prospects,
        'current_filters': {
            'search': search_query,
            'commercial': commercial_id,
            'state': state_filter
        }
    }
    return render(request, 'tenant_folder/conseil/pipeline.html', context)

@login_required(login_url="institut_app:login")
@ajax_required
@module_permission_required('con', 'change')
@transaction.atomic
def ApiToggleFavorite(request):
    prospect_id = request.POST.get('prospect_id')
    try:
        prospect = Prospets.objects.get(id=prospect_id)
        prospect.conseil_is_favorite = not prospect.conseil_is_favorite
        prospect.save()
        
        from t_crm.models import UserActionLog
        UserActionLog.objects.create(
            user=request.user,
            action_type='UPDATE',
            target_model='Prospets',
            target_id=str(prospect.id),
            details=f"Modification favori pipeline: {prospect.nom} (Favori: {prospect.conseil_is_favorite})",
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({'status': 'success', 'is_favorite': prospect.conseil_is_favorite})
    except Prospets.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Prospect non trouvé.'})

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def ExportPipelineCsv(request):
    import csv
    from django.http import HttpResponse
    
    prospects = Prospets.objects.filter(context='con', conseil_is_active=True)
    # Re-apply filters from session or GET if needed, simplicity for now:
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="pipeline_conseil.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Nom', 'Prénom', 'Entreprise', 'Étape', 'Budget', 'Probabilité', 'Date Closing'])
    
    for p in prospects:
        writer.writerow([p.nom, p.prenom, p.entreprise, p.get_conseil_pipeline_stage_display(), p.conseil_budget, p.conseil_probability, p.conseil_closing_date])
        
    return response

@login_required(login_url="institut_app:login")
@ajax_required
@module_permission_required('con', 'change')
@transaction.atomic
def ApiConvertProspectToDevis(request):
    from django.urls import reverse
    # Note: 'prospect_id' parameter actually comes as opportunite_id from the new pipeline
    opp_id = request.POST.get('prospect_id') 
    try:
        opp = Opportunite.objects.get(id=opp_id)
        prospect = opp.prospect
        
        # Assign the first enterprise by default for the devis
        # In a multi-tenant set, Entreprise.objects.first() should be the correct one.
        entreprise = Entreprise.objects.first()
        
        config = None
        if entreprise:
            config = ConseilConfiguration.objects.filter(entreprise=entreprise).first()
        if not config:
            config = ConseilConfiguration.objects.filter(entreprise=None).first()
            
        default_conditions = config.default_conditions_commerciales if config else ""

        new_devis = Devis.objects.create(
            client=prospect,
            opportunite=opp, # Link the devis to the opportunity
            date_emission=timezone.now().date(),
            entreprise=entreprise,
            etat='brouillon',
            conditions_commerciales=default_conditions
        )
        
        # Update Opportunity Stage
        opp.stage = 'negociation'
        opp.save()
        
        redirect_url = reverse('t_conseil:configure-devis', kwargs={'pk': new_devis.num_devis})
        
        return JsonResponse({
            'status': 'success', 
            'message': 'Opportunité convertie en devis avec succès.',
            'redirect_url': redirect_url
        })
    except Opportunite.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Opportunité non trouvée.'})

@login_required(login_url="institut_app:login")
@ajax_required
@module_permission_required('con', 'delete')
@transaction.atomic
def ApiDeleteDevis(request):
    num_devis = request.POST.get('num_devis')
    try:
        devis = Devis.objects.get(num_devis=num_devis)
        devis.delete()
        
        from t_crm.models import UserActionLog
        UserActionLog.objects.create(
            user=request.user,
            action_type='DELETE',
            target_model='Devis',
            target_id=str(num_devis),
            details=f"Suppression du devis: {num_devis}",
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({'status': 'success', 'message': 'Devis supprimé avec succès.'})
    except Devis.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Devis non trouvé.'})

@login_required(login_url="institut_app:login")
@login_required(login_url="institut_app:login")
@ajax_required
@module_permission_required('con', 'change')
@transaction.atomic
def ApiUpdatePipelineStage(request):
    if request.method == "POST":
        # 'prospect_id' coming from frontend is actually Opportunite ID now
        opp_id = request.POST.get('prospect_id') 
        new_stage = request.POST.get('new_stage')
        
        try:
            opp = Opportunite.objects.get(id=opp_id)
            opp.stage = new_stage
            opp.save()
            
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='UPDATE',
                target_model='Opportunite',
                target_id=str(opp.id),
                details=f"Mise à jour du statut de l'opportunité {opp.nom} vers {new_stage}",
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            return JsonResponse({'status': 'success', 'message': 'Statut mis à jour.'})
        except Opportunite.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Opportunité non trouvée.'})
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'})

@login_required(login_url="institut_app:login")
@ajax_required
@module_permission_required('con', 'add')
@transaction.atomic
def ApiCreateOpportunite(request):
    if request.method == "POST":
        prospect_id = request.POST.get('prospect_id')
        nom = request.POST.get('nom')
        budget = request.POST.get('budget', 0)
        
        if not prospect_id or not nom:
            return JsonResponse({'status': 'error', 'message': 'Prospect et Nom sont obligatoires.'})
            
        try:
            prospect = Prospets.objects.get(id=prospect_id)
            
            opp = Opportunite.objects.create(
                prospect=prospect,
                nom=nom,
                budget=budget,
                stage='entrant',
                commercial=request.user
            )
            
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='CREATE',
                target_model='Opportunite',
                target_id=str(opp.id),
                details=f"Création d'une opportunité: {nom} pour {prospect.entreprise}",
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            # Get initials
            initials = ""
            if prospect.nom: initials += prospect.nom[0].upper()
            if prospect.prenom: initials += prospect.prenom[0].upper()

            return JsonResponse({
                'status': 'success', 
                'message': 'Opportunité créée avec succès.', 
                'id': opp.id,
                'slug': prospect.slug,
                'initials': initials,
                'nom_complet': f"{prospect.nom} {prospect.prenom or ''}".strip(),
                'entreprise': prospect.entreprise or ''
            })
            
        except Prospets.DoesNotExist:
             return JsonResponse({'status': 'error', 'message': 'Prospect non trouvé.'})
             
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'})

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def ApiGetOpportunite(request):
    opp_id = request.GET.get('id')
    try:
        opp = Opportunite.objects.get(id=opp_id)
        return JsonResponse({
            'status': 'success',
            'data': {
                'id': opp.id,
                'nom': opp.nom,
                'budget': opp.budget,
                'probability': opp.probability,
                'closing_date': opp.closing_date.strftime('%Y-%m-%d') if opp.closing_date else '',
                'stage': opp.stage,
                'prospect_name': f"{opp.prospect.nom} {opp.prospect.prenom or ''}".strip()
            }
        })
    except Opportunite.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Opportunité non trouvée.'})

@login_required(login_url="institut_app:login")
@ajax_required
@module_permission_required('con', 'change')
@transaction.atomic
def ApiUpdateOpportunite(request):
    if request.method == "POST":
        opp_id = request.POST.get('id')
        nom = request.POST.get('nom')
        budget = request.POST.get('budget')
        probability = request.POST.get('probability')
        closing_date = request.POST.get('closing_date')
        
        try:
            opp = Opportunite.objects.get(id=opp_id)
            opp.nom = nom
            opp.budget = budget or 0
            opp.probability = probability or 0
            if closing_date:
                opp.closing_date = closing_date
            else:
                opp.closing_date = None
            
            opp.save()
            
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='UPDATE',
                target_model='Opportunite',
                target_id=str(opp.id),
                details=f"Mise à jour de l'opportunité: {nom}",
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            return JsonResponse({'status': 'success', 'message': 'Opportunité mise à jour.'})
        except Opportunite.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Opportunité non trouvée.'})
            
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'})

@module_permission_required('con', 'add')
def ConfigurationConseil(request):
    enterprise_id = request.GET.get('enterprise_id') or request.POST.get('enterprise_id')
    enterprise = None
    
    # Si enterprise_id est fourni mais vide (sélection "Configuration Globale"), enterprise reste None.
    if enterprise_id:
        try:
            enterprise = Entreprise.objects.get(id=enterprise_id)
        except Entreprise.DoesNotExist:
            pass
    elif 'enterprise_id' not in request.GET and 'enterprise_id' not in request.POST:
        # Par défaut au premier chargement
        enterprise = Entreprise.objects.first()

    config, created = ConseilConfiguration.objects.get_or_create(entreprise=enterprise)
    global_config, _ = ConseilConfiguration.objects.get_or_create(entreprise=None)
    
    # Global TVAs
    tvas = TvaConseil.objects.all().order_by('valeur')
    enterprises = Entreprise.objects.all()
    
    if request.method == "POST":
        action = request.POST.get('action')
        
        if action == 'save_config' or not action: # Default fallback if action missing
            try:
                global_config.default_tva_percent = request.POST.get('default_tva_percent') or 19.00
                global_config.show_tva_on_devis = request.POST.get('show_tva_on_devis') == 'on'
                global_config.show_tva_on_facture = request.POST.get('show_tva_on_facture') == 'on'
                global_config.save()
                
                config.enable_remise_global = request.POST.get('enable_remise_global') == 'on'
                config.default_remise_percent = request.POST.get('default_remise_percent') or 0.00
                config.show_remise_on_devis = request.POST.get('show_remise_on_devis') == 'on'
                config.show_remise_on_facture = request.POST.get('show_remise_on_facture') == 'on'
                
                # Numbering Configuration
                config.devis_prefix = request.POST.get('devis_prefix', 'DEV')
                config.devis_counter_width = request.POST.get('devis_counter_width') or 4
                config.facture_prefix = request.POST.get('facture_prefix', 'FAC')
                config.facture_counter_width = request.POST.get('facture_counter_width') or 4
                
                config.default_conditions_commerciales = request.POST.get('default_conditions_commerciales', '')
                config.payment_methods = request.POST.get('payment_methods', '')
                
                compte_bancaire_id = request.POST.get('compte_bancaire_defaut')
                if compte_bancaire_id:
                    from institut_app.models import BankAccount
                    try:
                        config.compte_bancaire_defaut = BankAccount.objects.get(id=compte_bancaire_id)
                    except BankAccount.DoesNotExist:
                        pass
                else:
                    config.compte_bancaire_defaut = None
                
                config.save()
                messages.success(request, "Configuration mise à jour avec succès.")
            except Exception as e:
                messages.error(request, f"Erreur lors de la mise à jour : {e}")
                
        elif action == 'add_tva':
            label = request.POST.get('tva_label')
            valeur = request.POST.get('tva_valeur')
            if label and valeur:
                try:
                    TvaConseil.objects.create(label=label, valeur=valeur)
                    messages.success(request, f"TVA '{label}' ajoutée avec succès.")
                except Exception as e:
                    messages.error(request, f"Erreur : {e}")
            else:
                messages.warning(request, "Veuillez remplir le libellé et la valeur.")
                
        elif action == 'delete_tva':
            tva_id = request.POST.get('tva_id')
            if tva_id:
                TvaConseil.objects.filter(id=tva_id).delete()
                messages.success(request, "TVA supprimée.")
            
        return redirect('t_conseil:ConfigurationConseil')

    from institut_app.models import BankAccount
    comptes_bancaires = BankAccount.objects.filter(is_archived=False)
    if enterprise:
        comptes_bancaires = comptes_bancaires.filter(entreprise=enterprise)
    
    context = {
        "tenant": request.tenant,
        'config': config,
        'global_config': global_config,
        'tvas': tvas,
        'enterprises': enterprises,
        'current_enterprise': enterprise,
        'comptes_bancaires': comptes_bancaires,
    }
    return render(request, 'tenant_folder/conseil/configuration.html', context)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def DetailsFacture(request, pk):
    facture = get_object_or_404(Facture, num_facture=pk)
    
    if facture.etat == 'brouillon':
        messages.warning(request, "Veuillez valider la facture avant de pouvoir consulter ses détails.")
        return redirect('t_conseil:configure-facture', pk=facture.num_facture)
        
    lignes_facture = facture.lignes_facture.all()
    config = ConseilConfiguration.objects.filter(entreprise=facture.entreprise).first() if facture.entreprise else ConseilConfiguration.objects.filter(entreprise=None).first()
    
    total_ht = 0
    tva_breakdown = {}
    for l in lignes_facture:
        total_ht += float(l.montant_ht)
        if facture.show_tva:
            rate = float(l.tva_percent)
            amt = float(l.montant_ht) * (rate / 100)
            tva_breakdown[rate] = tva_breakdown.get(rate, 0) + amt
            
    total_tva = sum(tva_breakdown.values())
    
    # Calculate Timbre
    timbre = float(facture.get_timbre())
    total_ttc = total_ht + total_tva + timbre
    
    sorted_tva = sorted([{'rate': r, 'amount': a} for r, a in tva_breakdown.items()], key=lambda x: x['rate'], reverse=True)
    
    # Amount in words
    total_in_words = ""
    try:
        total_in_words = num_to_words_fr(total_ttc)
    except Exception as e:
        print(f"Error converting: {e}")

    paiements_lies = facture.tresorerie_paiements.all().order_by('-date_paiement')

    context = {
        "tenant": request.tenant,
        "facture": facture,
        "lignes_facture": lignes_facture,
        "total_ht": total_ht,
        "total_tva": total_tva,
        "timbre": timbre,
        "total_ttc": total_ttc,
        "tva_breakdown": sorted_tva,
        "total_in_words": total_in_words,
        "config": config,
        "paiements_lies": paiements_lies,
    }
    return render(request, 'tenant_folder/conseil/details_facture.html', context)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'add')
def AddNewFacture(request):
    if request.method == "POST":
        form = NewFactureForms(request.POST)
        if form.is_valid():
            facture = form.save(commit=False)
            
            config = None
            if facture.entreprise:
                config = ConseilConfiguration.objects.filter(entreprise=facture.entreprise).first()
            if not config:
                config = ConseilConfiguration.objects.filter(entreprise=None).first()
            
            if config and config.default_conditions_commerciales:
                facture.conditions_commerciales = config.default_conditions_commerciales
                
            facture.save()
            
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='CREATE',
                target_model='Facture',
                target_id=str(facture.num_facture),
                details=f"Création d'une nouvelle facture: {facture.num_facture}",
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            return redirect('t_conseil:configure-facture', pk=facture.num_facture)
    else:
        form = NewFactureForms()
    
    context = {
        "tenant": request.tenant,
        "form": form,
        "prospects": Prospets.objects.filter(context='con'),
        "enterprises": Entreprise.objects.all(),
    }
    return render(request, 'tenant_folder/conseil/nouveau-facture.html', context)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'change')
def configure_facture(request, pk):
    facture = get_object_or_404(Facture, num_facture=pk)
    lignes_facture = facture.lignes_facture.all()
    
    # Get configuration for specific enterprise if set, otherwise first or default
    config = None
    if facture.entreprise:
        config = ConseilConfiguration.objects.filter(entreprise=facture.entreprise).first()
    
    if not config:
        config, _ = ConseilConfiguration.objects.get_or_create(entreprise=None)
        
    tvas = TvaConseil.objects.all().order_by('valeur')
        
    # Check if we can edit
    can_edit = (facture.etat == 'brouillon')
    
    from t_tresorerie.models import ParametreFinancier
    fin_config = ParametreFinancier.get_instance()
    
    context = {
        "tenant": request.tenant,
        "facture": facture,
        "lignes_facture": lignes_facture,
        "config": config,
        "fin_config": fin_config,
        "tvas": tvas,
        "enterprises": Entreprise.objects.all(),
        "can_edit": can_edit
    }
    return render(request, 'tenant_folder/conseil/configure-facture.html', context)

@login_required(login_url='institut_app:login')
@ajax_required
@module_permission_required('con', 'add')
@transaction.atomic
def ApiSaveFactureItems(request):
    import json
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)
        
    try:
        data = json.loads(request.body)
        facture_id = data.get('facture_id')
        items_data = data.get('items', [])
        show_tva = data.get('show_tva', False)
        show_remise = data.get('show_remise', False)
        entreprise_id = data.get('entreprise_id')
        conditions = data.get('conditions_commerciales', '')
        mode_paiement = data.get('mode_paiement', 'virement')
        date_emission = data.get('date_emission')
        date_echeance = data.get('date_echeance')
        
        try:
            facture = Facture.objects.get(num_facture=facture_id) 
        except Facture.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Facture introuvable.'})
        
        if facture.etat != 'brouillon':
            return JsonResponse({'status': 'error', 'message': 'Modification impossible : la facture n\'est plus en brouillon.'})

        # Update global fields
        facture.show_tva = show_tva
        facture.show_remise = show_remise
        facture.conditions_commerciales = conditions
        facture.mode_paiement = mode_paiement
        
        if date_emission:
            facture.date_emission = date_emission
        if date_echeance:
            facture.date_echeance = date_echeance
            
        if entreprise_id:
            try:
                facture.entreprise = Entreprise.objects.get(id=entreprise_id) 
            except Entreprise.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Entreprise introuvable.'})
        facture.save()
        
        # Clear existing items and recreate
        facture.lignes_facture.all().delete()
        
        total_ttc = 0
        total_ht = 0
        
        for item in items_data:
            t_id = item.get('thematique_id')
            description = item.get('description', '')
            long_description = item.get('long_description', '')
            qty = float(item.get('quantity', 0))
            unit_price = float(item.get('unitPrice', 0))
            remise_percent = float(item.get('remise_percent', 0))
            tva_percent = float(item.get('tva_percent', 0))
            
            thematique = None
            if t_id:
                try:
                    thematique = Thematiques.objects.get(id=t_id) 
                except Thematiques.DoesNotExist:
                    return JsonResponse({'status': 'error', 'message': 'Thematiques introuvable.'})
                
            montant_ht = (qty * unit_price) * (1 - (remise_percent / 100))
            tva_amount = montant_ht * (tva_percent / 100)
            
            LignesFacture.objects.create(
                facture=facture,
                thematique=thematique,
                description=description,
                long_description=long_description,
                quantite=qty,
                prix_unitaire=unit_price,
                montant_ht=montant_ht,
                remise_percent=remise_percent,
                tva_percent=tva_percent
            )
            
            total_ht += montant_ht
            total_ttc += (montant_ht + tva_amount)
            
        facture.total_ttc = total_ttc
        
        # --- Timbre Calculation ---
        montant_timbre = 0
        from t_tresorerie.models import ParametreFinancier
        config_fin = ParametreFinancier.get_instance()
        
        apply_timbre = False
        if getattr(facture, 'type_facture', 'standard') != 'avoir':
            if config_fin.activer_timbre or mode_paiement == 'esp':
                if not config_fin.timbre_cash_only or mode_paiement == 'esp':
                    apply_timbre = True
                
        if apply_timbre:
            import math
            import json
            try:
                bareme = json.loads(config_fin.timbre_bareme)
            except Exception:
                bareme = [
                    {"min_ttc": 0, "max_ttc": 300, "rate": 0.0, "is_exempt": True},
                    {"min_ttc": 301, "max_ttc": 30000, "rate": 1.0, "is_exempt": False},
                    {"min_ttc": 30001, "max_ttc": 100000, "rate": 1.5, "is_exempt": False},
                    {"min_ttc": 100001, "max_ttc": None, "rate": 2.0, "is_exempt": False}
                ]
            bareme = sorted(bareme, key=lambda b: b.get('min_ttc', 0))
            
            def calculate_timbre_for_bracket(ttc_val, bracket_rate, is_ex):
                if is_ex or bracket_rate == 0:
                    return 0
                nb_t = math.ceil(ttc_val / 100)
                raw_timbre = nb_t * float(bracket_rate)
                min_stamp = max(float(config_fin.timbre_min), 5.0)
                if raw_timbre < min_stamp:
                    raw_timbre = min_stamp
                return math.ceil(raw_timbre)
                
            selected_bracket = None
            for b in bareme:
                min_ttc = float(b.get('min_ttc', 0))
                max_ttc = b.get('max_ttc')
                rate_val = float(b.get('rate', 0.0))
                
                if max_ttc is not None:
                    if min_ttc <= float(total_ttc) <= float(max_ttc):
                        selected_bracket = b
                        break
                else:
                    if float(total_ttc) >= min_ttc:
                        selected_bracket = b
                        break
                        
            if not selected_bracket:
                selected_bracket = bareme[-1]
                
            bracket_rate = float(selected_bracket.get('rate', 0.0))
            is_ex_bracket = selected_bracket.get('is_exempt', bracket_rate == 0.0)
            
            montant_timbre = calculate_timbre_for_bracket(float(total_ttc), bracket_rate, is_ex_bracket)
            
        facture.montant_timbre = montant_timbre
        # ---------------------------

        facture.save()
        
        from t_crm.models import UserActionLog
        UserActionLog.objects.create(
            user=request.user,
            action_type='UPDATE',
            target_model='Facture',
            target_id=str(facture.num_facture),
            details=f"Mise à jour des éléments de la facture: {facture.num_facture}",
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({'status': 'success', 'message': 'Facture enregistrée avec succès.'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'approuv')
@transaction.atomic
def ApiValidateFacture(request):
    if request.method == 'POST':
        facture_id = request.POST.get('facture_id')
        try:
            facture = Facture.objects.get(num_facture=facture_id)
            
            # Update Facture Status to 'En attente de paiement'
            facture.etat = 'battente'
            facture.save()
            
            if getattr(facture, 'type_facture', 'standard') == 'avoir':
                from t_tresorerie.models import Depenses
                # Calculate total HT and TTC for Depense
                total_ht = sum(l.montant_ht for l in facture.lignes_facture.all())
                total_ttc = facture.total_ttc()
                source_num = facture.facture_source.num_facture if facture.facture_source else ''
                
                Depenses.objects.create(
                    label=f"Avoir sur facture {source_num}",
                    client=facture.client,
                    date_depense=facture.date_emission,
                    montant_ht=total_ht,
                    montant_ttc=total_ttc,
                    etat=False,
                    description=f"Généré automatiquement suite à la validation de l'avoir {facture.num_facture}",
                    entite=facture.entreprise,
                    reference=facture.num_facture
                )
            else:
                # Update Related Prospect Pipeline Stage
                if facture.client:
                    facture.client.conseil_pipeline_stage = 'recouvrement'
                    facture.client.save()

                # Update Linked Opportunity Stage to 'recouvrement'
                if facture.devis_source and hasattr(facture.devis_source, 'opportunite'):
                    opp = facture.devis_source.opportunite
                    if opp:
                        opp.stage = 'recouvrement'
                        opp.save()
            
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='UPDATE',
                target_model='Facture',
                target_id=str(facture.num_facture),
                details=f"Validation de la facture: {facture.num_facture}",
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            return JsonResponse({'status': 'success', 'message': 'Facture validée et en attente de paiement.'})
        except Facture.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Facture non trouvée.'}, status=404)
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'change')
@transaction.atomic
def ApiRevertFactureToDraft(request):
    if request.method == 'POST':
        facture_id = request.POST.get('facture_id')
        try:
            facture = Facture.objects.get(num_facture=facture_id)
            facture.etat = 'brouillon'
            facture.save()
            
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='UPDATE',
                target_model='Facture',
                target_id=str(facture.num_facture),
                details=f"Repassage en brouillon de la facture: {facture.num_facture}",
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            return JsonResponse({'status': 'success', 'message': 'Facture repassée en brouillon.'})
        except Facture.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Facture non trouvée.'}, status=404)
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

@login_required(login_url='institut_app:login')
@ajax_required
@module_permission_required('con', 'delete')
@transaction.atomic
def ApiDeleteFacture(request):
    facture_id = request.POST.get('facture_id')
    try:
        facture = Facture.objects.get(num_facture=facture_id)
        
        # We allow deletion of validated factures.
        
        from t_tresorerie.models import Paiements as TresoreriePaiements, OperationsBancaire, Rembourssements

        # 1. Delete associated Tresorerie Paiements and their Lettrages (OperationsBancaire)
        tresorerie_paiements = TresoreriePaiements.objects.filter(facture=facture)
        for t_p in tresorerie_paiements:
            OperationsBancaire.objects.filter(paiement=t_p).delete()
            t_p.delete()

        # 2. Delete OperationsBancaire linked to Conseil Paiement
        conseil_paiements = facture.paiements.all()
        for c_p in conseil_paiements:
            OperationsBancaire.objects.filter(conseil_paiement=c_p).delete()
            # The c_p itself will be cascade deleted when facture is deleted

        # 3. Delete associated Rembourssements
        Rembourssements.objects.filter(facture=facture).delete()

        # 4. Devis is NOT deleted. It stays in the system.
        # We must reset the pipeline stage of the client and opportunity
        if facture.client:
            new_stage = 'devis_envoye' if facture.devis_source else 'negociation'
            facture.client.conseil_pipeline_stage = new_stage
            facture.client.save()
            
        if facture.devis_source and hasattr(facture.devis_source, 'opportunite'):
            opp = facture.devis_source.opportunite
            if opp:
                opp.stage = 'devis_envoye'
                opp.save()

        # 5. Delete the Facture itself
        facture.delete()
        
        from t_crm.models import UserActionLog
        UserActionLog.objects.create(
            user=request.user,
            action_type='DELETE',
            target_model='Facture',
            target_id=str(facture_id),
            details=f"Suppression de la facture: {facture_id}",
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({'status': 'success', 'message': 'Facture et informations liées supprimées avec succès.'})
    except Facture.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Facture non trouvée.'})

@login_required(login_url='institut_app:login')
@ajax_required
@module_permission_required('con', 'add')
@transaction.atomic
def ApiCreateAvoir(request):
    facture_id = request.POST.get('facture_id')
    try:
        facture = Facture.objects.get(num_facture=facture_id)
        if getattr(facture, 'type_facture', 'standard') == 'avoir':
            return JsonResponse({'status': 'error', 'message': 'Impossible de créer un avoir depuis un avoir.'})
        
        avoir = Facture.objects.create(
            client=facture.client,
            devis_source=facture.devis_source,
            date_emission=timezone.now().date(),
            tva=facture.tva,
            show_tva=facture.show_tva,
            show_remise=facture.show_remise,
            mode_paiement=facture.mode_paiement,
            etat='brouillon',
            conditions_commerciales=facture.conditions_commerciales,
            module_source=facture.module_source,
            entreprise=facture.entreprise,
            type_facture='avoir',
            facture_source=facture
        )
        
        for ligne in facture.lignes_facture.all():
            LignesFacture.objects.create(
                facture=avoir,
                thematique=ligne.thematique,
                description=ligne.description,
                long_description=ligne.long_description,
                quantite=ligne.quantite,
                prix_unitaire=ligne.prix_unitaire,
                montant_ht=ligne.montant_ht,
                remise_percent=ligne.remise_percent,
                tva_percent=ligne.tva_percent,
                specialite=ligne.specialite
            )
            
        from django.urls import reverse
        redirect_url = reverse('t_conseil:configure-facture', kwargs={'pk': avoir.num_facture})
        
        return JsonResponse({'status': 'success', 'message': 'Facture d\'avoir générée avec succès.', 'redirect_url': redirect_url})
    except Facture.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Facture non trouvée.'})

@ajax_required
@module_permission_required('con', 'view')
def ApiFetchEnterpriseTvas(request):
    enterprise_id = request.GET.get('enterprise_id')
    if not enterprise_id:
        return JsonResponse({'status': 'error', 'message': 'ID Entreprise manquant.'})
    
    tvas = TvaConseil.objects.all().values('valeur', 'label', 'is_default')
    return JsonResponse({'status': 'success', 'tvas': list(tvas)})


@login_required(login_url='institut_app:login')
@module_permission_required('con', 'view')
def ListeDAS(request):
    """
    Renders the page to manage DAS mappings (Products/Services to PaymentCategory).
    """
    context = {
        'tenant': request.tenant,
        'mappings': DASMapping.objects.all().select_related('thematique', 'payment_category'),
        'thematiques': Thematiques.objects.filter(etat='active'),
        'payment_categories': PaymentCategory.objects.all(),
    }
    return render(request, 'tenant_folder/conseil/liste_das.html', context)


@login_required(login_url='institut_app:login')
@ajax_required
@module_permission_required('con', 'add')
@transaction.atomic
def ApiSaveDAS(request):
    """
    API to create or update a DAS mapping.
    """
    if request.method == "POST":
        das_id = request.POST.get('das_id')
        designation = request.POST.get('designation')
        thematique_id = request.POST.get('thematique_id')
        category_id = request.POST.get('category_id')

        if not designation or not thematique_id or not category_id:
            return JsonResponse({'status': 'error', 'message': 'Tous les champs sont obligatoires.'})

        try:
            thematique = Thematiques.objects.get(id=thematique_id)
            category = PaymentCategory.objects.get(id=category_id)

            if das_id:
                try:
                    mapping = DASMapping.objects.get(id=das_id) 
                except DASMapping.DoesNotExist:
                    return JsonResponse({'status': 'error', 'message': 'DASMapping introuvable.'})
                mapping.designation = designation
                mapping.thematique = thematique
                mapping.payment_category = category
                mapping.save()
                message = "Mapping DAS mis Ã  jour avec succès."
            else:
                DASMapping.objects.create(
                    designation=designation,
                    thematique=thematique,
                    payment_category=category
                )
                message = "Nouveau mapping DAS créé avec succès."

            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='UPDATE' if das_id else 'CREATE',
                target_model='DASMapping',
                target_id=str(mapping.id if das_id else category.id), # Note: using category id if creating as we don't return the new mapping id
                details=f"Création/Mise à jour d'un mapping DAS: {designation}",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            return JsonResponse({'status': 'success', 'message': message})

        except (Thematiques.DoesNotExist, PaymentCategory.DoesNotExist):
            return JsonResponse({'status': 'error', 'message': 'Produit ou catégorie introuvable.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'})


@login_required(login_url='institut_app:login')
@ajax_required
@module_permission_required('con', 'delete')
@transaction.atomic
def ApiDeleteDAS(request):
    """
    API to delete a DAS mapping.
    """
    if request.method == "POST":
        das_id = request.POST.get('das_id')
        try:
            mapping = DASMapping.objects.get(id=das_id)
            designation = mapping.designation
            mapping.delete()
            
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='DELETE',
                target_model='DASMapping',
                target_id=str(das_id),
                details=f"Suppression d'un mapping DAS: {designation}",
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            return JsonResponse({'status': 'success', 'message': 'Mapping supprimé avec succès.'})
        except DASMapping.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Mapping introuvable.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'})

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def DownloadFacturePDF(request, pk):
    try:
        facture = get_object_or_404(Facture, num_facture=pk)
        
        if facture.etat == 'brouillon':
            messages.warning(request, "La facture doit être validée avant de pouvoir générer le PDF.")
            return redirect('t_conseil:configure-facture', pk=facture.num_facture)
            
        lignes = facture.lignes_facture.all()
        
        # Calculate totals
        total_ht = sum(ligne.montant_ht for ligne in lignes)
        total_tva = sum(ligne.montant_ht * (ligne.tva_percent / 100) for ligne in lignes)
        timbre = float(facture.get_timbre())
        total_ttc = float(total_ht) + float(total_tva) + timbre
        
        # Convert total to words
        amount_in_words = num_to_words_fr(float(total_ttc))
        
        # Logo URL for WeasyPrint
        logo_url = ""
        if facture.entreprise and facture.entreprise.logo:
            logo_url = request.build_absolute_uri(facture.entreprise.logo.url)
            
        context = {
            'facture': facture,
            'total_ht': total_ht,
            'total_tva': total_tva,
            'timbre': timbre,
            'total_ttc': total_ttc,
            'amount_in_words': amount_in_words,
            'logo_url': logo_url,
        }
        
        html_string = render_to_string('pdf/facture_pdf.html', context)
        
        # Create PDF
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        pdf = html.write_pdf()
        
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Facture_{facture.num_facture}.pdf"'
        return response
        
    except Exception as e:
        return HttpResponse(f"Erreur lors de la génération du PDF: {str(e)}", status=500)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'add')
@transaction.atomic
def ApiSaveParticipants(request):
    import json
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        devis_id = data.get('devis_id')
        facture_id = data.get('facture_id')
        participants = data.get('participants', [])
        
        devis = None
        if devis_id:
            try:
                devis = Devis.objects.get(num_devis=devis_id) 
            except Devis.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Devis introuvable.'})
        
        facture = None
        if facture_id:
            try:
                facture = Facture.objects.get(num_facture=facture_id) 
            except Facture.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Facture introuvable.'})
            
        # Full sync: delete existing for this doc and re-add
        from django.db import transaction
        with transaction.atomic():
            if devis:
                Participant.objects.filter(devis=devis).delete()
            if facture:
                Participant.objects.filter(facture=facture).delete()
                
            for p in participants:
                Participant.objects.create(
                    devis=devis,
                    facture=facture,
                    nom=p.get('nom'),
                    prenom=p.get('prenom'),
                    email=p.get('email'),
                    telephone=p.get('telephone'),
                    date_naissance=p.get('date_naissance') if p.get('date_naissance') else None,
                    lieu_naissance=p.get('lieu_naissance'),
                    poste=p.get('poste'),
                    nin=p.get('nin')
                )
            
        from t_crm.models import UserActionLog
        UserActionLog.objects.create(
            user=request.user,
            action_type='CREATE',
            target_model='Participant',
            target_id=str(devis_id or facture_id),
            details=f"Sauvegarde en masse des participants pour Devis/Facture.",
            ip_address=request.META.get('REMOTE_ADDR')
        )
            
        return JsonResponse({'status': 'success', 'message': 'Participants enregistrés avec succès.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def ApiGetDegreeFormations(request):
    from t_formations.models import Specialites
    specialites = Specialites.objects.all().values('id', 'label', 'prix', 'code', 'created_at', 'formation__nom')
    return JsonResponse(list(specialites), safe=False)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'view')
def ApiLoadDegreeFormationsList(request):
    from t_formations.models import Formation
    formations = Formation.objects.all().values('id', 'nom', 'description', 'prix_formation', 'code', 'date_creation')
    return JsonResponse(list(formations), safe=False)

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'view')
def ApiGetSpecialiteDetails(request):
    """
    Returns full details for a specific specialty, including associated modules.
    """
    spec_id = request.GET.get('id')
    if not spec_id:
        return JsonResponse({'status': 'error', 'message': 'Specialite ID required'}, status=400)
    
    from t_formations.models import Specialites, Modules
    try:
        spec = Specialites.objects.get(id=spec_id)
        
        # Fetch associated modules
        modules = Modules.objects.filter(specialite=spec, is_archived=False).values('id', 'code', 'label', 'coef', 'duree')
        
        data = {
            'id': spec.id,
            'label': spec.label,
            'code': spec.code,
            'prix': str(spec.prix),
            'prix_double': str(spec.prix_double_diplomation) if spec.prix_double_diplomation else None,
            'duree': spec.duree,
            'nb_semestre': spec.nb_semestre,
            'branche': spec.branche,
            'abr': spec.abr,
            'formation_nom': spec.formation.nom if spec.formation else '-',
            'nb_tranche': spec.nb_tranche,
            'condition_access': spec.condition_access,
            'etat': spec.get_etat_display() if hasattr(spec, 'get_etat_display') else spec.etat,
            'modules': list(modules)
        }
        return JsonResponse({'status': 'success', 'data': data})
    except Specialites.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Spécialité introuvable'}, status=404)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'add')
@transaction.atomic
def ApiEnrollToGroup(request):
    from t_groupe.models import Groupe, GroupeLine
    from t_crm.models import Prospets
    import json
    
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
        
    try:
        data = json.loads(request.body)
        participant_id = data.get('participant_id') # From Participant model
        prospect_id = data.get('prospect_id') # From Prospets model
        groupe_id = data.get('groupe_id')
        
        try:
            groupe = Groupe.objects.get(id=groupe_id) 
        except Groupe.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Groupe introuvable.'})
        
        prospect = None
        if prospect_id:
            try:
                prospect = Prospets.objects.get(id=prospect_id) 
            except Prospets.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Prospets introuvable.'})
        elif participant_id:
            try:
                p = Participant.objects.get(id=participant_id) 
            except Participant.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Participant introuvable.'})
            prospect = Prospets.objects.filter(nom=p.nom, prenom=p.prenom, email=p.email).first()
            if not prospect:
                prospect = Prospets.objects.create(
                    nom=p.nom,
                    prenom=p.prenom,
                    email=p.email,
                    telephone=p.telephone,
                    date_naissance=p.date_naissance,
                    lieu_naissance=p.lieu_naissance,
                    nin=p.nin,
                    type_prospect='particulier',
                    statut='convertit',
                    context='con'
                )
        
        if prospect:
            from django.db import transaction
            with transaction.atomic():
                if not GroupeLine.objects.filter(groupe=groupe, student=prospect).exists():
                    GroupeLine.objects.create(groupe=groupe, student=prospect)
                    prospect.statut = 'convertit'
                    prospect.save()
                    
                    from t_crm.models import UserActionLog
                    UserActionLog.objects.create(
                        user=request.user,
                        action_type='CREATE',
                        target_model='GroupeLine',
                        target_id=str(groupe.id),
                        details=f"Inscription de {prospect.nom} {prospect.prenom} au groupe {groupe.nom}",
                        ip_address=request.META.get('REMOTE_ADDR')
                    )
                    
                    return JsonResponse({'status': 'success', 'message': f'{prospect.nom} {prospect.prenom} inscrit au groupe {groupe.nom} avec succès.'})
                else:
                    return JsonResponse({'status': 'error', 'message': 'DéjÃ  inscrit Ã  ce groupe.'})
        
        return JsonResponse({'status': 'error', 'message': 'Participant ou prospect introuvable.'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required(login_url='institut_app:login')
@ajax_required
@module_permission_required('con', 'view')
def ApiGetGroups(request):
    """
    Returns groups for a specific specialty.
    """
    specialite_id = request.GET.get('specialite_id')
    if not specialite_id:
        return JsonResponse({'status': 'error', 'message': 'Specialite ID required'}, status=400)
    
    from t_groupe.models import Groupe
    groups = Groupe.objects.filter(specialite_id=specialite_id).values('id', 'nom', 'promotion__label', 'etat')
    return JsonResponse({'status': 'success', 'groups': list(groups)})

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def ApiLoadProspectParticipants(request):
    prospect_id = request.GET.get('prospect_id')
    if not prospect_id:
        return JsonResponse({'status': 'error', 'message': 'Prospect ID required'}, status=400)
    
    participants = Participant.objects.filter(prospect_id=prospect_id).values(
        'id', 'nom', 'prenom', 'email', 'telephone', 'date_naissance', 'lieu_naissance', 'poste', 'nin'
    )
    return JsonResponse({'status': 'success', 'participants': list(participants)})

@login_required(login_url="institut_app:login")
@ajax_required
@module_permission_required('con', 'add')
@transaction.atomic
def ApiSaveParticipant(request):
    import json
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        p_id = data.get('id')
        prospect_id = data.get('prospect_id')
        doc_id = data.get('devis_id') # Could be devis_123 or facture_123
        
        is_facture = False
        actual_id = doc_id
        if doc_id and str(doc_id).startswith('devis_'):
            actual_id = str(doc_id).replace('devis_', '')
        elif doc_id and str(doc_id).startswith('facture_'):
            is_facture = True
            actual_id = str(doc_id).replace('facture_', '')
        
        if not prospect_id and not p_id and not doc_id:
            return JsonResponse({'status': 'error', 'message': 'Prospect ID ou Document ID requis'}, status=400)
            
        if p_id:
            try:
                participant = Participant.objects.get(id=p_id) 
            except Participant.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Participant introuvable.'})
        else:
            participant = Participant()
            if prospect_id:
                participant.prospect_id = prospect_id
            if actual_id:
                if is_facture:
                    participant.facture_id = actual_id
                else:
                    participant.devis_id = actual_id
            
        participant.nom = data.get('nom')
        participant.prenom = data.get('prenom')
        participant.email = data.get('email')
        participant.telephone = data.get('telephone')
        participant.date_naissance = data.get('date_naissance') if data.get('date_naissance') else None
        participant.lieu_naissance = data.get('lieu_naissance')
        participant.poste = data.get('poste')
        participant.nin = data.get('nin')
        participant.save()
        
        from t_crm.models import UserActionLog
        UserActionLog.objects.create(
            user=request.user,
            action_type='UPDATE' if p_id else 'CREATE',
            target_model='Participant',
            target_id=str(participant.id),
            details=f"Création/Mise à jour du participant: {participant.nom} {participant.prenom}",
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({'status': 'success', 'message': 'Participant enregistré avec succès.', 'id': participant.id})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required(login_url="institut_app:login")
@ajax_required
@module_permission_required('con', 'delete')
@transaction.atomic
def ApiDeleteParticipant(request):
    import json
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        p_id = data.get('id')
        if not p_id:
            return JsonResponse({'status': 'error', 'message': 'Participant ID required'}, status=400)
            
        Participant.objects.filter(id=p_id).delete()
        
        from t_crm.models import UserActionLog
        UserActionLog.objects.create(
            user=request.user,
            action_type='DELETE',
            target_model='Participant',
            target_id=str(p_id),
            details=f"Suppression du participant {p_id}",
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({'status': 'success', 'message': 'Participant supprimé avec succès.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required(login_url='institut_app:login')
@ajax_required
@module_permission_required('con', 'add')
@transaction.atomic
def ApiSaveEtsDetails(request):
    if request.method == 'POST':
        prospect_id = request.POST.get('prospect_id')
        try:
            prospect = Prospets.objects.get(id=prospect_id)
            prospect.entreprise = request.POST.get('entreprise')
            prospect.rc = request.POST.get('rc')
            prospect.nif = request.POST.get('nif')
            prospect.nis = request.POST.get('nis')
            prospect.art_imp = request.POST.get('art_imp')
            prospect.observation = request.POST.get('observation')
            prospect.save()
            
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='UPDATE',
                target_model='Prospets',
                target_id=str(prospect.id),
                details=f"Mise à jour des informations entreprise du prospect: {prospect.entreprise}",
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            return JsonResponse({'status': 'success', 'message': 'Informations entreprise mises Ã  jour.'})
        except Prospets.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Prospect non trouvé.'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required(login_url='institut_app:login')
@ajax_required
@module_permission_required('con', 'view')
def ApiLoadBankAccounts(request):
    prospect_id = request.GET.get('id_prospect')
    accounts = ProspectBankAccount.objects.filter(prospect_id=prospect_id).values(
        'id', 'bank_name', 'rib', 'swift', 'bank_address', 'is_active'
    )
    return JsonResponse(list(accounts), safe=False)

@login_required(login_url='institut_app:login')
@ajax_required
@module_permission_required('con', 'add')
@transaction.atomic
def ApiSaveBankAccount(request):
    if request.method == 'POST':
        account_id = request.POST.get('account_id')
        prospect_id = request.POST.get('prospect_id')
        
        if account_id:
            try:
                account = ProspectBankAccount.objects.get(id=account_id) 
            except ProspectBankAccount.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'ProspectBankAccount introuvable.'})
        else:
            try:
                prospect = Prospets.objects.get(id=prospect_id) 
            except Prospets.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Prospets introuvable.'})
            account = ProspectBankAccount(prospect=prospect)
            
        account.bank_name = request.POST.get('bank_name')
        account.rib = request.POST.get('rib')
        account.swift = request.POST.get('swift')
        account.bank_address = request.POST.get('bank_address')
        account.save()
        
        from t_crm.models import UserActionLog
        UserActionLog.objects.create(
            user=request.user,
            action_type='UPDATE' if account_id else 'CREATE',
            target_model='ProspectBankAccount',
            target_id=str(account.id),
            details=f"Mise à jour du compte bancaire {account.bank_name}",
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({'status': 'success', 'message': 'Compte bancaire enregistré.'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required(login_url='institut_app:login')
@ajax_required
@module_permission_required('con', 'delete')
@transaction.atomic
def ApiDeleteBankAccount(request):
    if request.method == 'POST':
        account_id = request.POST.get('account_id')
        try:
            account = ProspectBankAccount.objects.get(id=account_id)
            bank_name = account.bank_name
            account.delete()
            
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='DELETE',
                target_model='ProspectBankAccount',
                target_id=str(account_id),
                details=f"Suppression du compte bancaire: {bank_name}",
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            return JsonResponse({'status': 'success', 'message': 'Compte bancaire supprimé.'})
        except ProspectBankAccount.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Compte non trouvé.'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required(login_url='institut_app:login')
@module_permission_required('con', 'view')
def GroupedParticipants(request):
    """
    Displays participants grouped by their thematics from confirmed quotes.
    Handles both explicit Participant objects and individual Prospect clients.
    """
    from .models import Thematiques, Participant
    from t_formations.models import Specialites, Promos
    from t_crm.models import Prospets

    # Get active thematics
    thematiques = Thematiques.objects.filter(etat='active')
    
    grouped_data = []
    for th in thematiques:
        # 1. Explicit participants linked to confirmed quotes with this thematic
        participants_explicit = Participant.objects.filter(
            devis__lignes_devis__thematique=th,
            devis__etat='accepte'
        ).select_related('prospect', 'devis').distinct()
        
        # 2. Individual prospects who are clients of confirmed thematic quotes
        prospects_as_clients = Prospets.objects.filter(
            client_devis__lignes_devis__thematique=th,
            client_devis__etat='accepte',
            type_prospect='particulier'
        ).distinct()

        unified_participants = []
        seen_emails = set()

        # Add explicit participants
        for p in participants_explicit:
            p_data = {
                'id': p.id,
                'nom': p.nom,
                'prenom': p.prenom,
                'email': p.email,
                'telephone': p.telephone,
                'poste': p.poste,
                'prospect': p.prospect,
                'devis': p.devis,
                'is_prospect_direct': False
            }
            unified_participants.append(p_data)
            if p.email: seen_emails.add(p.email)

        # Add prospect clients if not already added
        for p in prospects_as_clients:
            if p.email and p.email in seen_emails:
                continue
            
            # Find the first accepted devis with this thematic for context
            devis = p.client_devis.filter(lignes_devis__thematique=th, etat='accepte').first()
            
            p_data = {
                'id': f"prospect_{p.id}",
                'nom': p.nom,
                'prenom': p.prenom,
                'email': p.email,
                'telephone': p.telephone,
                'poste': getattr(p, 'poste', '') or '',
                'prospect': p,
                'devis': devis,
                'is_prospect_direct': True
            }
            unified_participants.append(p_data)

        if unified_participants:
            grouped_data.append({
                'thematique': th,
                'participants': unified_participants
            })

    context = {
        'tenant': request.tenant,
        'grouped_data': grouped_data,
        'specialites': Specialites.objects.all(),
        'promos': Promos.objects.filter(etat='active'),
    }
    return render(request, 'tenant_folder/conseil/grouped_participants.html', context)


@login_required(login_url='institut_app:login')
@ajax_required
@module_permission_required('con', 'add')
@transaction.atomic
def ApiCreateGroupFromParticipants(request):
    """
    API to create a group in t_groupe from selected participants.
    Handles both Participant IDs and prospect IDs (prefixed with prospect_).
    """
    import json
    from t_groupe.models import Groupe, GroupeLine
    from t_crm.models import Prospets
    from t_formations.models import Thematiques, Promos
    from .models import Participant
    
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
        
    try:
        data = json.loads(request.body)
        participant_ids = data.get('participant_ids', [])
        group_name = data.get('group_name')
        thematique_id = data.get('thematique_id')
        promo_id = data.get('promo_id')
        
        if not group_name or not thematique_id:
            return JsonResponse({'status': 'error', 'message': 'Nom du groupe et thématique sont obligatoires.'}, status=400)
            
        # Create the group
        groupe = Groupe.objects.create(
            nom=group_name,
            thematique_id=thematique_id,
            promotion_id=promo_id,
            etat='brouillon',
        )
        
        count = 0
        for pid_str in participant_ids:
            try:
                student = None
                if str(pid_str).startswith('prospect_'):
                    prospect_id = pid_str.replace('prospect_', '')
                    try:
                        student = Prospets.objects.get(id=prospect_id) 
                    except Prospets.DoesNotExist:
                        return JsonResponse({'status': 'error', 'message': 'Prospets introuvable.'})
                else:
                    try:
                        participant = Participant.objects.get(id=pid_str) 
                    except Participant.DoesNotExist:
                        return JsonResponse({'status': 'error', 'message': 'Participant introuvable.'})
                    # Find or create a Prospets record for the individual participant
                    if participant.email:
                        student = Prospets.objects.filter(email=participant.email, is_ets_prospect=False).first()
                    
                    if not student:
                        student = Prospets.objects.create(
                            nom=participant.nom,
                            prenom=participant.prenom,
                            email=participant.email,
                            telephone=participant.telephone,
                            date_naissance=participant.date_naissance,
                            lieu_naissance=participant.lieu_naissance,
                            nin=participant.nin,
                            context='con',
                            is_ets_prospect=False,
                            type_prospect='particulier',
                            entreprise=participant.prospect.nom if participant.prospect else None,
                        )
                
                if student:
                    if not GroupeLine.objects.filter(groupe=groupe, student=student).exists():
                        GroupeLine.objects.create(groupe=groupe, student=student)
                        count += 1
            except Exception as e:
                print(f"Error adding participant {pid_str}: {e}")
                continue
        
        from t_crm.models import UserActionLog
        UserActionLog.objects.create(
            user=request.user,
            action_type='CREATE',
            target_model='Groupe',
            target_id=str(groupe.id),
            details=f"Création d'un groupe conseil en masse: {group_name} avec {count} participants",
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({
            'status': 'success', 
            'message': f'Groupe "{group_name}" créé avec succès. {count} participants ajoutés.',
            'group_id': groupe.id
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required(login_url="institut_app:login")
@module_permission_required('con', 'approuv')
@transaction.atomic
def ApiAcceptDevis(request):
    if request.method == 'POST':
        devis_id = request.POST.get('devis_id')
        try:
            from .models import Devis
            devis = Devis.objects.get(num_devis=devis_id)
            devis.etat = 'accepte'
            devis.save()
            
            # Update CRM Pipeline Stage
            if devis.opportunite:
                devis.opportunite.stage = 'facture'
                devis.opportunite.save()
            
            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='UPDATE',
                target_model='Devis',
                target_id=str(devis.num_devis),
                details=f"Acceptation du devis: {devis.num_devis}",
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            return JsonResponse({'status': 'success', 'message': 'Le devis a été accepté avec succès.'})
        except Devis.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Devis introuvable.'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Mauvaise méthode.'}, status=405)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'approuv')
@transaction.atomic
def ApiRejectDevis(request):
    if request.method == 'POST':
        devis_id = request.POST.get('devis_id')
        try:
            from .models import Devis
            devis = Devis.objects.get(num_devis=devis_id)
            devis.etat = 'rejete'
            devis.save()
            
            # Update CRM Pipeline Stage
            if devis.opportunite:
                devis.opportunite.stage = 'perdu'
                devis.opportunite.save()

            from t_crm.models import UserActionLog
            UserActionLog.objects.create(
                user=request.user,
                action_type='UPDATE',
                target_model='Devis',
                target_id=str(devis.num_devis),
                details=f"Rejet du devis: {devis.num_devis}",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            return JsonResponse({'status': 'success', 'message': 'Le devis a été rejeté.'})
        except Devis.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Devis introuvable.'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Mauvaise méthode.'}, status=405)

@login_required(login_url="institut_app:login")
@module_permission_required('con', 'view')
def PaiementsConseilListe(request):
    # Fetch all payments related to t_conseil
    paiements = Paiement.objects.all().order_by('-date_paiement', '-created_at')
    
    context = {
        'paiements': paiements,
    }
    return render(request, 'tenant_folder/conseil/liste_des_paiements.html', context)


@login_required(login_url="institut_app:login")
@module_permission_required('con', 'add')
def ApiCreateRendezVousPipeline(request):
    if request.method == "POST":
        from t_crm.models import RendezVous, Prospets
        prospect_id = request.POST.get('prospect_id')
        date_rendez_vous = request.POST.get('date_rendez_vous')
        heure_rendez_vous = request.POST.get('heure_rendez_vous')
        type_rdv = request.POST.get('type')
        object_rdv = request.POST.get('object')
        
        if not prospect_id or not date_rendez_vous or not heure_rendez_vous:
            return JsonResponse({'status': 'error', 'message': 'Veuillez remplir les champs obligatoires.'})
            
        try:
            prospect = Prospets.objects.get(id=prospect_id)
            RendezVous.objects.create(
                created_by=request.user,
                prospect=prospect,
                date_rendez_vous=date_rendez_vous,
                heure_rendez_vous=heure_rendez_vous,
                type=type_rdv,
                object=object_rdv,
                context='con',
                statut='en_attente'
            )
            return JsonResponse({'status': 'success', 'message': 'Rendez-vous planifié avec succès.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'})
