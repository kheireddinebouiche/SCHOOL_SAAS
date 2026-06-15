from django.utils.decorators import method_decorator
from institut_app.decorators import module_permission_required
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DetailView, FormView, DeleteView
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from .models import Contrat, FichePaie, ParametresPaie, TypeContrat, Rubrique, RubriqueContrat, LignePaie, ValidationFicheMensuelleFormateur
from .logic import PaieEngine
from django import forms
from decimal import Decimal
from t_formations.models import Formateurs
from t_rh.models import Employees, Presence
from institut_app.models import Entreprise
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
import openpyxl
from openpyxl import Workbook
from t_etudiants.models import SuiviCours, LigneRegistrePresence
from t_timetable.models import TimetableEntry
from datetime import datetime, timedelta
from django.db.models import Count, Sum, Q
import collections

@method_decorator(module_permission_required('rh', 'view'), name='dispatch')
class ContratListView(LoginRequiredMixin, ListView):
    model = Contrat
    context_object_name = 'contrats'

    def get_queryset(self):
        active_id = self.request.session.get('active_entreprise_id')
        if active_id == 'none':
            return super().get_queryset().filter(entreprise__isnull=True)
        if active_id:
            return super().get_queryset().filter(entreprise_id=active_id)
        return super().get_queryset().none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['entreprises'] = Entreprise.objects.all()
        context['active_entreprise_id'] = self.request.session.get('active_entreprise_id')
        return context

class ContratForm(forms.ModelForm):
    class Meta:
        model = Contrat
        exclude = ['entreprise', 'employee', 'prime_transport', 'prime_panier', 'created_at', 'updated_at']
        widgets = {
            'formateur': forms.Select(attrs={'class': 'form-select'}),
            'type_contrat': forms.Select(attrs={'class': 'form-select'}),
            'date_debut': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
            'date_fin': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
            'salaire_base': forms.NumberInput(attrs={'class': 'form-control'}),
            'salaire_horaire': forms.NumberInput(attrs={'class': 'form-control'}),
            'actif': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }
        
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for field_name, field in self.fields.items():
            if field_name not in ['formateur', 'type_contrat', 'date_debut']:
                field.required = False

@method_decorator(module_permission_required('rh', 'add'), name='dispatch')
class ContratCreateView(LoginRequiredMixin, CreateView):
    model = Contrat
    form_class = ContratForm
    template_name = 't_ressource_humaine/contrat_form.html'
    success_url = reverse_lazy('t_ressource_humaine:contrat_list')

    def get_template_names(self):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return ['t_ressource_humaine/_contrat_form.html']
        return [self.template_name]

    def form_valid(self, form):
        active_id = self.request.session.get('active_entreprise_id')
        if not active_id:
             return JsonResponse({'status': 'error', 'message': 'Veuillez sélectionner une entreprise avant de créer un contrat'})
        
        form.instance.entreprise_id = active_id if active_id != 'none' else None
        
        self.object = form.save()
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': 'Contrat créé avec succès'})
        return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            html = render_to_string(
                't_ressource_humaine/_contrat_form.html',
                {'form': form},
                request=self.request
            )
            return JsonResponse({'status': 'invalid', 'html': html})
        return super().form_invalid(form)

@method_decorator(module_permission_required('rh', 'change'), name='dispatch')
class ContratUpdateView(LoginRequiredMixin, UpdateView):
    model = Contrat
    form_class = ContratForm
    template_name = 't_ressource_humaine/contrat_form.html'
    success_url = reverse_lazy('t_ressource_humaine:contrat_list')

    def get_template_names(self):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return ['t_ressource_humaine/_contrat_form.html']
        return [self.template_name]

    def form_valid(self, form):
        self.object = form.save()
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': 'Contrat modifié avec succès'})
        return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            html = render_to_string(
                't_ressource_humaine/_contrat_form.html',
                {'form': form},
                request=self.request
            )
            return JsonResponse({'status': 'invalid', 'html': html})
        return super().form_invalid(form)

@method_decorator(module_permission_required('rh', 'view'), name='dispatch')
class ContratDetailView(LoginRequiredMixin, DetailView):
    model = Contrat
    template_name = 't_ressource_humaine/contrat_print.html'

class FichePaieGenerationForm(forms.Form):
    mois = forms.ChoiceField(choices=FichePaie.MOIS_CHOICES, widget=forms.Select(attrs={'class': 'form-select'}))
    annee = forms.IntegerField(initial=2024, widget=forms.NumberInput(attrs={'class': 'form-control'}))
    jours_travailles = forms.IntegerField(initial=22, min_value=0, max_value=31, widget=forms.NumberInput(attrs={'class': 'form-control'}))
    heures_travailles = forms.DecimalField(initial=0, min_value=0, required=False, widget=forms.NumberInput(attrs={'class': 'form-control'}))
    heures_absence = forms.DecimalField(initial=0, min_value=0, required=False, widget=forms.NumberInput(attrs={'class': 'form-control'}))
    primes_exceptionnelles = forms.DecimalField(initial=0, required=False, widget=forms.NumberInput(attrs={'class': 'form-control'}))
    avance_sur_salaire = forms.DecimalField(initial=0, required=False, widget=forms.NumberInput(attrs={'class': 'form-control'}))

@module_permission_required('rh', 'add')
def generer_paie(request, contrat_id):
    contrat = get_object_or_404(Contrat, pk=contrat_id)
    
    initial_jours = 22
    initial_heures_absence = 0
    mois_select = int(request.GET.get('mois', datetime.now().month))
    annee_select = int(request.GET.get('annee', datetime.now().year))
    initial_heures = request.GET.get('heures_travailles', 0)

    if contrat.employee:
        presences = Presence.objects.filter(
            employee=contrat.employee,
            date__month=mois_select,
            date__year=annee_select
        )
        days_present = presences.filter(status='present').count()
        days_late = presences.filter(status='late').count()
        days_half = presences.filter(status='half_day').count()
        initial_jours = days_present + days_late + (days_half * 0.5)
        initial_heures_absence = presences.filter(status='absent').count() * 8
    
    contract_rubrics_config = {rc.rubrique_id: rc for rc in RubriqueContrat.objects.filter(contrat=contrat)}
    contract_defaults = {}
    
    all_rubriques = Rubrique.objects.filter(actif=True).prefetch_related('eligible_types')
    rubriques_actives = []
    
    for r in all_rubriques:
        if r.eligible_types.exists() and not r.eligible_types.filter(label__icontains=contrat.type_contrat).exists():
            continue
        rc = contract_rubrics_config.get(r.id)
        if rc:
            if not rc.actif:
                continue
            contract_defaults[r.id] = rc.valeur
        rubriques_actives.append(r)
    
    if request.method == 'POST':
        form = FichePaieGenerationForm(request.POST)
        if form.is_valid():
            mois = int(form.cleaned_data['mois'])
            annee = form.cleaned_data['annee']
            jours = form.cleaned_data['jours_travailles']
            heures = form.cleaned_data['heures_travailles'] or 0
            h_absence = form.cleaned_data.get('heures_absence') or 0
            primes_ex = form.cleaned_data.get('primes_exceptionnelles') or 0
            avance = form.cleaned_data.get('avance_sur_salaire') or 0
            
            if FichePaie.objects.filter(contrat=contrat, mois=mois, annee=annee).exists():
                message = f"Une fiche de paie existe déjà pour {mois}/{annee}"
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'status': 'error', 'message': message})
                messages.error(request, message)
                return redirect('t_ressource_humaine:contrat_list')
            
            # Extract Dynamic Rubrics
            lignes_rubriques = []
            for rubrique in rubriques_actives:
                valeur_str = request.POST.get(f'rubrique_{rubrique.id}', '0')
                try:
                    valeur = Decimal(valeur_str)
                    if valeur != 0:
                        lignes_rubriques.append({'rubrique': rubrique, 'valeur': valeur})
                except:
                    pass
            
            result = PaieEngine.calculer_paie(contrat, jours, heures, heures_absence=h_absence, lignes_rubriques=lignes_rubriques)
            
            final_net = result['net_a_payer'] + Decimal(primes_ex) - Decimal(avance)

            fiche = FichePaie(
                contrat=contrat,
                entreprise=contrat.entreprise,
                mois=mois,
                annee=annee,
                jours_travailles=jours,
                heures_travailles=heures,
                heures_absence=h_absence,
                primes_exceptionnelles=primes_ex,
                retenues_absences=result['retenue_absences_montant'],
                avance_sur_salaire=avance,
                salaire_base_calcule=result['salaire_base_calcule'],
                base_ss=result['base_ss'],
                montant_ss=result['montant_ss'],
                salaire_imposable=result['salaire_imposable'],
                irg=result['irg'],
                net_a_payer=final_net,
                prime_panier=result['prime_panier'],
                prime_transport=result['prime_transport'],
                is_validated=True 
            )
            if contrat.entreprise:
                fiche.entreprise = contrat.entreprise
            fiche.save()
            
            # Save Lignes Paie
            for ligne in result['detail_lignes']:
                LignePaie.objects.create(
                    fiche_paie=fiche,
                    rubrique=ligne['rubrique'],
                    valeur_saisie=ligne['valeur_saisie'],
                    montant=ligne['montant']
                )
            
            msg = f"Fiche de paie générée : Net {fiche.net_a_payer} DA"
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success', 
                    'message': msg,
                    'show_modal': True,
                    'redirect_url': reverse_lazy('t_ressource_humaine:fiche_paie_detail', kwargs={'pk': fiche.pk})
                })
            messages.success(request, msg)
            return redirect('t_ressource_humaine:fiche_paie_detail', pk=fiche.pk)
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                html = render_to_string(
                    't_ressource_humaine/_generate_paie_form.html',
                    {'form': form, 'contrat': contrat, 'rubriques': rubriques_actives, 'contract_defaults': contract_defaults},
                    request=request
                )
                return JsonResponse({'status': 'invalid', 'html': html})
    else:
        form = FichePaieGenerationForm(initial={
            'mois': mois_select,
            'annee': annee_select,
            'jours_travailles': initial_jours,
            'heures_travailles': initial_heures,
            'heures_absence': initial_heures_absence
        })

    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 't_ressource_humaine/_generate_paie_form.html', 
                      {'form': form, 'contrat': contrat, 'rubriques': rubriques_actives, 'contract_defaults': contract_defaults})
        
    return render(request, 't_ressource_humaine/generate_paie.html', 
                  {'form': form, 'contrat': contrat, 'rubriques': rubriques_actives, 'contract_defaults': contract_defaults})

@module_permission_required('rh', 'change')
def modifier_paie(request, pk):
    fiche = get_object_or_404(FichePaie, pk=pk)
    contrat = fiche.contrat
    rubriques_actives = Rubrique.objects.filter(actif=True)
    
    # Load existing values for rubrics
    existing_lignes = {ligne.rubrique_id: ligne.valeur_saisie for ligne in fiche.lignes_paie.all()}
    
    if request.method == 'POST':
        form = FichePaieGenerationForm(request.POST)
        if form.is_valid():
            # We don't change month/year usually, or maybe we do? 
            # Ideally modification should just recalculate amounts based on new hours/variables
            # But the form includes month/year. We should keep consistency.
            
            mois = int(form.cleaned_data['mois'])
            annee = form.cleaned_data['annee']
            jours = form.cleaned_data['jours_travailles']
            heures = form.cleaned_data['heures_travailles'] or 0
            h_absence = form.cleaned_data.get('heures_absence') or 0
            primes_ex = form.cleaned_data.get('primes_exceptionnelles') or 0
            avance = form.cleaned_data.get('avance_sur_salaire') or 0
            
            # Check for duplicates if month/year changed to something that exists (excluding self)
            if (mois != fiche.mois or annee != fiche.annee) and FichePaie.objects.filter(contrat=contrat, mois=mois, annee=annee).exclude(pk=pk).exists():
                message = f"Une autre fiche de paie existe déjà pour {mois}/{annee}"
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'status': 'error', 'message': message})
                messages.error(request, message)
                return redirect('t_ressource_humaine:fiche_paie_list')

            # Extract Dynamic Rubrics
            lignes_rubriques = []
            for rubrique in rubriques_actives:
                valeur_str = request.POST.get(f'rubrique_{rubrique.id}', '0')
                try:
                    valeur = Decimal(valeur_str)
                    if valeur != 0:
                        lignes_rubriques.append({'rubrique': rubrique, 'valeur': valeur})
                except:
                    pass

            result = PaieEngine.calculer_paie(contrat, jours, heures, heures_absence=h_absence, lignes_rubriques=lignes_rubriques)
            
            final_net = result['net_a_payer'] + Decimal(primes_ex) - Decimal(avance)

            # Update existing fiche
            fiche.mois = mois
            fiche.annee = annee
            fiche.jours_travailles = jours
            fiche.heures_travailles = heures
            fiche.heures_absence = h_absence
            fiche.primes_exceptionnelles = primes_ex
            fiche.retenues_absences = result['retenue_absences_montant']
            fiche.avance_sur_salaire = avance
            fiche.salaire_base_calcule = result['salaire_base_calcule']
            fiche.base_ss = result['base_ss']
            fiche.montant_ss = result['montant_ss']
            fiche.salaire_imposable = result['salaire_imposable']
            fiche.irg = result['irg']
            fiche.net_a_payer = final_net
            fiche.prime_panier = result['prime_panier']
            fiche.prime_transport = result['prime_transport']
            
            fiche.save()
            
            # Sync Lignes Paie (Delete all and recreate - simplest for now)
            fiche.lignes_paie.all().delete()
            for ligne in result['detail_lignes']:
                LignePaie.objects.create(
                    fiche_paie=fiche,
                    rubrique=ligne['rubrique'],
                    valeur_saisie=ligne['valeur_saisie'],
                    montant=ligne['montant']
                )
            
            msg = f"Fiche de paie modifiée : Net {fiche.net_a_payer} DA"
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success', 
                    'message': msg,
                    'show_modal': True, # Re-open detail modal? or just close edit modal and refresh table?
                    # Let's say we refresh table.
                    'reload_table': True
                })
            messages.success(request, msg)
            return redirect('t_ressource_humaine:fiche_paie_detail', pk=fiche.pk)
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                html = render_to_string(
                    't_ressource_humaine/_generate_paie_form.html',
                    {'form': form, 'contrat': contrat, 'modifying': True, 'fiche_id': fiche.pk, 'rubriques': rubriques_actives, 'existing_lignes': existing_lignes},
                    request=request
                )
                return JsonResponse({'status': 'invalid', 'html': html})
    else:
        # Pre-fill form
        form = FichePaieGenerationForm(initial={
            'mois': fiche.mois,
            'annee': fiche.annee,
            'jours_travailles': fiche.jours_travailles,
            'heures_travailles': fiche.heures_travailles,
            'heures_absence': fiche.heures_absence,
            'primes_exceptionnelles': fiche.primes_exceptionnelles,
            'avance_sur_salaire': fiche.avance_sur_salaire
        })
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 't_ressource_humaine/_generate_paie_form.html', 
                      {'form': form, 'contrat': contrat, 'modifying': True, 'fiche_id': fiche.pk, 'rubriques': rubriques_actives, 'existing_lignes': existing_lignes})
        
    return render(request, 't_ressource_humaine/generate_paie.html', 
                  {'form': form, 'contrat': contrat, 'rubriques': rubriques_actives, 'existing_lignes': existing_lignes})

@module_permission_required('rh', 'delete')
def supprimer_paie(request, pk):
    fiche = get_object_or_404(FichePaie, pk=pk)
    
    if request.method == 'POST':
        fiche.delete()
        msg = "Fiche de paie supprimée avec succès."
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': msg, 'reload_table': True})
        messages.success(request, msg)
        return redirect('t_ressource_humaine:fiche_paie_list')
    
    # Render confirmation template
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 't_ressource_humaine/_fiche_paie_delete.html', {'fiche': fiche})
        
    return redirect('t_ressource_humaine:fiche_paie_list') # Fallback if direct access GET


@method_decorator(module_permission_required('rh', 'view'), name='dispatch')
class FichePaieDetailView(LoginRequiredMixin, DetailView):
    model = FichePaie
    template_name = 't_ressource_humaine/fiche_paie_print.html'
    context_object_name = 'fiche'

    def get_template_names(self):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return ['t_ressource_humaine/_fiche_paie_detail.html']
        return [self.template_name]

@method_decorator(module_permission_required('rh', 'view'), name='dispatch')
class FichePaieListView(LoginRequiredMixin, ListView):
    model = FichePaie
    template_name = 't_ressource_humaine/fiche_paie_list.html'
    context_object_name = 'fiches'

    def get(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            self.object_list = self.get_queryset()
            context = self.get_context_data()
            return render(request, 't_ressource_humaine/_fiche_paie_table.html', context)
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        active_id = self.request.session.get('active_entreprise_id')
        queryset = super().get_queryset().select_related('contrat__formateur').order_by('-annee', '-mois', '-generated_at')
        
        if active_id == 'none':
            queryset = queryset.filter(entreprise__isnull=True)
        elif active_id:
            queryset = queryset.filter(entreprise_id=active_id)
        else:
            queryset = queryset.none()
        contrat_id = self.request.GET.get('contrat')
        formateur_id = self.request.GET.get('formateur')
        annee = self.request.GET.get('annee')
        
        if contrat_id:
            queryset = queryset.filter(contrat_id=contrat_id)
        if formateur_id:
            queryset = queryset.filter(contrat__formateur_id=formateur_id)
        if annee:
            queryset = queryset.filter(annee=annee)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['entreprises'] = Entreprise.objects.all()
        context['active_entreprise_id'] = self.request.session.get('active_entreprise_id')
        
        contrat_id = self.request.GET.get('contrat')
        formateur_id = self.request.GET.get('formateur')
        annee_selected = self.request.GET.get('annee')
        
        if contrat_id:
            context['filter_contrat'] = get_object_or_404(Contrat, pk=contrat_id)
        
        context['formateurs'] = Formateurs.objects.all().order_by('nom')
        context['selected_formateur'] = int(formateur_id) if formateur_id and formateur_id.isdigit() else None
        context['selected_annee'] = int(annee_selected) if annee_selected and annee_selected.isdigit() else None
        
        # Get unique years from FichePaie for the year filter
        context['years'] = FichePaie.objects.values_list('annee', flat=True).distinct().order_by('-annee')
        
        return context

class ParametresPaieForm(forms.ModelForm):
    class Meta:
        model = ParametresPaie
        fields = ['taux_ss', 'jours_travailles_standard', 'heures_mensuelles_standard', 'seuil_exoneration_irg', 'taux_irg_vacataire']
        widgets = {
            'taux_ss': forms.NumberInput(attrs={'class': 'form-control', 'step': '0.0001'}),
            'jours_travailles_standard': forms.NumberInput(attrs={'class': 'form-control'}),
            'heures_mensuelles_standard': forms.NumberInput(attrs={'class': 'form-control', 'step': '0.01'}),
            'seuil_exoneration_irg': forms.NumberInput(attrs={'class': 'form-control'}),
            'taux_irg_vacataire': forms.NumberInput(attrs={'class': 'form-control', 'step': '0.0001'}),
        }

@module_permission_required('rh', 'change')
def config_paie(request):
    active_id = request.session.get('active_entreprise_id')
    if not active_id:
        messages.warning(request, "Veuillez sélectionner une entreprise ou 'Non-affectés' pour accéder à la configuration.")
        return redirect('t_ressource_humaine:contrat_list')
        
    if active_id == 'none':
        entreprise = None
    else:
        entreprise = get_object_or_404(Entreprise, pk=active_id)
        
    config = ParametresPaie.get_config(entreprise=entreprise)
    if request.method == 'POST':
        form = ParametresPaieForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, "Configuration de la paie mise à jour")
            return redirect('t_ressource_humaine:config_paie')
    else:
        form = ParametresPaieForm(instance=config)
    
    return render(request, 't_ressource_humaine/parametres_paie.html', {'form': form, 'config': config})

# -- Rubriques --

from .models import Rubrique, LignePaie, RubriqueContrat

@module_permission_required('rh', 'change')
def manage_rubriques_contrat(request, contrat_id):
    contrat = get_object_or_404(Contrat, pk=contrat_id)
    
    # Filter rubrics: Keep if eligible_types is empty or contains the contract's type
    all_rubriques = Rubrique.objects.filter(actif=True)
    rubriques_eligibles = []
    
    # We need to map the contract type (string or object) to the eligible_types
    # Since we are using t_rh.models.Contrats usually, let's see which model 'contrat' is.
    # The view uses 'Contrat' (from t_ressource_humaine).
    
    for r in all_rubriques:
        # If no specific types assigned, it's global
        if not r.eligible_types.exists():
            rubriques_eligibles.append(r)
        else:
            # Check if any of the eligible types match the contract's type
            # Note: t_ressource_humaine.Contrat.type_contrat is a string (CDI, CDD)
            # Rubrique.eligible_types are t_rh.TypesContrat objects.
            # We need to match by label.
            if r.eligible_types.filter(label__icontains=contrat.type_contrat).exists():
                rubriques_eligibles.append(r)
            
    existing_rc = {rc.rubrique_id: rc for rc in RubriqueContrat.objects.filter(contrat=contrat)}

    if request.method == 'POST':
        for rubrique in rubriques_eligibles:
            valeur = request.POST.get(f'valeur_{rubrique.id}')
            actif = request.POST.get(f'actif_{rubrique.id}') == 'on'
            
            # Update or create RubriqueContrat
            if valeur is not None:
                try:
                    valeur_dec = Decimal(valeur) if valeur else Decimal(0)
                    RubriqueContrat.objects.update_or_create(
                        contrat=contrat,
                        rubrique=rubrique,
                        defaults={'valeur': valeur_dec, 'actif': actif}
                    )
                except ValueError:
                    pass

        msg = "Primes du contrat mises à jour."
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': msg})
        messages.success(request, msg)
        return redirect('t_ressource_humaine:contrat_list')

    # Prepare data for template
    rubrique_data = []
    for r in rubriques_eligibles:
        rc = existing_rc.get(r.id)
        rubrique_data.append({
            'rubrique': r,
            'valeur': rc.valeur if rc else r.valeur_defaut,
            'actif': rc.actif if rc else True
        })
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 't_ressource_humaine/rubriques_contrat_form.html', 
                      {'contrat': contrat, 'rubrique_data': rubrique_data})
                      
    return render(request, 't_ressource_humaine/rubriques_contrat_form.html', 
                  {'contrat': contrat, 'rubrique_data': rubrique_data})



from t_rh.models import TypesContrat

class RubriqueForm(forms.ModelForm):
    eligible_types = forms.ModelMultipleChoiceField(
        queryset=TypesContrat.objects.all(),
        widget=forms.CheckboxSelectMultiple(attrs={'class': 'form-check-input'}),
        required=False,
        label="Types de contrats éligibles",
        help_text="Sélectionnez les types de contrats auxquels cette rubrique s'applique."
    )

    class Meta:
        model = Rubrique
        fields = ['libelle', 'type_rubrique', 'mode_calcul', 'valeur_defaut', 'eligible_types', 'est_cotisable', 'est_imposable', 'actif']
        widgets = {
            'libelle': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'ex: Prime de Panier'}),
            'type_rubrique': forms.Select(attrs={'class': 'form-select'}),
            'mode_calcul': forms.Select(attrs={'class': 'form-select'}),
            'valeur_defaut': forms.NumberInput(attrs={'class': 'form-control', 'step': '0.01'}),
            'est_cotisable': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'est_imposable': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'actif': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }

@method_decorator(module_permission_required('rh', 'view'), name='dispatch')
class RubriqueListView(LoginRequiredMixin, ListView):
    model = Rubrique
    template_name = 't_ressource_humaine/rubrique_list.html'
    context_object_name = 'rubriques'

@method_decorator(module_permission_required('rh', 'add'), name='dispatch')
class RubriqueCreateView(LoginRequiredMixin, CreateView):
    model = Rubrique
    form_class = RubriqueForm
    template_name = 't_ressource_humaine/rubrique_form.html'
    success_url = reverse_lazy('t_ressource_humaine:rubrique_list')

    def get_template_names(self):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return ['t_ressource_humaine/_rubrique_form.html']
        return [self.template_name]

    def form_valid(self, form):
        self.object = form.save()
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': 'Rubrique créée avec succès', 'reload_table': True})
        return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
             html = render_to_string(
                't_ressource_humaine/_rubrique_form.html',
                {'form': form},
                request=self.request
            )
             return JsonResponse({'status': 'invalid', 'html': html})
        return super().form_invalid(form)

@method_decorator(module_permission_required('rh', 'change'), name='dispatch')
class RubriqueUpdateView(LoginRequiredMixin, UpdateView):
    model = Rubrique
    form_class = RubriqueForm
    template_name = 't_ressource_humaine/rubrique_form.html'
    success_url = reverse_lazy('t_ressource_humaine:rubrique_list')

    def get_template_names(self):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return ['t_ressource_humaine/_rubrique_form.html']
        return [self.template_name]

    def form_valid(self, form):
        self.object = form.save()
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': 'Rubrique modifiée avec succès', 'reload_table': True})
        return super().form_valid(form)
        
    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
             html = render_to_string(
                't_ressource_humaine/_rubrique_form.html',
                {'form': form},
                request=self.request
            )
             return JsonResponse({'status': 'invalid', 'html': html})
        return super().form_invalid(form)

@method_decorator(module_permission_required('rh', 'delete'), name='dispatch')
class RubriqueDeleteView(LoginRequiredMixin, DeleteView):
    model = Rubrique
    success_url = reverse_lazy('t_ressource_humaine:rubrique_list')
    template_name = 't_ressource_humaine/_rubrique_delete.html'
    
    def form_valid(self, form):
        success_url = self.get_success_url()
        self.object.delete()
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': 'Rubrique supprimée avec succès', 'reload_table': True})
        return redirect(success_url)

@login_required
@module_permission_required('rh', 'add')
def init_conventional_primes(request):
    """
    Initialise les primes conventionnelles standards si elles n'existent pas.
    """
    primes = [
        {'libelle': 'Prime d\'Expérience Professionnelle (PEP)', 'type': 'GAIN', 'calcul': 'PERCENT', 'cotisable': True, 'imposable': True},
        {'libelle': 'Indemnité de Nuisance', 'type': 'GAIN', 'calcul': 'FIXE', 'cotisable': True, 'imposable': True},
        {'libelle': 'Prime de Rendement Individuel (PRI)', 'type': 'GAIN', 'calcul': 'PERCENT', 'cotisable': True, 'imposable': True},
        {'libelle': 'Prime de Rendement Collectif (PRC)', 'type': 'GAIN', 'calcul': 'PERCENT', 'cotisable': True, 'imposable': True},
        {'libelle': 'Indemnité de Panier', 'type': 'GAIN', 'calcul': 'FIXE', 'cotisable': False, 'imposable': False},
        {'libelle': 'Indemnité de Transport', 'type': 'GAIN', 'calcul': 'FIXE', 'cotisable': False, 'imposable': False},
        {'libelle': 'Heures Supplémentaires 50%', 'type': 'GAIN', 'calcul': 'HOURS', 'cotisable': True, 'imposable': True},
        {'libelle': 'Absence non justifiée', 'type': 'RETENUE', 'calcul': 'FIXE', 'cotisable': True, 'imposable': True},
    ]

    count = 0
    for p in primes:
        obj, created = Rubrique.objects.get_or_create(
            libelle=p['libelle'],
            defaults={
                'type_rubrique': p['type'],
                'mode_calcul': p['calcul'],
                'est_cotisable': p['cotisable'],
                'est_imposable': p['imposable'],
                'actif': True
            }
        )
        if created:
            count += 1

    return JsonResponse({
        'status': 'success',
        'message': f'{count} primes conventionnelles ont été initialisées.'
    })

@module_permission_required('rh', 'add')
def reset_rubriques(request):
    """
    Supprime toutes les rubriques non utilisées, corrige/met à jour
    les rubriques utilisées qui ont des paramètres erronés,
    et injecte la configuration réglementaire correcte.
    """
    from t_ressource_humaine.models import Rubrique, RubriqueContrat, LignePaie
    
    # Définir les rubriques conventionnelles cibles
    target_rubriques = [
        {
            'libelle': "Indemnité d'Expérience Professionnelle (IEP)",
            'type_rubrique': 'GAIN',
            'mode_calcul': 'PERCENT',
            'est_cotisable': True,
            'est_imposable': True,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Prime d'Expérience Professionnelle (PEP)",
            'type_rubrique': 'GAIN',
            'mode_calcul': 'PERCENT',
            'est_cotisable': True,
            'est_imposable': True,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Indemnité de Nuisance",
            'type_rubrique': 'GAIN',
            'mode_calcul': 'FIXE',
            'est_cotisable': True,
            'est_imposable': True,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Prime de Danger",
            'type_rubrique': 'GAIN',
            'mode_calcul': 'FIXE',
            'est_cotisable': True,
            'est_imposable': True,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Prime de Nuisance",
            'type_rubrique': 'GAIN',
            'mode_calcul': 'FIXE',
            'est_cotisable': True,
            'est_imposable': True,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Prime d'Astreinte",
            'type_rubrique': 'GAIN',
            'mode_calcul': 'FIXE',
            'est_cotisable': True,
            'est_imposable': True,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Indemnité de Zone",
            'type_rubrique': 'GAIN',
            'mode_calcul': 'FIXE',
            'est_cotisable': True,
            'est_imposable': True,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Prime de Rendement Individuel (PRI)",
            'type_rubrique': 'GAIN',
            'mode_calcul': 'PERCENT',
            'est_cotisable': True,
            'est_imposable': True,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Prime de Rendement Collectif (PRC)",
            'type_rubrique': 'GAIN',
            'mode_calcul': 'PERCENT',
            'est_cotisable': True,
            'est_imposable': True,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Heures Supplémentaires 50%",
            'type_rubrique': 'GAIN',
            'mode_calcul': 'HOURS',
            'est_cotisable': True,
            'est_imposable': True,
            'valeur_defaut': 1.50,
            'actif': True
        },
        {
            'libelle': "Heures Supplémentaires 75%",
            'type_rubrique': 'GAIN',
            'mode_calcul': 'HOURS',
            'est_cotisable': True,
            'est_imposable': True,
            'valeur_defaut': 1.75,
            'actif': True
        },
        {
            'libelle': "Indemnité de Panier (Non Cotisable)",
            'type_rubrique': 'GAIN',
            'mode_calcul': 'FIXE',
            'est_cotisable': False,
            'est_imposable': False,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Indemnité de Transport (Non Cotisable)",
            'type_rubrique': 'GAIN',
            'mode_calcul': 'FIXE',
            'est_cotisable': False,
            'est_imposable': False,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Prime de Panier (Cotisable)",
            'type_rubrique': 'GAIN',
            'mode_calcul': 'FIXE',
            'est_cotisable': True,
            'est_imposable': True,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Prime de Transport (Cotisable)",
            'type_rubrique': 'GAIN',
            'mode_calcul': 'FIXE',
            'est_cotisable': True,
            'est_imposable': True,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Absence non justifiée",
            'type_rubrique': 'RETENUE',
            'mode_calcul': 'FIXE',
            'est_cotisable': False,
            'est_imposable': False,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Retenue Absence",
            'type_rubrique': 'RETENUE',
            'mode_calcul': 'HOURS',
            'est_cotisable': False,
            'est_imposable': False,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Retenue Retard",
            'type_rubrique': 'RETENUE',
            'mode_calcul': 'HOURS',
            'est_cotisable': False,
            'est_imposable': False,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Acompte sur Salaire",
            'type_rubrique': 'RETENUE',
            'mode_calcul': 'FIXE',
            'est_cotisable': False,
            'est_imposable': False,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Cotisation Mutuelle",
            'type_rubrique': 'RETENUE',
            'mode_calcul': 'FIXE',
            'est_cotisable': False,
            'est_imposable': False,
            'valeur_defaut': 0,
            'actif': True
        },
        {
            'libelle': "Remboursement Prêt",
            'type_rubrique': 'RETENUE',
            'mode_calcul': 'FIXE',
            'est_cotisable': False,
            'est_imposable': False,
            'valeur_defaut': 0,
            'actif': True
        }
    ]

    deleted_count = 0
    updated_count = 0
    created_count = 0
    deactivated_count = 0

    target_names = [t['libelle'] for t in target_rubriques]

    # Nettoyer les rubriques existantes
    all_current = list(Rubrique.objects.all())
    for r in all_current:
        # Vérifier si elle est utilisée
        is_used = RubriqueContrat.objects.filter(rubrique=r).exists() or LignePaie.objects.filter(rubrique=r).exists()
        
        is_duplicate_or_invalid = (r.libelle not in target_names)
        if r.libelle in ["Indemnité de Panier", "Indemnité de Transport"]:
            is_duplicate_or_invalid = True
            
        if is_duplicate_or_invalid:
            if not is_used:
                r.delete()
                deleted_count += 1
            else:
                if r.actif:
                    r.actif = False
                    r.save()
                    deactivated_count += 1

    # Injecter ou mettre à jour les rubriques cibles
    for t in target_rubriques:
        existing = Rubrique.objects.filter(libelle=t['libelle']).first()
        if existing:
            changed = False
            if existing.type_rubrique != t['type_rubrique']:
                existing.type_rubrique = t['type_rubrique']
                changed = True
            if existing.mode_calcul != t['mode_calcul']:
                existing.mode_calcul = t['mode_calcul']
                changed = True
            if existing.est_cotisable != t['est_cotisable']:
                existing.est_cotisable = t['est_cotisable']
                changed = True
            if existing.est_imposable != t['est_imposable']:
                existing.est_imposable = t['est_imposable']
                changed = True
            if not existing.actif:
                existing.actif = True
                changed = True
            
            if changed:
                existing.save()
                updated_count += 1
        else:
            Rubrique.objects.create(
                libelle=t['libelle'],
                type_rubrique=t['type_rubrique'],
                mode_calcul=t['mode_calcul'],
                est_cotisable=t['est_cotisable'],
                est_imposable=t['est_imposable'],
                valeur_defaut=t['valeur_defaut'],
                actif=t['actif']
            )
            created_count += 1

    return JsonResponse({
        'status': 'success',
        'message': f'Réinitialisation terminée : {created_count} créées, {updated_count} mises à jour, {deleted_count} supprimées, {deactivated_count} désactivées.'
    })

# -- Config Paie (Existing) --

@module_permission_required('rh', 'view')
def select_entreprise_paie(request):
    if request.method == 'POST':
        entreprise_id = request.POST.get('entreprise_id')
        request.session['active_entreprise_id'] = entreprise_id
        if entreprise_id == 'none':
            messages.info(request, "Affichage des données non-affectées.")
        elif entreprise_id:
            messages.success(request, f"Entreprise sélectionnée avec succès.")
        else:
            messages.warning(request, "Aucune entreprise sélectionnée.")
        return redirect(request.META.get('HTTP_REFERER', 't_ressource_humaine:contrat_list'))
    return redirect('t_ressource_humaine:contrat_list')

# --- Import / Export Rubriques ---

@module_permission_required('rh', 'view')
def export_rubriques(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rubriques"

    # Define headers
    headers = [
        'ID', 'Libelle', 'Type', 'Mode de Calcul', 
        'Cotisable (Oui/Non)', 'Imposable (Oui/Non)', 'Actif (Oui/Non)', 'Types Contrat (CDI, CDD, VACATION)'
    ]
    ws.append(headers)

    rubriques = Rubrique.objects.all()
    for r in rubriques:
        ws.append([
            r.id,
            r.libelle,
            r.type_rubrique,
            r.mode_calcul,
            'Oui' if r.est_cotisable else 'Non',
            'Oui' if r.est_imposable else 'Non',
            'Oui' if r.actif else 'Non',
            ", ".join([t.label for t in r.eligible_types.all()])
        ])

    # Styling headers
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="rubriques_paie.xlsx"'
    wb.save(response)
    return response

@module_permission_required('rh', 'add')
def import_rubriques(request):
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
            count_created = 0
            count_updated = 0
            
            for row in rows:
                if not row[1]: # Libelle is mandatory
                    continue
                    
                rubrique_id = row[0]
                libelle = str(row[1])
                type_rubrique = str(row[2]).upper() if row[2] else 'GAIN'
                mode_calcul = str(row[3]).upper() if row[3] else 'FIXE'
                cotisable = True if str(row[4]).lower() in ['oui', 'true', '1'] else False
                imposable = True if str(row[5]).lower() in ['oui', 'true', '1'] else False
                actif = True if str(row[6]).lower() in ['oui', 'true', '1', 'none'] else False # Default to true
                
                types_contrat_str = str(row[7]) if row[7] else ""
                types_contrat = [t.strip().upper() for t in types_contrat_str.split(',') if t.strip()]
                
                # Validation of choices
                if type_rubrique not in ['GAIN', 'RETENUE']: type_rubrique = 'GAIN'
                if mode_calcul not in ['FIXE', 'PERCENT', 'HOURS']: mode_calcul = 'FIXE'
                
                defaults = {
                    'libelle': libelle,
                    'type_rubrique': type_rubrique,
                    'mode_calcul': mode_calcul,
                    'est_cotisable': cotisable,
                    'est_imposable': imposable,
                    'actif': actif,
                }
                
                if rubrique_id and str(rubrique_id).isdigit():
                    obj, created = Rubrique.objects.update_or_create(id=int(rubrique_id), defaults=defaults)
                    if created: count_created += 1
                    else: count_updated += 1
                else:
                    obj = Rubrique.objects.create(**defaults)
                    count_created += 1
                
                # Update ManyToMany relationships
                if types_contrat:
                    from t_rh.models import TypesContrat
                    type_objs = TypesContrat.objects.filter(label__in=types_contrat)
                    obj.eligible_types.set(type_objs)
                    
            return JsonResponse({
                'status': 'success', 
                'message': f'Importation réussie : {count_created} rubriques créées, {count_updated} mises à jour.'
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Erreur lors de l\'importation : {str(e)}'})
            
    return JsonResponse({'status': 'error', 'message': 'Fichier manquant ou méthode non autorisée.'})

@module_permission_required('rh', 'view')
def fiches_mensuelles(request):
    """
    View to display monthly course summary for trainers.
    """
    # Get filter params
    formateur_id = request.GET.get('formateur')
    annee_filter = request.GET.get('annee')
    mois_filter = request.GET.get('mois')

    # Base query for completed course sessions
    suivis_qs = SuiviCours.objects.filter(is_done=True).select_related(
        'ligne_presence',
        'ligne_presence__teacher',
        'ligne_presence__module',
        'ligne_presence__registre__groupe'
    )

    # Apply filters
    if formateur_id:
        suivis_qs = suivis_qs.filter(ligne_presence__teacher_id=formateur_id)
    if annee_filter:
        suivis_qs = suivis_qs.filter(date_seance__year=annee_filter)
    if mois_filter:
        suivis_qs = suivis_qs.filter(date_seance__month=mois_filter)

    suivis = suivis_qs.order_by('-date_seance')

    # Prepare data structure: Data -> Trainer -> Year -> Month -> List of sessions
    report_data = collections.defaultdict(lambda: collections.defaultdict(lambda: collections.defaultdict(list)))
    
    # Cache for durations to avoid redundant queries
    durations_cache = {}

    for suivi in suivis:
        teacher = suivi.ligne_presence.teacher
        if not teacher:
            continue
            
        date = suivi.date_seance
        if not date:
            continue
            
        year = date.year
        month = date.month
        
        # Calculate duration
        module = suivi.ligne_presence.module
        groupe = suivi.ligne_presence.registre.groupe if suivi.ligne_presence.registre else None
        
        cache_key = (module.id if module else None, groupe.id if groupe else None)
        if cache_key not in durations_cache:
            # Try to find a TimetableEntry for this module and group
            entries = TimetableEntry.objects.filter(
                cours=module,
                timetable__groupe=groupe
            ).distinct()
            
            total_minutes = 0
            for entry in entries:
                if entry.heure_debut and entry.heure_fin:
                    d = datetime.combine(datetime.today(), entry.heure_debut)
                    f = datetime.combine(datetime.today(), entry.heure_fin)
                    total_minutes += (f - d).total_seconds() / 60
            
            if entries.exists():
                avg_minutes = total_minutes / entries.count()
                durations_cache[cache_key] = avg_minutes / 60
            else:
                # Fallback to LigneRegistrePresence strings if possible
                try:
                    if suivi.ligne_presence.heure_debut and suivi.ligne_presence.heure_fin:
                        h1 = datetime.strptime(suivi.ligne_presence.heure_debut, "%H:%M")
                        h2 = datetime.strptime(suivi.ligne_presence.heure_fin, "%H:%M")
                        durations_cache[cache_key] = (h2 - h1).total_seconds() / 3600
                    else:
                        durations_cache[cache_key] = 0
                except:
                    durations_cache[cache_key] = 0
        
        duration = durations_cache[cache_key]
        
        session_info = {
            'date': date,
            'module': module,
            'groupe': groupe,
            'duration': round(duration, 2),
            'observation': suivi.observation
        }
        
        report_data[teacher][year][month].append(session_info)

    # Fetch validations
    validations = ValidationFicheMensuelleFormateur.objects.all()
    validated_set = {(v.formateur_id, v.annee, v.mois) for v in validations}

    # Fetch existing fiches de paie
    fiches_paie = FichePaie.objects.select_related('contrat').filter(contrat__formateur__isnull=False)
    paie_set = {(fp.contrat.formateur_id, fp.annee, fp.mois) for fp in fiches_paie if fp.contrat}

    # Flatten and format for template
    formatted_report = []
    total_heures_global = 0
    total_sessions_global = 0
    
    for teacher, years in report_data.items():
        contrat = Contrat.objects.filter(formateur=teacher, actif=True).first()
        teacher_report = {
            'teacher': teacher,
            'contrat_id': contrat.id if contrat else None,
            'years': []
        }
        for year, months in sorted(years.items(), reverse=True):
            year_report = {
                'year': year,
                'months': []
            }
            months_fr = {
                1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril',
                5: 'Mai', 6: 'Juin', 7: 'Juillet', 8: 'Août',
                9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre'
            }
            for month, sessions in sorted(months.items(), reverse=True):
                month_total_hours = round(sum(s['duration'] for s in sessions), 2)
                total_heures_global += month_total_hours
                total_sessions_global += len(sessions)
                
                month_report = {
                    'month': month,
                    'month_name': months_fr.get(month, ''),
                    'sessions': sessions,
                    'total_hours': month_total_hours,
                    'is_validated': (teacher.id, year, month) in validated_set,
                    'is_paid': (teacher.id, year, month) in paie_set
                }
                year_report['months'].append(month_report)
            teacher_report['years'].append(year_report)
        formatted_report.append(teacher_report)

    # Metadata for filters
    all_formateurs = Formateurs.objects.all().order_by('nom')
    all_years = SuiviCours.objects.filter(is_done=True, date_seance__isnull=False).values_list('date_seance__year', flat=True).distinct().order_by('-date_seance__year')

    context = {
        'report': formatted_report,
        'all_formateurs': all_formateurs,
        'all_years': all_years,
        'months_choices': [
            (1, 'Janvier'), (2, 'Février'), (3, 'Mars'), (4, 'Avril'), 
            (5, 'Mai'), (6, 'Juin'), (7, 'Juillet'), (8, 'Août'),
            (9, 'Septembre'), (10, 'Octobre'), (11, 'Novembre'), (12, 'Décembre')
        ],
        'selected_formateur': int(formateur_id) if formateur_id and formateur_id.isdigit() else None,
        'selected_annee': int(annee_filter) if annee_filter and annee_filter.isdigit() else None,
        'selected_mois': int(mois_filter) if mois_filter and mois_filter.isdigit() else None,
        'metrics': {
            'total_formateurs': len(report_data.keys()),
            'total_heures': total_heures_global,
            'total_sessions': total_sessions_global
        }
    }
    
    return render(request, 'tenant_folder/rh/fiches_mensuelles.html', context)

@module_permission_required('rh', 'add')
def valider_fiche_mensuelle_formateur(request):
    if request.method == 'POST':
        formateur_id = request.POST.get('formateur_id')
        mois = int(request.POST.get('mois'))
        annee = int(request.POST.get('annee'))
        total_heures = request.POST.get('total_heures', 0)
        
        try:
            formateur = Formateurs.objects.get(id=formateur_id)
            ValidationFicheMensuelleFormateur.objects.get_or_create(
                formateur=formateur,
                mois=mois,
                annee=annee,
                defaults={
                    'total_heures': total_heures,
                    'validated_by': request.user
                }
            )
            return JsonResponse({'status': 'success', 'message': 'Fiche mensuelle validée avec succès.', 'reload_page': True})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
            
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée'})

@method_decorator(module_permission_required('rh', 'view'), name='dispatch')
class FormateurFichePaieListView(LoginRequiredMixin, ListView):
    model = FichePaie
    template_name = 't_ressource_humaine/fiche_paie_list_formateur.html'
    context_object_name = 'fiches'

    def get(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            self.object_list = self.get_queryset()
            context = self.get_context_data()
            return render(request, 't_ressource_humaine/_fiche_paie_table.html', context)
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        active_id = self.request.session.get('active_entreprise_id')
        queryset = super().get_queryset().select_related('contrat__formateur').filter(contrat__formateur__isnull=False).order_by('-annee', '-mois', '-generated_at')
        
        if active_id == 'none':
            queryset = queryset.filter(entreprise__isnull=True)
        elif active_id:
            queryset = queryset.filter(entreprise_id=active_id)
            
        formateur_id = self.request.GET.get('formateur')
        annee = self.request.GET.get('annee')
        
        if formateur_id:
            queryset = queryset.filter(contrat__formateur_id=formateur_id)
        if annee:
            queryset = queryset.filter(annee=annee)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import Entreprise
        context['entreprises'] = Entreprise.objects.all()
        context['active_entreprise_id'] = self.request.session.get('active_entreprise_id')
        
        formateur_id = self.request.GET.get('formateur')
        annee_selected = self.request.GET.get('annee')
        
        context['formateurs'] = Formateurs.objects.all().order_by('nom')
        context['selected_formateur'] = int(formateur_id) if formateur_id and formateur_id.isdigit() else None
        context['selected_annee'] = int(annee_selected) if annee_selected and annee_selected.isdigit() else None
        context['years'] = FichePaie.objects.filter(contrat__formateur__isnull=False).values_list('annee', flat=True).distinct().order_by('-annee')
        
        return context

@module_permission_required('rh', 'view')
def assistant_paie_formateur(request):
    month = int(request.GET.get('month', datetime.now().month))
    year = int(request.GET.get('year', datetime.now().year))
    action = request.GET.get('action')
    
    validations = ValidationFicheMensuelleFormateur.objects.filter(mois=month, annee=year).select_related('formateur')
    
    payroll_data = []
    fiches_existantes = FichePaie.objects.filter(mois=month, annee=year, contrat__formateur__isnull=False).values_list('contrat__formateur_id', flat=True)
    
    for val in validations:
        if val.formateur_id in fiches_existantes:
            continue
            
        contrat = Contrat.objects.filter(formateur=val.formateur, actif=True).first()
        if not contrat:
            continue
            
        heures = Decimal(val.total_heures)
        
        contract_rubrics_config = {rc.rubrique_id: rc for rc in RubriqueContrat.objects.filter(contrat=contrat)}
        all_rubriques = Rubrique.objects.filter(actif=True).prefetch_related('eligible_types')
        lignes_rubriques = []
        for r in all_rubriques:
            if r.eligible_types.exists() and not r.eligible_types.filter(label__icontains=contrat.type_contrat).exists():
                continue
            rc = contract_rubrics_config.get(r.id)
            if rc and rc.actif:
                lignes_rubriques.append({'rubrique': r, 'valeur': rc.valeur})
                
        res = PaieEngine.calculer_paie(
            contrat, 
            jours_travailles=0,
            heures_travailles=heures,
            lignes_rubriques=lignes_rubriques
        )
        
        payroll_data.append({
            'formateur': val.formateur,
            'contrat_obj': contrat,
            'heures': heures,
            'result': res
        })
        
    if action == 'valider':
        from django.db import transaction
        with transaction.atomic():
            for data in payroll_data:
                contrat = data['contrat_obj']
                res = data['result']
                fiche, created = FichePaie.objects.update_or_create(
                    contrat=contrat,
                    mois=month,
                    annee=year,
                    defaults={
                        'entreprise': contrat.entreprise,
                        'heures_travailles': data['heures'],
                        'jours_travailles': 0,
                        'salaire_base_calcule': res['salaire_base_calcule'],
                        'montant_ss': res['montant_ss'],
                        'base_ss': res['base_ss'],
                        'salaire_imposable': res['salaire_imposable'],
                        'irg': res['irg'],
                        'net_a_payer': res['net_a_payer'],
                        'prime_panier': res['prime_panier'],
                        'prime_transport': res['prime_transport'],
                        'is_validated': True
                    }
                )
                
                LignePaie.objects.filter(fiche_paie=fiche).delete()
                for ligne in res['detail_lignes']:
                    LignePaie.objects.create(
                        fiche_paie=fiche,
                        rubrique=ligne['rubrique'],
                        valeur_saisie=ligne['valeur_saisie'],
                        montant=ligne['montant']
                    )
        messages.success(request, f"Les paies de {len(payroll_data)} formateurs ont été générées avec succès.")
        return redirect('t_ressource_humaine:formateur_fiche_paie_list')

    totals = {
        'net_a_payer': sum(data['result']['net_a_payer'] for data in payroll_data),
        'irg': sum(data['result']['irg'] for data in payroll_data),
        'montant_ss': sum(data['result']['montant_ss'] for data in payroll_data),
        'salaire_base_calcule': sum(data['result']['salaire_base_calcule'] for data in payroll_data),
        'count': len(payroll_data)
    }

    context = {
        'month': month,
        'year': year,
        'payroll_data': payroll_data,
        'totals': totals,
        'months_choices': [
            (1, 'Janvier'), (2, 'Février'), (3, 'Mars'), (4, 'Avril'),
            (5, 'Mai'), (6, 'Juin'), (7, 'Juillet'), (8, 'Août'),
            (9, 'Septembre'), (10, 'Octobre'), (11, 'Novembre'), (12, 'Décembre')
        ]
    }
    return render(request, 'tenant_folder/rh/paie/assistant_paie_formateur.html', context)

