import json
from decimal import Decimal
import io
import openpyxl
from openpyxl import Workbook
from django.shortcuts import render, redirect, get_object_or_404
from datetime import date
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, FileResponse
from .models import GlobalPaymentCategory, GlobalDepensesCategory, GlobalPaymentType, PostesBudgetaire, SaaSNotification
from .forms import GlobalPaymentCategoryForm, GlobalDepensesCategoryForm, GlobalPaymentTypeForm
from .utils import sync_global_categories
from django.contrib import messages
from app.models import Institut
from django_tenants.utils import schema_context
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from institut_app.utils_notifications import send_notification_to_module_level

@login_required(login_url='login')
def update_tenant_type(request):
    if request.method == 'POST':
        tenant_id = request.POST.get('tenant_id')
        new_type = request.POST.get('tenant_type')
        
        tenant = get_object_or_404(Institut, id=tenant_id)
        tenant.tenant_type = new_type
        tenant.save()
        
        return JsonResponse({'status': 'success', 'message': f'Le type de l\'institut "{tenant.nom}" a été mis à jour.'})
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

from t_tresorerie.models import PaymentCategory, DepensesCategory, PaymentType
from t_crm.models import Prospets, Opportunite, FicheDeVoeux
from t_formations.models import Promos
from django.db.models import Sum, Count
from institut_app.models import Entreprise
from .models import BudgetCampaign, BudgetLine, PostesBudgetaire, BudgetLineDetail
from .budget_utils import get_campaign_realization_data, month_ratio

@login_required
def sync_payment_types(request):
    if request.method == 'POST':
        try:
            sync_global_categories()
            return JsonResponse({'success': True, 'message': 'Synchronisation terminée avec succès.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Erreur lors de la synchronisation : {str(e)}'})
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée.'}, status=405)
from t_crm.models import Prospets, Opportunite
from django.db.models import Sum, Count
from institut_app.models import Entreprise
from .models import BudgetCampaign, BudgetLine, PostesBudgetaire, BudgetLineDetail

@login_required(login_url='login')
def index(request):
    from .models import BudgetCampaign, BudgetLine, BudgetExtensionRequest, BudgetExtensionItem
    from django.db.models import Sum
    from .budget_utils import get_campaign_realization_data
    
    # 1. Get Filters
    tenant_id = request.GET.get('tenant_id')
    campaign_id = request.GET.get('campaign_id')
    selected_trimester = request.GET.get('trimester', 'all')
    selected_month = request.GET.get('month', 'all')
    view_mode = request.GET.get('view_mode', 'prorata') # 'prorata' or 'full'
    
    # 2. Get Campaigns & Active Campaign
    campaigns = BudgetCampaign.objects.all().order_by('-date_debut')
    if campaign_id:
        active_campaign = campaigns.filter(id=campaign_id).first()
    else:
        active_campaign = campaigns.filter(is_active=True).first()
    
    if not active_campaign:
        return render(request, 'public_folder/configuration_index.html', {
            'title': 'Pilotage Budgétaire',
            'active_campaign': None,
            'campaigns': campaigns,
            'all_instituts': Institut.objects.filter(is_visible=True).exclude(schema_name='public'),
            'selected_tenant': None,
            'selected_trimester': selected_trimester,
            'selected_month': selected_month,
            'view_mode': view_mode,
            'stats': None,
        })

    # 3. Get Tenants
    all_instituts = Institut.objects.filter(is_visible=True).exclude(schema_name='public')
    if tenant_id and tenant_id != 'all':
        selected_tenant = all_instituts.filter(id=tenant_id).first()
        tenants_to_process = [selected_tenant] if selected_tenant else all_instituts
    else:
        selected_tenant = None
        tenants_to_process = all_instituts

    # 4. Determine As-Of Date for ratios
    as_of_date = date.today()
    if selected_month and selected_month != 'all':
        m = int(selected_month)
        y = active_campaign.date_debut.year if m >= 7 else active_campaign.date_debut.year + 1
        import calendar
        last_day = calendar.monthrange(y, m)[1]
        as_of_date = date(y, m, last_day)

    # 5. Get Realization Data
    realization_data = get_campaign_realization_data(active_campaign, tenants_to_process, as_of_date=as_of_date)
    
    # 6. Calculate Summary Stats
    totals = realization_data['totals']
    
    # OBJECTIF ASSIGNÉ
    if selected_tenant:
        line = BudgetLine.objects.filter(campaign=active_campaign, institut=selected_tenant).first()
        target_revenue = line.montant if line else 0
    else:
        # Global target from campaign, or fallback to sum of all institute objectives
        target_revenue = active_campaign.target_revenue or BudgetLine.objects.filter(campaign=active_campaign).aggregate(Sum('montant'))['montant__sum'] or 0
    
    # TOTAL PRÉVU (VALIDÉ)
    # If global, we take the sum of all validated budget lines (envelopes)
    # If filtered, we take the sum of detailed plans (dispatched)
    if selected_tenant:
        total_prevu = totals['dispatched_recette']
    else:
        total_prevu = BudgetLine.objects.filter(
            campaign=active_campaign, 
            statut='validated'
        ).aggregate(Sum('montant'))['montant__sum'] or 0
    
    # ÉCART OBJECTIF / PRÉVU
    ecart_objectif = total_prevu - target_revenue
    
    # RALLONGES APPROUVÉES
    approved_extensions_query = BudgetExtensionItem.objects.filter(
        request__campaign=active_campaign, 
        request__status='approved'
    )
    if selected_tenant:
        approved_extensions_query = approved_extensions_query.filter(request__institut=selected_tenant)
        
    approved_extensions = approved_extensions_query.aggregate(
        total=Sum('requested_amount') - Sum('old_amount')
    )['total'] or 0

    # 6. Preparation for Campus Breakdown
    campus_performance = []
    campus_consommation = []
    
    # Calculate avg_ratio for unallocated portion
    avg_ratio = sum(Decimal(str(realization_data['ratios'].get(t, 0))) for t in ['t1', 't2', 't3', 't4']) / 4
    
    # Pre-fetch budget goals to avoid N+1
    budget_goals = {bl.institut_id: bl.montant for bl in BudgetLine.objects.filter(campaign=active_campaign)}
    
    for inst in tenants_to_process:
        inst_data = get_campaign_realization_data(active_campaign, [inst])
        inst_totals = inst_data['totals']
        
        if selected_month and selected_month != 'all':
            m_key = int(selected_month)
            m_data = inst_data['monthly_totals'][m_key]
            
            # Recettes
            inst_target_m = budget_goals.get(inst.id, 0) / 12
            # Monthly unallocated: approx gap / 12
            inst_unallocated_m = max(0, Decimal(str(inst_target_m)) - m_data['plan_r'])
            # Month ratio handled by as_of_date context
            inst_prorata_r = m_data['pro_r'] + (inst_unallocated_m * Decimal(str(month_ratio(m_key, as_of_date))))
            taux_r = (m_data['real_r'] / inst_prorata_r * 100) if inst_prorata_r > 0 else 0
            
            # Depenses
            inst_prorata_d = m_data['pro_d']
            taux_d = (m_data['real_d'] / inst_prorata_d * 100) if inst_prorata_d > 0 else 0
            
            # Define values for campus lists
            real_r = m_data['real_r']
            real_d = m_data['real_d']
            prev_d = inst_prorata_d
        elif selected_trimester and selected_trimester != 'all':
            t_key = selected_trimester
            t_data = inst_data['trimester_totals'][t_key]
            
            # Recettes
            inst_target_t = budget_goals.get(inst.id, 0) / 4
            inst_unallocated_t = max(0, Decimal(str(inst_target_t)) - t_data['full_r'])
            t_ratio = realization_data['ratios'].get(t_key, 0)
            inst_prorata_r = t_data['pro_r'] + (inst_unallocated_t * Decimal(str(t_ratio)))
            taux_r = (t_data['real_r'] / inst_prorata_r * 100) if inst_prorata_r > 0 else 0
            
            # Depenses
            inst_prorata_d = t_data['pro_d']
            taux_d = (t_data['real_d'] / inst_prorata_d * 100) if inst_prorata_d > 0 else 0
            
            # Define values for campus lists
            real_r = t_data['real_r']
            real_d = t_data['real_d']
            prev_d = inst_prorata_d
        else:
            # Chiffre d'affaires - include unallocated gap for consistency with global gauge
            inst_target = budget_goals.get(inst.id, 0)
            inst_unallocated = max(0, Decimal(str(inst_target)) - inst_totals['dispatched_recette'])
            inst_prorata_r = inst_totals['pro_rata_recette'] + (inst_unallocated * avg_ratio)
            
            real_r = inst_totals['realized_recette']
            taux_r = (real_r / inst_prorata_r * 100) if inst_prorata_r > 0 else 0
            
            # Dépenses - Use pro-rata for "on track" percentage
            prev_d = inst_totals['pro_rata_depense']
            real_d = inst_totals['realized_depense']
            taux_d = (real_d / prev_d * 100) if prev_d > 0 else 0
        
        campus_performance.append({
            'id': inst.id,
            'name': inst.nom,
            'budget': inst_prorata_r,
            'realise': real_r,
            'taux': round(taux_r, 1)
        })
        
        campus_consommation.append({
            'id': inst.id,
            'name': inst.nom,
            'budget': prev_d,
            'realise': real_d,
            'taux': round(taux_d, 1)
        })

    # 7. Detail par Trimestre
    trimestre_details_recette = []
    trimestre_details_depense = []
    
    # Calculate how much of the target_revenue is NOT yet dispatched to specific postes
    total_dispatched_recette_details = 0
    for group in realization_data['combined_postes']:
        for dp in group['display_postes']:
            if dp['poste'].type == 'recette':
                total_dispatched_recette_details += dp['global']['full_prevu']
    
    unallocated_recette = max(0, target_revenue - total_dispatched_recette_details)
    unallocated_per_t = unallocated_recette / 4

    for t in ['t1', 't2', 't3', 't4']:
        t_ratio = realization_data['ratios'].get(t, 0)
        t_full_r = Decimal('0')
        t_prorata_r = Decimal('0')
        t_realise_r = 0
        
        t_full_d = Decimal('0')
        t_prorata_d = Decimal('0')
        t_realise_d = 0
        
        for group in realization_data['combined_postes']:
            for dp in group['display_postes']:
                if dp['poste'].type == 'recette':
                    t_full_r += dp[t]['full_prevu']
                    t_prorata_r += dp[t]['prevu']
                    t_realise_r += dp[t]['realise']
                else:
                    t_full_d += dp[t]['full_prevu']
                    t_prorata_d += dp[t]['prevu']
                    t_realise_d += dp[t]['realise']
        
        # Add unallocated portion to prorata_budget for revenue
        t_unallocated_prorata = Decimal(str(unallocated_per_t)) * Decimal(str(t_ratio))
        t_prorata_r_total = t_prorata_r + t_unallocated_prorata

        trimestre_details_recette.append({
            'label': t.upper(),
            'full_budget': t_full_r,
            'prorata_budget': t_prorata_r_total,
            'realise': t_realise_r,
            'ecart': t_realise_r - t_prorata_r_total,
            'taux': (t_realise_r / t_prorata_r_total * 100) if t_prorata_r_total > 0 else 0
        })
        
        trimestre_details_depense.append({
            'label': t.upper(),
            'full_budget': t_full_d,
            'prorata_budget': t_prorata_d,
            'realise': t_realise_d,
            'restant': t_full_d - t_realise_d,
            'taux': (t_realise_d / t_prorata_d * 100) if t_prorata_d > 0 else 0
        })

    # Filter tables if a specific trimester is selected
    if selected_trimester and selected_trimester != 'all':
        trimestre_details_recette = [t for t in trimestre_details_recette if t['label'].lower() == selected_trimester]
        trimestre_details_depense = [t for t in trimestre_details_depense if t['label'].lower() == selected_trimester]

    # 8. Final Global KPIs (Corrected to include unallocated portion in denominator)
    if selected_month and selected_month != 'all':
        m_key = int(selected_month)
        m_all_data = realization_data['monthly_totals'][m_key]
        
        current_realized_r = m_all_data['real_r']
        current_realized_d = m_all_data['real_d']
        
        # Global Target for the month
        current_target_r = target_revenue / 12
        current_unallocated_m = max(0, Decimal(str(current_target_r)) - m_all_data['plan_r'])
        current_prorata_r = m_all_data['pro_r'] + (current_unallocated_m * Decimal(str(month_ratio(m_key, as_of_date))))
        current_prorata_d = m_all_data['pro_d']
        
        taux_realisation_global = (current_realized_r / current_prorata_r * 100) if current_prorata_r > 0 else 0
        taux_consommation_global = (current_realized_d / current_prorata_d * 100) if current_prorata_d > 0 else 0
        
        global_budget_r = m_all_data['plan_r']
        global_budget_d = m_all_data['plan_d']
    elif selected_trimester and selected_trimester != 'all':
        t_key = selected_trimester
        t_all_data = realization_data['trimester_totals'][t_key]
        
        current_realized_r = t_all_data['real_r']
        current_realized_d = t_all_data['real_d']
        
        # Global Target for the trimester
        current_target_r = target_revenue / 4
        current_unallocated_r = max(0, Decimal(str(current_target_r)) - t_all_data['full_r'])
        t_ratio = realization_data['ratios'].get(t_key, 0)
        current_prorata_r = t_all_data['pro_r'] + (current_unallocated_r * Decimal(str(t_ratio)))
        
        current_prorata_d = t_all_data['pro_d']
        
        taux_realisation_global = (current_realized_r / current_prorata_r * 100) if current_prorata_r > 0 else 0
        taux_consommation_global = (current_realized_d / current_prorata_d * 100) if current_prorata_d > 0 else 0
        
        global_budget_r = t_all_data['full_r']
        global_budget_d = t_all_data['full_d']
    else:
        current_prorata_r = sum(t['prorata_budget'] for t in trimestre_details_recette)
        current_prorata_d = sum(t['prorata_budget'] for t in trimestre_details_depense)

        current_realized_r = totals['realized_recette']
        current_realized_d = totals['realized_depense']
        
        taux_realisation_global = (current_realized_r / current_prorata_r * 100) if current_prorata_r > 0 else 0
        taux_consommation_global = (current_realized_d / current_prorata_d * 100) if current_prorata_d > 0 else 0
        
        global_budget_r = totals['dispatched_recette']
        global_budget_d = totals['dispatched_depense']

    # 8. Override Prorata if View Mode is 'Full'
    if view_mode == 'full':
        current_prorata_r = global_budget_r + unallocated_recette
        current_prorata_d = global_budget_d
        
        taux_realisation_global = (current_realized_r / current_prorata_r * 100) if current_prorata_r > 0 else 0
        taux_consommation_global = (current_realized_d / current_prorata_d * 100) if current_prorata_d > 0 else 0
        
        # Also update trimester details for 'Full' view
        for t_r in trimestre_details_recette:
            t_r['prorata_budget'] = t_r['full_budget'] + Decimal(str(unallocated_recette / 4))
            t_r['taux'] = (t_r['realise'] / t_r['prorata_budget'] * 100) if t_r['prorata_budget'] > 0 else 0
            t_r['ecart'] = t_r['realise'] - t_r['prorata_budget']
            
        for t_d in trimestre_details_depense:
            t_d['prorata_budget'] = t_d['full_budget']
            t_d['taux'] = (t_d['realise'] / t_d['prorata_budget'] * 100) if t_d['prorata_budget'] > 0 else 0
            t_d['restant'] = t_d['full_budget'] - t_d['realise']

        for cp in campus_performance:
            # Need to re-calculate inst_target if possible, or use pro-rata as best effort
            # For simplicity in 'Full' mode, we use the dispatched target
            pass # Keep them as is or refine if user insists on full campus target

    # 9. Margin Calculations (Universal)
    global_marge_budget = current_prorata_r - current_prorata_d
    global_marge_realise = current_realized_r - current_realized_d
    taux_marge_realisation = (global_marge_realise / global_marge_budget * 100) if global_marge_budget != 0 else 0

    # 9. Trimester Details for Margin
    trimestre_details_marge = []
    for t_key in ['t1', 't2', 't3', 't4']:
        # Find matching details in revenue and expense lists
        t_r_item = next((t for t in trimestre_details_recette if t['label'].lower() == t_key), None)
        t_d_item = next((t for t in trimestre_details_depense if t['label'].lower() == t_key), None)
        
        if t_r_item and t_d_item:
            marge_b = t_r_item['prorata_budget'] - t_d_item['prorata_budget']
            marge_r = t_r_item['realise'] - t_d_item['realise']
            
            trimestre_details_marge.append({
                'label': t_r_item['label'],
                'prevu': marge_b,
                'realise': marge_r,
                'ecart': marge_r - marge_b,
                'taux': (marge_r / marge_b * 100) if marge_b != 0 else 0
            })

    # 10. Campus Performance for Margin
    campus_marge = []
    for cp_r in campus_performance:
        cp_d = next((c for c in campus_consommation if c['name'] == cp_r['name']), {'budget': 0, 'realise': 0})
        marge_b = cp_r['budget'] - cp_d['budget']
        marge_r = cp_r['realise'] - cp_d['realise']
        campus_marge.append({
            'name': cp_r['name'],
            'budget': marge_b,
            'realise': marge_r,
            'taux': round((marge_r / marge_b * 100), 1) if marge_b != 0 else 0
        })

    # Filter tables if a specific trimester is selected
    if selected_trimester and selected_trimester != 'all':
        trimestre_details_marge = [t for t in trimestre_details_marge if t['label'].lower() == selected_trimester]

    MONTHS = [
        (7, 'Juillet'), (8, 'Août'), (9, 'Septembre'), (10, 'Octobre'),
        (11, 'Novembre'), (12, 'Décembre'), (1, 'Janvier'),
        (2, 'Février'), (3, 'Mars'), (4, 'Avril'),
        (5, 'Mai'), (6, 'Juin')
    ]

    return render(request, 'public_folder/configuration_index.html', {
        'title': 'Pilotage Budgétaire',
        'active_campaign': active_campaign,
        'campaigns': campaigns,
        'all_instituts': all_instituts,
        'selected_tenant': selected_tenant,
        'selected_trimester': selected_trimester,
        'selected_month': selected_month,
        'month_list': MONTHS,
        'stats': {
            'objectif_global': target_revenue,
            'total_prevu': total_prevu,
            'ecart_objectif': ecart_objectif,
            'rallonges_approuvees': approved_extensions,
            
            'ca_global': {
                'taux': round(taux_realisation_global, 1),
                'budget': global_budget_r,
                'realise': current_realized_r,
                'ecart': current_realized_r - global_budget_r,
                'trimestres': trimestre_details_recette,
                'campus': campus_performance
            },
            'depenses_globales': {
                'taux': round(taux_consommation_global, 1),
                'budget': global_budget_d,
                'realise': current_realized_d,
                'ecart': current_realized_d - global_budget_d,
                'trimestres': trimestre_details_depense,
                'campus': campus_consommation
            },
            'marge_globale': {
                'taux': round(taux_marge_realisation, 1),
                'budget': global_marge_budget,
                'realise': global_marge_realise,
                'ecart': global_marge_realise - global_marge_budget,
                'trimestres': trimestre_details_marge,
                'campus': campus_marge
            },
            'unallocated_recette': unallocated_recette,
            'view_mode': view_mode,
        }
    })

@login_required(login_url='login')
def mise_en_route(request):
    from .models import BudgetCampaign, BudgetLine, BudgetExtensionRequest
    
    # 1. Base Data
    campaigns = BudgetCampaign.objects.all().order_by('-date_debut')
    total_instituts = Institut.objects.filter(is_visible=True).exclude(schema_name='public').count()
    
    # 2. Process Campaigns (Progress Calculation)
    for campaign in campaigns:
        # Progress = % of institutes with VALIDATED budget
        validated_count = BudgetLine.objects.filter(campaign=campaign, statut='validated').count()
        campaign.progress = (validated_count / total_instituts * 100) if total_instituts > 0 else 0
        campaign.validated_count = validated_count
        
    # 3. Active Campaign Tracking
    active_campaign = campaigns.filter(is_active=True).first()
    active_stats = None
    if active_campaign:
        lines = BudgetLine.objects.filter(campaign=active_campaign)
        active_stats = {
            'validated': lines.filter(statut='validated').count(),
            'submitted': lines.filter(statut='submitted').count(),
            'draft': lines.filter(statut='draft').count(),
            'rejected': lines.filter(statut='rejected').count(),
            'total': total_instituts,
            'pending': total_instituts - lines.count()
        }
    
    # 4. Global Stats
    global_stats = {
        'total_campaigns': campaigns.count(),
        'active_count': campaigns.filter(is_active=True).count(),
        'pending_extensions': BudgetExtensionRequest.objects.filter(status='pending').count(),
        'total_instituts': total_instituts
    }
    
    return render(request, 'associe_app/mise_en_route.html', {
        'title': 'Mise en Route - Budget',
        'campaigns': campaigns,
        'active_campaign': active_campaign,
        'active_stats': active_stats,
        'global_stats': global_stats,
    })

@login_required(login_url='login')
def configuration_budget(request):
    return render(request, 'public_folder/configuration.html', {'title': 'Configuration Budget'})

@login_required(login_url='login')
def configuration_structure(request):
    return render(request, 'public_folder/configuration.html', {'title': 'Configuration Structure'})

# --- Global Payment Type Views ---

@login_required(login_url='login')
def global_payment_type_list(request):
    types = GlobalPaymentType.objects.all().order_by('name')
    types_list = []
    for t in types:
        cats = t.payment_categories.all()
        types_list.append({
            'id': t.id,
            'name': t.name,
            'payment_categories': [{'id': c.id, 'name': c.name} for c in cats],
            'category_ids': [c.id for c in cats]
        })
    
    import json
    types_json = json.dumps(types_list)
    
    form = GlobalPaymentTypeForm()
    return render(request, 'associe_app/global_payment_type_list.html', {
        'types': types, 
        'types_json': types_json,
        'form': form,
        'title': 'Types de Paiement Globaux'
    })

@login_required(login_url='login')
def global_payment_type_create(request):
    if request.method == 'POST':
        form = GlobalPaymentTypeForm(request.POST)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success', 'message': 'Type de paiement créé avec succès.'})
            return redirect('associe_app:global_payment_type_list')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors.get_json_data()})
    return redirect('associe_app:global_payment_type_list')

@login_required(login_url='login')
def global_payment_type_edit(request, pk):
    payment_type = get_object_or_404(GlobalPaymentType, pk=pk)
    if request.method == 'POST':
        form = GlobalPaymentTypeForm(request.POST, instance=payment_type)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success', 'message': 'Type de paiement modifié avec succès.'})
            return redirect('associe_app:global_payment_type_list')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors.get_json_data()})
    
    # For loading form data via AJAX
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'name': payment_type.name,
            'payment_categories': list(payment_type.payment_categories.values_list('id', flat=True))
        })
    return redirect('associe_app:global_payment_type_list')

@login_required(login_url='login')
def global_payment_type_delete(request, pk):
    payment_type = get_object_or_404(GlobalPaymentType, pk=pk)
    if request.method == 'POST':
        payment_type.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': 'Type de paiement supprimé avec succès.'})
    return redirect('associe_app:global_payment_type_list')

# --- Helper Functions for Real-Time Serialization ---

def _get_payment_categories_list():
    categories = GlobalPaymentCategory.objects.all().order_by('name')
    categories_list = []
    for cat in categories:
        categories_list.append({
            'id': cat.id,
            'name': cat.name,
            'parent_id': cat.parent.id if cat.parent else None,
            'type': cat.category_type,
            'type_display': cat.get_category_type_display(),
            'description': cat.description or ''
        })
    return categories_list

def _get_depenses_categories_list():
    categories = GlobalDepensesCategory.objects.all().order_by('name')
    categories_list = []
    for cat in categories:
        categories_list.append({
            'id': cat.id,
            'name': cat.name,
            'parent_id': cat.parent.id if cat.parent else None,
            'payment_category': cat.payment_category.name if cat.payment_category else '-',
            'payment_category_id': cat.payment_category.id if cat.payment_category else None,
            'description': cat.description or ''
        })
    return categories_list

# --- Global Payment Category Views ---

@login_required(login_url='login')
def global_payment_category_list(request):
    categories_list = _get_payment_categories_list()
    import json
    categories_json = json.dumps(categories_list)
    form = GlobalPaymentCategoryForm()
    return render(request, 'associe_app/global_payment_category_list.html', {
        'categories_json': categories_json,
        'form': form,
        'title': 'Catégories de Paiement Globales'
    })

@login_required(login_url='login')
def global_payment_category_create(request):
    if request.method == 'POST':
        form = GlobalPaymentCategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success', 
                    'message': 'Catégorie de paiement créée avec succès.',
                    'categories': _get_payment_categories_list(),
                    'new_id': category.id,
                    'parent_id': category.parent.id if category.parent else None
                })
            return redirect('associe_app:global_payment_category_list')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors.get_json_data()})
    return redirect('associe_app:global_payment_category_list')

@login_required(login_url='login')
def global_payment_category_edit(request, pk):
    category = get_object_or_404(GlobalPaymentCategory, pk=pk)
    if request.method == 'POST':
        form = GlobalPaymentCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success', 
                    'message': 'Catégorie de paiement modifiée avec succès.',
                    'categories': _get_payment_categories_list()
                })
            return redirect('associe_app:global_payment_category_list')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors.get_json_data()})
    
    # For loading form data via AJAX
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'name': category.name,
            'parent': category.parent.id if category.parent else '',
            'category_type': category.category_type,
            'description': category.description or ''
        })
    return redirect('associe_app:global_payment_category_list')

@login_required(login_url='login')
def global_payment_category_delete(request, pk):
    category = get_object_or_404(GlobalPaymentCategory, pk=pk)
    if request.method == 'POST':
        category.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'success', 
                'message': 'Catégorie de paiement supprimée avec succès.',
                'categories': _get_payment_categories_list()
            })
    return redirect('associe_app:global_payment_category_list')

@login_required(login_url='login')
def export_payment_categories(request):
    categories = GlobalPaymentCategory.objects.all().order_by('id')
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Categories"
    
    headers = ['ID', 'Nom', 'Parent_ID', 'Parent_Nom', 'Type', 'Description']
    ws.append(headers)
    
    for cat in categories:
        ws.append([
            cat.id,
            cat.name,
            cat.parent.id if cat.parent else '',
            cat.parent.name if cat.parent else '',
            cat.category_type,
            cat.description or ''
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = FileResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="categories_paiement.xlsx"'
    return response

@login_required(login_url='login')
def import_payment_categories(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        filename = file.name.lower()
        data = []
        
        try:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row): continue
                    data.append(dict(zip(headers, row)))
            elif filename.endswith('.json'):
                data = json.load(file)
            else:
                return JsonResponse({'status': 'error', 'message': 'Format non supporté.'})
            
            # Hierarchical Import Logic (Simple multi-pass)
            success_count = 0
            # Pass 1: Roots and updates
            for row in data:
                parent_id = row.get('Parent_ID') or row.get('parent_id')
                # If no parent or parent exists, we can create/update
                parent = None
                if parent_id:
                    parent = GlobalPaymentCategory.objects.filter(id=parent_id).first()
                
                GlobalPaymentCategory.objects.update_or_create(
                    id=row.get('ID') or row.get('id'),
                    defaults={
                        'name': row.get('Nom') or row.get('name'),
                        'parent': parent,
                        'category_type': row.get('Type') or row.get('category_type') or 'standard',
                        'description': row.get('Description') or row.get('description') or ''
                    }
                )
                success_count += 1
                
            return JsonResponse({
                'status': 'success', 
                'message': f'{success_count} catégories importées.',
                'categories': _get_payment_categories_list()
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Erreur: {str(e)}'})
            
    return JsonResponse({'status': 'error', 'message': 'Requête invalide.'})

# --- Global Depenses Category Views ---

@login_required(login_url='login')
def global_depenses_category_list(request):
    categories_list = _get_depenses_categories_list()
    import json
    categories_json = json.dumps(categories_list)
    form = GlobalDepensesCategoryForm()
    return render(request, 'associe_app/global_depenses_category_list.html', {
        'categories_json': categories_json,
        'form': form,
        'title': 'Catégories de Dépenses Globales'
    })

@login_required(login_url='login')
def global_depenses_category_create(request):
    if request.method == 'POST':
        form = GlobalDepensesCategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success', 
                    'message': 'Catégorie de dépense créée avec succès.',
                    'categories': _get_depenses_categories_list(),
                    'new_id': category.id,
                    'parent_id': category.parent.id if category.parent else None
                })
            return redirect('associe_app:global_depenses_category_list')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors.get_json_data()})
    return redirect('associe_app:global_depenses_category_list')

@login_required(login_url='login')
def export_depenses_categories(request):
    categories = GlobalDepensesCategory.objects.all().order_by('id')
    wb = Workbook()
    ws = wb.active
    ws.title = "Depenses Categories"
    
    headers = ['ID', 'Nom', 'Parent_ID', 'Parent_Nom', 'Payment_Category_ID', 'Payment_Category_Nom', 'Description']
    ws.append(headers)
    
    for cat in categories:
        ws.append([
            cat.id,
            cat.name,
            cat.parent.id if cat.parent else '',
            cat.parent.name if cat.parent else '',
            cat.payment_category.id if cat.payment_category else '',
            cat.payment_category.name if cat.payment_category else '',
            cat.description or ''
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = FileResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="categories_depenses.xlsx"'
    return response

@login_required(login_url='login')
def import_depenses_categories(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        filename = file.name.lower()
        data = []
        
        try:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row): continue
                    data.append(dict(zip(headers, row)))
            elif filename.endswith('.json'):
                data = json.load(file)
            else:
                return JsonResponse({'status': 'error', 'message': 'Format non supporté.'})
            
            success_count = 0
            for row in data:
                parent_id = row.get('Parent_ID') or row.get('parent_id')
                pay_cat_id = row.get('Payment_Category_ID') or row.get('payment_category_id')
                
                parent = None
                if parent_id:
                    parent = GlobalDepensesCategory.objects.filter(id=parent_id).first()
                
                pay_cat = None
                if pay_cat_id:
                    pay_cat = GlobalPaymentCategory.objects.filter(id=pay_cat_id).first()

                GlobalDepensesCategory.objects.update_or_create(
                    id=row.get('ID') or row.get('id'),
                    defaults={
                        'name': row.get('Nom') or row.get('name'),
                        'parent': parent,
                        'payment_category': pay_cat,
                        'description': row.get('Description') or row.get('description') or ''
                    }
                )
                success_count += 1
                
            return JsonResponse({
                'status': 'success', 
                'message': f'{success_count} catégories importées.',
                'categories': _get_depenses_categories_list()
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Erreur: {str(e)}'})
            
    return JsonResponse({'status': 'error', 'message': 'Requête invalide.'})

@login_required(login_url='login')
def global_depenses_category_edit(request, pk):
    category = get_object_or_404(GlobalDepensesCategory, pk=pk)
    if request.method == 'POST':
        form = GlobalDepensesCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success', 
                    'message': 'Catégorie de dépense modifiée avec succès.',
                    'categories': _get_depenses_categories_list()
                })
            return redirect('associe_app:global_depenses_category_list')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors.get_json_data()})
    
    # For loading form data via AJAX
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'name': category.name,
            'parent': category.parent.id if category.parent else '',
            'payment_category': category.payment_category.id if category.payment_category else '',
            'description': category.description or ''
        })
    return redirect('associe_app:global_depenses_category_list')

@login_required(login_url='login')
def global_depenses_category_delete(request, pk):
    category = get_object_or_404(GlobalDepensesCategory, pk=pk)
    if request.method == 'POST':
        category.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'success', 
                'message': 'Catégorie de dépense supprimée avec succès.',
                'categories': _get_depenses_categories_list()
            })
    return redirect('associe_app:global_depenses_category_list')

@login_required(login_url='login')
def sync_categories_view(request):
    try:
        sync_global_categories()
        msg = "Synchronisation effectuée avec succès sur tous les tenants."
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': msg})
        messages.success(request, msg)
    except Exception as e:
        msg = f"Erreur lors de la synchronisation : {str(e)}"
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': msg})
        messages.error(request, msg)
    
    return redirect(request.META.get('HTTP_REFERER', 'associe_app:configuration_index'))

@login_required(login_url='login')
def tenant_data_list(request):
    tenants = Institut.objects.filter(is_visible=True).exclude(schema_name='public')
    return render(request, 'associe_app/tenant_visualization.html', {
        'tenants': tenants,
        'title': 'Visualisation des Données Tenants'
    })

@login_required(login_url='login')
def tenant_workspace(request, tenant_id):
    tenant = get_object_or_404(Institut, id=tenant_id)
    
    from django.apps import apps
    from django.conf import settings
    
    grouped_models = {}
    for app_name in settings.TENANT_APPS:
        if app_name.startswith('django.') or app_name in ['phonenumber_field', 'django_countries', 'taggit', 'pdf_editor']:
            continue
        try:
            app_config = apps.get_app_config(app_name.split('.')[-1])
            app_label_name = str(app_config.verbose_name).capitalize() if app_config.verbose_name else app_config.label.capitalize()
            
            for model in app_config.get_models():
                if app_label_name not in grouped_models:
                    grouped_models[app_label_name] = []
                grouped_models[app_label_name].append({
                    'id': f"{app_config.label}.{model.__name__}",
                    'name': str(model._meta.verbose_name).capitalize()
                })
        except LookupError:
            pass
            
    # Sort models within groups
    for app_label_name in grouped_models:
        grouped_models[app_label_name] = sorted(grouped_models[app_label_name], key=lambda x: x['name'])
    
    # Sort groups by name alphabetically
    grouped_models = dict(sorted(grouped_models.items()))
    
    return render(request, 'associe_app/tenant_workspace.html', {
        'tenant': tenant,
        'grouped_models': grouped_models,
        'title': f'Espace de Travail - {tenant.nom}'
    })

@login_required(login_url='login')
def get_tenant_categories(request, tenant_id):
    tenant = get_object_or_404(Institut, id=tenant_id)
    payment_categories = []
    depenses_categories = []
    crm_stats = {}
    
    with schema_context(tenant.schema_name):
        # Fetch payment categories
        p_cats = PaymentCategory.objects.all()
        for pc in p_cats:
            payment_categories.append({
                'id': pc.id,
                'name': pc.name,
                'parent': pc.parent.name if pc.parent else '-',
                'type': pc.get_category_type_display()
            })
            
        # Fetch depenses categories
        d_cats = DepensesCategory.objects.all()
        for dc in d_cats:
            depenses_categories.append({
                'id': dc.id,
                'name': dc.name,
                'parent': dc.parent.name if dc.parent else '-',
                'payment_category': dc.payment_category.name if dc.payment_category else '-'
            })

        # Fetch payment types
        p_types_data = []
        p_types = PaymentType.objects.all()
        for pt in p_types:
            p_types_data.append({
                'id': pt.id,
                'name': pt.name,
                'global_id': pt.global_id,
                'categories': [c.name for c in pt.payment_categories.all()]
            })

        # CRM Statistics
        all_prospects = Prospets.objects.all()
        crm_stats = {
            'total_prospects': all_prospects.count(),
            'by_status': list(all_prospects.values('statut').annotate(count=Count('statut'))),
            'by_etat': list(all_prospects.values('etat').annotate(count=Count('etat'))),
            'active_opportunities': Opportunite.objects.filter(is_active=True).count(),
            'total_budget': float(Opportunite.objects.filter(is_active=True).aggregate(total=Sum('budget'))['total'] or 0)
        }

    return JsonResponse({
        'status': 'success',
        'tenant_name': tenant.nom,
        'payment_categories': payment_categories,
        'depenses_categories': depenses_categories,
        'payment_types': p_types_data,
        'crm_stats': crm_stats
    })

@login_required(login_url='login')
def delete_tenant_payment_type(request, tenant_id, payment_type_id):
    tenant = get_object_or_404(Institut, id=tenant_id)
    print(f"Attempting to delete PT {payment_type_id} for tenant {tenant.nom} ({tenant.schema_name})")
    if request.method == 'POST':
        try:
            with schema_context(tenant.schema_name):
                p_type = get_object_or_404(PaymentType, id=payment_type_id)
                name = p_type.name
                p_type.delete()
                print(f"Successfully deleted {name}")
            return JsonResponse({
                'status': 'success', 
                'message': f'Type de paiement "{name}" supprimé du tenant "{tenant.nom}".'
            })
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"Error during deletion for tenant {tenant.nom}: {error_trace}")
            return JsonResponse({
                'status': 'error', 
                'message': str(e),
                'traceback': error_trace
            })
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

@login_required(login_url='login')
def crm_statistics(request):
    from t_crm.models import DocumentsDemandeInscription
    tenants = Institut.objects.filter(is_visible=True).exclude(schema_name='public').order_by('nom')
    is_print = request.GET.get('print') == 'true'
    institutes_stats = []
    
    # Prospect Status List
    statuses = ['visiteur', 'prinscrit', 'instance', 'convertit', 'annuler']
    global_stats = {
        'total_prospects': 0,
        'active_opportunities': 0,
        'total_budget': 0,
        'status_counts': {s: 0 for s in statuses}
    }

    from django.utils import timezone
    from datetime import date
    
    today = timezone.now().date()
    if today.month >= 7:
        year_n_start = today.year
    else:
        year_n_start = today.year - 1

    n_t1_start = date(year_n_start, 7, 1)
    n_t1_end = date(year_n_start, 9, 30)
    n_t2_start = date(year_n_start, 10, 1)
    n_t2_end = date(year_n_start, 12, 31)
    n_t3_start = date(year_n_start + 1, 1, 1)
    n_t3_end = date(year_n_start + 1, 3, 31)
    n_t4_start = date(year_n_start + 1, 4, 1)
    n_t4_end = date(year_n_start + 1, 6, 30)

    year_n_1_start = year_n_start - 1
    n_1_t1_start = date(year_n_1_start, 7, 1)
    n_1_t1_end = date(year_n_1_start, 9, 30)
    n_1_t2_start = date(year_n_1_start, 10, 1)
    n_1_t2_end = date(year_n_1_start, 12, 31)
    n_1_t3_start = date(year_n_1_start + 1, 1, 1)
    n_1_t3_end = date(year_n_1_start + 1, 3, 31)
    n_1_t4_start = date(year_n_1_start + 1, 4, 1)
    n_1_t4_end = date(year_n_1_start + 1, 6, 30)

    def get_quarter(d):
        if n_t1_start <= d <= n_t1_end: return 'n', 't1'
        if n_t2_start <= d <= n_t2_end: return 'n', 't2'
        if n_t3_start <= d <= n_t3_end: return 'n', 't3'
        if n_t4_start <= d <= n_t4_end: return 'n', 't4'
        if n_1_t1_start <= d <= n_1_t1_end: return 'n_1', 't1'
        if n_1_t2_start <= d <= n_1_t2_end: return 'n_1', 't2'
        if n_1_t3_start <= d <= n_1_t3_end: return 'n_1', 't3'
        if n_1_t4_start <= d <= n_1_t4_end: return 'n_1', 't4'
        return None, None

    kpi_keys = ['prospects', 'opportunites'] + statuses
    comparison_data = {
        kpi: {
            't1': {'n': 0, 'n_1': 0},
            't2': {'n': 0, 'n_1': 0},
            't3': {'n': 0, 'n_1': 0},
            't4': {'n': 0, 'n_1': 0},
            'global': {'n': 0, 'n_1': 0}
        } for kpi in kpi_keys
    }

    for tenant in tenants:
        try:
            with schema_context(tenant.schema_name):
                # Basic counts
                all_prospects = Prospets.objects.all()
                prospects_count = all_prospects.count()
                
                tenant_status_counts = {s: all_prospects.filter(statut=s).count() for s in statuses}
                docs_count = DocumentsDemandeInscription.objects.all().count()
                
                opportunities = Opportunite.objects.filter(is_active=True)
                opportunities_count = opportunities.count()
                budget_sum = float(opportunities.aggregate(total=Sum('budget'))['total'] or 0)

                # Status breakdown for global summation
                for s in statuses:
                    global_stats['status_counts'][s] += tenant_status_counts[s]

                # Promo Breakdown
                promo_stats = []
                promos = Promos.objects.filter(is_archived=False).order_by('-begin_year')[:10] # Last 10 promos
                for promo in promos:
                    p_ids = FicheDeVoeux.objects.filter(promo=promo).values_list('prospect_id', flat=True)
                    promo_prospects = Prospets.objects.filter(id__in=p_ids)
                    promo_total = promo_prospects.count()
                    if promo_total > 0:
                        s_counts = {s: promo_prospects.filter(statut=s).count() for s in statuses}
                        promo_stats.append({
                            'label': promo.label,
                            'total': promo_total,
                            'status_counts': s_counts
                        })

                institutes_stats.append({
                    'id': tenant.id,
                    'name': tenant.nom,
                    'prospects_count': prospects_count,
                    'status_counts': tenant_status_counts,
                    'docs_count': docs_count,
                    'opportunities_count': opportunities_count,
                    'budget_sum': budget_sum,
                    'promo_stats': promo_stats
                })

                global_stats['total_prospects'] += prospects_count
                global_stats['active_opportunities'] += opportunities_count
                global_stats['total_budget'] += budget_sum

                # Comparison logic
                comp_prospects = Prospets.objects.filter(created_at__date__gte=n_1_t1_start, created_at__date__lte=n_t4_end)
                for p in comp_prospects:
                    year, quarter = get_quarter(p.created_at.date())
                    if year:
                        comparison_data['prospects'][quarter][year] += 1
                        comparison_data['prospects']['global'][year] += 1
                        if p.statut in comparison_data:
                            comparison_data[p.statut][quarter][year] += 1
                            comparison_data[p.statut]['global'][year] += 1

                comp_opportunites = Opportunite.objects.filter(created_at__date__gte=n_1_t1_start, created_at__date__lte=n_t4_end)
                for o in comp_opportunites:
                    year, quarter = get_quarter(o.created_at.date())
                    if year:
                        comparison_data['opportunites'][quarter][year] += 1
                        comparison_data['opportunites']['global'][year] += 1
        except Exception as e:
            continue

    context = {
        'tenants': tenants, # For the selector
        'institutes_stats': institutes_stats,
        'global_stats': global_stats,
        'statuses': statuses,
        'comparison_data': comparison_data,
        'year_n': f"{year_n_start}-{year_n_start+1}",
        'year_n_1': f"{year_n_1_start}-{year_n_1_start+1}",
        'title': 'Statistiques CRM Globales'
    }
    
    if is_print:
        return render(request, 'associe_app/crm_statistics_print.html', context)
        
    return render(request, 'associe_app/crm_statistics.html', context)

@login_required(login_url='login')
def get_crm_stats_api(request, tenant_id):
    """Dynamic API for CRM stats of a specific institute"""
    tenant = get_object_or_404(Institut, id=tenant_id)
    statuses = ['visiteur', 'prinscrit', 'instance', 'convertit', 'annuler']
    
    try:
        with schema_context(tenant.schema_name):
            all_prospects = Prospets.objects.all()
            total_prospects = all_prospects.count()
            
            opportunities = Opportunite.objects.filter(is_active=True)
            active_opportunities = opportunities.count()
            total_budget = float(opportunities.aggregate(total=Sum('budget'))['total'] or 0)
            
            status_breakdown = []
            for s in statuses:
                count = all_prospects.filter(statut=s).count()
                status_breakdown.append({
                    'statut': s,
                    'label': s.capitalize(), # Map these better in frontend or backend
                    'count': count,
                    'percentage': (count / total_prospects * 100) if total_prospects > 0 else 0
                })

            # Promo Breakdown
            promo_stats = []
            promos = Promos.objects.filter(is_archived=False).order_by('-begin_year')
            for promo in promos:
                fiches = FicheDeVoeux.objects.filter(promo=promo).select_related('prospect', 'specialite', 'specialite__formation')
                promo_total = fiches.count()
                
                if promo_total > 0:
                    # Global status breakdown for the promo
                    s_counts = []
                    for s in statuses:
                        count = fiches.filter(prospect__statut=s).count()
                        s_counts.append({
                            'statut': s,
                            'count': count,
                            'percentage': (count / promo_total * 100) if promo_total > 0 else 0
                        })
                    
                    # Detailed breakdown by Formation / Specialite
                    formation_data = {}
                    for f in fiches:
                        spec = f.specialite
                        if not spec: continue
                        form = spec.formation
                        form_name = form.nom if form else "Sans Formation"
                        
                        if form_name not in formation_data:
                            formation_data[form_name] = {}
                        
                        spec_name = spec.label
                        if spec_name not in formation_data[form_name]:
                            formation_data[form_name][spec_name] = {st: 0 for st in statuses}
                            formation_data[form_name][spec_name]['total'] = 0
                        
                        status = f.prospect.statut
                        if status in statuses:
                            formation_data[form_name][spec_name][status] += 1
                            formation_data[form_name][spec_name]['total'] += 1
                    
                    # Convert to list for JSON
                    form_list = []
                    for fn, specs in formation_data.items():
                        spec_list = []
                        for sn, s_counts_spec in specs.items():
                            spec_list.append({
                                'label': sn,
                                'total': s_counts_spec['total'],
                                'status_counts': {st: s_counts_spec[st] for st in statuses}
                            })
                        form_list.append({
                            'label': fn,
                            'specialites': spec_list
                        })

                    promo_stats.append({
                        'id': promo.id,
                        'label': promo.label,
                        'session': promo.session,
                        'total': promo_total,
                        'status_breakdown': s_counts,
                        'formations': form_list
                    })

            return JsonResponse({
                'status': 'success',
                'tenant_name': tenant.nom,
                'total_prospects': total_prospects,
                'active_opportunities': active_opportunities,
                'total_budget': total_budget,
                'status_breakdown': status_breakdown,
                'promo_stats': promo_stats
            })
    except Exception as e:
        import traceback
        return JsonResponse({'status': 'error', 'message': str(e), 'trace': traceback.format_exc()})

@login_required(login_url='login')
def purge_tenant_categories(request, tenant_id):
    tenant = get_object_or_404(Institut, id=tenant_id)
    
    if request.method == 'POST':
        try:
            with schema_context(tenant.schema_name):
                # Delete all synchronized categories
                # Note: We are deleting from t_tresorerie.models, not associe_app.models
                DepensesCategory.objects.all().delete()
                PaymentCategory.objects.all().delete()
                
            return JsonResponse({
                'status': 'success', 
                'message': f'Données du tenant "{tenant.nom}" purgées avec succès.'
            })
        except Exception as e:
            return JsonResponse({
                'status': 'error', 
                'message': f'Erreur lors de la purge: {str(e)}'
            })
            
            return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

from django.db.models import ProtectedError
from django.apps import apps

@login_required(login_url='login')
def reset_tenant_table(request):
    if request.method == 'POST':
        tenant_id = request.POST.get('tenant_id')
        model_identifier = request.POST.get('model_identifier')
        
        if not tenant_id or not model_identifier:
             return JsonResponse({'status': 'error', 'message': 'Paramètres manquants.'})
             
        tenant = get_object_or_404(Institut, id=tenant_id)
        
        try:
            app_label, model_name = model_identifier.split('.')
            model = apps.get_model(app_label, model_name)
        except (ValueError, LookupError):
            return JsonResponse({'status': 'error', 'message': 'Modèle introuvable.'})
            
        try:
            with schema_context(tenant.schema_name):
                from django.db import connection
                table_name = model._meta.db_table
                with connection.cursor() as cursor:
                    cursor.execute(f'TRUNCATE TABLE "{tenant.schema_name}"."{table_name}" CASCADE;')
            return JsonResponse({'status': 'success', 'message': f'La table a été vidée avec succès, incluant toutes les données dépendantes (Cascade).'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Erreur lors du vidage de la table: {str(e)}'})
            
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

@login_required(login_url='login')
def preview_tenant_table_reset(request):
    if request.method == 'GET':
        tenant_id = request.GET.get('tenant_id')
        model_identifier = request.GET.get('model_identifier')
        
        if not tenant_id or not model_identifier:
            return JsonResponse({'status': 'error', 'message': 'Paramètres manquants.'})
             
        tenant = get_object_or_404(Institut, id=tenant_id)
        
        try:
            app_label, model_name = model_identifier.split('.')
            base_model = apps.get_model(app_label, model_name)
        except (ValueError, LookupError):
            return JsonResponse({'status': 'error', 'message': 'Modèle introuvable.'})

        def get_dependent_models(model, seen=None):
            from django.contrib.contenttypes.models import ContentType
            if seen is None:
                seen = set()
            if model in seen:
                return seen
            seen.add(model)
            for f in model._meta.get_fields():
                if (f.one_to_many or f.one_to_one) and f.auto_created and not f.concrete:
                    rel_model = f.related_model
                    if rel_model and not rel_model._meta.proxy:
                        get_dependent_models(rel_model, seen)
            return seen

        dependent_models = get_dependent_models(base_model)
        
        table_info = []
        try:
            with schema_context(tenant.schema_name):
                for m in dependent_models:
                    count = m.objects.all().count()
                    if count > 0 or m == base_model:
                        table_info.append({
                            'name': str(m._meta.verbose_name).capitalize(),
                            'table': m._meta.db_table,
                            'count': count,
                            'is_target': m == base_model
                        })
            
            # Sort target first, then by count desc
            table_info = sorted(table_info, key=lambda x: (not x['is_target'], -x['count']))
            
            return JsonResponse({'status': 'success', 'data': table_info})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Erreur lors du calcul de la preview: {str(e)}'})
            
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

@login_required(login_url='login')
def delete_tenant(request, tenant_id):
    """
    Securely delete a tenant (Institut) and its associated data (schema, domains).
    This action is irreversible.
    """
    if request.method == 'POST':
        tenant = get_object_or_404(Institut, id=tenant_id)
        
        # Safety check: Prevent deleting public schema
        if tenant.schema_name == 'public':
            return JsonResponse({
                'status': 'error', 
                'message': 'Le schéma public ne peut pas être supprimé.'
            }, status=403)
            
        try:
            name = tenant.nom
            schema = tenant.schema_name
            
            # 1. Delete the Institut instance (this should trigger domains deletion if CASCADE)
            # and potentially the schema if configured, but we'll do it explicitly to be sure.
            tenant.delete()
            
            # 2. Explicitly drop the schema (django-tenants doesn't always do it on instance delete)
            from django.db import connection
            with connection.cursor() as cursor:
                cursor.execute(f'DROP SCHEMA IF EXISTS "{schema}" CASCADE;')
            
            return JsonResponse({
                'status': 'success', 
                'message': f'L\'institut "{name}" et son schéma "{schema}" ont été supprimés avec succès.'
            })
        except Exception as e:
            return JsonResponse({
                'status': 'error', 
                'message': f'Erreur lors de la suppression de l\'institut: {str(e)}'
            })
            
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

@login_required(login_url='login')
def list_orphaned_schemas(request):
    """
    Lists all PostgreSQL schemas that are NOT in the Institut table 
    and are NOT system schemas.
    """
    from django.db import connection
    
    # 1. Get all schemas from DB
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT schema_name 
            FROM information_schema.schemata 
            WHERE schema_name NOT IN ('public', 'information_schema', 'pg_catalog', 'pg_toast')
            AND schema_name NOT LIKE 'pg_temp_%'
            AND schema_name NOT LIKE 'pg_toast_temp_%'
            ORDER BY schema_name;
        """)
        db_schemas = [row[0] for row in cursor.fetchall()]
    
    # 2. Get all schemas registered in Institut model
    registered_schemas = set(Institut.objects.values_list('schema_name', flat=True))
    
    # 3. Identify orphaned schemas
    orphaned_schemas = [s for s in db_schemas if s not in registered_schemas]
    
    return render(request, 'associe_app/schema_management.html', {
        'orphaned_schemas': orphaned_schemas,
        'title': 'Gestion des Schémas Orphelins'
    })

@login_required(login_url='login')
def delete_orphaned_schema(request):
    """
    Permanently delete an orphaned PostgreSQL schema.
    """
    if request.method == 'POST':
        schema = request.POST.get('schema')
        
        # Safety checks
        system_schemas = {'public', 'information_schema', 'pg_catalog', 'pg_toast'}
        if not schema or schema in system_schemas or schema.startswith('pg_'):
            return JsonResponse({'status': 'error', 'message': 'Action non autorisée sur ce schéma.'}, status=403)
            
        # Ensure it's not actually registered
        if Institut.objects.filter(schema_name=schema).exists():
             return JsonResponse({'status': 'error', 'message': 'Ce schéma est actuellement utilisé par un institut actif.'}, status=400)

        try:
            from django.db import connection
            with connection.cursor() as cursor:
                cursor.execute(f'DROP SCHEMA "{schema}" CASCADE;')
                
            return JsonResponse({
                'status': 'success', 
                'message': f'Le schéma "{schema}" a été supprimé avec succès.'
            })
        except Exception as e:
            return JsonResponse({
                'status': 'error', 
                'message': f'Erreur lors de la suppression du schéma : {str(e)}'
            })
            
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

# --- Postes Budgetaires Views ---

from .models import PostesBudgetaire
from .forms import PostesBudgetaireForm

@login_required(login_url='login')
def postes_budgetaires_list(request):
    postes = PostesBudgetaire.objects.all().order_by('order', 'type', 'label')
    postes_list = []
    for poste in postes:
        postes_list.append({
            'id': poste.id,
            'name': poste.label, # Using 'name' for consistency with other trees
            'label': poste.label,
            'type': poste.type,
            'type_display': poste.get_type_display(),
            'parent_id': poste.parent.id if poste.parent else None,
            'description': poste.description,
            'order': poste.order,
            'depense_categories': [c.id for c in poste.depense_categories.all()],
            'depense_categories_labels': [c.name for c in poste.depense_categories.all()],
            'payment_categories': [c.id for c in poste.payment_categories.all()],
            'payment_categories_labels': [c.name for c in poste.payment_categories.all()]
        })
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse(postes_list, safe=False)

    form = PostesBudgetaireForm()
    return render(request, 'associe_app/postes_budgetaires_list.html', {
        'postes': postes, 
        'form': form,
        'title': 'Plan Comptable - Postes Budgétaires'
    })

@login_required(login_url='login')
def postes_budgetaire_create(request):
    if request.method == 'POST':
        form = PostesBudgetaireForm(request.POST)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success', 'message': 'Poste budgétaire créé avec succès.'})
            return redirect('associe_app:postes_budgetaires_list')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors.get_json_data()})
    return redirect('associe_app:postes_budgetaires_list')

@login_required(login_url='login')
def postes_budgetaire_edit(request, pk):
    poste = get_object_or_404(PostesBudgetaire, pk=pk)
    if request.method == 'POST':
        form = PostesBudgetaireForm(request.POST, instance=poste)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success', 'message': 'Poste budgétaire modifié avec succès.'})
            return redirect('associe_app:postes_budgetaires_list')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors.get_json_data()})
    
    # For loading form data via AJAX
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'label': poste.label,
            'type': poste.type,
            'description': poste.description,
            'order': poste.order,
            'parent': poste.parent.id if poste.parent else '',
            'depense_categories': list(poste.depense_categories.values_list('id', flat=True)),
            'payment_categories': list(poste.payment_categories.values_list('id', flat=True))
        })
    return redirect('associe_app:postes_budgetaires_list')

@login_required(login_url='login')
def postes_budgetaire_delete(request, pk):
    poste = get_object_or_404(PostesBudgetaire, pk=pk)
    if request.method == 'POST':
        poste.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': 'Poste budgétaire supprimé avec succès.'})
    return redirect('associe_app:postes_budgetaires_list')

# --- Budget Campaign Views ---

@login_required(login_url='login')
def budget_campaign_list(request):
    campaigns = BudgetCampaign.objects.all().order_by('-date_debut')
    active_count = BudgetCampaign.objects.filter(is_active=True).count()
    return render(request, 'associe_app/budget_campaign_list.html', {
        'campaigns': campaigns,
        'active_count': active_count,
        'title': 'Campagnes Budgétaires'
    })


def perform_bulk_notification(campaign):
    """Helper to notify all tenants about a campaign"""
    tenants = Institut.objects.filter(is_visible=True).exclude(schema_name='public')
    message = f"Une nouvelle campagne budgétaire '{campaign.name}' est maintenant disponible. Veuillez préparer votre proposition."
    link = "/direction/mes-campagnes-budgetaires/" 
    
    success_count = 0
    from django.utils import timezone
    from datetime import date
    
    today = timezone.now().date()
    if today.month >= 7:
        year_n_start = today.year
    else:
        year_n_start = today.year - 1

    n_t1_start = date(year_n_start, 7, 1)
    n_t1_end = date(year_n_start, 9, 30)
    n_t2_start = date(year_n_start, 10, 1)
    n_t2_end = date(year_n_start, 12, 31)
    n_t3_start = date(year_n_start + 1, 1, 1)
    n_t3_end = date(year_n_start + 1, 3, 31)
    n_t4_start = date(year_n_start + 1, 4, 1)
    n_t4_end = date(year_n_start + 1, 6, 30)

    year_n_1_start = year_n_start - 1
    n_1_t1_start = date(year_n_1_start, 7, 1)
    n_1_t1_end = date(year_n_1_start, 9, 30)
    n_1_t2_start = date(year_n_1_start, 10, 1)
    n_1_t2_end = date(year_n_1_start, 12, 31)
    n_1_t3_start = date(year_n_1_start + 1, 1, 1)
    n_1_t3_end = date(year_n_1_start + 1, 3, 31)
    n_1_t4_start = date(year_n_1_start + 1, 4, 1)
    n_1_t4_end = date(year_n_1_start + 1, 6, 30)

    def get_quarter(d):
        if n_t1_start <= d <= n_t1_end: return 'n', 't1'
        if n_t2_start <= d <= n_t2_end: return 'n', 't2'
        if n_t3_start <= d <= n_t3_end: return 'n', 't3'
        if n_t4_start <= d <= n_t4_end: return 'n', 't4'
        if n_1_t1_start <= d <= n_1_t1_end: return 'n_1', 't1'
        if n_1_t2_start <= d <= n_1_t2_end: return 'n_1', 't2'
        if n_1_t3_start <= d <= n_1_t3_end: return 'n_1', 't3'
        if n_1_t4_start <= d <= n_1_t4_end: return 'n_1', 't4'
        return None, None

    kpi_keys = ['prospects', 'opportunites', 'pipeline'] + statuses
    comparison_data = {
        kpi: {
            't1': {'n': 0, 'n_1': 0},
            't2': {'n': 0, 'n_1': 0},
            't3': {'n': 0, 'n_1': 0},
            't4': {'n': 0, 'n_1': 0},
            'global': {'n': 0, 'n_1': 0}
        } for kpi in kpi_keys
    }

    for tenant in tenants:
        try:
            with schema_context(tenant.schema_name):
                send_notification_to_module_level('ger', [1, 2, 3, 4], message, link)
                success_count += 1
        except Exception as e:
            print(f"Error notifying tenant {tenant.nom}: {str(e)}")
    return success_count

@login_required(login_url='login')
def budget_campaign_activate(request, campaign_slug):
    campaign = get_object_or_404(BudgetCampaign, slug=campaign_slug)
    if request.method == 'POST':
        try:
            # Toggle the active status
            campaign.is_active = not campaign.is_active
            campaign.save()
            
            action_message = "activée" if campaign.is_active else "désactivée"
            
            # Auto-notify if activated
            notif_msg = ""
            if campaign.is_active:
                count = perform_bulk_notification(campaign)
                notif_msg = f" (Notifications envoyées à {count} instituts)"
            
            messages.success(request, f"Campagne '{campaign.name}' {action_message} avec succès.{notif_msg}")
            
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success', 'message': f"Campagne '{campaign.name}' {action_message} avec succès.{notif_msg}"})
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la modification du statut de la campagne: {e}")
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': f"Erreur lors de la modification du statut de la campagne: {e}"})
                
    return redirect('associe_app:budget_campaign_list')

@login_required(login_url='login')
def notify_tenants_of_campaign(request, campaign_slug):
    """Manual notification trigger"""
    campaign = get_object_or_404(BudgetCampaign, slug=campaign_slug)
    if request.method == 'POST':
        success_count = perform_bulk_notification(campaign)
        msg = f"Notification envoyée avec succès à {success_count} instituts."
        messages.success(request, msg)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': msg})
    return redirect('associe_app:budget_campaign_list')

@login_required(login_url='login')
def budget_campaign_delete(request, campaign_slug):
    campaign = get_object_or_404(BudgetCampaign, slug=campaign_slug)
    if request.method == 'POST':
        campaign.delete()
        messages.success(request, f"Campagne '{campaign.name}' supprimée avec succès.")
    return redirect('associe_app:budget_campaign_list')

@login_required(login_url='login')
def budget_campaign_instituts(request, campaign_slug):
    campaign = get_object_or_404(BudgetCampaign, slug=campaign_slug)
    instituts = Institut.objects.filter(is_visible=True).exclude(schema_name='public')
    
    # 1. Process Form Submission (Bulk Save)
    if request.method == 'POST':
        try:
            for key, value in request.POST.items():
                if key.startswith('budget_'):
                    # Format: budget_INSTITUTID
                    parts = key.split('_')
                    if len(parts) == 2:
                        inst_id = parts[1]
                        # Robust cleaning for accounting format
                        clean_value = value.replace(' ', '').replace('\xa0', '').replace('\u202f', '')
                        if ',' in clean_value and '.' in clean_value:
                            if clean_value.rfind(',') > clean_value.rfind('.'):
                                clean_value = clean_value.replace('.', '').replace(',', '.')
                            else:
                                clean_value = clean_value.replace(',', '')
                        else:
                            clean_value = clean_value.replace(',', '.')
                        
                        try:
                            amount = float(clean_value) if clean_value else 0
                            if amount >= 0:
                                BudgetLine.objects.update_or_create(
                                    campaign=campaign,
                                    institut_id=inst_id,
                                    defaults={'montant': amount}
                                )
                        except (ValueError, TypeError):
                            continue
            messages.success(request, "Objectifs globaux enregistrés avec succès.")
        except Exception as e:
             messages.error(request, f"Erreur lors de l'enregistrement: {e}")
        return redirect('associe_app:budget_campaign_instituts', campaign_slug=campaign.slug)

    # 2. Build Data for Display
    instituts_data = []
    configured_instituts = 0
    
    # Prefetch extensions to avoid N+1
    from .models import BudgetExtensionRequest
    pending_extensions = BudgetExtensionRequest.objects.filter(
        campaign=campaign, 
        status='pending'
    ).values_list('institut_id', 'id')
    
    # Create a map: institut_id -> request_id
    pending_extensions_map = {inst_id: req_id for inst_id, req_id in pending_extensions}

    for inst in instituts:
        line = BudgetLine.objects.filter(campaign=campaign, institut=inst).first()
        is_configured = line is not None
        if is_configured:
            configured_instituts += 1
            
        instituts_data.append({
            'institut': inst,
            'is_configured': is_configured,
            'montant': line.montant if line else 0,
            'statut': line.statut if line else 'none',
            'pending_extension_id': pending_extensions_map.get(inst.id)
        })
        
    total_inst = len(instituts)
    progress = (configured_instituts / total_inst * 100) if total_inst > 0 else 0
    
    return render(request, 'associe_app/budget_campaign_instituts.html', {
        'campaign': campaign,
        'instituts_data': instituts_data,
        'configured_count': configured_instituts,
        'pending_count': total_inst - configured_instituts,
        'progress': progress,
        'title': f'Objectifs par Institut - {campaign.name}'
    })

@login_required(login_url='login')
def budget_campaign_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        
        # Automatic Date Logic: Fiscal Year starts July 1st
        today = date.today()
        if today.month < 7: # Jan - June
            start_year = today.year - 1
        else: # July - Dec
            start_year = today.year
            
        date_debut = date(start_year, 7, 1)
        date_fin = date(start_year + 1, 6, 30)
        
        default_name = f"Budget {start_year}-{start_year + 1}"
        
        if not name:
            name = default_name
            
        # Create or Get (to avoid duplicates for same period/name)
        campaign, created = BudgetCampaign.objects.get_or_create(
            name=name,
            defaults={
                'date_debut': date_debut, 
                'date_fin': date_fin, 
                'is_active': False
            }
        )
        
        if created:
            messages.success(request, f"Campagne '{name}' créée avec succès. Vous pouvez la configurer avant de la lancer.")
        else:
            messages.info(request, f"La campagne '{name}' existe déjà.")
            
    return redirect('associe_app:budget_campaign_list')

@login_required(login_url='login')
def budget_campaign_review(request, campaign_slug, institut_id):
    """
    Review and Validate/Reject a tenant's budget dispatch proposal.
    """
    campaign = get_object_or_404(BudgetCampaign, slug=campaign_slug)
    institut = get_object_or_404(Institut, id=institut_id)
    budget_line = get_object_or_404(BudgetLine, campaign=campaign, institut=institut)
    
    # Restriction: Admin cannot see details until submitted
    is_visible_to_admin = budget_line.statut in ['submitted', 'validated', 'rejected']
    
    details = BudgetLineDetail.objects.none()
    structured_postes = []
    total_dispatched = 0
    total_recette_propose = 0
    total_depense_propose = 0
    entreprises = []
    row_quarters = {}
    allocations = {}

    if is_visible_to_admin:
        # 1. Global Items (Public)
        all_postes = PostesBudgetaire.objects.prefetch_related('payment_categories', 'depense_categories').order_by('-type', 'order', 'label')
        
        structured_postes = []
        for p in all_postes:
            if p.parent is None:
                children = [child for child in all_postes if child.parent_id == p.id]
                display_postes = []
                # If parent has direct categories or no children, add it to display
                if p.payment_categories.exists() or p.depense_categories.exists() or not children:
                    display_postes.append(p)
                display_postes.extend(children)
                
                structured_postes.append({
                    'parent_poste': p,
                    'is_standalone': not children,
                    'display_postes': display_postes
                })

        details = BudgetLineDetail.objects.filter(campaign=campaign, institut=institut)
        # Map: {poste_id}_none_0_0 -> BudgetLineDetail (Aggregated)
        allocations = {}
        for d in details:
            # Aggregate everything at the poste level to match dispatch view
            key = f"{d.poste_id}_none_0_0"
            
            if key not in allocations:
                # We need a copy because we might modify montant
                import copy
                allocations[key] = copy.copy(d)
                # Ensure the copy reflects 'none' category for display logic match
                allocations[key].payment_category_id = None
                allocations[key].depense_category_id = None
            else:
                allocations[key].montant += d.montant

            # Add category specific totals
            if d.payment_category_id:
                cat_key = f"{d.poste_id}_pay_{d.payment_category_id}_0"
                if cat_key not in allocations:
                    import copy
                    allocations[cat_key] = copy.copy(d)
                else:
                    allocations[cat_key].montant += d.montant
            elif d.depense_category_id:
                cat_key = f"{d.poste_id}_dep_{d.depense_category_id}_0"
                if cat_key not in allocations:
                    import copy
                    allocations[cat_key] = copy.copy(d)
                else:
                    allocations[cat_key].montant += d.montant
        
        total_dispatched = details.aggregate(Sum('montant'))['montant__sum'] or 0
        total_recette_propose = details.filter(poste__type='recette').aggregate(Sum('montant'))['montant__sum'] or 0
        total_depense_propose = details.filter(poste__type='depense').aggregate(Sum('montant'))['montant__sum'] or 0

        # 2. Tenant Specific Entities
        with schema_context(institut.schema_name):
            entreprises = list(Entreprise.objects.all().order_by('designation'))

    if request.method == "POST":
        action = request.POST.get('action')
        commentaire = request.POST.get('commentaire', '')
        
        if action == 'save_quartely':
            # Save quarterly percentages (Row Level)
            # Input format: tX_POSTE_CAT
            updated_details = []
            
            # Create a map of existing details for quick lookup: (poste_id, pay_cat_id, dep_cat_id) -> list of details
            details_map = {}
            for d in details:
                key = (d.poste_id, d.payment_category_id, d.depense_category_id)
                if key not in details_map:
                    details_map[key] = []
                details_map[key].append(d)

            for key, value in request.POST.items():
                if key.startswith('t1_') or key.startswith('t2_') or key.startswith('t3_') or key.startswith('t4_'):
                    # Format: tX_POSTE_CAT
                    parts = key.split('_') 
                    # Expect ['t1', '12', '5'] (cat exists) or ['t1', '12', 'None']
                    if len(parts) >= 3:
                        quarter = parts[0] # t1, t2, t3, t4
                        poste_id = parts[1]
                        cat_id_str = parts[2]
                        
                        try:
                            val = float(value.replace(',', '.')) if value else 0.0
                            val = min(max(val, 0), 100) # Clamp 0-100
                            
                            real_cat_id = int(cat_id_str) if cat_id_str.lower() != 'none' and cat_id_str != '' else None
                            
                            # Update ALL details for this row (Poste + Category)
                            # In this simplified view, we check both payment and depense categories
                            row_details = []
                            if real_cat_id:
                                # Try both payment and depense since we don't know the type from the key alone here
                                row_details.extend(details_map.get((int(poste_id), real_cat_id, None), []))
                                row_details.extend(details_map.get((int(poste_id), None, real_cat_id), []))
                            else:
                                # Get ALL details for this poste, regardless of category
                                # because 'none' means all categories for this poste in the single-row UI
                                for d_key, d_list in details_map.items():
                                    if d_key[0] == int(poste_id):
                                        row_details.extend(d_list)
                            
                            for detail in row_details:
                                if quarter == 't1': detail.t1_percent = val
                                elif quarter == 't2': detail.t2_percent = val
                                elif quarter == 't3': detail.t3_percent = val
                                elif quarter == 't4': detail.t4_percent = val
                                updated_details.append(detail)
                                
                        except ValueError:
                            continue
            
            # Bulk update if possible, or save individually
            # Django bulk_update is efficient
            if updated_details:
                BudgetLineDetail.objects.bulk_update(updated_details, ['t1_percent', 't2_percent', 't3_percent', 't4_percent'])
                
            messages.success(request, "Les pourcentages trimestriels ont été enregistrés.")
            return redirect('associe_app:budget_campaign_review', campaign_slug=campaign.slug, institut_id=institut.id)

        elif action == 'validate':
            budget_line.statut = 'validated'
            budget_line.commentaire = '' # Clear old comments
            messages.success(request, f"Le budget pour {institut.nom} a été validé.")
        elif action == 'reject':
            budget_line.statut = 'rejected'
            budget_line.commentaire = commentaire
        elif action == 'reset':
             # Reset Logic
            BudgetLineDetail.objects.filter(campaign=campaign, institut=institut).delete()
            budget_line.statut = 'draft' # Or whatever starting status is appropriate
            budget_line.commentaire = ''
            messages.info(request, f"La proposition de budget pour {institut.nom} a été réinitialisée.")
        
        budget_line.save()

        # Notify the institute's managers about the decision
        try:
            if action in ['validate', 'reject']:
                campaign_name = campaign.name
                if action == 'validate':
                    notif_message = f"Votre budget pour la campagne '{campaign_name}' a été validé par l'administration."
                else:
                    # Truncate comment if too long for notification
                    clean_comment = commentaire[:100] + "..." if len(commentaire) > 100 else commentaire
                    notif_message = f"Votre budget pour la campagne '{campaign_name}' a été rejeté. Motif : {clean_comment}"
                
                # Link to the dispatch page in the tenant app
                notif_link = f"/direction/dispatch-budget/{campaign.slug}/"
                
                with schema_context(institut.schema_name):
                    # Notify all users with access to 'ger' module
                    send_notification_to_module_level(
                        module_name='ger',
                        role_levels=[1, 2, 3, 4], # All levels having access to the module
                        message=notif_message,
                        link=notif_link
                    )
        except Exception as ne:
            print(f"Error sending decision notification to institute: {ne}")

        return redirect('associe_app:budget_campaign_instituts', campaign_slug=campaign.slug)

    # Prepare row_quarters for template (Collective for all categories of a poste)
    if is_visible_to_admin:
        for d in details:
            # Key: "poste_id_None" (to match template with row_key=pid|add:"_None")
            row_key = f"{d.poste_id}_None"
            
            # We assume all details for the same poste have the same percentages
            if row_key not in row_quarters:
                row_quarters[row_key] = {
                    't1': d.t1_percent,
                    't2': d.t2_percent,
                    't3': d.t3_percent,
                    't4': d.t4_percent
                }

    # 3. Realization Data for this specific institute
    realization_data = get_campaign_realization_data(campaign, [institut])
    
    # 4. Previous Campaign Data (N-1)
    previous_campaign = BudgetCampaign.objects.filter(date_fin__lte=campaign.date_debut).order_by('-date_fin').first()
    prev_realization_data = None
    prev_data_dict = {}
    if previous_campaign:
        prev_realization_data = get_campaign_realization_data(previous_campaign, [institut])
        for group in prev_realization_data['combined_postes']:
            for item in group['display_postes']:
                prev_data_dict[str(item['poste'].id)] = item

    context = {
        'campaign': campaign,
        'institut': institut,
        'budget_line': budget_line,
        'structured_postes': structured_postes,
        'total_dispatched': total_dispatched,
        'total_recette_propose': total_recette_propose,
        'total_depense_propose': total_depense_propose,
        'remaining': total_recette_propose - total_depense_propose,
        'entreprises': entreprises,
        'row_quarters': row_quarters,
        'allocations': allocations,
        'is_visible_to_admin': is_visible_to_admin,
        'realization_data': realization_data,
        'combined_postes': realization_data.get('combined_postes', []) if realization_data else [],
        'realization_totals': realization_data.get('totals', {}) if realization_data else {},
        'previous_campaign': previous_campaign,
        'prev_data_dict': prev_data_dict,
        'title': f'Révision Budget - {institut.nom}',
    }
    return render(request, 'associe_app/budget_campaign_review.html', context)

@login_required(login_url="login")
def extension_requests_list(request):
    from .models import BudgetExtensionRequest
    
    requests = BudgetExtensionRequest.objects.filter(institut__is_visible=True).select_related('campaign', 'institut').order_by('-created_at')
    
    context = {
        'tenants': tenants, # For the selector
        'institutes_stats': institutes_stats,
        'global_stats': global_stats,
        'statuses': statuses,
        'comparison_data': comparison_data,
        'year_n': f"{year_n_start}-{year_n_start+1}",
        'year_n_1': f"{year_n_1_start}-{year_n_1_start+1}",
        'title': 'Statistiques CRM Globales'
    }
    return render(request, 'associe_app/extension_requests_list.html', context)

@login_required(login_url="login")
def review_extension(request, request_id):
    from .models import BudgetExtensionRequest, BudgetExtensionItem, BudgetLine, BudgetLineDetail
    from django.db import transaction
    from institut_app.models import Entreprise # Import purely for type hinting if needed, but actual query is dynamic
    # We need to import Entreprise dynamically from the tenant model or similar if possible, 
    # or just use the model known in the context.
    # Actually Entreprise is defined in `institut_app.models` usually, or `tenant_app`. 
    # Let's check where Entreprise is defined. 
    # In `institut_app/views.py` it was used. Let's assume it's available or we use raw SQL/ORM within context.
    # We imported `Entreprise` in `institut_app/views.py` from where? 
    # Likely `from app.models import Entreprise` if valid, or `from institut_app.models import Entreprise`.
    # Based on checking `institut_app/views.py` imports earlier, it wasn't shown explicitly but used.
    # Let's check `app/models.py` or similar later if it fails. 
    # For now, I will use `from app.models import Entreprise` inside the context manager block if needed, 
    # or rely on `details.entreprise_id`.
    
    # Correction: `Entreprise` is likely in `institut_app` or `app` but specific to tenant.
    # I'll use `from app.models import Entreprise` assuming it's the tenant model.
    from institut_app.models import Entreprise 

    ext_request = get_object_or_404(BudgetExtensionRequest, id=request_id)
    items = ext_request.items.select_related('poste', 'payment_category').order_by('poste__label')
    
    if request.method == "POST":
        action = request.POST.get('action')
        comment = request.POST.get('admin_comment', '')
        
        if action == 'approve':
            try:
                with transaction.atomic():
                    # Update BudgetLineDetails
                    for item in items:
                        BudgetLineDetail.objects.update_or_create(
                            campaign=ext_request.campaign,
                            institut=ext_request.institut,
                            poste=item.poste,
                            payment_category=item.payment_category,
                            entreprise_id=item.entreprise_id,
                            defaults={'montant': item.old_amount + item.requested_amount}
                        )
                    
                    
                    ext_request.status = 'approved'
                    ext_request.admin_comment = comment
                    ext_request.save()
                    
                    messages.success(request, f"La demande de rallonge pour {ext_request.institut} a été approuvée.")
                    
            except Exception as e:
                messages.error(request, f"Erreur lors de l'approbation : {str(e)}")
        
        elif action == 'reject':
            if not comment.strip():
                messages.error(request, "Veuillez indiquer le motif du rejet.")
                return redirect('associe_app:review_extension', request_id=request_id)
                
            ext_request.status = 'rejected'
            ext_request.admin_comment = comment
            ext_request.save()
            messages.warning(request, f"La demande de rallonge pour {ext_request.institut} a été rejetée.")
            
        # Send Notification to the Institute (Users with 'ger' module access)
        if action in ['approve', 'reject']:
            try:
                from django.urls import reverse
                with schema_context(ext_request.institut.schema_name):
                    from institut_app.models import UserModuleRole
                    from institut_app.utils_notifications import send_notification_to_user
                    from django.contrib.auth import get_user_model
                    
                    status_fr = "approuvée" if action == 'approve' else "rejetée"
                    message = f"Votre demande de rallonge pour la campagne {ext_request.campaign} a été {status_fr}."
                    if comment:
                        message += f" Commentaire : {comment}"
                    
                    # Link to the extension request page in the tenant dashboard
                    link = reverse('institut_app:request_extension', 
                                   kwargs={'campaign_slug': ext_request.campaign.slug}, 
                                   urlconf='app.urls')
                    
                    # Target users who have access to the 'ger' module
                    users_to_notify = UserModuleRole.objects.filter(module__name='ger').values_list('user_id', flat=True).distinct()
                    TenantUser = get_user_model()
                    
                    print(f"DEBUG: Notifying {len(users_to_notify)} users in tenant {ext_request.institut.schema_name}")
                    for user_id in users_to_notify:
                        try:
                            user = TenantUser.objects.get(id=user_id)
                            send_notification_to_user(user, message, link)
                        except Exception as u_err:
                            print(f"DEBUG: Failed to notify user {user_id}: {u_err}")
                            
            except Exception as e:
                print(f"DEBUG: Error in extension review notification: {e}")

        return redirect('associe_app:extension_requests_list')

    # Fetch enterprise names mapping
    entreprise_map = {0: 'Global'}
    try:
        with schema_context(ext_request.institut.schema_name):
            ents = Entreprise.objects.all()
            for e in ents:
                entreprise_map[e.id] = e.designation
    except Exception as e:
        print(f"Error fetching entreprises: {e}")

    context = {
        'tenants': tenants, # For the selector
        'institutes_stats': institutes_stats,
        'global_stats': global_stats,
        'statuses': statuses,
        'comparison_data': comparison_data,
        'year_n': f"{year_n_start}-{year_n_start+1}",
        'year_n_1': f"{year_n_1_start}-{year_n_1_start+1}",
        'title': 'Statistiques CRM Globales'
    }
    context['title'] = f'Révision de Rallonge - {ext_request.institut.nom}'
    return render(request, 'associe_app/review_extension.html', context)



@login_required(login_url='login')
def export_postes_budgetaires(request):
    postes = PostesBudgetaire.objects.all().order_by('order', 'type', 'label')
    wb = Workbook()
    ws = wb.active
    ws.title = "Postes Budgetaires"
    
    headers = [
        'ID', 'Label', 'Type', 'Order', 'Description', 'Parent_ID', 'Parent_Nom',
        'Depense_Categories_IDs', 'Payment_Categories_IDs'
    ]
    ws.append(headers)
    
    for p in postes:
        depense_cats = ",".join([str(c.id) for c in p.depense_categories.all()])
        payment_cats = ",".join([str(c.id) for c in p.payment_categories.all()])
        
        ws.append([
            p.id,
            p.label,
            p.type,
            p.order,
            p.description,
            p.parent.id if p.parent else '',
            p.parent.label if p.parent else '',
            depense_cats,
            payment_cats
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = FileResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="postes_budgetaires.xlsx"'
    return response

@login_required(login_url='login')
def import_postes_budgetaires(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        filename = file.name.lower()
        data = []
        
        try:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row): continue
                    data.append(dict(zip(headers, row)))
            elif filename.endswith('.json'):
                data = json.load(file)
            else:
                return JsonResponse({'status': 'error', 'message': 'Format non supporté.'})
            
            # Pass 1: Create or Update items (without parent/categories for now)
            item_map = {}
            for row in data:
                item_id = row.get('ID') or row.get('id')
                label = row.get('Label') or row.get('label')
                p_type = row.get('Type') or row.get('type')
                order = row.get('Order') or row.get('order') or 0
                desc = row.get('Description') or row.get('description') or ''
                
                poste, created = PostesBudgetaire.objects.update_or_create(
                    id=item_id if item_id else None,
                    defaults={
                        'label': label,
                        'type': p_type,
                        'order': int(order),
                        'description': desc
                    }
                )
                item_map[item_id or poste.id] = (poste, row)

            # Pass 2: Link Parents and Categories
            for item_id, (poste, row) in item_map.items():
                parent_id = row.get('Parent_ID') or row.get('parent_id')
                depense_cats_str = str(row.get('Depense_Categories_IDs') or row.get('depense_categories_ids') or '')
                payment_cats_str = str(row.get('Payment_Categories_IDs') or row.get('payment_categories_ids') or '')

                if parent_id:
                    poste.parent = PostesBudgetaire.objects.filter(id=parent_id).first()
                else:
                    poste.parent = None
                
                poste.save()

                if depense_cats_str:
                    ids = [int(i.strip()) for i in depense_cats_str.split(',') if i.strip().isdigit()]
                    poste.depense_categories.set(GlobalDepensesCategory.objects.filter(id__in=ids))
                else:
                    poste.depense_categories.clear()

                if payment_cats_str:
                    ids = [int(i.strip()) for i in payment_cats_str.split(',') if i.strip().isdigit()]
                    poste.payment_categories.set(GlobalPaymentCategory.objects.filter(id__in=ids))
                else:
                    poste.payment_categories.clear()

            return JsonResponse({'status': 'success', 'message': f'{len(data)} postes budgétaires traités avec succès.'})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f"Erreur lors de l'import : {str(e)}"})
            
    return JsonResponse({'status': 'error', 'message': 'Requête invalide.'})

# --- User Management Views ---

@login_required(login_url='login')
def user_list(request):
    users = User.objects.all().order_by('-date_joined')
    return render(request, 'associe_app/users/user_list.html', {
        'users': users,
        'title': 'Gestion des Utilisateurs'
    })

@login_required(login_url='login')
def user_create(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        is_staff = request.POST.get('is_staff') == 'on'
        is_active = request.POST.get('is_active') == 'on'
        is_superuser = request.POST.get('is_superuser') == 'on'
        
        if User.objects.filter(username=username).exists():
            messages.error(request, "Ce nom d'utilisateur existe déjà.")
        else:
            User.objects.create(
                username=username,
                email=email,
                password=make_password(password),
                first_name=first_name,
                last_name=last_name,
                is_staff=is_staff,
                is_active=is_active,
                is_superuser=is_superuser
            )
            messages.success(request, "Utilisateur créé avec succès.")
            return redirect('associe_app:user_list')
            
    return render(request, 'associe_app/users/user_form.html', {'title': 'Nouvel Utilisateur'})

@login_required(login_url='login')
def user_edit(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user.username = request.POST.get('username')
        user.email = request.POST.get('email')
        user.first_name = request.POST.get('first_name')
        user.last_name = request.POST.get('last_name')
        user.is_staff = request.POST.get('is_staff') == 'on'
        user.is_active = request.POST.get('is_active') == 'on'
        user.is_superuser = request.POST.get('is_superuser') == 'on'
        
        password = request.POST.get('password')
        if password:
            user.password = make_password(password)
            
        user.save()
        messages.success(request, "Utilisateur mis à jour avec succès.")
        return redirect('associe_app:user_list')
        
    return render(request, 'associe_app/users/user_form.html', {
        'edit_user': user,
        'title': "Modifier l'Utilisateur"
    })

@login_required(login_url='login')
def user_delete(request, pk):
    user = get_object_or_404(User, pk=pk)
    if user.is_superuser:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': "Impossible de supprimer un super-utilisateur."})
        messages.error(request, "Impossible de supprimer un super-utilisateur.")
    else:
        user.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': "Utilisateur supprimé."})
        messages.success(request, "Utilisateur supprimé.")
    return redirect('associe_app:user_list')

@login_required(login_url='login')
def user_toggle_status(request, pk):
    user = get_object_or_404(User, pk=pk)
    if user.is_superuser and request.user != user:
         messages.error(request, "Vous ne pouvez pas désactiver un autre super-utilisateur.")
    else:
        user.is_active = not user.is_active
        user.save()
        status = "activé" if user.is_active else "désactivé"
        messages.success(request, f"Utilisateur {status}.")
        
    return redirect('associe_app:user_list')

@login_required(login_url='login')
def get_tenant_sync_list(request):
    tenants = Institut.objects.exclude(schema_name='public').values('nom', 'schema_name')
    return JsonResponse({'status': 'success', 'tenants': list(tenants)})

@login_required(login_url='login')
def sync_single_tenant_view(request):
    if request.method == 'POST':
        schema_name = request.POST.get('schema_name')
        if schema_name:
            from .utils import sync_tenant_categories
            try:
                sync_tenant_categories(schema_name)
                return JsonResponse({'status': 'success', 'message': f'Tenant {schema_name} synchronisé.'})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Requête invalide.'})

@login_required
def api_mark_saas_notification_read(request):
    if request.method == 'POST':
        notif_id = request.POST.get('id')
        try:
            notification = SaaSNotification.objects.get(id=notif_id, user=request.user)
            notification.is_read = True
            notification.save()
            return JsonResponse({'status': 'success'})
        except SaaSNotification.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Notification not found'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required
def api_mark_all_saas_notifications_read(request):
    if request.method == 'POST':
        SaaSNotification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required
def api_delete_saas_notification(request):
    if request.method == 'POST':
        notif_id = request.POST.get('id')
        try:
            notification = SaaSNotification.objects.get(id=notif_id, user=request.user)
            notification.delete()
            return JsonResponse({'status': 'success'})
        except SaaSNotification.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Notification not found'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required(login_url='login')
def satisfaction_pending(request):
    return render(request, 'associe_app/satisfaction_pending.html', {'title': 'Satisfaction'})

@login_required(login_url='login')
def crm_user_logs(request):
    from app.models import Institut
    from t_crm.models import UserActionLog
    from django_tenants.utils import schema_context
    from django.contrib.auth.models import User
    
    filter_tenant = request.GET.get('tenant')
    filter_user = request.GET.get('user')
    filter_action = request.GET.get('action')

    if filter_tenant:
        tenants = Institut.objects.filter(id=filter_tenant, is_visible=True).exclude(schema_name='public')
    else:
        tenants = Institut.objects.filter(is_visible=True).exclude(schema_name='public').order_by('nom')
        
    all_tenants = Institut.objects.filter(is_visible=True).exclude(schema_name='public').order_by('nom')
    all_users = User.objects.all().order_by('username')
    action_choices = UserActionLog.ACTION_CHOICES
    
    institutes_logs = []

    from django.utils import timezone
    from datetime import date
    
    today = timezone.now().date()
    if today.month >= 7:
        year_n_start = today.year
    else:
        year_n_start = today.year - 1

    n_t1_start = date(year_n_start, 7, 1)
    n_t1_end = date(year_n_start, 9, 30)
    n_t2_start = date(year_n_start, 10, 1)
    n_t2_end = date(year_n_start, 12, 31)
    n_t3_start = date(year_n_start + 1, 1, 1)
    n_t3_end = date(year_n_start + 1, 3, 31)
    n_t4_start = date(year_n_start + 1, 4, 1)
    n_t4_end = date(year_n_start + 1, 6, 30)

    year_n_1_start = year_n_start - 1
    n_1_t1_start = date(year_n_1_start, 7, 1)
    n_1_t1_end = date(year_n_1_start, 9, 30)
    n_1_t2_start = date(year_n_1_start, 10, 1)
    n_1_t2_end = date(year_n_1_start, 12, 31)
    n_1_t3_start = date(year_n_1_start + 1, 1, 1)
    n_1_t3_end = date(year_n_1_start + 1, 3, 31)
    n_1_t4_start = date(year_n_1_start + 1, 4, 1)
    n_1_t4_end = date(year_n_1_start + 1, 6, 30)

    def get_quarter(d):
        if n_t1_start <= d <= n_t1_end: return 'n', 't1'
        if n_t2_start <= d <= n_t2_end: return 'n', 't2'
        if n_t3_start <= d <= n_t3_end: return 'n', 't3'
        if n_t4_start <= d <= n_t4_end: return 'n', 't4'
        if n_1_t1_start <= d <= n_1_t1_end: return 'n_1', 't1'
        if n_1_t2_start <= d <= n_1_t2_end: return 'n_1', 't2'
        if n_1_t3_start <= d <= n_1_t3_end: return 'n_1', 't3'
        if n_1_t4_start <= d <= n_1_t4_end: return 'n_1', 't4'
        return None, None

    kpi_keys = ['prospects', 'opportunites', 'pipeline'] + statuses
    comparison_data = {
        kpi: {
            't1': {'n': 0, 'n_1': 0},
            't2': {'n': 0, 'n_1': 0},
            't3': {'n': 0, 'n_1': 0},
            't4': {'n': 0, 'n_1': 0},
            'global': {'n': 0, 'n_1': 0}
        } for kpi in kpi_keys
    }

    for tenant in tenants:
        try:
            with schema_context(tenant.schema_name):
                logs = UserActionLog.objects.all().select_related('user').order_by('-created_at')
                
                if filter_user:
                    logs = logs.filter(user_id=filter_user)
                if filter_action:
                    logs = logs.filter(action_type=filter_action)
                    
                logs_data = []
                for log in logs:
                    logs_data.append({
                        'user': log.user.username if log.user else 'Système',
                        'action': log.get_action_type_display(),
                        'target_model': log.target_model,
                        'details': log.details,
                        'date': log.created_at,
                    })
                
                if logs_data or not (filter_user or filter_action):
                    institutes_logs.append({
                        'id': tenant.id,
                        'name': tenant.nom,
                        'logs': logs_data
                    })
        except Exception as e:
            continue

    context = {
        'tenants': tenants, # For the selector
        'institutes_stats': institutes_stats,
        'global_stats': global_stats,
        'statuses': statuses,
        'comparison_data': comparison_data,
        'year_n': f"{year_n_start}-{year_n_start+1}",
        'year_n_1': f"{year_n_1_start}-{year_n_1_start+1}",
        'title': 'Statistiques CRM Globales'
    }

    if request.GET.get('export_csv') == '1':
        import csv
        import io
        from django.http import HttpResponse
        
        output = io.StringIO()
        writer = csv.writer(output, delimiter=',', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(['Institut', 'Utilisateur', 'Action', 'Cible', 'Détails', 'Date'])
        
        for tenant_log in institutes_logs:
            tenant_name = tenant_log['name']
            for log in tenant_log['logs']:
                writer.writerow([
                    tenant_name,
                    log['user'],
                    log['action'],
                    log['target_model'],
                    str(log['details']).replace('\n', ' ').replace('\r', '') if log['details'] else '',
                    log['date'].strftime('%Y-%m-%d %H:%M:%S') if log['date'] else ''
                ])
                
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="crm_user_logs.csv"'
        response.write('\ufeff'.encode('utf8'))  # BOM pour Excel
        response.write(output.getvalue())
        return response

    if request.GET.get('export_excel') == '1':
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CRM Logs"
        
        headers = ['Institut', 'Utilisateur', 'Action', 'Cible', 'Détails', 'Date']
        ws.append(headers)
        
        for col_num, cell in enumerate(ws[1], 1):
            cell.font = openpyxl.styles.Font(bold=True)
            
        for tenant_log in institutes_logs:
            tenant_name = tenant_log['name']
            for log in tenant_log['logs']:
                ws.append([
                    tenant_name,
                    log['user'],
                    log['action'],
                    log['target_model'],
                    log['details'],
                    log['date'].strftime('%Y-%m-%d %H:%M:%S') if log['date'] else ''
                ])
                
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="crm_user_logs.xlsx"'
        wb.save(response)
        return response

    return render(request, 'associe_app/crm_user_logs.html', context)


@login_required(login_url='login')
def platform_usage_rate(request):
    from app.models import Institut
    from t_crm.models import UserActionLog
    from django_tenants.utils import schema_context
    from django.utils import timezone
    from django.db.models import Count, Q, Max

    tenants = Institut.objects.filter(is_visible=True).exclude(schema_name='public').order_by('nom')
    all_tenants = tenants # For the modal selection

    is_print = request.GET.get('print') == 'true'
    if is_print:
        tenant_ids = request.GET.getlist('tenants')
        if tenant_ids:
            tenants = tenants.filter(id__in=tenant_ids)

    usage_data = []

    now = timezone.now()

    from django.utils import timezone
    from datetime import date
    
    today = timezone.now().date()
    if today.month >= 7:
        year_n_start = today.year
    else:
        year_n_start = today.year - 1

    n_t1_start = date(year_n_start, 7, 1)
    n_t1_end = date(year_n_start, 9, 30)
    n_t2_start = date(year_n_start, 10, 1)
    n_t2_end = date(year_n_start, 12, 31)
    n_t3_start = date(year_n_start + 1, 1, 1)
    n_t3_end = date(year_n_start + 1, 3, 31)
    n_t4_start = date(year_n_start + 1, 4, 1)
    n_t4_end = date(year_n_start + 1, 6, 30)

    year_n_1_start = year_n_start - 1
    n_1_t1_start = date(year_n_1_start, 7, 1)
    n_1_t1_end = date(year_n_1_start, 9, 30)
    n_1_t2_start = date(year_n_1_start, 10, 1)
    n_1_t2_end = date(year_n_1_start, 12, 31)
    n_1_t3_start = date(year_n_1_start + 1, 1, 1)
    n_1_t3_end = date(year_n_1_start + 1, 3, 31)
    n_1_t4_start = date(year_n_1_start + 1, 4, 1)
    n_1_t4_end = date(year_n_1_start + 1, 6, 30)

    def get_quarter(d):
        if n_t1_start <= d <= n_t1_end: return 'n', 't1'
        if n_t2_start <= d <= n_t2_end: return 'n', 't2'
        if n_t3_start <= d <= n_t3_end: return 'n', 't3'
        if n_t4_start <= d <= n_t4_end: return 'n', 't4'
        if n_1_t1_start <= d <= n_1_t1_end: return 'n_1', 't1'
        if n_1_t2_start <= d <= n_1_t2_end: return 'n_1', 't2'
        if n_1_t3_start <= d <= n_1_t3_end: return 'n_1', 't3'
        if n_1_t4_start <= d <= n_1_t4_end: return 'n_1', 't4'
        return None, None

    kpi_keys = ['prospects', 'opportunites', 'pipeline'] + statuses
    comparison_data = {
        kpi: {
            't1': {'n': 0, 'n_1': 0},
            't2': {'n': 0, 'n_1': 0},
            't3': {'n': 0, 'n_1': 0},
            't4': {'n': 0, 'n_1': 0},
            'global': {'n': 0, 'n_1': 0}
        } for kpi in kpi_keys
    }

    for tenant in tenants:
        try:
            with schema_context(tenant.schema_name):
                creation_date = tenant.date_creation
                if not creation_date:
                    continue
                
                delta = now - creation_date
                days_active = max(delta.days, 1)

                user_actions = UserActionLog.objects.values(
                    'user__username', 'user__first_name', 'user__last_name'
                ).annotate(
                    total_actions=Count('id'),
                    login_count=Count('id', filter=Q(action_type='LOGIN')),
                    last_login=Max('created_at', filter=Q(action_type='LOGIN'))
                ).order_by('-total_actions')

                login_logs = UserActionLog.objects.filter(action_type='LOGIN').values('user__username', 'created_at').order_by('-created_at')
                user_login_dates = {}
                for log in login_logs:
                    uname = log['user__username'] if log['user__username'] else 'Système'
                    if uname not in user_login_dates:
                        user_login_dates[uname] = []
                    user_login_dates[uname].append(log['created_at'])

                users_data = []
                for ua in user_actions:
                    username = ua['user__username'] if ua['user__username'] else 'Système'
                    total = ua['total_actions']
                    rate = round(total / days_active, 2)
                    
                    if rate >= 10:
                        status = 'Excellent'
                        color = 'success'
                    elif rate >= 2:
                        status = 'Actif'
                        color = 'primary'
                    elif rate > 0:
                        status = 'Faible'
                        color = 'warning'
                    else:
                        status = 'Inactif'
                        color = 'danger'

                    users_data.append({
                        'username': username,
                        'first_name': ua['user__first_name'],
                        'last_name': ua['user__last_name'],
                        'total_actions': total,
                        'login_count': ua['login_count'],
                        'last_login': ua['last_login'],
                        'login_dates': user_login_dates.get(username, []),
                        'rate': rate,
                        'status': status,
                        'color': color
                    })

                usage_data.append({
                    'id': tenant.id,
                    'name': tenant.nom,
                    'days_active': days_active,
                    'users': users_data,
                    'creation_date': creation_date
                })
        except Exception as e:
            continue

    context = {
        'tenants': tenants, # For the selector
        'institutes_stats': institutes_stats,
        'global_stats': global_stats,
        'statuses': statuses,
        'comparison_data': comparison_data,
        'year_n': f"{year_n_start}-{year_n_start+1}",
        'year_n_1': f"{year_n_1_start}-{year_n_1_start+1}",
        'title': 'Statistiques CRM Globales'
    }
    
    if is_print:
        return render(request, 'associe_app/platform_usage_rate_print.html', context)
    return render(request, 'associe_app/platform_usage_rate.html', context)
